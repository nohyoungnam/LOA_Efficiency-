from __future__ import annotations

# ==============================================================================
# v124 early CPU guard: RapidOCR/ONNX/OpenCV 스레드 제한
# - 반드시 pandas/streamlit/내부 modules import 전에 실행되어야 합니다.
# - 웹 배포 기본은 CPU 점유율을 낮추는 1-thread 모드입니다.
# ==============================================================================
import os as _loa_early_os

def _loa_apply_cpu_limits_early() -> None:
    mode = str(_loa_early_os.environ.get("LOA_OCR_CPU_MODE", "web_low")).strip().lower()
    if mode in {"local_fast", "fast", "desktop_fast"}:
        default_threads = "4"
        force = str(_loa_early_os.environ.get("LOA_FORCE_CPU_LIMIT", "0")).strip().lower() in {"1", "true", "yes", "on"}
    elif mode in {"balanced", "local_balanced"}:
        default_threads = "2"
        force = True
    else:
        default_threads = "1"
        force = True
    threads = str(_loa_early_os.environ.get("LOA_OCR_THREADS", default_threads)).strip() or default_threads
    try:
        n = max(1, min(8, int(float(threads))))
    except Exception:
        n = int(default_threads)
    threads = str(n)
    _loa_early_os.environ["LOA_EFFECTIVE_OCR_THREADS"] = threads
    _loa_early_os.environ.setdefault("LOA_OCR_CPU_MODE", mode or "web_low")
    for var in (
        "RAPIDOCR_THREADS",
        "OMP_NUM_THREADS",
        "ORT_NUM_THREADS",
        "OPENBLAS_NUM_THREADS",
        "MKL_NUM_THREADS",
        "VECLIB_MAXIMUM_THREADS",
        "NUMEXPR_NUM_THREADS",
        "NUMEXPR_MAX_THREADS",
    ):
        if force or not _loa_early_os.environ.get(var):
            _loa_early_os.environ[var] = threads
    _loa_early_os.environ.setdefault("OMP_WAIT_POLICY", "PASSIVE")
    _loa_early_os.environ.setdefault("KMP_BLOCKTIME", "0")
    _loa_early_os.environ.setdefault("KMP_SETTINGS", "0")
    _loa_early_os.environ.setdefault("CUDA_VISIBLE_DEVICES", "")
    # OpenCV는 나중에 import되더라도 여기서 한 번 제한을 시도합니다.
    try:
        import cv2 as _loa_cv2  # type: ignore
        _loa_cv2.setNumThreads(n)
        try:
            _loa_cv2.ocl.setUseOpenCL(False)
        except Exception:
            pass
    except Exception:
        pass

_loa_apply_cpu_limits_early()


import json
import math
import io
import os
import re
import zipfile
import html
import time
from datetime import datetime
from pathlib import Path
from typing import Any, Dict

import pandas as pd
import streamlit as st
import yaml
from PIL import Image

try:
    import altair as alt
except Exception:  # noqa: BLE001
    alt = None

from modules.api_parser import summarize_all
from modules.calculators import (
    binomial_metrics,
    compute_efficiency,
    corrected_dps,
    clamp_percent,
    format_korean_number,
    parse_korean_number,
    grade_efficiency,
    infer_hurricane_brutal_counts,
    normalize_battle_df,
    summarize_battle,
)
from modules.lostark_api import LostArkApiClient, serializable_bundle
from modules import direction as _dir
from modules.direction import (
    AttackDirectionType,
    DirectionalEngravingState,
    DirectionSettings,
    DirectionValidationError,
    resolve_engraving_state,
    evaluate_skill_direction,
    classify_skill_attack_direction,
    calc_direction_data_confidence,
)
from modules.ocr_engine import (
    easyocr_available,
    empty_battle_table,
    image_from_upload,
)
from modules.fixed_grid_ocr import (
    apply_summary_overrides,
    get_easyocr_reader,
    correct_battle_skill_names,
    correct_battle_skill_names_with_icons,
    make_attack_ocr_debug,
    make_summary_ocr_debug,
    parse_attack_fixed_grid,
    parse_summary_fixed_grid,
    save_learned_battle_icons_v32,
    _load_local_skill_db_candidates_v80,
)

ROOT = Path(__file__).parent
CONFIG_DIR = ROOT / "configs"
DATA_DIR = ROOT / "data"
SAMPLE_DIR = ROOT / "samples"
EXPORT_DIR = ROOT / "exports"
EXPORT_DIR.mkdir(exist_ok=True)
APP_CALC_VERSION = "v99_integrated_debug"


def strip_api_text(value: Any) -> str:
    """API Tooltip HTML/JSON 문자열을 일반 텍스트로 정리합니다.

    v56에서 이 함수가 일부 내부 함수 안에만 정의되어, 사이드바 API 검색 시
    _compact_text()가 먼저 호출되면 NameError가 발생할 수 있었습니다.
    전역 유틸로 올려 모든 파서/요약 함수에서 공통 사용합니다.
    """
    text = str(value or "")
    text = re.sub(r"<BR\s*/?>", "\n", text, flags=re.I)
    text = re.sub(r"<[^>]+>", "", text)
    text = (
        text.replace("&nbsp;", " ")
        .replace("&lt;", "<")
        .replace("&gt;", ">")
        .replace("&amp;", "&")
        .replace("&#39;", "'")
        .replace("&quot;", '"')
    )
    # Lost Ark API Tooltip이 JSON 문자열 형태로 들어와도 최소한 보기 좋은 텍스트로 축약합니다.
    text = text.replace("\r", " ").replace("\n", " ")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


# 클래스별 대표 직업/세팅명 후보입니다.
# API가 별도 "직업" 필드를 주지 않기 때문에 아크패시브 깨달음/각인 원문에서 이 이름을 찾아 추정합니다.
# 새 직업이 추가되어도, 먼저 아크패시브 깨달음 노드명을 자동 후보로 보여주고 이 표는 fallback/검증용으로만 씁니다.
CLASS_JOB_KEYWORDS: dict[str, list[str]] = {
    "버서커": ["광기", "광전사의 비기"],
    "디스트로이어": ["분노의 망치", "중력 수련"],
    "워로드": ["고독한 기사", "전투 태세"],
    "홀리나이트": ["심판자", "축복의 오라"],
    "슬레이어": ["포식자", "처단자"],
    "배틀마스터": ["초심", "오의 강화"],
    "인파이터": ["극의: 체술", "충격 단련"],
    "기공사": ["무상신공", "역천지체"],
    "창술사": ["절제", "절정"],
    "스트라이커": ["오의난무", "일격필살"],
    "브레이커": ["수라의 길", "권왕파천무"],
    "데빌헌터": ["강화 무기", "핸드거너"],
    "블래스터": ["화력 강화", "포격 강화"],
    "호크아이": ["죽음의 습격", "두 번째 동료"],
    "스카우터": ["진화의 유산", "아르데타인의 기술"],
    "건슬링어": ["피스메이커", "사냥의 시간"],
    "바드": ["절실한 구원", "진실된 용맹"],
    "서머너": ["상급 소환사", "넘치는 교감"],
    "아르카나": ["황후의 은총", "황제의 칙령"],
    "소서리스": ["점화", "환류"],
    "블레이드": ["잔재된 기운", "버스트"],
    "데모닉": ["멈출 수 없는 충동", "완벽한 억제"],
    "리퍼": ["달의 소리", "갈증"],
    "소울이터": ["만월의 집행자", "그믐의 경계"],
    "도화가": ["만개", "회귀"],
    "기상술사": ["이슬비", "질풍노도"],
    "환수사": ["야성", "환수 각성"],
}


def _compact_text(value: Any) -> str:
    return re.sub(r"\s+", " ", strip_api_text(str(value or ""))).strip()


def _flatten_values_with_path(obj: Any, path: str = "") -> list[tuple[str, str]]:
    """API bundle을 path/value 문자열로 펼칩니다. UI 분류와 직업 추정용입니다."""
    rows: list[tuple[str, str]] = []
    if isinstance(obj, dict):
        for k, v in obj.items():
            kp = f"{path}.{k}" if path else str(k)
            rows.extend(_flatten_values_with_path(v, kp))
    elif isinstance(obj, list):
        for i, v in enumerate(obj):
            rows.extend(_flatten_values_with_path(v, f"{path}[{i}]"))
    else:
        if obj is not None:
            text = _compact_text(obj)
            if text:
                rows.append((path, text))
    return rows


def _joined_df_text(df: Any) -> str:
    if not isinstance(df, pd.DataFrame) or df.empty:
        return ""
    vals = []
    for col in df.columns:
        try:
            vals.extend(_compact_text(x) for x in df[col].tolist())
        except Exception:
            pass
    return " ".join(x for x in vals if x)


def _filter_df_contains(df: Any, *keywords: str) -> pd.DataFrame:
    if not isinstance(df, pd.DataFrame) or df.empty:
        return pd.DataFrame()
    if not keywords:
        return df.copy()
    mask = pd.Series(False, index=df.index)
    for col in df.columns:
        col_text = df[col].astype(str)
        for kw in keywords:
            mask = mask | col_text.str.contains(re.escape(kw), case=False, na=False)
    return df[mask].copy()







# v56 -------------------------------------------------------------------------
# 직업/세팅명 추정은 아래 순서로 자동화합니다.
# 1) API가 JobName/ArkPassive 명칭을 직접 주면 사용
# 2) 아크패시브 깨달음 활성 노드명에서 클래스별 후보와 매칭
# 3) 직업 각인/각인 원문에서 후보와 매칭
# 4) 실패하면 깨달음 활성 노드 후보를 그대로 노출
# 아이덴티티 효과 자체는 modules/api_skill_estimator.py에서 API tooltip 자동 파서 + configs/class_rules.yaml 보완룰로 계산합니다.
JOB_GENERIC_WORDS_V56 = {
    "깨달음", "진화", "도약", "아크패시브", "전투 특성", "전투특성", "스킬", "효과", "노드", "레벨",
    "피해", "치명", "치명타", "공격", "공격력", "적중률", "피해량", "추가", "증가", "감소",
}


def _v56_load_class_rules_for_jobs() -> dict[str, Any]:
    path = CONFIG_DIR / "class_rules.yaml"
    if not path.exists():
        return {}
    try:
        return yaml.safe_load(path.read_text(encoding="utf-8")) or {}
    except Exception:
        return {}


def _v56_class_job_keywords() -> dict[str, list[str]]:
    merged: dict[str, list[str]] = {k: list(v) for k, v in CLASS_JOB_KEYWORDS.items()}
    rules = _v56_load_class_rules_for_jobs()
    for class_name, block in (rules.get("classes") or {}).items():
        jobs = list((block.get("jobs") or {}).keys())
        if jobs:
            base = merged.setdefault(str(class_name), [])
            for job in jobs:
                if str(job) not in base:
                    base.append(str(job))
    return merged


def _v58_job_match_keywords(class_name: str, job_name: str) -> list[str]:
    """직업명 외에 깨달음/아이덴티티 스킬명으로도 직업을 추정하기 위한 키워드.

    v125: 아르카나는 "황후의 기사", "루인", "카드"처럼 두 세팅 모두에 섞일 수 있는
    단어가 많아 직업 판정에 쓰면 황제 캐릭터도 황후로 오탐할 수 있습니다.
    따라서 아르카나 직업 판정은 직업명 전체(황후의 은총/황제의 칙령)만 사용하고,
    실제 확정은 아래 깨달음 1티어 정확 일치 로직이 담당합니다.
    """
    rules = _v56_load_class_rules_for_jobs()
    block = ((rules.get("classes") or {}).get(class_name) or {}).get("jobs", {}).get(job_name, {})
    keys: list[str] = [str(job_name)]

    raw_keys = list(block.get("engraving_keywords") or []) + list(block.get("identity_keywords") or []) + list(block.get("arkpassive_keywords") or [])
    if class_name == "아르카나":
        broad_arcana_tokens = {"황후", "황제", "루인", "카드", "카드 덱"}
        raw_keys = [k for k in raw_keys if str(k).strip() not in broad_arcana_tokens]

    for k in raw_keys:
        ks = str(k).strip()
        if ks and ks not in keys:
            keys.append(ks)
    return keys


def _v56_node_is_inactive(obj: Any) -> bool:
    if not isinstance(obj, dict):
        return False
    false_tokens = {"false", "0", "none", "inactive", "disabled", "비활성", "미사용", "미선택"}
    true_keys = ["IsSelected", "IsActive", "Selected", "Active", "Enabled", "Use", "IsEnable"]
    for k in true_keys:
        if k in obj:
            v = obj.get(k)
            if isinstance(v, bool):
                return not v
            if str(v).strip().lower() in false_tokens:
                return True
    # 로스트아크 API/가공 JSON에서 노드 레벨/포인트가 0이면 보통 비활성입니다.
    for k in ["Level", "level", "Lv", "lv", "Point", "Points", "point", "points"]:
        if k in obj:
            try:
                if float(str(obj.get(k)).replace(",", "")) <= 0:
                    return True
            except Exception:
                pass
    return False


def _v56_iter_dict_nodes(obj: Any, path: str = "") -> list[tuple[str, dict[str, Any]]]:
    rows: list[tuple[str, dict[str, Any]]] = []
    if isinstance(obj, dict):
        if not _v56_node_is_inactive(obj):
            rows.append((path, obj))
            for k, v in obj.items():
                rows.extend(_v56_iter_dict_nodes(v, f"{path}.{k}" if path else str(k)))
    elif isinstance(obj, list):
        for i, v in enumerate(obj):
            rows.extend(_v56_iter_dict_nodes(v, f"{path}[{i}]"))
    return rows


def _v56_text_from_node(obj: dict[str, Any]) -> str:
    parts: list[str] = []
    for k in ["Name", "name", "NodeName", "nodeName", "Title", "title", "Description", "description", "Effect", "effect", "Tooltip", "tooltip"]:
        if k in obj and obj.get(k) not in (None, ""):
            parts.append(_compact_text(obj.get(k)))
    return " ".join(p for p in parts if p)


def _v56_node_name(obj: dict[str, Any]) -> str:
    for k in ["Name", "name", "NodeName", "nodeName", "Title", "title"]:
        v = obj.get(k)
        if v:
            text = _compact_text(v).strip()
            text = re.sub(r"Lv\.?\s*\d+|레벨\s*\d+|\d+\s*티어|\d+\s*단계", "", text).strip(" -:·[]()")
            if text:
                return text
    return ""


def _v58_is_arkpassive_path(path: str) -> bool:
    path_l = str(path or "").lower()
    return any(k in path_l for k in ["arkpassive", "ark_passive", "ark passive", "arkpassives"])


def _v58_is_enlightenment_path_or_text(path: str, text: str) -> bool:
    joined = f"{path} {text}".lower()
    return ("깨달음" in text) or ("enlight" in joined) or ("enlightenment" in joined)


def _v58_is_valid_job_candidate_name(name: str) -> bool:
    name = str(name or "").strip()
    if not name or name in JOB_GENERIC_WORDS_V56:
        return False
    if len(name) < 2 or len(name) > 18:
        return False
    # 장비/보석/수치/일반 효과 문구는 직업 후보에서 제외합니다.
    bad_tokens = [
        "도래한", "결전", "반지", "귀걸이", "목걸이", "팔찌", "어빌리티", "스톤", "보석", "겁화", "작열",
        "피해", "증가", "감소", "치명", "치피", "공격력", "스택", "초", "레벨", "포인트", "랭크", "Lv", "Level",
    ]
    if any(tok in name for tok in bad_tokens):
        return False
    if re.search(r"\d|%", name):
        return False
    return True


def _v56_candidate_names_from_active_enlightenment(bundle: dict[str, Any]) -> list[dict[str, str]]:
    """v58: 직업 후보는 아크패시브/깨달음 영역에서만 추출합니다.

    이전 버전은 전체 API 원문 안에 '깨달음'이라는 단어가 한 번이라도 있으면 장비/악세 Tooltip까지
    후보로 훑어서 '도래한 결전의 반지' 같은 장비명이 직업 후보로 섞였습니다. 이제 ArkPassive 영역만
    보고, 장비/각인/보석 경로는 후보에서 제외합니다.
    """
    data_candidates: list[tuple[str, Any]] = []
    if isinstance(bundle, dict):
        for key in ["arkpassive", "ArkPassive", "ark_passive", "ArkPassiveEffects"]:
            payload = bundle.get(key)
            if isinstance(payload, dict) and "data" in payload:
                data_candidates.append((key, payload.get("data")))
            elif payload is not None:
                data_candidates.append((key, payload))

    out: list[dict[str, str]] = []
    seen = set()
    forbidden_path_tokens = ["ArmoryEquipment", "Equipment", "Accessory", "ArmoryGem", "Gem", "Engraving", "Card", "Bracelet"]
    for root_key, data in data_candidates:
        for path, node in _v56_iter_dict_nodes(data):
            full_path = f"{root_key}.{path}" if path else str(root_key)
            if any(tok.lower() in full_path.lower() for tok in forbidden_path_tokens):
                continue
            text = _v56_text_from_node(node)
            joined = f"{full_path} {text}"
            if not _v58_is_arkpassive_path(full_path):
                continue
            if not _v58_is_enlightenment_path_or_text(full_path, joined):
                continue
            name = _v56_node_name(node)
            if not _v58_is_valid_job_candidate_name(name):
                continue
            key = (name, full_path)
            if key in seen:
                continue
            seen.add(key)
            out.append({"name": name, "path": full_path, "text": text[:360]})
    return out



def _v125_level_value_from_any(value: Any) -> float | None:
    """아크패시브 노드의 레벨/포인트가 0인지, 1 이상인지 느슨하게 판정합니다."""
    if value is None:
        return None
    if isinstance(value, bool):
        return 1.0 if value else 0.0
    text = _compact_text(value)
    if not text:
        return None
    low = text.lower()
    if any(tok in low for tok in ["inactive", "disabled", "unselected", "false"]):
        return 0.0
    if any(tok in text for tok in ["비활성", "미사용", "미선택"]):
        return 0.0
    # 레벨/포인트 필드에서는 숫자 하나만 있어도 활성 여부로 쓸 수 있습니다.
    nums = re.findall(r"-?\d+(?:\.\d+)?", text.replace(",", ""))
    if not nums:
        return None
    try:
        return max(float(x) for x in nums)
    except Exception:
        return None


def _v125_node_selected_state(node: dict[str, Any]) -> bool | None:
    """노드 dict가 활성인지 판정합니다. True/False/None(모름)을 반환합니다."""
    if not isinstance(node, dict):
        return None
    for k in ["IsSelected", "IsActive", "Selected", "Active", "Enabled", "Use", "IsEnable"]:
        if k not in node:
            continue
        v = node.get(k)
        if isinstance(v, bool):
            return v
        text = str(v).strip().lower()
        if text in {"true", "1", "yes", "on", "active", "enabled", "selected"}:
            return True
        if text in {"false", "0", "no", "off", "inactive", "disabled", "unselected"}:
            return False
    for k in ["Level", "level", "Lv", "lv", "Point", "Points", "point", "points", "Grade", "grade"]:
        if k in node:
            lv = _v125_level_value_from_any(node.get(k))
            if lv is not None:
                return lv > 0
    return None


def _v125_row_selected_state(row: Any) -> bool | None:
    """parse_arkpassive()가 만든 표 행 기준 활성/비활성 여부를 판정합니다."""
    try:
        level = row.get("레벨/등급")
    except Exception:
        level = None
    lv = _v125_level_value_from_any(level)
    if lv is not None:
        return lv > 0
    try:
        blob = f"{row.get('경로', '')} {row.get('설명', '')}"
    except Exception:
        blob = ""
    blob_l = str(blob).lower()
    if any(tok in blob_l for tok in ["inactive", "disabled", "unselected", "false"]):
        return False
    if any(tok in str(blob) for tok in ["비활성", "미사용", "미선택"]):
        return False
    return None


def _v125_is_tier1_enlightenment_context(path: str, text: str) -> bool:
    blob = f"{path} {text}"
    low = blob.lower()
    has_enlight = ("깨달음" in blob) or ("enlight" in low)
    has_tier1 = bool(re.search(r"(?:^|[^0-9])1\s*(?:티어|tier|단계|t)(?:[^0-9]|$)", blob, flags=re.I))
    # 로스트아크 API 구조에 따라 path에 Tier 번호가 없고, 깨달음 Effects만 노출되는 경우도 있습니다.
    return has_enlight or has_tier1


def _v125_exact_job_from_enlightenment_tier1(
    summary: dict[str, Any] | None,
    bundle: dict[str, Any] | None,
    class_name: str,
    known: list[str],
) -> dict[str, Any] | None:
    """v125: 깨달음 1티어에 찍힌 직업 노드명을 최우선으로 직업을 확정합니다.

    특히 아르카나는 '황후의 기사' 노드나 '루인/카드' 키워드 때문에 황제의 칙령을
    황후의 은총으로 잘못 판정하는 경우가 있어, 직업명 전체가 노드명과 정확히 일치할 때만
    높은 점수의 확정 근거로 사용합니다.
    """
    if not class_name or not known:
        return None
    known_set = {str(x).strip() for x in known if str(x).strip()}
    if not known_set:
        return None

    hits: list[dict[str, Any]] = []

    # 1) 원본 ArkPassive dict: IsSelected/Level/Point 같은 활성 필드를 볼 수 있으면 가장 신뢰합니다.
    if isinstance(bundle, dict):
        data_candidates: list[tuple[str, Any]] = []
        for key in ["arkpassive", "ArkPassive", "ark_passive", "ArkPassiveEffects"]:
            payload = bundle.get(key)
            if isinstance(payload, dict) and "data" in payload:
                data_candidates.append((key, payload.get("data")))
            elif payload is not None:
                data_candidates.append((key, payload))
        for root_key, data in data_candidates:
            for path, node in _v56_iter_dict_nodes(data):
                if not isinstance(node, dict):
                    continue
                name = _v56_node_name(node)
                if name not in known_set:
                    continue
                selected = _v125_node_selected_state(node)
                if selected is False:
                    continue
                full_path = f"{root_key}.{path}" if path else str(root_key)
                text = _v56_text_from_node(node)
                tier1_ctx = _v125_is_tier1_enlightenment_context(full_path, text)
                hits.append({
                    "job": name,
                    "score": 1000 if tier1_ctx else 940,
                    "source": f"깨달음 1티어 직업 노드 정확 일치: {full_path}",
                    "selected_known": selected is True,
                })

    # 2) summarize_all()의 아크패시브 표: 원본 dict에서 못 잡힌 경우 보완합니다.
    ark_df = (summary or {}).get("arkpassive") if isinstance(summary, dict) else None
    if isinstance(ark_df, pd.DataFrame) and not ark_df.empty and "이름" in ark_df.columns:
        for _, row in ark_df.iterrows():
            name = _compact_text(row.get("이름"))
            if name not in known_set:
                continue
            selected = _v125_row_selected_state(row)
            if selected is False:
                continue
            path = _compact_text(row.get("경로"))
            text = f"{row.get('레벨/등급', '')} {row.get('설명', '')}"
            tier1_ctx = _v125_is_tier1_enlightenment_context(path, text)
            hits.append({
                "job": name,
                "score": 900 if tier1_ctx else 840,
                "source": f"아크패시브 표의 깨달음 1티어 직업 노드 정확 일치: {path or '-'}",
                "selected_known": selected is True,
            })

    if not hits:
        return None

    # 같은 직업은 최고 점수만 남깁니다.
    best_by_job: dict[str, dict[str, Any]] = {}
    for hit in hits:
        job = str(hit.get("job") or "")
        cur = best_by_job.get(job)
        if cur is None or (float(hit.get("score") or 0), bool(hit.get("selected_known"))) > (float(cur.get("score") or 0), bool(cur.get("selected_known"))):
            best_by_job[job] = hit
    ranked = sorted(best_by_job.values(), key=lambda x: (float(x.get("score") or 0), bool(x.get("selected_known"))), reverse=True)
    if not ranked:
        return None

    # 둘 이상의 직업이 완전히 같은 확정도로 잡히면 임의 선택하지 않고 기존 점수 로직에 넘깁니다.
    if len(ranked) >= 2:
        top_key = (float(ranked[0].get("score") or 0), bool(ranked[0].get("selected_known")))
        second_key = (float(ranked[1].get("score") or 0), bool(ranked[1].get("selected_known")))
        if top_key == second_key:
            return None
    return ranked[0]

def _infer_job_from_arkpassive(summary: dict[str, Any] | None, bundle: dict[str, Any] | None) -> dict[str, Any]:  # type: ignore[override]
    """v56: API/깨달음 활성 노드/각인 원문을 모두 사용해 직업명을 자동 추정합니다."""
    summary = summary or {}
    bundle = bundle or {}
    profile = summary.get("profile_summary") if isinstance(summary.get("profile_summary"), dict) else {}
    class_name = str(profile.get("클래스") or "").strip()
    class_jobs = _v56_class_job_keywords()
    known = class_jobs.get(class_name, [])
    all_known = sorted({x for xs in class_jobs.values() for x in xs}, key=len, reverse=True)
    candidates = known or all_known

    raw_pairs = _flatten_values_with_path(bundle)
    ark_df = summary.get("arkpassive")
    ark_text = _joined_df_text(ark_df)
    bundle_text = " ".join(t for _, t in raw_pairs)
    active_enlight = _v56_candidate_names_from_active_enlightenment(bundle)
    active_text = " ".join((x.get("name", "") + " " + x.get("text", "")) for x in active_enlight)
    engrave_pairs = [(p, t) for p, t in raw_pairs if "engraving" in p.lower() or "각인" in t]
    engrave_text = " ".join(t for _, t in engrave_pairs)

    scored: list[dict[str, Any]] = []

    # v125: 아르카나 등 직업명이 깨달음 1티어 노드명에 정확히 찍히는 클래스는
    # 루인/카드/황후의 기사 같은 보조 키워드보다 이 값을 최우선으로 사용합니다.
    exact_tier1_job = _v125_exact_job_from_enlightenment_tier1(summary, bundle, class_name, known)
    if exact_tier1_job:
        scored.append({
            "직업 후보": exact_tier1_job.get("job"),
            "점수": exact_tier1_job.get("score", 1000),
            "근거": exact_tier1_job.get("source", "깨달음 1티어 직업 노드 정확 일치"),
        })

    # API가 직접 직업/세팅명을 주는 경우 최우선.
    direct_keys = ["JobName", "jobName", "Job", "job", "SpecializationName", "specializationName", "EngravingClassName"]
    for p, t in raw_pairs:
        if any(k.lower() in p.lower() for k in direct_keys):
            cand = t.strip()
            if cand and cand not in JOB_GENERIC_WORDS_V56:
                scored.append({"직업 후보": cand, "점수": 120, "근거": f"API 직접 필드: {p}"})

    for name in candidates:
        score = 0
        reasons: list[str] = []
        if not name:
            continue
        match_keys = _v58_job_match_keywords(class_name, name)
        if any(k and k in active_text for k in match_keys):
            score += 95
            reasons.append("깨달음/아이덴티티 키워드")
        if any(k and k in ark_text for k in match_keys):
            score += 25
            reasons.append("아크패시브 표")
        if any(k and k in engrave_text for k in match_keys):
            score += 35
            reasons.append("각인 원문")
        # API 원문 전체는 약한 근거로만 사용합니다. 장비/보석 명칭은 직업 후보로 확정하지 않습니다.
        if any(k and k in bundle_text for k in match_keys):
            score += 5
            reasons.append("API 원문")
        if name in known:
            score += 5
        if score:
            scored.append({"직업 후보": name, "점수": score, "근거": ", ".join(dict.fromkeys(reasons))})

    # v58: 신직업 fallback은 클래스 후보표가 없을 때만 낮은 점수로 표시합니다.
    # 장비/악세/보석 명칭이 직업 후보로 섞이는 문제를 막기 위해 일반 상황에서는 후보표 밖 이름을 확정하지 않습니다.
    if not known:
        for item in active_enlight:
            name = item.get("name", "").strip()
            if not _v58_is_valid_job_candidate_name(name) or name in candidates:
                continue
            scored.append({"직업 후보": name, "점수": 20, "근거": f"깨달음 활성 노드 후보: {item.get('path', '')}"})

    # 같은 후보가 여러 번 잡히면 최고 점수 + 근거 병합.
    merged: dict[str, dict[str, Any]] = {}
    for row in scored:
        name = str(row.get("직업 후보") or "")
        if not name:
            continue
        cur = merged.get(name)
        if cur is None or float(row.get("점수", 0) or 0) > float(cur.get("점수", 0) or 0):
            merged[name] = dict(row)
        elif cur is not None:
            cur["근거"] = ", ".join(dict.fromkeys([*(str(cur.get("근거") or "").split(", ")), *(str(row.get("근거") or "").split(", "))]))
    scored = list(merged.values())
    scored.sort(key=lambda r: (float(r.get("점수", 0) or 0), str(r.get("직업 후보", "") in known), len(str(r.get("직업 후보", "")))), reverse=True)

    job = scored[0]["직업 후보"] if scored else ""
    source = scored[0]["근거"] if scored else "추정 실패"

    return {
        "class_name": class_name,
        "job_name": job,
        "source": source,
        "candidates": pd.DataFrame(scored) if scored else pd.DataFrame(columns=["직업 후보", "점수", "근거"]),
        "active_enlightenment_nodes": pd.DataFrame(active_enlight) if active_enlight else pd.DataFrame(columns=["name", "path", "text"]),
    }


def enrich_summary_with_identity(summary: dict[str, Any] | None, bundle: dict[str, Any] | None) -> dict[str, Any] | None:  # type: ignore[override]
    if not isinstance(summary, dict):
        return summary
    info = _infer_job_from_arkpassive(summary, bundle or {})
    profile = summary.setdefault("profile_summary", {})
    if isinstance(profile, dict):
        profile["직업"] = info.get("job_name") or profile.get("직업") or ""
        profile["직업 추정 근거"] = info.get("source") or ""
    summary["job_detection"] = info
    rows = [
        {"구분": "클래스", "값": info.get("class_name") or "-", "근거": "Profile CharacterClassName"},
        {"구분": "직업/세팅", "값": info.get("job_name") or "추정 실패", "근거": info.get("source") or "-"},
        {"구분": "아이덴티티 효과", "값": "깨달음 자동 파서 + 직업 룰", "근거": "아크패시브/깨달음 Tooltip 우선, 실패 시 configs/class_rules.yaml"},
    ]
    summary["profile_identity_summary"] = pd.DataFrame(rows)
    return summary

# =====================================================================
# 아르카나 카드 아이덴티티 확률 보정 (황제 / 황후의 기사 직접피해 카드)
# =====================================================================
# 카드 아이덴티티는 매 판 뽑기 RNG로 직접피해 카드(황제·황후의 기사)가 몇 번
# 나오느냐에 따라 딜 편차가 큽니다. data/arcana_card_probabilities.xlsx의
# 각인별 '카드 뜰 확률'과 전투분석기 '카드 사용 횟수'를 이용해, 이 두 카드의
# 사용 횟수를 이항분포 기대값(N×p)으로 정규화하고 0.1% 하한(불운 최소)을 계산합니다.
# 파일 수정 시각(mtime)을 키로 캐시해, 엑셀 확률을 고치면 앱 재시작 없이
# 다음 검색/재실행 때 자동 반영되도록 합니다. (mtime, 결과) 튜플로 보관.
_ARCANA_DAMAGE_CARDS_CACHE: tuple[float, dict[str, dict[str, float]]] | None = None


def _load_arcana_damage_cards() -> dict[str, dict[str, float]]:
    """각인(시트명)별 직접피해 카드의 '카드 뜰 확률'(0~1)을 반환합니다.

    반환 예: {"황제의 칙령": {"황제": 0.154, "황후의 기사": 0.075},
              "황후의 은총": {"황후의 기사": 0.079}}
    - 계산분류가 '직접피해'로 시작하는 카드만(황제/황후의 기사) 뽑습니다.
    - 확률은 F열 '카드가 뜰 확률_%'(퍼센트, 예: 15.4)를 우선 읽어 100으로 나눕니다.
      F열이 없으면 E열 '확률_rate'(0~1 소수)로 폴백합니다.
      → 엑셀에서 F열 숫자만 고치면 바로 반영됩니다.
    - 파일 수정 시각이 바뀌면 캐시를 자동 갱신하므로 앱 재시작이 필요 없습니다.
    """
    global _ARCANA_DAMAGE_CARDS_CACHE
    path = DATA_DIR / "arcana_card_probabilities.xlsx"
    try:
        mtime = path.stat().st_mtime
    except Exception:
        mtime = 0.0
    if _ARCANA_DAMAGE_CARDS_CACHE is not None and _ARCANA_DAMAGE_CARDS_CACHE[0] == mtime:
        return _ARCANA_DAMAGE_CARDS_CACHE[1]
    out: dict[str, dict[str, float]] = {}
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)
        pct_col = "카드가 뜰 확률_%"
        rate_col = "확률_rate"
        obs_col = "실측횟수"

        def _num(v):
            try:
                return float(v)
            except (TypeError, ValueError):
                return None

        for ws in wb.worksheets:
            header = [str(c.value or "").strip() for c in ws[1]]
            idx = {h: i for i, h in enumerate(header)}
            if "카드명" not in idx or "계산분류" not in idx:
                continue
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            # 시트 전체 실측횟수 합계(확률 = 실측/합계) → 수식 캐시가 없을 때의 폴백 기준
            total_obs = 0.0
            if obs_col in idx:
                for row in rows:
                    ov = _num(row[idx[obs_col]])
                    if ov and ov > 0:
                        total_obs += ov
            cards: dict[str, float] = {}
            for row in rows:
                cat = str(row[idx["계산분류"]] or "").strip()
                if not cat.startswith("직접피해"):
                    continue
                name = str(row[idx["카드명"]] or "").strip()
                p = None
                # 1순위: F열 '카드가 뜰 확률_%'(퍼센트) → 100으로 나눔
                if pct_col in idx:
                    fv = _num(row[idx[pct_col]])
                    if fv is not None:
                        p = fv / 100.0
                # 2순위: E열 '확률_rate'(0~1 소수)
                if p is None and rate_col in idx:
                    p = _num(row[idx[rate_col]])
                # 3순위: 실측횟수 / 시트 실측합계 (수식 캐시가 비어도 항상 계산됨)
                if p is None and obs_col in idx and total_obs > 0:
                    ov = _num(row[idx[obs_col]])
                    if ov is not None:
                        p = ov / total_obs
                if name and p is not None and 0.0 < p <= 1.0:
                    cards[name] = p
            if cards:
                out[str(ws.title).strip()] = cards
    except Exception:
        out = {}
    _ARCANA_DAMAGE_CARDS_CACHE = (mtime, out)
    return out


def _arcana_engraving_from_api(summary: dict[str, Any] | None) -> str | None:
    """API 각인 목록에서 아르카나 직업 각인(황제의 칙령 / 황후의 은총)을 직접 읽습니다.

    직업 추정(_infer_job_from_arkpassive)은 깨달음 노드 텍스트 기반이라, '황후의 기사'
    깨달음 노드를 찍은 황제의 칙령 캐릭터를 '황후의 은총'으로 오탐할 수 있습니다.
    실제 장착 각인 이름을 최우선으로 읽어 이 오탐을 방지합니다.
    """
    try:
        import pandas as _pd
        eng_df = (summary or {}).get("engravings")
        if isinstance(eng_df, _pd.DataFrame) and not eng_df.empty:
            cols = [c for c in ("이름", "레벨/설명") if c in eng_df.columns]
            blob = " ".join(str(v) for c in cols for v in eng_df[c].tolist())
            if "황제의 칙령" in blob:
                return "황제의 칙령"
            if "황후의 은총" in blob:
                return "황후의 은총"
    except Exception:
        pass
    return None


def _arcana_knight_arkpassive_level(summary: dict[str, Any] | None) -> int:
    """아크패시브 '깨달음 - 황후의 기사' 노드 레벨을 반환합니다(없으면 0).

    황후의 기사 카드는 이 노드를 1레벨 이상 찍어야 사용할 수 있으므로, 확률 보정 대상에
    포함할지 결정하는 데 씁니다. 노드가 API에 노출됐지만 레벨 숫자를 못 읽으면(할당된 노드만
    보통 노출되므로) 1레벨로 간주합니다.
    """
    try:
        import pandas as _pd
        ark = (summary or {}).get("arkpassive")
        if not isinstance(ark, _pd.DataFrame) or ark.empty or "이름" not in ark.columns:
            return 0
        best = 0
        found = False
        for _, r in ark.iterrows():
            nm = str(r.get("이름") or "")
            if "황후의 기사" not in nm:
                continue
            found = True
            lvl_txt = f"{r.get('레벨/등급', '')} {r.get('설명', '')}"
            m = re.search(r"(\d+)", str(lvl_txt))
            if m:
                best = max(best, int(m.group(1)))
        if found and best == 0:
            best = 1  # 할당된 노드만 API에 노출되므로 레벨 미상이면 최소 1로 간주
        return best
    except Exception:
        return 0


def _arcana_card_context() -> dict[str, Any] | None:
    """현재 캐릭터가 아르카나면 각인/직접피해 카드 확률을 반환, 아니면 None.

    - 각인(시트 선택)은 사용자 수동 지정 > API 각인 목록 > 직업 추정 순으로 결정합니다.
    - 황후의 기사 카드는 아크패시브 '황후의 기사'가 1레벨 이상일 때만 포함합니다.
    - 황제의 칙령이면 '황제' 카드는 시트에 항상 존재하므로 그대로 포함됩니다.
    """
    try:
        summary = st.session_state.get("api_summary") or {}
        prof = summary.get("profile_summary") if isinstance(summary, dict) else {}
        cls = str((prof or {}).get("클래스") or "").strip()
        if cls != "아르카나":
            return None
        table = _load_arcana_damage_cards()

        # 각인 결정: 수동 override 우선 → API 각인 목록 → 직업 추정.
        override = str(st.session_state.get("arcana_engraving_override", "") or "").strip()
        engraving = ""
        source = ""
        if override in ("황제의 칙령", "황후의 은총"):
            engraving, source = override, "수동 지정"
        if not engraving:
            api_eng = _arcana_engraving_from_api(summary)
            if api_eng:
                engraving, source = api_eng, "API 각인 목록"
        if not engraving:
            engraving = str((prof or {}).get("직업") or "").strip() or str((summary.get("job_detection") or {}).get("job_name") or "").strip()
            source = "직업 추정"

        cards = dict(table.get(engraving) or {})
        if not cards:
            return None

        # 황후의 기사 카드 게이팅: 깨달음 '황후의 기사' 노드 1레벨 이상일 때만.
        knight_level = _arcana_knight_arkpassive_level(summary)
        knight_gated = False
        if "황후의 기사" in cards and knight_level < 1:
            cards.pop("황후의 기사", None)
            knight_gated = True

        return {
            "engraving": engraving,
            "engraving_source": source,
            "cards": cards,
            "knight_of_empress_level": knight_level,
            "knight_of_empress_gated": knight_gated,
        }
    except Exception:
        return None


def _binom_expected_count(n: int, p: float) -> float:
    return float(n) * float(p)


def _binom_lower_count(n: int, p: float, alpha: float = 0.001) -> int:
    """이항분포 Binomial(n, p)에서 누적확률이 alpha 이상이 되는 최소 정수 횟수.

    '불운하게 이보다 더 적게 나올 확률 < alpha'인 하한값입니다. (예: 0.1% 하한)
    """
    try:
        from math import comb
        n = int(n)
        if n <= 0 or p <= 0.0:
            return 0
        if p >= 1.0:
            return n
        q = 1.0 - p
        cum = 0.0
        for k in range(0, n + 1):
            cum += comb(n, k) * (p ** k) * (q ** (n - k))
            if cum >= alpha:
                return k
        return n
    except Exception:
        return 0


_ARCANA_CARD_TEMPLATE_CACHE: tuple | None = None


def _arcana_card_label_template():
    """data/Arcana_Card_data.png에서 '카드 사용 횟수' 라벨 영역만 잘라 반환합니다.

    반환: (label_gray_ndarray, full_width, full_height) 또는 (None, 0, 0).
    라벨(밝은 글자)만 매칭에 써야 값 숫자(95/272 등)가 달라도 위치를 찾습니다.
    """
    global _ARCANA_CARD_TEMPLATE_CACHE
    if _ARCANA_CARD_TEMPLATE_CACHE is not None:
        return _ARCANA_CARD_TEMPLATE_CACHE
    try:
        import cv2
        import numpy as np
        p = DATA_DIR / "Arcana_Card_data.png"
        if not p.exists():
            _ARCANA_CARD_TEMPLATE_CACHE = (None, 0, 0)
            return _ARCANA_CARD_TEMPLATE_CACHE
        full = cv2.imread(str(p), cv2.IMREAD_GRAYSCALE)
        if full is None:
            _ARCANA_CARD_TEMPLATE_CACHE = (None, 0, 0)
            return _ARCANA_CARD_TEMPLATE_CACHE
        ht, wt = full.shape[:2]
        col_max = full.max(axis=0)
        bright = np.where(col_max > 120)[0]
        right = int(bright.max()) + 40 if len(bright) else 400
        right = max(200, min(right, wt))
        label = full[:, 0:right]
        _ARCANA_CARD_TEMPLATE_CACHE = (label, wt, ht)
    except Exception:
        _ARCANA_CARD_TEMPLATE_CACHE = (None, 0, 0)
    return _ARCANA_CARD_TEMPLATE_CACHE


def _detect_arcana_card_uses(summary_img) -> dict[str, Any] | None:
    """종합정보 이미지에서 '카드 사용 횟수' 값을 템플릿 매칭 + OCR로 읽습니다.

    라벨 템플릿을 멀티스케일로 찾아(순서 변경돼도 위치 추적) 그 행 오른쪽 영역의
    숫자 토큰 중 가장 오른쪽(값은 우측 정렬) 값을 반환합니다. 확신이 없으면 None.
    반환: {"value": int, "score": float} 또는 None.
    """
    try:
        import cv2
        import numpy as np
        label, wt, ht = _arcana_card_label_template()
        if label is None or summary_img is None:
            return None
        scene = np.array(summary_img.convert("L"))
        H, W = scene.shape[:2]
        lh, lw = label.shape[:2]
        best = (0.0, 1.0, None)  # (score, scale, loc)
        for s in np.linspace(0.30, 1.80, 32):
            w = int(lw * s)
            h = int(lh * s)
            if w < 12 or h < 8 or w > W or h > H:
                continue
            lab_s = cv2.resize(label, (w, h))
            res = cv2.matchTemplate(scene, lab_s, cv2.TM_CCOEFF_NORMED)
            _, mx, _, ml = cv2.minMaxLoc(res)
            if mx > best[0]:
                best = (mx, s, ml)
        score, s, loc = best
        if loc is None or score < 0.72:
            return None
        x0, y0 = loc
        row_w = wt * s
        row_h = ht * s
        # 라벨 오른쪽(값 영역)만 OCR: 라벨 폭 이후 ~ 행 끝
        nx1 = int(max(0, x0 + lw * s * 0.9))
        nx2 = int(min(W, x0 + row_w))
        ny1 = int(max(0, y0 - row_h * 0.15))
        ny2 = int(min(H, y0 + row_h * 1.15))
        if nx2 - nx1 < 10 or ny2 - ny1 < 6:
            return None
        crop = scene[ny1:ny2, nx1:nx2]
        # 작은 글자 대비 3배 확대
        crop3 = cv2.resize(crop, (crop.shape[1] * 3, crop.shape[0] * 3), interpolation=cv2.INTER_CUBIC)
        from modules.fixed_grid_ocr import get_ocr_reader
        reader = get_ocr_reader()
        tokens = []
        try:
            try:
                results = reader.readtext(crop3, detail=1, paragraph=False, allowlist="0123456789")
            except TypeError:
                results = reader.readtext(crop3, detail=1, paragraph=False)
            for item in results or []:
                # item: (box, text, conf) 형태
                if isinstance(item, (list, tuple)) and len(item) >= 2:
                    box = item[0]
                    txt = str(item[1])
                else:
                    box, txt = None, str(item)
                digits = "".join(ch for ch in txt if ch.isdigit())
                if not digits:
                    continue
                try:
                    cx = float(np.mean([pt[0] for pt in box])) if box else 0.0
                except Exception:
                    cx = 0.0
                tokens.append((cx, int(digits)))
        except Exception:
            return None
        if not tokens:
            return None
        # 값은 우측 정렬 → 가장 오른쪽 숫자 토큰 선택
        tokens.sort(key=lambda t: t[0])
        val = tokens[-1][1]
        if 0 < val < 100000:
            return {"value": int(val), "score": round(float(score), 3)}
        return None
    except Exception:
        return None


def safe_int_value(value, default: int = 0) -> int:
    """None, NaN, 빈 문자열, OCR 실패값을 number_input에 넣을 수 있는 int로 변환."""
    if value is None:
        return default

    try:
        if pd.isna(value):
            return default
    except Exception:
        pass

    text = str(value).strip().replace(",", "")
    if text in ["", "-", "None", "nan", "NaN"]:
        return default

    try:
        return int(float(text))
    except Exception:
        return default


def safe_percent_value(value, default: float | None = None) -> float | None:
    """OCR 퍼센트를 0~100 범위, 소수 둘째 자리로 정리."""
    cleaned = clamp_percent(value)
    return default if cleaned is None else cleaned


def fmt_percent(value: Any) -> str:
    cleaned = safe_percent_value(value)
    return "-" if cleaned is None else f"{cleaned:.2f}%"


def fmt_ratio_percent(value: Any) -> str:
    """0~1 비율 값을 화면용 퍼센트 문자열로 변환합니다."""
    try:
        if value is None or pd.isna(value):
            return "-"
    except Exception:
        if value is None:
            return "-"
    try:
        return f"{float(value) * 100:.2f}%"
    except Exception:
        return "-"


def normalize_name(value: Any) -> str:
    return "".join(str(value or "").split()).lower()



def _normalize_special_battle_name_v86(value: Any) -> str:
    """v86: 공용 룬/파일명 fallback 표시명을 최종 한글 표시명으로 통일합니다."""
    raw = str(value or "").strip()
    if not raw:
        return raw
    compact = re.sub(r"[\s:_\-]+", "", raw.lower())
    mapping = {
        "poisonrune": "스킬룬 중독",
        "poisonrune.png": "스킬룬 중독",
        "poison": "스킬룬 중독",
        "중독룬": "스킬룬 중독",
        "스킬룬중독": "스킬룬 중독",
        "스킬룬poison": "스킬룬 중독",
        "bleedrune": "스킬룬 출혈",
        "bleedrune.png": "스킬룬 출혈",
        "bleed": "스킬룬 출혈",
        "출혈룬": "스킬룬 출혈",
        "스킬룬출혈": "스킬룬 출혈",
        "스킬룬bleed": "스킬룬 출혈",
    }
    if compact in mapping:
        return mapping[compact]
    # 기존 표기 '스킬룬 : 출혈' / '스킬룬: 출혈'도 공백형으로 통일
    m = re.match(r"^스킬룬\s*[:：]?\s*(.+)$", raw)
    if m:
        return f"스킬룬 {m.group(1).strip()}"
    return raw

def sanitize_battle_table(df: pd.DataFrame) -> pd.DataFrame:
    """OCR 결과표에서 퍼센트가 100.0096처럼 튀는 값을 화면 표시 전에 정리."""
    if df is None or df.empty:
        return df
    out = df.copy()
    for _name_col in ["이름", "name", "스킬명"]:
        if _name_col in out.columns:
            out[_name_col] = out[_name_col].apply(_normalize_special_battle_name_v86)
    rate_cols = [
        "back_attack_rate", "back_attack_share", "head_attack_rate", "head_attack_share", "crit_rate", "crit_share", "cooldown_rate", "share_rate",
        "백어택 적중률", "백어택 비중", "헤드어택 적중률", "헤드어택 비중", "치명타 적중률", "치명타 비중", "쿨타임 비율", "피해량 지분",
    ]
    for col in rate_cols:
        if col in out.columns:
            out[col] = out[col].apply(lambda v: safe_percent_value(v))
    if "casts" in out.columns:
        out["casts"] = out["casts"].apply(lambda v: safe_int_value(v, 0) if v not in [None, "", "-"] else None)
    if "사용 횟수" in out.columns:
        out["사용 횟수"] = out["사용 횟수"].apply(lambda v: safe_int_value(v, 0) if v not in [None, "", "-"] else None)
    return out



def filter_adopted_skills_df(df: pd.DataFrame, include_names: "set | None" = None) -> pd.DataFrame:
    """스킬 탭 표시용. 레벨 1은 미채용으로 보고 숨깁니다.

    API 구조가 바뀌어 레벨 컬럼이 없거나 전부 1 이하로 읽히면 원본을 그대로 보여줍니다.
    include_names: 레벨 1이어도 항상 포함할 스킬명 집합 (예: 초각성 스킬).
    """
    if df is None or df.empty:
        return pd.DataFrame()
    if "레벨" not in df.columns:
        return df
    out = df.copy()
    levels = pd.to_numeric(out["레벨"], errors="coerce")
    mask = levels > 1
    if include_names:
        name_col = "스킬명" if "스킬명" in out.columns else ("name" if "name" in out.columns else None)
        if name_col:
            mask = mask | out[name_col].astype(str).isin(include_names)
    adopted = out[mask].copy()
    return adopted if not adopted.empty else out



def get_ocr_skill_name_candidates() -> list[str]:
    """전투분석기 OCR 이름 보정용 스킬명 후보를 API 요약에서 가져옵니다.

    우선 최종 계산표/기준 계산표의 스킬명을 쓰고, 없으면 스킬 탭의 채용 스킬명을 사용합니다.
    도약/초각성 설명에서 레벨 1이지만 계산표에 포함된 스킬도 후보에 들어가도록 계산표를 우선합니다.
    """
    summary = st.session_state.get("api_summary") or {}
    candidates: list[str] = []
    for key in ["skill_crit_estimates", "lostbuilds_base_skill_estimates", "skills"]:
        df = summary.get(key)
        if not isinstance(df, pd.DataFrame) or df.empty:
            continue
        use_df = df
        if key == "skills":
            use_df = filter_adopted_skills_df(df)
        col = "스킬명" if "스킬명" in use_df.columns else "name" if "name" in use_df.columns else None
        if not col:
            continue
        for name in use_df[col].dropna().astype(str).tolist():
            name = name.strip()
            if name and name not in candidates:
                candidates.append(name)
    return candidates




def apply_quick_name_editor(kind: str) -> None:
    """data_editor에서 텍스트 셀이 엔터로 바로 확정되지 않는 환경을 위한 빠른 수정 폼.

    스킬명을 입력하고 Enter를 누르면 form submit으로 해당 행 이름을 바로 바꿉니다.
    """
    key = f"{kind}_table"
    df = st.session_state.get(key)
    if not isinstance(df, pd.DataFrame) or df.empty:
        return
    name_col = "이름" if "이름" in df.columns else "name" if "name" in df.columns else None
    if not name_col:
        return

    options = []
    for i, row in df.reset_index().iterrows():
        name = str(row.get(name_col) or "").strip() or "빈 이름"
        dmg = str(row.get("피해량") or row.get("damage_text") or "").strip()
        label = f"{int(row['index'])}: {name}" + (f" / {dmg}" if dmg else "")
        options.append((int(row["index"]), label))
    if not options:
        return

    with st.expander("스킬명 빠른 수정 - Enter 적용", expanded=False):
        st.caption("표 안에서 입력 확정이 불편하면 여기서 행을 고르고 새 이름을 입력한 뒤 Enter를 누르면 바로 적용됩니다.")
        with st.form(f"{kind}_quick_name_form", clear_on_submit=False):
            label_to_idx = {label: idx for idx, label in options}
            selected_label = st.selectbox("수정할 행", list(label_to_idx.keys()), key=f"{kind}_quick_name_row")
            current_idx = label_to_idx[selected_label]
            current_name = str(df.loc[current_idx, name_col] or "") if current_idx in df.index else ""
            new_name = st.text_input("새 스킬명", value=current_name, key=f"{kind}_quick_name_value")
            submitted = st.form_submit_button("적용")
        if submitted:
            updated = df.copy()
            if current_idx in updated.index:
                updated.loc[current_idx, name_col] = new_name.strip()
                st.session_state[key] = prepare_battle_editor_df(sanitize_battle_table(updated))
                st.success(f"{current_idx}번 행 이름을 '{new_name.strip()}'로 수정했어.")
                st.rerun()


def display_battle_table(df: pd.DataFrame) -> pd.DataFrame:
    """결과 대시보드용으로 컬럼명과 단위를 보기 좋게 정리."""
    if df is None or df.empty:
        return pd.DataFrame()
    display = df.copy()
    display["피해량"] = display["damage"].apply(format_korean_number)
    display["피해 지분"] = display["share_rate"].apply(fmt_percent)
    display["치명타 적중"] = display["crit_rate"].apply(fmt_percent)
    if "crit_share" in display.columns:
        display["치명타 비중"] = display["crit_share"].apply(fmt_percent)
    display["백어택 적중"] = display["back_attack_rate"].apply(fmt_percent)
    if "back_attack_share" in display.columns:
        display["백어택 비중"] = display["back_attack_share"].apply(fmt_percent)
    if "head_attack_rate" in display.columns:
        display["헤드어택 적중"] = display["head_attack_rate"].apply(fmt_percent)
    if "head_attack_share" in display.columns:
        display["헤드어택 비중"] = display["head_attack_share"].apply(fmt_percent)
    display["쿨타임"] = display["cooldown_rate"].apply(fmt_percent)
    cols = ["name", "피해량", "피해 지분", "백어택 적중", "백어택 비중", "헤드어택 적중", "헤드어택 비중", "치명타 적중", "치명타 비중", "casts", "쿨타임"]
    display = display[[c for c in cols if c in display.columns]]
    return display.rename(columns={"name": "스킬명", "casts": "사용 횟수"})


def get_skill_crit_estimates() -> pd.DataFrame:
    summary = st.session_state.get("api_summary") or {}
    df = summary.get("skill_crit_estimates")
    return df if isinstance(df, pd.DataFrame) else pd.DataFrame()


def selected_api_crit_columns() -> tuple[str, str]:
    """사이드바 선택값에 따라 스킬별 API 치명/치피 컬럼을 결정."""
    basis = st.session_state.get("api_crit_basis", "백어택 기준(조건부 포함)")
    if basis == "상시 옵션만":
        return "예상 치명 확률(상시)(%)", "예상 치피(상시)(%)"
    if basis == "조건부 포함":
        return "예상 치명 확률(조건부 포함)(%)", "예상 치피(조건부 포함)(%)"
    return "예상 치명 확률(백어택 기준)(%)", "예상 치피(조건부 포함)(%)"


def api_estimated_value_for_battle(norm_df: pd.DataFrame, column: str, fallback: float) -> float:
    """전투분석기 스킬 피해 비중으로 API 추정값을 가중 평균."""
    estimate_df = get_skill_crit_estimates()
    if estimate_df.empty or norm_df is None or norm_df.empty or column not in estimate_df.columns:
        return fallback
    lookup = {}
    for _, row in estimate_df.iterrows():
        try:
            lookup[normalize_name(row.get("스킬명"))] = float(row.get(column))
        except Exception:
            pass
    weighted = []
    for _, row in norm_df.iterrows():
        damage = row.get("damage")
        if damage is None or (isinstance(damage, float) and math.isnan(damage)) or damage <= 0:
            continue
        value = lookup.get(normalize_name(row.get("name")))
        if value is not None and not math.isnan(value):
            weighted.append((float(value), float(damage)))
    if not weighted:
        vals = pd.to_numeric(estimate_df.get(column), errors="coerce").dropna()
        return float(vals.mean()) if not vals.empty else fallback
    total_w = sum(w for _, w in weighted)
    return sum(v * w for v, w in weighted) / total_w if total_w > 0 else fallback


def api_estimated_crit_damage_for_battle(norm_df: pd.DataFrame) -> float:
    _crit_col, damage_col = selected_api_crit_columns()
    return api_estimated_value_for_battle(norm_df, damage_col, 200.0)


def api_estimated_crit_rate_for_battle(norm_df: pd.DataFrame) -> float:
    crit_col, _damage_col = selected_api_crit_columns()
    return api_estimated_value_for_battle(norm_df, crit_col, 0.0)




# ==============================
# v19: 전분 결과 / 개선분석 계산
# ==============================
HELPER_BATTLE_COLUMNS = ["OCR 원본 이름", "이름 보정 점수", "이름 보정 여부", "_ocr_row_index", "_icon_score", "_text_score", "_name_match_text_score", "_name_match_icon_score", "_name_match_reason", "_icon_match_name", "_icon_match_score", "_icon_match_source", "_icon_match_top3", "_unmatched_aggregated"]
STANDARD_BATTLE_COLUMN_ORDER = [
    "이름", "피해량", "초당 피해량", "피해량 지분",
    "백어택 적중률", "백어택 비중", "헤드어택 적중률", "헤드어택 비중",
    "치명타 적중률", "치명타 비중", "사용 횟수", "쿨타임 비율",
]


def strip_battle_helper_columns(df: pd.DataFrame) -> pd.DataFrame:
    """OCR 보정용 내부 컬럼은 사용자가 검수하는 표에서 제거합니다."""
    if df is None or df.empty:
        return df if isinstance(df, pd.DataFrame) else pd.DataFrame()
    out = df.copy()
    out = out.drop(columns=[c for c in HELPER_BATTLE_COLUMNS if c in out.columns], errors="ignore")
    return out




def _apply_data_editor_state_edits(df: pd.DataFrame, editor_key: str) -> pd.DataFrame:
    """Streamlit data_editor의 edited_rows 상태를 직접 반영합니다.

    일부 환경에서는 Enter가 셀 확정이 아니라 다음 칸 이동처럼 동작하면서
    반환 DF와 세션 상태 반영 타이밍이 어긋날 수 있습니다.
    이 함수는 widget state의 edited_rows를 한 번 더 적용해서
    표 안에서 입력한 스킬명/수치가 최대한 그대로 남도록 합니다.
    """
    if df is None:
        return pd.DataFrame()
    out = df.copy()
    state = st.session_state.get(editor_key)
    edits = None
    if isinstance(state, dict):
        edits = state.get("edited_rows") or state.get("editedCells")
    if not edits:
        return out
    try:
        for row_key, changes in edits.items():
            row_idx = int(row_key)
            if row_idx < 0 or row_idx >= len(out) or not isinstance(changes, dict):
                continue
            for col, value in changes.items():
                if col in out.columns:
                    out.at[out.index[row_idx], col] = value
    except Exception:
        return out
    return out


def _finite_float(value: Any, default: float | None = None) -> float | None:
    try:
        if value is None or pd.isna(value):
            return default
    except Exception:
        pass
    try:
        v = float(str(value).replace(",", "").replace("%", ""))
        if math.isnan(v) or math.isinf(v):
            return default
        return v
    except Exception:
        return default


def _first_number(row: pd.Series | dict | None, cols: list[str], default: float | None = None) -> float | None:
    if row is None:
        return default
    for col in cols:
        try:
            if col in row:
                v = _finite_float(row.get(col), None)  # type: ignore[arg-type]
                if v is not None:
                    return v
        except Exception:
            continue
    return default


def _expected_crit_multiplier(crit_rate_percent: float | None, crit_damage_percent: float | None) -> float:
    crit_rate = max(0.0, min(100.0, float(crit_rate_percent or 0.0))) / 100.0
    crit_damage = max(100.0, float(crit_damage_percent or 200.0)) / 100.0
    return 1.0 + crit_rate * (crit_damage - 1.0)


def _adjust_damage_to_expected_crit(
    damage: float | None,
    observed_crit_rate: float | None,
    observed_crit_share: float | None,
    expected_crit_rate: float | None,
    crit_damage_percent: float | None,
) -> float:
    """치명 운을 API 기대 치명률 기준으로 평균화합니다.

    신규 전투분석기의 `치명타 비중`은 단순 적중 횟수가 아니라
    `치명타 피해량 / 총 피해량`에 가까운 값입니다. 그래서 비중이 있으면
    관측 피해를 먼저 비치명 기준 피해량으로 환산한 뒤 API 기대 치명률을
    다시 곱합니다. 비중이 없을 때만 기존 치명타 적중률 방식으로 보정합니다.
    """
    if damage is None or damage <= 0:
        return 0.0
    if expected_crit_rate is None:
        return float(damage)

    crit_damage_mul = max(1.0, float(crit_damage_percent or 200.0) / 100.0)
    expected_mult = _expected_crit_multiplier(expected_crit_rate, crit_damage_percent)

    # v136: 치명운 보정은 '치명 운'을 API 기대치로 평균화하는 것이지, 치명이 없는 스킬을
    # 기대 치명률만큼 부풀리는 게 아닙니다. 관측 치명 비중/적중률이 0이면 그 스킬은
    # 치명이 안 터지는(또는 치명 대상이 아닌) 스킬(예: 기본 공격)이므로 보정하지 않습니다.
    # (수백 히트에서 치명 0%는 '운 나쁨'이 아니라 '해당 없음'입니다.)
    share = _finite_float(observed_crit_share, None)
    rate = _finite_float(observed_crit_rate, None)
    if (share is None or share <= 0.05) and (rate is None or rate <= 0.05):
        return float(damage)

    # 극단적 왜곡 방지: 치명운 보정 배율은 ±60% 밴드로 제한합니다.
    # 정상 스킬의 실제 치명 운 편차는 이 안에 들어오고, 손상된 행만 걸러집니다.
    def _clamp_adj(adjusted: float) -> float:
        lo, hi = float(damage) * 0.6, float(damage) * 1.6
        return max(lo, min(hi, adjusted))

    if share is not None and share > 0.05:
        share_rate = max(0.0, min(100.0, share)) / 100.0
        # D = noncrit_base + crit_base * C, crit_damage_share = crit_base*C / D
        # base_total = D * ((1 - S) + S / C)
        base_equivalent = float(damage) * ((1.0 - share_rate) + share_rate / crit_damage_mul)
        return _clamp_adj(base_equivalent * expected_mult)

    obs_mult = _expected_crit_multiplier(rate, crit_damage_percent)
    if obs_mult <= 0:
        return float(damage)
    return _clamp_adj(float(damage) * expected_mult / obs_mult)


def _skill_role(share_rate: float | None, name: str) -> str:
    n = str(name or "")
    share = float(share_rate or 0.0)
    # '기본 공격'은 정확히 공용 기본공격일 때만 보조/기타로 분류합니다.
    # '수라결 기본 공격'(브레이커 수라의 길 전용)처럼 클래스 고유 기본공격은 일반 스킬로 취급합니다.
    _is_generic_basic = n.strip() in ("기본 공격", "기본공격")
    if "기타" in n or "스킬룬" in n or _is_generic_basic:
        return "보조/기타"
    if share >= 8:
        return "주력기"
    if share >= 3:
        return "보조딜"
    return "보조/기타"


def _direction_type_from_attack_type(text: Any) -> str:
    t = str(text or "")
    if "백" in t:
        return "백어택"
    if "헤드" in t:
        return "헤드어택"
    return "없음"



def _collect_text_from_obj(obj: Any, limit: int = 250_000) -> str:
    """API bundle/summary 안의 텍스트를 얕게 모아 각인 채용 여부를 감지합니다."""
    parts: list[str] = []

    def walk(x: Any) -> None:
        if sum(len(p) for p in parts) > limit:
            return
        if isinstance(x, str):
            if x:
                parts.append(x)
        elif isinstance(x, dict):
            for k, v in x.items():
                if isinstance(k, str):
                    parts.append(k)
                walk(v)
        elif isinstance(x, (list, tuple)):
            for v in x:
                walk(v)
        elif isinstance(x, pd.DataFrame):
            try:
                parts.append(x.to_csv(index=False))
            except Exception:
                pass
        elif x is not None and not isinstance(x, (int, float, bool)):
            # v48: 각인/보석/아크패시브 툴팁은 2000자를 넘는 경우가 많습니다.
            # 방향 각인 감지에는 긴 문자열도 필요하므로 안전하게 일부만 잘라 수집합니다.
            txt = str(x)
            if txt:
                parts.append(txt[:20000])

    walk(obj)
    return "\n".join(parts)


def _detect_directional_engravings() -> list[dict[str, Any]]:
    """API에서 기습의 대가/결투의 대가 채용 여부와 조건부 방향 피해 보너스를 추정합니다.

    v50:
    - `기술의 대가` 오인식 alias 제거. 화면에는 항상 `기습의 대가` / `결투의 대가`로 표시합니다.
    - 각인 Tooltip의 `백어택/헤드어택 성공 시 ... 추가로 N%` 조건부 보너스를 우선 읽습니다.
    - 전설/유물 등급에 따른 상시 피해 수치는 별도로 보존하고, 효율 계산은 조건부 방향 보너스 기준으로 합니다.
    """
    summary = st.session_state.get("api_summary") or {}
    bundle = st.session_state.get("api_bundle") or {}
    text_blob = _collect_text_from_obj({"summary": summary, "bundle": bundle})

    def as_df(x: Any) -> pd.DataFrame:
        return x if isinstance(x, pd.DataFrame) else pd.DataFrame()

    source_frames: list[tuple[str, pd.DataFrame]] = []
    for key, label in [
        ("damage_sources", "피해군 감지 출처"),
        ("base_damage_sources", "기준 계산 출처"),
        ("arkgrid_damage_sources", "아크그리드 계산 출처"),
        ("crit_sources", "치명/치피 출처"),
        ("engravings", "각인"),
    ]:
        df = as_df(summary.get(key))
        if not df.empty:
            source_frames.append((label, df))

    def frame_text(df: pd.DataFrame) -> str:
        try:
            return df.to_csv(index=False)
        except Exception:
            return str(df)

    frame_blob = "\n".join(frame_text(df) for _, df in source_frames)
    all_text = f"{text_blob}\n{frame_blob}"

    def has_exact(alias: str) -> bool:
        return bool(alias and alias in all_text)

    def _numbers_from_columns(df: pd.DataFrame, row_mask: pd.Series, columns: list[str]) -> list[float]:
        vals: list[float] = []
        m = df[row_mask]
        if m.empty:
            return vals
        for col in columns:
            if col not in m.columns:
                continue
            for v in pd.to_numeric(m[col], errors="coerce").dropna().tolist():
                try:
                    fv = float(v)
                    if fv > 0:
                        vals.append(fv)
                except Exception:
                    pass
        return vals

    def read_bonus(display_name: str, attack_word: str) -> tuple[float, float, str, str]:
        conditional_bonus: float | None = None
        global_bonus: float | None = None
        source = ""
        grade_text = _extract_effect_grade_from_sources(display_name)

        # 1) 정리된 DataFrame에서 먼저 탐색합니다.
        for label, df in source_frames:
            if df.empty:
                continue
            row_mask = pd.Series([False] * len(df), index=df.index)
            for col in df.columns:
                try:
                    row_mask = row_mask | df[col].astype(str).str.contains(re.escape(display_name), na=False)
                except Exception:
                    continue
            if not row_mask.any():
                continue
            source = source or label
            try:
                row_text = "\n".join(df[row_mask].astype(str).agg(" ".join, axis=1).tolist())
                grade_text = grade_text or _extract_grade_stage(row_text, max_stage=4)
            except Exception:
                pass
            cond_vals = _numbers_from_columns(df, row_mask, [
                "조건부 방향 피해(%)", "방향성 피해(조건부)(%)", "백어택 피해(%)", "헤드어택 피해(%)", "방향성 피해(%)",
            ])
            # 조건부 보너스는 보통 15/25처럼 큰 값입니다. 상시 4.8/7.6이 섞이면 더 큰 값을 우선합니다.
            if cond_vals:
                conditional_bonus = max(cond_vals)
            glob_vals = _numbers_from_columns(df, row_mask, [
                "전역 피해(%)", "적에게 주는 피해(상시)(%)", "적에게 주는 피해(%)", "피해량 증가(%)", "피해 증가(%)",
            ])
            if glob_vals:
                # 전역 피해는 조건부보다 작고, 여러 값이 있을 땐 장착 등급의 최종값에 가까운 최대값을 사용합니다.
                global_bonus = max(glob_vals)
            if conditional_bonus is not None:
                break

        # 2) Tooltip 원문에서 조건부 방향 보너스를 직접 추출합니다.
        windows: list[str] = []
        for mm in re.finditer(re.escape(display_name), all_text):
            windows.append(all_text[max(0, mm.start() - 400): mm.start() + 1800])
        if not windows and has_exact(display_name):
            windows.append(all_text)

        for window in windows:
            if not grade_text:
                grade_text = grade_text or _extract_grade_stage(window, max_stage=4)
            if conditional_bonus is None:
                patterns = [
                    rf"{attack_word}\s*성공\s*시[^%]{{0,90}}?(?:피해량|피해)[^%]{{0,40}}?(?:추가로\s*)?(\d+(?:\.\d+)?)\s*%",
                    rf"{attack_word}[^%]{{0,90}}?(?:피해량|피해)[^%]{{0,40}}?(?:추가로\s*)?(\d+(?:\.\d+)?)\s*%",
                    rf"방향성\s*공격[^%]{{0,90}}?(?:피해량|피해)[^%]{{0,40}}?(\d+(?:\.\d+)?)\s*%",
                ]
                found: list[float] = []
                for pat in patterns:
                    for m in re.finditer(pat, window, re.S):
                        try:
                            found.append(float(m.group(1)))
                        except Exception:
                            pass
                if found:
                    conditional_bonus = max(found)
                    source = source or "API tooltip"
            if global_bonus is None:
                # `백어택 성공 시` 뒤의 조건부 문장은 제외하고, 각인 상시 피해 증가만 읽습니다.
                glob_window = re.split(rf"{attack_word}\s*성공\s*시", window, maxsplit=1)[0]
                found_g: list[float] = []
                for m in re.finditer(r"적에게\s*주는\s*피해(?:량)?[^%]{0,40}?(\d+(?:\.\d+)?)\s*%", glob_window, re.S):
                    try:
                        found_g.append(float(m.group(1)))
                    except Exception:
                        pass
                if found_g:
                    global_bonus = max(found_g)
                    source = source or "API tooltip"
            if conditional_bonus is not None and global_bonus is not None:
                break

        return float(conditional_bonus or 0.0), float(global_bonus or 0.0), source or "API 텍스트 감지", grade_text

    specs = [
        {"display": "기습의 대가", "attack": "백어택", "attack_word": "백어택"},
        {"display": "결투의 대가", "attack": "헤드어택", "attack_word": "헤드어택"},
    ]

    out: list[dict[str, Any]] = []
    for spec in specs:
        if not has_exact(spec["display"]):
            continue
        b, g, src, grade = read_bonus(spec["display"], spec["attack_word"])
        out.append({
            "각인": spec["display"],
            "공격타입": spec["attack"],
            "조건부 방향 피해(%)": b,
            "전역 피해(%)": g,
            "각인 등급": grade,
            "감지 출처": src,
        })
    return out

def _directional_gain_from_damage_share(share_percent: float | None, bonus_percent: float | None) -> tuple[float | None, float | None]:
    """방향 조건 피해 비중과 각인 방향 보너스로 실질 딜증/효율을 계산합니다.

    share_percent는 보너스가 적용된 뒤의 총 피해 중 조건 충족 피해 비중입니다.
    따라서 조건부 보너스 B를 제거한 피해를 역산해 실제 증가율을 계산합니다.
    """
    if share_percent is None or bonus_percent is None or bonus_percent <= 0:
        return None, None
    s = max(0.0, min(100.0, float(share_percent))) / 100.0
    b = max(0.0, float(bonus_percent)) / 100.0
    without_bonus_ratio = (1.0 - s) + s / (1.0 + b)
    if without_bonus_ratio <= 0:
        return None, None
    gain = (1.0 / without_bonus_ratio) - 1.0
    efficiency = gain / b if b > 0 else None
    return gain, efficiency


def _build_directional_engraving_efficiency(result_df: pd.DataFrame, engravings: list[dict[str, Any]]) -> pd.DataFrame:
    if result_df is None or result_df.empty or not engravings:
        return pd.DataFrame()
    rows: list[dict[str, Any]] = []
    total_damage = pd.to_numeric(result_df.get("실전 관측"), errors="coerce").fillna(0).sum()
    for eng in engravings:
        atk = eng.get("공격타입")
        share_col = "백어택 비중" if atk == "백어택" else "헤드어택 비중"
        sub = result_df[result_df.get("공격타입").eq(atk)].copy() if "공격타입" in result_df.columns else pd.DataFrame()
        eligible_damage = pd.to_numeric(sub.get("실전 관측"), errors="coerce").fillna(0).sum() if not sub.empty else 0.0
        if not sub.empty and share_col in sub.columns and "실전 관측" in sub.columns:
            sh = pd.to_numeric(sub[share_col], errors="coerce")
            dmg = pd.to_numeric(sub["실전 관측"], errors="coerce").fillna(0)
            valid = sh.notna() & (dmg > 0)
            target_skill_share = float((sh[valid] * dmg[valid]).sum() / dmg[valid].sum()) if valid.any() else None
            total_condition_share = float((sh[valid] / 100.0 * dmg[valid]).sum() / total_damage * 100.0) if valid.any() and total_damage > 0 else None
        else:
            target_skill_share = None
            total_condition_share = None
        bonus = _finite_float(eng.get("조건부 방향 피해(%)"), 0.0) or 0.0
        target_gain, target_bonus_eff = _directional_gain_from_damage_share(target_skill_share, bonus)
        actual_gain, overall_bonus_eff = _directional_gain_from_damage_share(total_condition_share, bonus)
        rows.append({
            "각인": eng.get("각인"),
            "각인 등급": eng.get("각인 등급"),
            "조건": atk,
            "대상 스킬 기준 활용률": target_skill_share,
            "대상 스킬 기준 각인 효율": target_bonus_eff,
            "전체 피해 기준 조건 비중": total_condition_share,
            "조건부 방향 피해(%)": bonus,
            "전역 피해(%)": eng.get("전역 피해(%)"),
            "대상 스킬 기준 실질 딜증": target_gain,
            "각인 실질 딜증": actual_gain,
            "조건부 보너스 활용 효율": overall_bonus_eff,
            "대상 스킬 피해": eligible_damage,
            "감지 출처": eng.get("감지 출처"),
        })
    return pd.DataFrame(rows)

# v137: 트라이포드에 따라 스킬 구조 방향이 바뀌는 목록.
# data/loa_skill_type_change_list.xlsx (클래스/스킬명/트라이포드/변경되는 스킬 타입)를 읽어서,
# 해당 트라이포드를 '선택'한 스킬은 방향 타입을 덮어씁니다.
# 파일을 못 읽으면 아래 임베드 기본값을 폴백으로 씁니다(파일이 진실의 원천, 상수는 안전망).
_SKILL_TYPE_OVERRIDE_FALLBACK: list[tuple[str, str, str]] = [
    ("방패 밀치기", "충격 방패", "방향성 제거"), ("배쉬", "충격 방패", "방향성 제거"),
    ("방패 격동", "충격 방패", "방향성 제거"), ("휩쓸기", "도전자", "방향성 제거"),
    ("유성 낙하", "도전자", "방향성 제거"), ("연속전격", "권풍", "방향성 제거"),
    ("징벌의 파도", "도전자", "방향성 제거"), ("진 파공권", "도전자", "방향성 제거"),
    ("청월난무", "지면 강타", "방향성 제거"), ("파천섬광", "도전자", "방향성 제거"),
    ("리벤지 블로우", "전방위 타격", "방향성 제거"), ("블레이즈 스윕", "전방위 타격", "방향성 제거"),
    ("섬광 베기", "집행자", "방향성 제거"), ("집행자의 검", "집행자", "방향성 제거"),
    ("정의 집행", "집행자", "방향성 제거"), ("신성검", "집행자", "방향성 제거"),
    ("퀵 샷", "사면초가", "방향성 제거"), ("잔혹한 추적자", "사면초가", "방향성 제거"),
    ("이퀄리브리엄", "사면초가", "방향성 제거"), ("데스파이어", "작열 난사", "방향성 제거"),
    ("심판의 날", "심판의 시간", "방향성 제거"), ("샷건 연사", "사면초가", "방향성 제거"),
    ("절멸의 탄환", "사면초가", "방향성 제거"), ("마탄의 사수", "사면초가", "방향성 제거"),
    ("레인 오브 불릿", "사면초가", "방향성 제거"), ("선풍용류각", "용아", "방향성 제거"),
    ("소울 시너스", "부서진 나침반", "방향성 제거"), ("데스 위핑", "부서진 나침반", "방향성 제거"),
    ("데굴방아", "환영 돌진", "방향성 제거"), ("단죄의 쇄도", "폭풍 쇄도", "방향성 제거"),
    ("단죄의 성흔", "순례자", "방향성 제거"), ("단죄의 속삭임", "순례자", "방향성 제거"),
    ("계시의 집행", "순례자", "방향성 제거"), ("렌딩 피니셔", "정면 승부", "헤드 어택으로 변경"),
]


@st.cache_data(show_spinner=False)
def _load_skill_type_overrides() -> dict[str, list[dict[str, str]]]:
    """트라이포드 기반 스킬 타입 변경 목록을 {스킬명(normalized): [{트라이포드, 변경}]}로 로드."""
    rows: list[tuple[str, str, str]] = []
    try:
        import os
        path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "loa_skill_type_change_list.xlsx")
        xdf = pd.read_excel(path)
        for _, r in xdf.iterrows():
            sk = str(r.get("스킬명") or "").strip()
            tp = str(r.get("트라이포드") or "").strip()
            ch = str(r.get("변경되는 스킬 타입") or "").strip()
            if sk and tp and ch:
                rows.append((sk, tp, ch))
    except Exception:
        rows = []
    if not rows:
        rows = list(_SKILL_TYPE_OVERRIDE_FALLBACK)
    out: dict[str, list[dict[str, str]]] = {}
    for sk, tp, ch in rows:
        out.setdefault(normalize_name(sk), []).append({"트라이포드": tp, "변경": ch})
    return out


_SKILL_DIR_OVERRIDE_CACHE: dict[str, dict[str, str]] | None = None


def _load_skill_direction_overrides() -> dict[str, dict[str, str]]:
    """스킬 방향(백어택/헤드어택/무방향) 수동 보정 마스터 표를 로드합니다.

    유지보수를 위해 방향 보정 데이터를 한 파일로 통합했습니다:
      data/skill_direction_overrides.xlsx  (시트: '방향보정', 열: 스킬명 / 공격타입 / 구분 / 비고)
    - 아덴/아이덴티티 스킬(API 목록에 없음) + 일반 스킬 방향 누락분(허리케인소드 등) +
      무방향 강제(와일드스톰프 등)을 모두 여기서 관리합니다.
    - 신캐릭터가 나오면 이 엑셀에 행만 추가하면 됩니다.

    반환: {스킬명(normalized): {"raw": '백어택'|'헤드어택'|'백/헤드'|'무방향', "name": 표시명, "구분": ...}}
          '공격 타입 없음'(회복/보호막 등)은 방향 계산 대상이 아니라 제외합니다.
    구파일 fallback: 마스터 파일이 없으면 data/lostark_identity_skills.xlsx를 읽습니다.
    """
    global _SKILL_DIR_OVERRIDE_CACHE
    if _SKILL_DIR_OVERRIDE_CACHE is not None:
        return _SKILL_DIR_OVERRIDE_CACHE

    def _norm_atype(at: str) -> str:
        # 실제 방향 주입은 백/헤드만 합니다. 무방향/공격 타입 없음은 주입하지 않고
        # 전투분석기 관측값 또는 API 방향을 그대로 따르게 둡니다(무방향 강제 안 함).
        at = str(at or "").strip()
        if "백" in at and "헤드" in at:
            return "백/헤드"
        if "백" in at:
            return "백어택"
        if "헤드" in at:
            return "헤드어택"
        return ""  # 무방향/공격 타입 없음 → 주입 대상 아님

    out: dict[str, dict[str, str]] = {}
    import os
    base = os.path.dirname(os.path.abspath(__file__))
    master = os.path.join(base, "data", "skill_direction_overrides.xlsx")
    legacy = os.path.join(base, "data", "lostark_identity_skills.xlsx")
    try:
        if os.path.exists(master):
            xdf = pd.read_excel(master, sheet_name="방향보정")
            name_col, type_col = "스킬명", "공격타입"
        else:
            xdf = pd.read_excel(legacy, sheet_name="아이덴티티스킬")
            name_col, type_col = "아이덴티티 스킬", "공격 타입"
        for _, r in xdf.iterrows():
            sk = str(r.get(name_col) or "").strip()
            if not sk:
                continue
            raw = _norm_atype(r.get(type_col))
            if raw == "":
                continue  # 공격 타입 없음 → 방향 계산 제외
            out[normalize_name(sk)] = {"raw": raw, "name": sk, "구분": str(r.get("구분") or "").strip()}
    except Exception:
        out = {}
    _SKILL_DIR_OVERRIDE_CACHE = out
    return out


def _selected_tripods_by_skill(bundle: Any) -> dict[str, set[str]]:
    """빌드(combat_skills)에서 스킬별 '선택된' 트라이포드 이름 집합을 뽑습니다."""
    out: dict[str, set[str]] = {}
    data = (bundle.get("combat_skills") or {}).get("data") if isinstance(bundle, dict) else None
    if not isinstance(data, list):
        return out
    for s in data:
        if not isinstance(s, dict):
            continue
        name = str(s.get("Name") or s.get("name") or "").strip()
        if not name:
            continue
        trips = s.get("Tripods") or s.get("tripods") or []
        sel: set[str] = set()
        if isinstance(trips, list):
            for t in trips:
                if not isinstance(t, dict):
                    continue
                is_sel = t.get("IsSelected", t.get("isSelected", t.get("selected", t.get("Selected"))))
                if is_sel:
                    tn = str(t.get("Name") or t.get("name") or "").strip()
                    if tn:
                        sel.add(tn)
        out[normalize_name(name)] = sel
    return out


def _apply_tripod_type_override(name: str, attack_type_raw: str, overrides: dict, selected_tripods: dict) -> str:
    """이 스킬에 트라이포드 기반 방향 변경이 적용되면 공격타입 원문을 덮어씁니다.

    - '방향성 제거' → "" (무방향/NONE)
    - '헤드 어택으로 변경' → "헤드어택"
    - '백 어택으로 변경' → "백어택"
    """
    key = normalize_name(name)
    rules = overrides.get(key)
    if not rules:
        return attack_type_raw
    sel = selected_tripods.get(key, set())
    for rule in rules:
        tp = rule.get("트라이포드", "")
        ch = rule.get("변경", "")
        # 선택 트라이포드 정보가 있으면 정확히 매칭, 없으면(빌드 트포 데이터 부재) 규칙 미적용.
        if tp and any(tp in s or s in tp for s in sel):
            if "방향성 제거" in ch:
                return ""
            if "헤드" in ch:
                return "헤드어택"
            if "백" in ch:
                return "백어택"
    return attack_type_raw


def _build_api_skill_lookup() -> dict[str, dict[str, Any]]:
    """API 최종 계산표에서 전분 보정에 필요한 스킬별 치명/치피/공격타입을 가져옵니다."""
    summary = st.session_state.get("api_summary") or {}
    bundle = st.session_state.get("api_bundle") or {}
    # 렌더 1회 안에서 스킬 정보/디버그 뷰가 이 함수를 여러 번 부릅니다.
    # 같은 summary/bundle이면 결과가 동일하므로 세션에 캐시해 재계산을 피합니다.
    _tok = (id(summary), id(bundle))
    _c = st.session_state.get("_api_skill_lookup_cache_v145")
    if isinstance(_c, dict) and _c.get("tok") == _tok:
        return _c.get("lookup") or {}
    df = summary.get("skill_crit_estimates")
    if not isinstance(df, pd.DataFrame) or df.empty:
        df = summary.get("lostbuilds_base_skill_estimates")
    lookup: dict[str, dict[str, Any]] = {}
    if not isinstance(df, pd.DataFrame) or df.empty or "스킬명" not in df.columns:
        return lookup
    type_overrides = _load_skill_type_overrides()
    selected_tripods = _selected_tripods_by_skill(bundle)
    for _, row in df.iterrows():
        name = str(row.get("스킬명") or "").strip()
        if not name:
            continue
        attack_type_raw = str(row.get("공격타입") or "")
        # v137: 트라이포드로 방향성이 바뀌는 스킬은 여기서 덮어씁니다(예: 파천섬광+도전자 → 무방향).
        attack_type_raw = _apply_tripod_type_override(name, attack_type_raw, type_overrides, selected_tripods)
        attack_type = _direction_type_from_attack_type(attack_type_raw)
        crit_rate = _first_number(row, [
            "예상 치명 확률(백어택 기준)(%)",
            "예상 치명 확률(조건부 포함)(%)",
            "기준 치명(백어택)(%)",
            "치명 증가 합계(조건부)(%)",
        ], None)
        crit_damage = _first_number(row, [
            "예상 치피(조건부 포함)(%)",
            "기준 치피(%)",
            "예상 치피(상시)(%)",
        ], 200.0)
        evolution_damage = _first_number(row, ["진화형 피해(조건부)(%)", "기준 진화형 피해(%)"], 0.0)
        final_multiplier = _first_number(row, ["예상 최종 배율(조건부)", "기준 최종 배율"], 1.0)
        lookup[normalize_name(name)] = {
            "스킬명": name,
            "공격타입": attack_type,
            "공격타입_원본": attack_type_raw,
            "api_crit_rate": crit_rate,
            "api_crit_damage": crit_damage,
            "api_evolution_damage": evolution_damage,
            "api_final_multiplier": final_multiplier,
            "스킬역할": _skill_role(None, name),
        }

    # 스킬 방향 수동 보정(마스터 표): 아덴/아이덴티티 스킬(API에 없음) 방향 주입 +
    # 일반 스킬 방향 누락분 교정 + 무방향 강제. 치명 데이터는 유지합니다.
    for nkey, info in _load_skill_direction_overrides().items():
        raw_dir = info.get("raw") or ""
        if not raw_dir:
            continue
        directional = ("백" in raw_dir) or ("헤드" in raw_dir)
        if nkey in lookup:
            # 이미 API 스킬로 잡혔으면 방향만 표 기준으로 교정(치명/치피는 그대로 둠).
            # 무방향 항목이면 방향을 없음으로 고정합니다.
            lookup[nkey]["공격타입_원본"] = raw_dir
            lookup[nkey]["공격타입"] = _direction_type_from_attack_type(raw_dir)
        elif directional:
            # API에 없는 방향성 스킬(아덴 등) → 방향만 주입(치명은 관측값 폴백).
            _iname = info.get("name") or nkey
            lookup[nkey] = {
                "스킬명": _iname,
                "공격타입": _direction_type_from_attack_type(raw_dir),
                "공격타입_원본": raw_dir,
                "api_crit_rate": None,
                "api_crit_damage": 200.0,
                "api_evolution_damage": 0.0,
                "api_final_multiplier": 1.0,
                "스킬역할": _skill_role(None, _iname),
                "_방향주입": True,
            }
        # 무방향인데 lookup에 없으면: 미매칭 스킬은 어차피 NONE이라 별도 처리 불필요.
    st.session_state["_api_skill_lookup_cache_v145"] = {"tok": _tok, "lookup": lookup}
    return lookup


def _aggregate_battle_rows(df: pd.DataFrame) -> dict[str, dict[str, Any]]:
    norm = normalize_battle_df(df)
    result: dict[str, dict[str, Any]] = {}
    if norm is None or norm.empty:
        return result
    rate_cols = ["back_attack_rate", "back_attack_share", "head_attack_rate", "head_attack_share", "crit_rate", "crit_share", "cooldown_rate", "share_rate"]
    for _, row in norm.iterrows():
        name = str(row.get("name") or "").strip()
        key = normalize_name(name)
        damage = _finite_float(row.get("damage"), 0.0) or 0.0
        if not key or damage <= 0:
            continue
        item = result.setdefault(key, {
            "name": name,
            "damage": 0.0,
            "dps": 0.0,
            "casts": 0,
            "_rate_weight": 0.0,
            **{c: None for c in rate_cols},
        })
        old_damage = float(item.get("damage") or 0.0)
        new_damage = old_damage + damage
        item["damage"] = new_damage
        dps = _finite_float(row.get("dps"), 0.0) or 0.0
        item["dps"] = float(item.get("dps") or 0.0) + dps
        casts = _finite_float(row.get("casts"), None)
        if casts is not None:
            item["casts"] = int(item.get("casts") or 0) + int(casts)
        for c in rate_cols:
            v = _finite_float(row.get(c), None)
            if v is None:
                continue
            prev = item.get(c)
            if prev is None or old_damage <= 0:
                item[c] = v
            else:
                item[c] = (float(prev) * old_damage + v * damage) / new_damage
    for item in result.values():
        item.pop("_rate_weight", None)
    return result


def _is_rune_or_misc(name: str) -> bool:
    n = str(name or "")
    # '기본 공격'은 정확히 공용 기본공격일 때만 룬/기타로 봅니다.
    # '수라결 기본 공격'(브레이커 수라의 길 전용)은 실제 스킬이므로 여기서 제외합니다.
    if n.strip() in ("기본 공격", "기본공격"):
        return True
    return any(x in n for x in ["기타", "스킬룬", "맥스웰", "라그나"])


def _detect_battle_data_anomalies(
    real: dict[str, dict[str, Any]],
    dummy: dict[str, dict[str, Any]],
    real_elapsed: float | None = None,
    dummy_elapsed: float | None = None,
) -> list[dict[str, Any]]:
    """전투분석기 입력값(주로 OCR 결과)에서 '값이 꼬였을 가능성이 높은' 행을 감지합니다.

    값은 절대 바꾸지 않습니다. 내부적으로 모순되거나(예: 치명 적중률 0%인데 비중이 큼)
    사용횟수가 허수 대비 비정상적으로 낮은 행을 찾아 사용자가 편집 표에서 직접
    고칠 수 있도록 경고 목록만 반환합니다.
    """
    warnings: list[dict[str, Any]] = []

    def _contradiction(rate: float | None, share: float | None) -> bool:
        # 한쪽은 사실상 0인데 다른 쪽은 큰 값 → 물리적으로 불가능(OCR 오독 신호).
        if rate is None or share is None:
            return False
        return (rate <= 1.0 and share >= 20.0) or (share <= 1.0 and rate >= 20.0)

    # v143 서버 교차검증용: 요약(종합정보) 총피해량을 읽어둡니다.
    def _summary_total(kind: str) -> float | None:
        m = st.session_state.get(f"{kind}_meta") or {}
        return _finite_float(m.get("total_damage"), None)
    real_total = _summary_total("real")
    dummy_total = _summary_total("dummy")

    _by_name: dict[str, list[str]] = {}
    def _add(name: str, reason: str) -> None:
        if not name:
            return
        _by_name.setdefault(name, [])
        if reason not in _by_name[name]:
            _by_name[name].append(reason)

    for key, r in real.items():
        name = str(r.get("name") or "").strip()
        if not name:
            continue
        d = dummy.get(key, {})

        cr = _finite_float(r.get("crit_rate"), None)
        cs = _finite_float(r.get("crit_share"), None)
        if _contradiction(cr, cs):
            _add(name, f"치명타 적중률({cr:.0f}%)과 비중({cs:.0f}%)이 모순됩니다")

        br = _finite_float(r.get("back_attack_rate"), None)
        bs = _finite_float(r.get("back_attack_share"), None)
        if _contradiction(br, bs):
            _add(name, f"백어택 적중률({br:.0f}%)과 비중({bs:.0f}%)이 모순됩니다")

        hr = _finite_float(r.get("head_attack_rate"), None)
        hs = _finite_float(r.get("head_attack_share"), None)
        if _contradiction(hr, hs):
            _add(name, f"헤드어택 적중률({hr:.0f}%)과 비중({hs:.0f}%)이 모순됩니다")

        rc = _finite_float(r.get("casts"), None)
        dc = _finite_float(d.get("casts"), None)
        # 실전 사용횟수가 1 이하인데 허수에서는 여러 번 쓴 스킬 → OCR이 앞자리를 놓쳤을 가능성.
        # 단, 가디언 토벌처럼 실전이 아주 짧으면 실제로 1회만 쓴 게 정상이므로,
        # 전투시간 비율로 '기대 사용횟수'를 환산해 그보다 크게 모자랄 때만 경고합니다.
        if rc is not None and rc <= 1 and dc is not None and dc >= 5:
            _tr = (float(real_elapsed) / float(dummy_elapsed)) if (real_elapsed and dummy_elapsed and dummy_elapsed > 0) else 1.0
            _expected_rc = dc * _tr
            if _expected_rc >= 5:   # 짧은 전투를 감안해도 여러 번 나왔어야 하는 경우만 오독 의심
                _add(name, f"실전 사용횟수({rc:.0f})가 허수({dc:.0f}) 대비 비정상적으로 낮습니다")

    # v143 서버 교차검증: 스킬 1개 피해가 '요약 총피해량'을 넘으면 자릿수 오독이 확실합니다
    # (예: 허리케인 2,083억을 72,083억으로 읽음). 합계가 총피해량을 크게 넘어도 경고합니다.
    for side_kind, side, tot in (("실전", real, real_total), ("허수", dummy, dummy_total)):
        if not tot or tot <= 0:
            continue
        ssum = 0.0
        for v in side.values():
            nm = str(v.get("name") or "").strip()
            dv = _finite_float(v.get("damage"), 0.0) or 0.0
            ssum += dv
            if nm and dv > tot * 1.05:
                _add(nm, f"{side_kind} 피해량({format_korean_number(dv)})이 총피해량({format_korean_number(tot)})보다 큼 — 자릿수 OCR 오독 의심")
        if ssum > tot * 1.25:
            _add(f"[{side_kind} 전체]", f"스킬 피해 합({format_korean_number(ssum)})이 총피해량({format_korean_number(tot)})을 크게 초과 — 일부 스킬 자릿수 오독 의심")

    for name, reasons in _by_name.items():
        warnings.append({"스킬명": name, "사유": reasons})
    return warnings




# ==============================
# v54: 특수 전투 효과 감지 / DPS 범위 보조표
# ==============================
def _text_windows(blob: str, keyword: str, left: int = 500, right: int = 1800) -> list[str]:
    if not blob or not keyword:
        return []
    wins: list[str] = []
    for m in re.finditer(re.escape(keyword), blob):
        wins.append(blob[max(0, m.start() - left): m.end() + right])
    return wins


def _max_percent_near(text: str, keywords: list[str], window: int = 90) -> float | None:
    vals: list[float] = []
    clean = re.sub(r"<[^>]+>", " ", str(text or ""))
    clean = re.sub(r"\\s+", " ", clean)
    for kw in keywords:
        k = re.escape(kw)
        pats = [
            rf"{k}[^0-9+\-]{{0,{window}}}([+\-]?\d+(?:\.\d+)?)\s*%",
            rf"([+\-]?\d+(?:\.\d+)?)\s*%[^가-힣A-Za-z0-9]{{0,{window}}}{k}",
        ]
        for pat in pats:
            for m in re.finditer(pat, clean, re.S):
                try:
                    vals.append(float(m.group(1)))
                except Exception:
                    pass
    if not vals:
        return None
    # 각인 Tooltip 안에는 낮은 단계/높은 단계 수치가 같이 잡힐 수 있으므로 최종 후보로 최대값을 우선 사용합니다.
    return max(vals)


def _extract_grade_stage(text: str, max_stage: int | None = 4) -> str:
    clean = re.sub(r"<[^>]+>", " ", str(text or ""))
    clean = re.sub(r"\s+", " ", clean).strip()
    if not clean:
        return ""

    grade_re = r"(고대|유물|전설|영웅|희귀|고급)"
    candidates: list[tuple[int, str, int]] = []

    # 예: "유물 4단계", "유물 4", "유물 Lv.4".
    # API/Tooltip 원문에는 다른 레벨/품질 숫자가 많이 섞이므로 각인 단계는 1~4 범위만 우선 인정합니다.
    for m in re.finditer(grade_re + r"\s*(?:각인|효과|등급)?[^0-9]{0,18}(\d+)\s*(?:단계|Lv\.?|레벨)?", clean):
        try:
            stage = int(m.group(2))
        except Exception:
            continue
        if stage <= 0:
            continue
        if max_stage is not None and stage > max_stage:
            continue
        # 텍스트 뒤쪽의 "유물 4"가 실제 장착 각인 단계로 붙는 경우가 많아서 뒤쪽 후보를 우선합니다.
        candidates.append((m.start(), m.group(1), stage))
    if candidates:
        _, grade, stage = candidates[-1]
        return f"{grade} {stage}단계"

    m = re.search(grade_re, clean)
    return m.group(1) if m else ""


def _extract_effect_grade_from_sources(effect_name: str) -> str:
    """정리된 피해군/각인 출처표에서 각인 등급과 단계를 우선 읽습니다.

    Tooltip 전체를 훑으면 장비 품질, 스톤 Lv, 다른 각인 단계가 섞일 수 있으므로
    이름이 정확히 일치하는 출처 행의 설명을 먼저 봅니다.
    "아드레날린 ... 유물 4" 같은 문장을 "유물 4단계"로 정규화합니다.
    """
    summary = st.session_state.get("api_summary") or {}
    target = normalize_name(effect_name)
    if not target:
        return ""
    for key in ["damage_sources", "base_damage_sources", "crit_sources", "arkgrid_damage_sources", "engravings"]:
        df = summary.get(key)
        if not isinstance(df, pd.DataFrame) or df.empty:
            continue
        for _, row in df.iterrows():
            row_name = str(row.get("이름") or row.get("각인") or row.get("효과") or "")
            if normalize_name(row_name) != target:
                continue
            parts = []
            for col in ["각인/효과 등급", "레벨/설명", "설명", "비고", "감지 출처"]:
                if col in row.index and row.get(col) is not None:
                    parts.append(str(row.get(col)))
            grade = _extract_grade_stage(" ".join(parts), max_stage=4)
            if grade:
                return grade
    return ""


def _engraving_own_text(name: str) -> str:
    """특정 각인(이름)의 '자기 자신' Description/Tooltip 텍스트만 반환합니다.

    blob 전체 윈도우(_text_windows)를 쓰면 인접한 다른 각인(예: 슬레이어 포식자 +30% 치적)의
    수치가 섞여 아드레날린 치명타 적중률이 30%로 잘못 잡히는 문제가 있어, 해당 각인 본인의
    API Description/Tooltip 으로 검색 범위를 한정합니다.
    """
    bundle = st.session_state.get("api_bundle") or {}
    eng_wrap = bundle.get("engravings") if isinstance(bundle, dict) else None
    data = None
    if isinstance(eng_wrap, dict):
        data = eng_wrap.get("data", eng_wrap)
    if not isinstance(data, dict):
        return ""
    parts: list[str] = []
    for key in ("Effects", "ArkPassiveEffects"):
        for e in (data.get(key) or []):
            if not isinstance(e, dict):
                continue
            nm = str(e.get("Name") or "")
            if name and (name in nm or nm in name):
                tip = e.get("Tooltip")
                tip_s = " ".join(tip) if isinstance(tip, list) else str(tip or "")
                parts.append(f"{e.get('Description') or ''} {tip_s}")
    return " ".join(p for p in parts if p.strip())


def _detect_special_combat_effects() -> list[dict[str, Any]]:
    """아드레날린/예리한 둔기/뭉툭한 가시 같은 특수 효과를 API 텍스트에서 검수용으로 감지합니다.

    계산의 1차 근거는 기존 피해군/치명 출처표이고, 여기서는 결과 화면에서 검수하고
    예측 DPS 범위에 참고할 수 있도록 정리된 값을 반환합니다.
    """
    summary = st.session_state.get("api_summary") or {}
    bundle = st.session_state.get("api_bundle") or {}
    blob = _collect_text_from_obj({"summary": summary, "bundle": bundle})

    rows: list[dict[str, Any]] = []

    def row_for(name: str) -> dict[str, Any] | None:
        # 1순위: 해당 각인 본인의 Description/Tooltip(주변 각인 수치 오염 방지).
        text = _engraving_own_text(name)
        if not text:
            wins = _text_windows(blob, name)
            if not wins:
                return None
            text = "\n".join(wins)
        out: dict[str, Any] = {
            "이름": name,
            "각인/효과 등급": _extract_effect_grade_from_sources(name) or _extract_grade_stage(text, max_stage=4),
            "치명타 적중률 증가(%)": None,
            "공격력 증가(%)": None,
            "치명타 피해 증가(%)": None,
            "치명타 적중률 고정(%)": None,
            "진화형 피해(%)": None,
            "랜덤 페널티 확률(%)": None,
            "랜덤 페널티 감소율(%)": None,
            "평균 피해 손실(%)": None,
            "기대 피해 계수(%)": None,
            "치명타 확률 상한(%)": None,
            "초과 전환율(%)": None,
            "진화 피해 상한(%)": None,
            "감지 출처": "API tooltip",
        }
        if name == "아드레날린":
            # 여러 단계 수치가 한 Tooltip에 같이 들어오면 더하지 않고 최종 후보 최대값을 사용합니다.
            out["치명타 적중률 증가(%)"] = _max_percent_near(text, ["치명타 적중률", "치명타 확률", "치적"])
            per_stack = None
            stacks = None
            m = re.search(r"공격력(?:이|을)?[^0-9]{0,40}([+\-]?\d+(?:\.\d+)?)\s*%[^\n]{0,180}?최대\s*(\d+)\s*중첩", text, re.S)
            if m:
                try:
                    per_stack = float(m.group(1)); stacks = float(m.group(2))
                except Exception:
                    pass
            if per_stack and stacks:
                out["공격력 증가(%)"] = per_stack * stacks
            else:
                out["공격력 증가(%)"] = _max_percent_near(text, ["공격력"])
        elif name == "예리한 둔기":
            out["치명타 피해 증가(%)"] = _max_percent_near(text, ["치명타 피해량", "치명타 피해", "치피"])
            # 예둔 페널티는 표현이 여러 가지라 감지되면 보조 지표로만 사용합니다.
            pm = re.search(r"(\d+(?:\.\d+)?)\s*%[^\n]{0,80}?(?:확률|일정 확률)[^\n]{0,120}?(\d+(?:\.\d+)?)\s*%[^\n]{0,50}?(?:감소|낮아)", text, re.S)
            if not pm:
                pm = re.search(r"(?:확률|일정 확률)[^\n]{0,80}?(\d+(?:\.\d+)?)\s*%[^\n]{0,120}?(?:피해|주는 피해)[^\n]{0,40}?(\d+(?:\.\d+)?)\s*%[^\n]{0,50}?(?:감소|낮아)", text, re.S)
            if pm:
                try:
                    out["랜덤 페널티 확률(%)"] = float(pm.group(1))
                    out["랜덤 페널티 감소율(%)"] = float(pm.group(2))
                except Exception:
                    pass
            # v137: 예리한 둔기 기본 수치(고정) — 툴팁 파싱이 실패해도 알려진 값으로 채웁니다.
            # 공격 시 10% 확률로 그 타격 피해 20% 감소(치명타와 무관), 치명타 피해량 +50%.
            if out["랜덤 페널티 확률(%)"] is None:
                out["랜덤 페널티 확률(%)"] = 10.0
            if out["랜덤 페널티 감소율(%)"] is None:
                out["랜덤 페널티 감소율(%)"] = 20.0
            if out["치명타 피해 증가(%)"] is None:
                out["치명타 피해 증가(%)"] = 50.0
            _p = (out["랜덤 페널티 확률(%)"] or 0.0) / 100.0
            _r = (out["랜덤 페널티 감소율(%)"] or 0.0) / 100.0
            out["평균 피해 손실(%)"] = round(_p * _r * 100.0, 2)          # 예: 10%×20% = 2%
            out["기대 피해 계수(%)"] = round((1.0 - _p * _r) * 100.0, 2)   # 예: 98%
        elif name == "뭉툭한 가시":
            # 뭉툭한 가시(아크패시브): 진화형 피해 +기본%, 치명타 확률 상한 X%,
            # 상한 초과 치확 × 전환율% → 진화형 피해로 전환, 이 노드 진화 피해 상한 Y%.
            # 스킬마다 치적이 다르므로 진화형 피해가 스킬별로 다르게 들어갑니다.
            # 툴팁에서 각 수치를 읽되, 못 읽으면 레벨2 기준값(15/80/150/75)을 폴백으로 씁니다.
            m_cap = re.search(r"치명타(?:가 발생할)?\s*확률[^%]{0,20}?최대\s*(\d+(?:\.\d+)?)\s*%\s*로\s*제한", text, re.S)
            m_conv = re.search(r"초과한[^%]{0,30}?확률의\s*(\d+(?:\.\d+)?)\s*%\s*가?\s*진화형", text, re.S)
            m_evocap = re.search(r"노드에 의한 진화형 피해는?\s*최대\s*(\d+(?:\.\d+)?)\s*%", text, re.S)
            m_base = re.search(r"진화형 피해가?\s*(\d+(?:\.\d+)?)\s*%\s*증가", text, re.S)
            out["진화형 피해(%)"] = float(m_base.group(1)) if m_base else 15.0            # 기본 진화형 피해
            out["치명타 확률 상한(%)"] = float(m_cap.group(1)) if m_cap else 80.0
            out["초과 전환율(%)"] = float(m_conv.group(1)) if m_conv else 150.0
            out["진화 피해 상한(%)"] = float(m_evocap.group(1)) if m_evocap else 75.0
            # 참고용: 치명타 확률 상한을 '치적 고정'처럼도 노출합니다.
            out["치명타 적중률 고정(%)"] = out["치명타 확률 상한(%)"]
        return out

    for key in ["아드레날린", "예리한 둔기", "뭉툭한 가시"]:
        r = row_for(key)
        if r:
            rows.append(r)
    return rows


def _build_expected_dps_range(result_df: pd.DataFrame, real_elapsed: float | None, special_effects: list[dict[str, Any]] | None = None) -> pd.DataFrame:
    """치명/예둔/뭉가 계열 랜덤성 기준의 이론 DPS 범위를 보조 지표로 만듭니다.

    v62:
    - 치적 자체의 변동성은 "아드레날린"이 아니라 "치명타 적중률"로 표시합니다.
    - 예리한 둔기/뭉툭한 가시/뭉가+예둔은 감지될 때 별도 행으로 보여줍니다.
      미채용 상태에서 억지 가정값을 계산하지 않기 위해 실제 감지된 조합만 표시합니다.
    """
    if result_df is None or result_df.empty or not real_elapsed or real_elapsed <= 0:
        return pd.DataFrame()
    included = result_df[result_df.get("비교포함", 0).eq(1)].copy() if "비교포함" in result_df.columns else result_df.copy()
    if included.empty:
        return pd.DataFrame()

    effects_by_name = {str(e.get("이름") or ""): e for e in (special_effects or [])}
    has_keen = "예리한 둔기" in effects_by_name
    has_blunt = "뭉툭한 가시" in effects_by_name

    def totals(use_keen_penalty: bool = False) -> tuple[float, float, float]:
        penalty_rate = 0.0
        penalty_reduce = 0.0
        if use_keen_penalty:
            keen = effects_by_name.get("예리한 둔기") or {}
            penalty_rate = _finite_float(keen.get("랜덤 페널티 확률(%)"), 0.0) or 0.0
            penalty_reduce = _finite_float(keen.get("랜덤 페널티 감소율(%)"), 0.0) or 0.0

        total_min = 0.0
        total_avg = 0.0
        total_max = 0.0
        for _, row in included.iterrows():
            avg_damage = _finite_float(row.get("치명타 보정 값"), _finite_float(row.get("실전 평균"), 0.0)) or 0.0
            p = max(0.0, min(100.0, _finite_float(row.get("실전 기대치명"), _finite_float(row.get("치명타 적중률"), 0.0)) or 0.0)) / 100.0
            cd = max(100.0, _finite_float(row.get("총 치피"), 200.0) or 200.0) / 100.0
            crit_mul = 1.0 + p * (cd - 1.0)
            base = avg_damage / crit_mul if crit_mul > 0 else avg_damage
            min_damage = base
            max_damage = base * cd
            if penalty_rate > 0 and penalty_reduce > 0:
                # 예둔 페널티가 최악으로 계속 걸리는 극단값 보조 표시.
                min_damage *= max(0.0, 1.0 - penalty_reduce / 100.0)
            total_min += min_damage
            total_avg += avg_damage
            total_max += max_damage
        return total_min, total_avg, total_max

    rows: list[dict[str, Any]] = []

    def add_row(name: str, use_keen_penalty: bool, note: str) -> None:
        mn, avg, mx = totals(use_keen_penalty=use_keen_penalty)
        rows.append({
            "고려 효과": name,
            "예상 최소 피해": mn,
            "예상 평균 피해": avg,
            "예상 최대 피해": mx,
            "예상 최소 DPS": mn / real_elapsed,
            "예상 평균 DPS": avg / real_elapsed,
            "예상 최대 DPS": mx / real_elapsed,
            "주의": note,
        })

    if has_keen:
        add_row("예리한 둔기", True, "예둔 페널티/치명타 분산을 단순화한 이론 범위")
    else:
        add_row("치명타 적중률", False, "치명타 적중률 분산을 단순화한 이론 범위")

    if has_blunt:
        add_row("뭉툭한 가시", False, "뭉툭한 가시 치적 고정/진피는 API 계산표 반영값 기준")
    if has_blunt and has_keen:
        add_row("뭉툭한 가시 + 예리한 둔기", True, "뭉가 반영값에 예둔 페널티/치명타 분산을 함께 고려한 범위")

    return pd.DataFrame(rows)


def _build_probability_correction_v143(result_df: pd.DataFrame, real_elapsed: float | None, special_effects: list[dict[str, Any]] | None) -> dict[str, Any] | None:
    """확률 효과 보정 요약. 치명타 확률 분산에 따른 '현실적' 실전 DPS 범위를 계산합니다.

    v144 변경점:
    - 예리한 둔기는 제외합니다. 허수아비 타격 때도 예둔 랜덤 확률이 동일하게 들어가고(=허수/실전 양쪽에
      같은 기대손실이 반영됨), 실측으로 분리해낼 방법이 없어 효율 비율에 순영향이 없기 때문입니다.
    - 최소 DPS는 '0회 치명(불가능)'이 아니라, 스킬별 타수를 반영한 치명 분포에서
      누적확률 0.1% 지점(약 −3.09σ)까지만 집계합니다. 그보다 더 불운한 경우는 현실적으로 불가능하므로 버립니다.
    - 최대 DPS는 이론상 상한(모든 타격 치명)을 그대로 둡니다.
    """
    if result_df is None or result_df.empty or not real_elapsed or real_elapsed <= 0:
        return None
    effects = {str(e.get("이름") or ""): e for e in (special_effects or [])}
    has_keen = "예리한 둔기" in effects   # 표시 트리거 용도로만 사용, 계산에는 반영하지 않습니다.
    has_blunt = "뭉툭한 가시" in effects

    # 아르카나 카드 아이덴티티 확률 보정 준비.
    # 황제/황후의 기사(직접피해 카드)의 관측 사용횟수를 '카드 뜰 확률 × 카드 사용 횟수(N)'
    # 기대값으로 정규화하고, 이항분포 0.1% 하한으로 불운 최소 딜을 계산합니다.
    card_ctx = _arcana_card_context()
    try:
        card_uses = int(st.session_state.get("arcana_card_uses", 0) or 0)
        if card_uses <= 0:  # 수동 입력이 없으면 이미지에서 감지한 값으로 폴백
            card_uses = int(st.session_state.get("arcana_card_uses_ocr", 0) or 0)
    except Exception:
        card_uses = 0
    card_cards = (card_ctx or {}).get("cards") or {}
    card_active = bool(card_ctx) and card_uses > 0 and bool(card_cards)

    if not has_keen and not has_blunt and not card_active:
        return None  # 확률 효과를 하나도 안 쓰면 이 섹션 자체를 숨깁니다.

    included = result_df[result_df.get("비교포함", 0).eq(1)].copy() if "비교포함" in result_df.columns else result_df.copy()
    if included.empty:
        included = result_df.copy()

    # 카드 직접피해 행은 치명 변동 루프에서 빼고, 아래 카드 모델로 따로 처리합니다.
    card_names_norm = {normalize_name(k) for k in card_cards} if card_active else set()

    total_mean = 0.0          # 기대(평균) 피해
    total_floor = 0.0         # 이론상 절대 최소(모든 타격 비치명)
    total_max = 0.0           # 이론상 절대 최대(모든 타격 치명)
    total_var = 0.0           # 치명 분포에 의한 총 피해 분산(타수 반영)
    for _, row in included.iterrows():
        if card_names_norm and normalize_name(row.get("스킬명")) in card_names_norm:
            continue  # 카드 직접피해 스킬은 카드 확률 모델로 별도 처리
        avg = _finite_float(row.get("치명타 보정 값"), _finite_float(row.get("실전 평균"), 0.0)) or 0.0
        p = max(0.0, min(100.0, _finite_float(row.get("실전 기대치명"), _finite_float(row.get("치명타 적중률"), 0.0)) or 0.0)) / 100.0
        cd = max(100.0, _finite_float(row.get("총 치피"), 200.0) or 200.0) / 100.0
        crit_mul = 1.0 + p * (cd - 1.0)
        base = avg / crit_mul if crit_mul > 0 else avg    # 비치명 기준 총 피해
        n = _finite_float(row.get("실전 사용횟수"), None)
        total_mean += base * (1.0 + p * (cd - 1.0))       # = avg
        total_floor += base
        total_max += base * cd
        # 타격당 치명 보너스 = (base/n)*(cd-1), 치명 발생 확률 p → 스킬 총 피해 분산
        if n and n >= 1:
            per_hit_bonus = (base / n) * (cd - 1.0)
            total_var += n * (per_hit_bonus ** 2) * p * (1.0 - p)

    # 누적확률 0.1% 지점(단측). 정규근사: mean − z·σ, z≈3.09. 절대 최소보다 낮게는 내려가지 않게 클램프.
    Z_001 = 3.0902
    sd = math.sqrt(total_var) if total_var > 0 else 0.0
    noncard_min = max(total_floor, total_mean - Z_001 * sd)

    # ── 아르카나 카드 확률 모델 ──
    # 각 직접피해 카드: 1회 평균 딜 = (해당 스킬 확률보정 관측 딜) / (관측 사용횟수).
    # 기대(=최대) 딜 = (N×p) × 1회 평균,  최소 딜 = 이항분포 0.1% 하한 횟수 × 1회 평균.
    card_mean = 0.0
    card_min = 0.0
    card_detail: list[dict[str, Any]] = []
    if card_active:
        has_skill_col = "스킬명" in result_df.columns
        norm_series = result_df["스킬명"].astype(str).map(normalize_name) if has_skill_col else None
        for card_name, p in card_cards.items():
            nn = normalize_name(card_name)
            obs_count = 0.0
            obs_total = 0.0
            per_use = 0.0
            if norm_series is not None:
                match = result_df[norm_series == nn]
                if not match.empty:
                    mrow = match.iloc[0]
                    obs_count = _finite_float(mrow.get("실전 사용횟수"), 0.0) or 0.0
                    obs_total = _finite_float(mrow.get("치명타 보정 값"), _finite_float(mrow.get("실전 평균"), 0.0)) or 0.0
                    if obs_count >= 1:
                        per_use = obs_total / obs_count
            exp_count = _binom_expected_count(card_uses, p)
            min_count = _binom_lower_count(card_uses, p, 0.001)
            exp_dmg = exp_count * per_use
            min_dmg = min_count * per_use
            card_mean += exp_dmg
            card_min += min_dmg
            card_detail.append({
                "카드": card_name,
                "확률(%)": round(p * 100, 2),
                "관측 횟수": int(round(obs_count)),
                "1회 평균 딜": per_use,
                "기대 횟수": exp_count,
                "최소 횟수": int(min_count),
                "기대 딜": exp_dmg,
                "최소 딜": min_dmg,
            })

    total_mean += card_mean          # 기대: 카드는 N×p 기대값으로 반영
    total_max += card_mean           # 최대: 카드는 중심(기대값) 고정 → 위쪽 편차 없음
    total_min = noncard_min + card_min  # 최소: 치명 불운 + 카드 0.1% 하한

    names = [n for n in ["뭉툭한 가시"] if n in effects]   # 예둔은 목록에서 제외
    if card_active:
        names = names + [f"카드 아이덴티티({card_ctx.get('engraving')})"]
    return {
        "채용": names,
        "평균 DPS": total_mean / real_elapsed,
        "최소 DPS": total_min / real_elapsed,
        "최대 DPS": total_max / real_elapsed,
        "최소 백분위(%)": 0.1,
        "카드 보정": ({
            "각인": card_ctx.get("engraving") if card_ctx else None,
            "각인 근거": card_ctx.get("engraving_source") if card_ctx else None,
            "황후의 기사 노드레벨": card_ctx.get("knight_of_empress_level") if card_ctx else None,
            "황후의 기사 제외됨": card_ctx.get("knight_of_empress_gated") if card_ctx else None,
            "카드 사용 횟수": card_uses,
            "실경과(초)": real_elapsed,
            "항목": card_detail,
            "기대 딜 합": card_mean,
            "최소 딜 합": card_min,
        } if card_active else None),
    }


def _blunt_thorn_params_from_effects(special_effects: list[dict[str, Any]] | None) -> dict[str, float] | None:
    """감지된 특수효과 목록에서 뭉툭한 가시가 채용됐으면 그 파라미터를 반환합니다.

    채용 안 했으면 None → 뭉가 관련 보정을 아예 적용하지 않습니다.
    """
    for e in (special_effects or []):
        if str(e.get("이름")) == "뭉툭한 가시":
            return {
                "base_evo": _finite_float(e.get("진화형 피해(%)"), 15.0) or 15.0,
                "crit_cap": _finite_float(e.get("치명타 확률 상한(%)"), 80.0) or 80.0,
                "conversion": _finite_float(e.get("초과 전환율(%)"), 150.0) or 150.0,
                "evo_cap": _finite_float(e.get("진화 피해 상한(%)"), 75.0) or 75.0,
            }
    return None


def _blunt_thorn_evolution_percent(skill_crit_percent: float | None, params: dict[str, float] | None) -> float | None:
    """뭉툭한 가시 채용 시, 스킬 치명타 적중률(이론값)에 따른 진화형 피해(%)를 계산합니다.

    상한 초과 치확(%) × 전환율/100 이 진화형 피해로 전환되고, 기본 진화형 피해를 더한 뒤
    이 노드의 진화 피해 상한으로 자릅니다. 스킬마다 치적이 다르므로 값이 달라집니다.
    예) 치적 120%, 상한 80%, 전환 150%, 상한 75% → 15 + (120-80)*1.5 = 75 → min(75,75)=75%.
        치적 80% → 15%.
    """
    if params is None or skill_crit_percent is None:
        return None
    excess = max(0.0, float(skill_crit_percent) - float(params["crit_cap"]))
    evo = float(params["base_evo"]) + excess * float(params["conversion"]) / 100.0
    return min(evo, float(params["evo_cap"]))


def build_result_and_improvement_tables(
    real_table: pd.DataFrame,
    dummy_table: pd.DataFrame,
    real_elapsed: float | None,
    dummy_elapsed: float | None,
    target_increase_rate: float,
    crit_synergy_count: float = 0.0,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, dict[str, Any]]:
    """엑셀의 결과/개선분석 시트를 Streamlit 안에서 계산합니다.

    핵심 흐름:
    - 실전/허수 관측 피해를 스킬별로 정리
    - 전분 치명률과 API 기대 치명률/치피를 이용해 치명운 제거 평균 피해 계산
    - 실전 피해를 치명운 기준으로 평균화한 뒤 스킬 1회당 보정 피해를 계산
    - 허수 전분의 CPM/사용횟수 페이스에 실전 1회 보정 피해를 얹어 허수페이스 기대 피해 계산
    - 허수페이스 대비 실전 보정 피해 손실을 개선 후보로 정렬
    """
    api_lookup = _build_api_skill_lookup()
    real = _aggregate_battle_rows(real_table)
    dummy = _aggregate_battle_rows(dummy_table)
    data_warnings = _detect_battle_data_anomalies(real, dummy, real_elapsed, dummy_elapsed)
    # v137: 뭉툭한 가시 채용 시 치명타 확률 상한(예: 80%)과 스킬별 진화형 피해를 계산에 반영합니다.
    _special_effects_early = _detect_special_combat_effects()
    _blunt_params = _blunt_thorn_params_from_effects(_special_effects_early)
    real_elapsed = float(real_elapsed or 0.0)
    dummy_elapsed = float(dummy_elapsed or 0.0)
    time_ratio = (real_elapsed / dummy_elapsed) if real_elapsed > 0 and dummy_elapsed > 0 else 1.0
    directional_engravings = _detect_directional_engravings()
    # v152: 파티 치명타 적중률 시너지 보정.
    # 실전은 파티 시너지로 치적이 올라가지만(예: 80→90), 허수/전분은 시너지 없이 측정된 값입니다.
    # '치적 시너지 인원 수'(최대 2명, 1명당 +10%p)를 실전 기대 치명률에만 더해, 시너지로 오른
    # 치명 피해가 '치명운'으로 오해되어 깎여나가지 않도록 실전 관측 피해를 올바른 기대치로 평균화합니다.
    _crit_synergy_bonus = max(0.0, min(2.0, float(crit_synergy_count or 0.0))) * 10.0

    # v136: 방향 각인 상태(기습/결투)를 전역으로 1번만 판정하고, 감지된 조건부 방향
    # 피해(%)로 각인 배율을 설정합니다. 둘 다 감지되면 validation error를 잡아
    # UI 경고만 남기고 각인 없음으로 폴백해 분석은 계속합니다.
    _amb_bonus = None
    _brw_bonus = None
    for _eng in directional_engravings:
        if str(_eng.get("각인") or "") == "기습의 대가":
            _amb_bonus = _finite_float(_eng.get("조건부 방향 피해(%)"), None)
        elif str(_eng.get("각인") or "") == "결투의 대가":
            _brw_bonus = _finite_float(_eng.get("조건부 방향 피해(%)"), None)
    _dir_settings = DirectionSettings().with_engraving_bonuses(_amb_bonus, _brw_bonus)
    _dir_conflict_warning = ""
    try:
        _global_eng_state = resolve_engraving_state(_amb_bonus is not None, _brw_bonus is not None)
    except DirectionValidationError as _e:
        _dir_conflict_warning = str(_e)
        _global_eng_state = DirectionalEngravingState.NONE

    keys = list(dict.fromkeys(list(real.keys()) + list(dummy.keys())))
    rows: list[dict[str, Any]] = []
    for key in keys:
        r = real.get(key, {})
        d = dummy.get(key, {})
        api = api_lookup.get(key, {})
        name = str(api.get("스킬명") or r.get("name") or d.get("name") or "").strip()
        # v136: 방향 평가는 modules.direction으로 분리했습니다. 여기서는 화면 관측값만
        # 모으고, 실제 방향 판정/보너스/각인 효율/손실은 real_avg 계산 후에 수행합니다.
        # 백/헤드 비중은 '피해 비중'(퍼센트)이고, 빈칸이면 None(데이터 없음)입니다.
        _rb_share = _finite_float(r.get("back_attack_share"), None)
        _rh_share = _finite_float(r.get("head_attack_share"), None)
        _rb_rate = _finite_float(r.get("back_attack_rate"), None)
        _rh_rate = _finite_float(r.get("head_attack_rate"), None)
        # v138: 방향 판정 = 전투분석기 관측 우선 + API 폴백.
        # 전투분석기에 백/헤드가 '한 번이라도' 관측되면(적중률 or 비중 > 0) 그 방향으로 확정합니다.
        # 관측이 전혀 없으면(빈칸/0) API 방향을 따릅니다 — 방향성 스킬이어도 그 판에 못 맞춰
        # 0이 될 수 있으니 0을 '방향 없음'으로 단정하지 않기 위함입니다.
        def _pos(v: Any) -> bool:
            f = _finite_float(v, None)
            return f is not None and f > 0.0
        _obs_back = _pos(r.get("back_attack_rate")) or _pos(r.get("back_attack_share")) or _pos(d.get("back_attack_rate")) or _pos(d.get("back_attack_share"))
        _obs_head = _pos(r.get("head_attack_rate")) or _pos(r.get("head_attack_share")) or _pos(d.get("head_attack_rate")) or _pos(d.get("head_attack_share"))
        _api_dir_raw = api.get("공격타입_원본") or api.get("공격타입") or ""
        if _obs_back and _obs_head:
            _eff_dir_raw = "백/헤드"
        elif _obs_back:
            _eff_dir_raw = "백어택"
        elif _obs_head:
            _eff_dir_raw = "헤드어택"
        else:
            _eff_dir_raw = _api_dir_raw
        api_crit_rate = _finite_float(api.get("api_crit_rate"), None)
        api_crit_damage = _finite_float(api.get("api_crit_damage"), 200.0) or 200.0
        api_evo = _finite_float(api.get("api_evolution_damage"), 0.0) or 0.0
        real_damage = _finite_float(r.get("damage"), 0.0) or 0.0
        dummy_damage = _finite_float(d.get("damage"), 0.0) or 0.0
        real_observed_crit = _finite_float(r.get("crit_rate"), api_crit_rate)
        dummy_observed_crit = _finite_float(d.get("crit_rate"), api_crit_rate)
        real_observed_crit_share = _finite_float(r.get("crit_share"), None)
        dummy_observed_crit_share = _finite_float(d.get("crit_share"), None)
        # v137: 백어택 +10% 치적은 API에 미포함이므로, 실전/허수 각각의 '백어택 적중률'만큼
        # 자동으로 더해 기대 치명률을 계산합니다. 허수는 보통 백 100%(항상 뒤를 침)이고
        # 실전은 100%가 아니라서, 실전과 허수의 기대 치명률이 서로 다릅니다(상쇄되지 않음).
        _early_type = classify_skill_attack_direction(_eff_dir_raw)
        _back_capable = _early_type in (AttackDirectionType.BACK_ONLY, AttackDirectionType.DUAL)
        _bcb = (_dir_settings.back_crit_rate_bonus or 0.10) * 100.0  # 보통 10%p
        _db_rate = _finite_float(d.get("back_attack_rate"), None)
        _real_back_frac = (max(0.0, min(1.0, _rb_rate / 100.0)) if _rb_rate is not None else 1.0) if _back_capable else 0.0
        _dummy_back_frac = (max(0.0, min(1.0, _db_rate / 100.0)) if _db_rate is not None else 1.0) if _back_capable else 0.0
        # 실전 기대 치명률에만 파티 치적 시너지를 더합니다(허수는 시너지 없이 측정된 값이므로 제외).
        _real_eff_crit = (api_crit_rate + _crit_synergy_bonus + _bcb * _real_back_frac) if api_crit_rate is not None else None
        _dummy_eff_crit = (api_crit_rate + _bcb * _dummy_back_frac) if api_crit_rate is not None else None

        # 뭉툭한 가시 채용 시, 기대 치명률은 상한(예: 80%)으로 자릅니다.
        # 상한 초과 치확은 치명타로 터지지 않고 진화형 피해로 전환되기 때문입니다.
        _real_exp_for_adj = _real_eff_crit
        _dummy_exp_for_adj = _dummy_eff_crit
        if _blunt_params is not None:
            if _real_exp_for_adj is not None:
                _real_exp_for_adj = min(_real_exp_for_adj, _blunt_params["crit_cap"])
            if _dummy_exp_for_adj is not None:
                _dummy_exp_for_adj = min(_dummy_exp_for_adj, _blunt_params["crit_cap"])
        real_avg = _adjust_damage_to_expected_crit(real_damage, real_observed_crit, real_observed_crit_share, _real_exp_for_adj, api_crit_damage)
        dummy_avg = _adjust_damage_to_expected_crit(dummy_damage, dummy_observed_crit, dummy_observed_crit_share, _dummy_exp_for_adj, api_crit_damage)

        # v136: 방향 평가 (구조 방향 + 착용 각인 + 기본 보너스 + 각인 효율 + 손실 + 신뢰도)
        _base_crit_frac = (api_crit_rate if api_crit_rate is not None else (real_observed_crit or 0.0)) / 100.0
        _crit_dmg_mul = (api_crit_damage or 200.0) / 100.0
        dir_eval = evaluate_skill_direction(
            api_attack_text=_eff_dir_raw,
            observed_back_share=(_rb_share / 100.0) if _rb_share is not None else None,
            observed_head_share=(_rh_share / 100.0) if _rh_share is not None else None,
            base_crit_rate=_base_crit_frac,
            crit_damage_multiplier=_crit_dmg_mul,
            current_crit_adjusted_damage=real_avg,
            global_engraving_state=_global_eng_state,
            settings=_dir_settings,
        )
        skill_type = dir_eval.skill_type
        # 레거시 '공격타입' 컬럼: 목표 방향(있으면) 또는 구조 기반 대표 방향으로 채웁니다.
        if dir_eval.target_direction == "BACK":
            attack_type = "백어택"
        elif dir_eval.target_direction == "HEAD":
            attack_type = "헤드어택"
        elif skill_type == AttackDirectionType.BACK_ONLY:
            attack_type = "백어택"
        elif skill_type == AttackDirectionType.HEAD_ONLY:
            attack_type = "헤드어택"
        elif skill_type == AttackDirectionType.DUAL:
            attack_type = "백/헤드"
        else:
            attack_type = "없음"
        _struct_label = {
            AttackDirectionType.BACK_ONLY: "백어택 전용",
            AttackDirectionType.HEAD_ONLY: "헤드어택 전용",
            AttackDirectionType.DUAL: "백/헤드 가능",
            AttackDirectionType.NONE: "무방향",
        }[skill_type]
        _eng_label = {
            DirectionalEngravingState.AMBUSH_MASTER: "기습의 대가",
            DirectionalEngravingState.MASTER_BRAWLER: "결투의 대가",
            DirectionalEngravingState.NONE: "없음",
        }[dir_eval.engraving_state]
        _target_label = {"BACK": "백어택", "HEAD": "헤드어택", None: "없음"}[dir_eval.target_direction]

        # v137: 뭉툭한 가시 진화형 피해 = 위에서 구한 '실전 실효 치명률'(_real_eff_crit,
        # API 기본 치적 + 백어택 적중률만큼의 +10%)로 계산합니다. 상한(80%) 초과분이
        # 진화형 피해로 전환되므로 상한 자르기 전 값을 씁니다.
        _blunt_evo = _blunt_thorn_evolution_percent(_real_eff_crit, _blunt_params)
        if _blunt_evo is not None:
            api_evo = _blunt_evo

        # v46: 허수페이스는 허수 피해량을 그대로 시간 환산하지 않습니다.
        # 실전에서 관측된 스킬 1회당 피해를 치명운만 평균화한 뒤,
        # 허수 전분의 사용횟수/CPM 페이스에 얹어서 계산합니다.
        # 이렇게 해야 레이드 버프/보스 상태/실전 피해 규모는 실전값을 쓰고,
        # 허수 대비 수행률은 사용횟수·쿨소화·포지션 차이만 비교할 수 있습니다.
        real_casts_for_pace = _finite_float(r.get("casts"), None)
        dummy_casts_for_pace = _finite_float(d.get("casts"), None)
        real_avg_per_cast = real_avg / real_casts_for_pace if real_casts_for_pace and real_casts_for_pace > 0 else None
        dummy_expected_casts = None
        if dummy_elapsed > 0 and real_elapsed > 0 and dummy_casts_for_pace is not None:
            dummy_expected_casts = dummy_casts_for_pace / dummy_elapsed * real_elapsed
        if real_avg_per_cast is not None and dummy_expected_casts is not None:
            dummy_pace = real_avg_per_cast * dummy_expected_casts
        else:
            # 사용횟수 정보가 없는 항목은 예전 방식으로만 fallback합니다.
            dummy_pace = dummy_avg * time_ratio if dummy_avg > 0 else 0.0

        observed_vs_avg = real_damage / real_avg if real_avg > 0 else None
        dummy_observed_vs_avg = dummy_damage / dummy_avg if dummy_avg > 0 else None
        real_avg_vs_pace = real_avg / dummy_pace if dummy_pace > 0 else None
        observed_vs_pace = real_damage / dummy_pace if dummy_pace > 0 else None
        cooldown_ratio = None
        if _finite_float(r.get("cooldown_rate"), None) is not None and _finite_float(d.get("cooldown_rate"), None):
            cooldown_ratio = (_finite_float(r.get("cooldown_rate"), 0.0) or 0.0) / max(1e-9, _finite_float(d.get("cooldown_rate"), 0.0) or 0.0)
        use_count_ratio = None
        real_cpm = None
        dummy_cpm = None
        if real_elapsed > 0 and _finite_float(r.get("casts"), None) is not None:
            real_cpm = (_finite_float(r.get("casts"), 0.0) or 0.0) / real_elapsed * 60.0
        if dummy_elapsed > 0 and _finite_float(d.get("casts"), None) is not None:
            dummy_cpm = (_finite_float(d.get("casts"), 0.0) or 0.0) / dummy_elapsed * 60.0
        if real_cpm is not None and dummy_cpm and dummy_cpm > 0:
            use_count_ratio = real_cpm / dummy_cpm
        # 레거시 포지션 지표: 목표 방향의 관측 적중률/비중(퍼센트). 표시 연속성 유지용.
        position_rate = None
        if attack_type == "백어택":
            position_rate = _rb_rate
        elif attack_type == "헤드어택":
            position_rate = _rh_rate
        back_attack_share = _rb_share
        head_attack_share = _rh_share
        position_share = _rb_share if attack_type == "백어택" else _rh_share if attack_type == "헤드어택" else None
        crit_share = _finite_float(r.get("crit_share"), None)

        # v136: 방향 각인 효율/실질 딜증은 dir_eval에서 가져옵니다.
        engraving_name = _eng_label if dir_eval.engraving_state != DirectionalEngravingState.NONE else ""
        engraving_efficiency = dir_eval.directional_engraving_efficiency
        # 실질 딜증(각인이 실제로 준 조건부 딜증) = 각인 효율 × 이론상 최대 딜증
        if dir_eval.engraving_state == DirectionalEngravingState.AMBUSH_MASTER:
            _theo = _dir_settings.ambush_master_multiplier - 1.0
        elif dir_eval.engraving_state == DirectionalEngravingState.MASTER_BRAWLER:
            _theo = _dir_settings.master_brawler_multiplier - 1.0
        else:
            _theo = None
        engraving_real_gain = (engraving_efficiency * _theo) if (engraving_efficiency is not None and _theo) else None

        # v136: raw(보너스 제거) 방향 비중 — 퍼센트로 환산해 표시합니다.
        raw_back_pct = dir_eval.raw_back_share * 100.0 if dir_eval.raw_back_share is not None else None
        raw_head_pct = dir_eval.raw_head_share * 100.0 if dir_eval.raw_head_share is not None else None
        raw_neutral_pct = dir_eval.raw_neutral_share * 100.0 if dir_eval.raw_neutral_share is not None else None
        basic_dir_util = dir_eval.basic_direction_bonus_utilization
        direction_loss_damage = dir_eval.direction_loss_damage
        engraving_loss_damage = dir_eval.directional_engraving_loss_damage
        stagger_bonus = dir_eval.stagger_bonus_expected
        # 방향 데이터 신뢰도(스킬 단위): 구조상 방향이 있는데 관측 데이터가 있으면 True.
        has_dir_data = bool(dir_eval.has_direction_data)
        needs_dir_data = skill_type in (
            AttackDirectionType.BACK_ONLY, AttackDirectionType.HEAD_ONLY, AttackDirectionType.DUAL,
        )

        include = True
        exclude_reason = ""
        if real_damage <= 0:
            include = False
            exclude_reason = "실전 피해 없음"
        elif dummy_pace <= 0:
            include = False
            exclude_reason = "허수 기준 없음"
        elif _is_rune_or_misc(name):
            include = False
            exclude_reason = "보조/기타 항목"
        loss = max(0.0, dummy_pace - real_avg) if include else 0.0
        rows.append({
            "스킬명": name,
            "공격타입": attack_type,
            "스킬 구조 타입": _struct_label,
            "착용 방향 각인": _eng_label,
            "목표 방향": _target_label,
            "기습/결투 각인": engraving_name,
            "기습/결투 효율": engraving_efficiency,
            "기습/결투 실질 딜증": engraving_real_gain,
            "기본 방향 보너스 활용률": basic_dir_util,
            "방향 각인 효율": engraving_efficiency,
            "방향 손실 피해": direction_loss_damage,
            "방향 각인 손실 피해": engraving_loss_damage,
            "백어택 raw 비중": raw_back_pct,
            "헤드어택 raw 비중": raw_head_pct,
            "무방향 raw 비중": raw_neutral_pct,
            "무력화 기대 보너스": stagger_bonus,
            "방향 데이터 있음": has_dir_data,
            "방향 데이터 필요": needs_dir_data,
            "실전 관측": real_damage,
            "실전 평균": real_avg,
            "치명타 보정 값": real_avg,
            "실전 사용횟수": real_casts_for_pace,
            "허수 관측": dummy_damage,
            "허수 평균": dummy_avg,
            "허수페이스": dummy_pace,
            "허수 페이스": dummy_pace,
            "허수 기준 예상 횟수": dummy_expected_casts,
            "실전 관측/평균": observed_vs_avg,
            "허수 관측/평균": dummy_observed_vs_avg,
            "실전평균/허수페이스": real_avg_vs_pace,
            "데미지 보정값 / 허수페이스": real_avg_vs_pace,
            "실전관측/허수페이스": observed_vs_pace,
            "쿨소화율": cooldown_ratio,
            "사용횟수율": use_count_ratio,
            "실전 포지션률": position_rate,
            "실전 포지션비중": position_share,
            "실전 포지션 비중": position_share,
            "백어택/헤드어택 비중": position_share,
            "방향 비중": position_share,
            "백어택 비중": back_attack_share,
            "헤드어택 비중": head_attack_share,
            "치명타 적중률": real_observed_crit,
            "치명타 비중": crit_share,
            "치명 보정 방식": "비중" if real_observed_crit_share is not None else "적중률",
            "포지션/치적 손실률": max(0.0, 1.0 - (position_rate or 0.0) / 100.0) if attack_type in ["백어택", "헤드어택"] else 0.0,
            "치명운": (observed_vs_avg - 1.0) if observed_vs_avg is not None else None,
            "실전 기대치명": api_crit_rate,
            "허수 기대치명": api_crit_rate,
            "총 치피": api_crit_damage,
            "진화형 피해": api_evo,
            "비교포함": 1 if include else 0,
            "제외사유": exclude_reason,
            "손실피해": loss,
            "현재 대비 상승여지": 0.0,
            "수행률": real_avg_vs_pace,
            "스킬 역할": _skill_role(_finite_float(r.get("share_rate"), None), name),
            "실전 피해 지분": _finite_float(r.get("share_rate"), None),
        })
    result_df = pd.DataFrame(rows)
    if result_df.empty:
        summary_df = pd.DataFrame()
        improvement_df = pd.DataFrame()
        return result_df, improvement_df, summary_df, {}

    included = result_df[result_df["비교포함"].eq(1)].copy()
    total_real_observed_all = float(result_df["실전 관측"].sum())
    total_real_avg_all = float(result_df["실전 평균"].sum())
    total_real_observed = float(included["실전 관측"].sum()) if not included.empty else total_real_observed_all
    total_real_avg = float(included["실전 평균"].sum()) if not included.empty else total_real_avg_all
    total_dummy_observed = float(included["허수 관측"].sum()) if not included.empty else float(result_df["허수 관측"].sum())
    total_dummy_avg = float(included["허수 평균"].sum()) if not included.empty else float(result_df["허수 평균"].sum())
    total_dummy_pace = float(included["허수페이스"].sum()) if not included.empty else float(result_df["허수페이스"].sum())
    total_loss = max(0.0, total_dummy_pace - total_real_avg)
    current_gap_rate = total_loss / total_real_avg if total_real_avg > 0 else 0.0
    target_needed = total_real_avg * float(target_increase_rate or 0.0)
    possible = "가능" if target_needed <= total_loss + 1e-9 else "목표 초과"

    if total_real_avg > 0:
        result_df["현재 대비 상승여지"] = result_df["손실피해"] / total_real_avg
    else:
        result_df["현재 대비 상승여지"] = 0.0

    # v136: 방향 데이터 신뢰도.
    # 커버리지 = (방향 데이터가 필요하고 실제로 있는 스킬의 실전 관측 피해) / (방향 데이터가 필요한 스킬의 실전 관측 피해)
    def _row_conf(row: Any) -> str:
        if not bool(row.get("방향 데이터 필요")):
            return "해당 없음"
        return "있음" if bool(row.get("방향 데이터 있음")) else "없음"
    result_df["방향 데이터 신뢰도"] = result_df.apply(_row_conf, axis=1)
    _need_mask = result_df["방향 데이터 필요"].astype(bool)
    _have_mask = _need_mask & result_df["방향 데이터 있음"].astype(bool)
    _need_dmg = float(pd.to_numeric(result_df.loc[_need_mask, "실전 관측"], errors="coerce").fillna(0).sum())
    _have_dmg = float(pd.to_numeric(result_df.loc[_have_mask, "실전 관측"], errors="coerce").fillna(0).sum())
    direction_data_coverage = (_have_dmg / _need_dmg) if _need_dmg > 0 else None
    direction_data_confidence = calc_direction_data_confidence(direction_data_coverage)

    special_effects = _detect_special_combat_effects()
    expected_dps_range_df = _build_expected_dps_range(result_df, real_elapsed, special_effects)
    probability_correction = _build_probability_correction_v143(result_df, real_elapsed, special_effects)

    summary_rows = [
        ["실전 관측 피해", total_real_observed_all, "전투분석기 실전 총피해. 보조/기타 포함"],
        ["치명타 보정 값", total_real_avg_all, "실전 피해를 API 기대 치명률 기준으로 평균화한 값. 보조/기타 포함"],
        ["비교 포함 실전 보정 피해", total_real_avg, "허수와 비교 가능한 스킬만 합산"],
        ["허수 관측 피해", total_dummy_observed, "전투분석기 허수 피해 합계"],
        ["허수 평균 피해", total_dummy_avg, "치명운을 제거한 허수 평균 피해"],
        ["허수페이스 기대 피해", total_dummy_pace, "허수 피해량 시간환산이 아니라 실전 1회 보정 피해 × 허수 사용횟수 페이스"],
        ["실전 관측 / 실전 평균", total_real_observed / total_real_avg if total_real_avg > 0 else None, "실전 치명운 영향"],
        ["허수 관측 / 허수 평균", total_dummy_observed / total_dummy_avg if total_dummy_avg > 0 else None, "허수 치명운 영향"],
        ["실전 평균 / 허수페이스", total_real_avg / total_dummy_pace if total_dummy_pace > 0 else None, "핵심 수행률"],
        ["실전 관측 / 허수페이스", total_real_observed / total_dummy_pace if total_dummy_pace > 0 else None, "관측값 기준 수행률"],
        ["가중 쿨소화율", _weighted_metric(result_df, "쿨소화율", "실전 관측"), "피해량 가중 쿨타임 비율"],
        ["가중 사용횟수율", _weighted_metric(result_df, "사용횟수율", "실전 관측"), "피해량 가중 CPM/사용횟수율"],
        ["총 상승 여지", total_loss, "허수페이스 - 실전 평균"],
        ["현재 대비 상승 여지", current_gap_rate, "총 상승 여지 / 실전 평균"],
        ["목표 상승률", float(target_increase_rate or 0.0), "사이드바 목표 입력"],
        ["목표 상승에 필요한 피해", target_needed, "실전 평균 피해 × 목표 상승률"],
        ["목표 가능 여부", possible, "목표가 허수페이스 한계 안인지 확인"],
        ["기본 방향 보너스 활용률(가중)", _weighted_metric(result_df, "기본 방향 보너스 활용률", "실전 관측"), "백/헤드 자체 보너스를 피해량 가중으로 얼마나 살렸는지"],
        ["방향 각인 효율(가중)", _weighted_metric(result_df, "방향 각인 효율", "실전 관측"), "기습/결투 각인을 피해량 가중으로 실제 딜로 전환한 비율"],
        ["방향 손실 피해 합계", float(pd.to_numeric(result_df["방향 손실 피해"], errors="coerce").fillna(0).clip(lower=0).sum()), "방향만 최적으로 들어갔을 때의 추가 가능 피해 합"],
        ["방향 데이터 커버리지", direction_data_coverage, "방향 평가가 필요한 피해 중 실제 방향 데이터가 있는 비율"],
        ["방향 데이터 신뢰도", direction_data_confidence, "커버리지 90%↑ 높음 / 60%↑ 보통 / 그 미만 낮음"],
    ]

    directional_engravings = _detect_directional_engravings()
    directional_eff_df = _build_directional_engraving_efficiency(result_df, directional_engravings)
    if not directional_eff_df.empty:
        for _, erow in directional_eff_df.iterrows():
            eng_name = str(erow.get("각인") or "방향 각인")
            summary_rows.append([f"{eng_name} 대상 스킬 조건 성공률", erow.get("대상 스킬 기준 활용률"), "대상 백/헤드 스킬 안에서 조건을 만족한 피해 비중"])
            summary_rows.append([f"{eng_name} 대상 스킬 각인 효율", erow.get("대상 스킬 기준 각인 효율"), "조건부 방향 보너스를 대상 스킬 기준 이론 최대치 대비 얼마나 살렸는지"])
            summary_rows.append([f"{eng_name} 조건부 보너스", erow.get("조건부 방향 피해(%)"), "API 각인 Tooltip에서 읽은 백/헤드 성공 시 추가 피해"])
            summary_rows.append([f"{eng_name} 전체 조건 비중", erow.get("전체 피해 기준 조건 비중"), "전체 실전 피해 중 각인 조건을 만족한 피해 비중"])
            summary_rows.append([f"{eng_name} 실질 딜증", erow.get("각인 실질 딜증"), "전체 피해 기준으로 역산한 각인 조건부 보너스 실질 딜증"])
    
    summary_df = pd.DataFrame(summary_rows, columns=["핵심 지표", "값", "해석"])

    improve = result_df.copy()
    def point(row: pd.Series) -> str:
        if int(row.get("비교포함", 0)) != 1:
            return "제외"
        if _finite_float(row.get("사용횟수율"), 1.0) is not None and (_finite_float(row.get("사용횟수율"), 1.0) or 0) < 0.85:
            return "사용횟수/쿨 밀림"
        if _finite_float(row.get("쿨소화율"), 1.0) is not None and (_finite_float(row.get("쿨소화율"), 1.0) or 0) < 0.85:
            return "쿨타임 소화"
        if row.get("공격타입") in ["백어택", "헤드어택"] and _finite_float(row.get("실전 포지션률"), 100.0) is not None and (_finite_float(row.get("실전 포지션률"), 100.0) or 0) < 85:
            return "포지션/백·헤드"
        if _finite_float(row.get("수행률"), 1.0) is not None and (_finite_float(row.get("수행률"), 1.0) or 0) < 0.9:
            return "딜각/사이클"
        if _finite_float(row.get("현재 대비 상승여지"), 0.0) and (_finite_float(row.get("현재 대비 상승여지"), 0.0) or 0) > 0.005:
            return "소폭 보완"
        return "유지"

    def action(p: str) -> str:
        return {
            "제외": "비교 제외",
            "사용횟수/쿨 밀림": "허수 대비 사용횟수 낮음: 주력기 우선순위·패턴 전 딜각 확인",
            "쿨타임 소화": "쿨타임이 밀림: 패턴 직후 즉시 사용 가능한 위치/동선 확인",
            "포지션/백·헤드": "백/헤드 적중률 낮음: 주력기 사용 전 포지션 고정",
            "딜각/사이클": "딜각/사이클 손실: 기믹 전후 주력기 배치 확인",
            "소폭 보완": "큰 구멍은 아니지만 반복 개선 대상",
            "유지": "현재 흐름 유지",
        }.get(p, "검수")

    improve["개선포인트"] = improve.apply(point, axis=1)
    improve["추천 액션"] = improve["개선포인트"].apply(action)
    improve["목표 달성 필요 개선률"] = improve["손실피해"].apply(lambda x: target_needed / x if x and x > 0 else None)
    improve["목표 판정"] = improve.apply(lambda r: "제외" if int(r.get("비교포함", 0)) != 1 else ("목표 초과" if possible == "목표 초과" else ("일부 개선만으로 목표 가능" if r.get("손실피해", 0) >= target_needed else "보조 개선 후보")), axis=1)
    improve = improve.sort_values(["비교포함", "손실피해"], ascending=[False, False]).reset_index(drop=True)
    improve.insert(0, "우선순위", [i + 1 if row["비교포함"] == 1 else None for i, row in improve.iterrows()])

    result_cols = [
        "스킬명", "공격타입", "스킬 구조 타입", "착용 방향 각인", "목표 방향",
        "실전 관측", "허수 페이스", "치명타 보정 값", "데미지 보정값 / 허수페이스",
        "쿨소화율", "사용횟수율",
        "기본 방향 보너스 활용률", "방향 각인 효율", "기습/결투 실질 딜증",
        "방향 손실 피해", "방향 각인 손실 피해",
        "백어택 raw 비중", "헤드어택 raw 비중", "무방향 raw 비중",
        "무력화 기대 보너스", "방향 데이터 신뢰도",
        "실전 포지션률", "실전 포지션 비중",
        "백어택/헤드어택 비중", "치명타 비중", "치명운", "제외사유",
    ]
    improvement_cols = [
        "우선순위", "스킬명", "공격타입", "기존 피해량", "허수페이스 기대", "손실피해",
        "현재 대비 상승여지", "치명운", "개선포인트", "추천 액션", "목표 달성 필요 개선률",
    ]
    improve = improve.rename(columns={"실전 평균": "기존 피해량", "허수페이스": "허수페이스 기대"})
    return (
        result_df[[c for c in result_cols if c in result_df.columns]].copy(),
        improve[[c for c in improvement_cols if c in improve.columns]].copy(),
        summary_df,
        {
            "total_real_observed": total_real_observed,
            "total_real_observed_all": total_real_observed_all,
            "total_real_avg": total_real_avg,
            "total_real_avg_all": total_real_avg_all,
            "total_dummy_pace": total_dummy_pace,
            "efficiency": total_real_avg / total_dummy_pace * 100 if total_dummy_pace > 0 else None,
            "direction_data_coverage": direction_data_coverage,
            "direction_data_confidence": direction_data_confidence,
            "direction_weighted_basic_util": _weighted_metric(result_df, "기본 방향 보너스 활용률", "실전 관측"),
            "direction_weighted_engraving_eff": _weighted_metric(result_df, "방향 각인 효율", "실전 관측"),
            "global_engraving_state": _global_eng_state.value,
            "direction_conflict_warning": _dir_conflict_warning,
            "data_warnings": data_warnings,
            "total_loss": total_loss,
            "current_gap_rate": current_gap_rate,
            "target_needed": target_needed,
            "target_rate": target_increase_rate,
            "possible": possible,
            "directional_engraving_efficiency": directional_eff_df.to_dict("records") if 'directional_eff_df' in locals() and isinstance(directional_eff_df, pd.DataFrame) else [],
            "special_combat_effects": special_effects if 'special_effects' in locals() else [],
            "expected_dps_range": expected_dps_range_df.to_dict("records") if 'expected_dps_range_df' in locals() and isinstance(expected_dps_range_df, pd.DataFrame) else [],
            "probability_correction": probability_correction if 'probability_correction' in locals() else None,
            "crit_synergy_count": int(max(0.0, min(2.0, float(crit_synergy_count or 0.0)))),
            "crit_synergy_bonus": _crit_synergy_bonus,
        },
    )


def _weighted_metric(df: pd.DataFrame, value_col: str, weight_col: str) -> float | None:
    if df is None or df.empty or value_col not in df.columns or weight_col not in df.columns:
        return None
    valid = df[[value_col, weight_col]].copy()
    valid[value_col] = pd.to_numeric(valid[value_col], errors="coerce")
    valid[weight_col] = pd.to_numeric(valid[weight_col], errors="coerce")
    valid = valid.dropna()
    valid = valid[valid[weight_col] > 0]
    if valid.empty:
        return None
    return float((valid[value_col] * valid[weight_col]).sum() / valid[weight_col].sum())


def format_analysis_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """화면 표시는 엑셀처럼 억 단위/퍼센트로 읽기 좋게 변환합니다."""
    if df is None or df.empty:
        return pd.DataFrame()
    out = df.copy()
    money_cols = ["실전 관측", "실전 평균", "치명타 보정 값", "허수 관측", "허수 평균", "허수페이스", "허수 페이스", "기존 피해량", "허수페이스 기대", "손실피해", "방향 손실 피해", "방향 각인 손실 피해", "대상 스킬 피해", "예상 최소 피해", "예상 평균 피해", "예상 최대 피해", "예상 최소 DPS", "예상 평균 DPS", "예상 최대 DPS"]
    ratio_cols = ["실전 관측/평균", "허수 관측/평균", "실전평균/허수페이스", "데미지 보정값 / 허수페이스", "실전관측/허수페이스", "쿨소화율", "사용횟수율", "현재 대비 상승여지", "수행률", "포지션/치적 손실률", "치명운", "목표 달성 필요 개선률", "각인 실질 딜증", "조건부 보너스 활용 효율", "기습/결투 효율", "기습/결투 실질 딜증", "대상 스킬 기준 각인 효율", "대상 스킬 기준 실질 딜증", "기본 방향 보너스 활용률", "방향 각인 효율", "무력화 기대 보너스"]
    percent_cols = ["실전 기대치명", "허수 기대치명", "실전 포지션률", "실전 포지션비중", "실전 포지션 비중", "방향 비중", "백어택 비중", "헤드어택 비중", "백어택 raw 비중", "헤드어택 raw 비중", "무방향 raw 비중", "치명타 적중률", "치명타 비중", "총 치피", "진화형 피해", "대상 스킬 기준 활용률", "전체 피해 기준 조건 비중", "조건부 방향 피해(%)", "전역 피해(%)", "치명타 적중률 증가(%)", "공격력 증가(%)", "치명타 피해 증가(%)", "랜덤 페널티 확률(%)", "랜덤 페널티 감소율(%)"]
    for c in money_cols:
        if c in out.columns:
            out[c] = out[c].apply(format_korean_number)
    for c in ratio_cols:
        if c in out.columns:
            out[c] = out[c].apply(lambda v: "-" if v is None or (isinstance(v, float) and math.isnan(v)) else f"{float(v) * 100:.2f}%")
    for c in percent_cols:
        if c in out.columns:
            out[c] = out[c].apply(lambda v: "-" if v is None or (isinstance(v, float) and math.isnan(v)) else f"{float(v):.2f}%")
    if "비교포함" in out.columns:
        out["비교포함"] = out["비교포함"].map({1: "포함", 0: "제외"}).fillna(out["비교포함"])
    return out


def inject_user_dashboard_css() -> None:
    """사용자용 화면을 더 깔끔하게 보이도록 최소한의 전역 스타일을 적용합니다."""
    st.markdown(
        """
        <style>
        .block-container {
            padding-top: 2.0rem;
            padding-bottom: 4rem;
            max-width: 1880px;
        }
        section[data-testid="stSidebar"] {
            background: #202129;
            border-right: 1px solid rgba(255,255,255,0.08);
        }
        section[data-testid="stSidebar"] div[data-testid="stMarkdownContainer"] p,
        section[data-testid="stSidebar"] label {
            font-size: 0.88rem;
        }
        div[data-testid="stMetric"] {
            background: rgba(255,255,255,0.035);
            border: 1px solid rgba(255,255,255,0.075);
            border-radius: 14px;
            padding: 16px 18px;
            min-height: 104px;
        }
        div[data-testid="stMetric"] label {
            color: rgba(255,255,255,0.72) !important;
        }
        div[data-testid="stMetricValue"] {
            font-weight: 800;
        }
        .loa-hero {
            padding: 24px 28px;
            border: 1px solid rgba(255,255,255,0.08);
            background: linear-gradient(135deg, rgba(255,78,90,0.16), rgba(67,139,255,0.08));
            border-radius: 20px;
            margin-bottom: 18px;
        }
        .loa-hero h1 {
            margin: 0 0 8px 0;
            font-size: 2.2rem;
            letter-spacing: -0.04em;
        }
        .loa-hero p {
            margin: 0;
            color: rgba(255,255,255,0.76);
            line-height: 1.55;
            font-size: 1.02rem;
        }
        .loa-section-note {
            color: rgba(255,255,255,0.62);
            margin-top: -0.3rem;
            margin-bottom: 1.0rem;
        }
        .stTabs [data-baseweb="tab-list"] {
            gap: 8px;
            border-bottom: 1px solid rgba(255,255,255,0.08);
        }
        .stTabs [data-baseweb="tab"] {
            padding: 12px 14px;
            border-radius: 10px 10px 0 0;
        }
        div[data-testid="stDataFrame"] {
            border-radius: 12px;
            overflow: hidden;
        }
        div[data-testid="stDataFrame"] div[role="gridcell"],
        div[data-testid="stDataFrame"] div[role="columnheader"] {
            line-height: 1.35;
        }
        div[data-testid="stTextInput"] input {
            text-overflow: clip;
        }
        div[data-testid="stDataFrame"] {
            font-size: 0.92rem;
        }
        div[data-testid="stDataFrame"] div[role="columnheader"] {
            white-space: normal !important;
            min-height: 36px;
        }
        .loa-wide-note {
            color: rgba(255,255,255,0.64);
            font-size: 0.92rem;
            margin-bottom: 0.75rem;
        }
        .loa-character-wrap {
            display: grid;
            grid-template-columns: 220px minmax(0, 1fr);
            gap: 18px;
            align-items: stretch;
            margin: 0.25rem 0 1.2rem 0;
        }
        .loa-character-image-box {
            border: 1px solid rgba(255,255,255,0.10);
            background: rgba(255,255,255,0.035);
            border-radius: 16px;
            padding: 12px;
            min-height: 260px;
        }
        .loa-character-card {
            border: 1px solid rgba(255,255,255,0.10);
            background: rgba(255,255,255,0.035);
            border-radius: 16px;
            padding: 20px 22px;
            min-height: 260px;
        }
        .loa-character-name {
            font-size: 2.0rem;
            font-weight: 900;
            line-height: 1.15;
            letter-spacing: -0.04em;
            margin-bottom: 0.35rem;
        }
        .loa-character-sub {
            color: rgba(255,255,255,0.72);
            font-size: 1.04rem;
            margin-bottom: 1.1rem;
        }
        .loa-character-grid {
            display: grid;
            grid-template-columns: repeat(5, minmax(120px, 1fr));
            gap: 10px;
        }
        .loa-character-pill {
            border: 1px solid rgba(255,255,255,0.08);
            background: rgba(0,0,0,0.18);
            border-radius: 12px;
            padding: 12px 14px;
        }
        .loa-character-pill span {
            display: block;
            color: rgba(255,255,255,0.56);
            font-size: 0.82rem;
            margin-bottom: 0.25rem;
        }
        .loa-character-pill b {
            font-size: 1.08rem;
        }
        .loa-metric-note {
            border: 1px solid rgba(255,255,255,0.08);
            background: rgba(255,255,255,0.028);
            border-radius: 14px;
            padding: 14px 16px;
            color: rgba(255,255,255,0.70);
            line-height: 1.65;
            font-size: 0.94rem;
        }
        .loa-metric-note b { color: rgba(255,255,255,0.94); }
        @media (max-width: 980px) {
            .loa-character-wrap { grid-template-columns: 1fr; }
            .loa-character-grid { grid-template-columns: repeat(2, minmax(120px, 1fr)); }
        }
        h1, h2, h3 {
            letter-spacing: -0.035em;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

def sidebar_controls() -> None:
    """좌측 패널은 실제 사용자 입력에 필요한 최소 항목만 둡니다."""
    calc_config = load_yaml(str(CONFIG_DIR / "calculation_presets.yaml"))
    defaults = calc_config.get("defaults", {})

    # 화면에서 숨긴 계산 기본값. 내부 계산식과 기존 함수 호환을 위해 session_state에는 유지합니다.
    st.session_state.setdefault("api_crit_basis", "백어택 기준(조건부 포함)")
    st.session_state.setdefault("target_crit_percent", float(defaults.get("target_crit_percent", 100.0)))
    st.session_state.setdefault("target_back_percent", float(defaults.get("target_back_attack_percent", 100.0)))
    st.session_state.setdefault("back_bonus_percent", float(defaults.get("back_attack_damage_bonus_percent", 5.0)))

    with st.sidebar:
        st.markdown("### LOA 실전 분석")
        pass  # 버전 변경 안내 제거
        st.caption("캐릭터를 불러오고, 목표 상승률만 정한 뒤 전투분석기 이미지를 넣으면 됩니다.")

        # 디버그 도구(OCR 원문/인식 과정 보기)는 평소엔 숨기고, 문제 진단 시에만 켭니다.
        st.session_state["show_debug_tools"] = st.checkbox(
            "🔧 디버그 도구 표시",
            value=bool(st.session_state.get("show_debug_tools", False)),
            help="OCR 원문/인식 과정 보기, 디버그 ZIP 생성 등을 표시합니다. 문제 진단이 필요할 때만 켜세요.",
        )

        token = st.text_input(
            "Lost Ark API Key",
            type="password",
            key="api_token",
            help="developer-lostark에서 발급받은 JWT/API KEY를 넣으세요.",
        )
        character_name = st.text_input(
            "캐릭터명",
            value=st.session_state.get("character_name", ""),
            key="character_name",
            placeholder="캐릭터명을 입력하세요",
        )
        st.number_input(
            "목표 상승률(%)",
            value=float(st.session_state.get("target_gain_percent", defaults.get("target_gain_percent", 5.0))),
            min_value=0.0,
            max_value=100.0,
            step=0.5,
            key="target_gain_percent",
            help="개선분석에서 목표로 삼을 총 딜 상승률입니다.",
        )

        fetch = st.button("검색", type="primary", use_container_width=True, key="sidebar_fetch_api")

        # v89: 검색 버튼 바로 아래에 성능 로그 버튼을 인라인으로 고정합니다.
        # 이전 run.bat가 구버전 폴더로 이동하던 문제를 함께 수정했으므로 이 문구가 보이면 최신 코드가 실행 중입니다.
        st.markdown("**⏱ 성능 병목 분석**")
        if st.button("실전 성능 타임라인 ZIP 생성", key="sidebar_real_perf_timeline_v89_clean", use_container_width=True):
            if not easyocr_available():
                st.error("OCR 엔진이 설치되어 있지 않아 성능 로그를 만들 수 없습니다.")
            else:
                summary_image = _get_uploaded_image_from_session_v88("real", "summary")
                attack_image = _get_uploaded_image_from_session_v88("real", "attack")
                if summary_image is None and attack_image is None:
                    st.warning("먼저 실전 전투분석기 종합정보/공격정보 이미지를 업로드하고 OCR을 한 번 실행해 주세요.")
                else:
                    with st.spinner("실전 OCR + 아이콘 매칭 성능 타임라인 수집 중..."):
                        report, perf_zip, perf_parsed = _run_ocr_perf_timeline_debug(
                            "real",
                            "실전 전투분석기",
                            summary_image,
                            attack_image,
                            gpu=bool(st.session_state.get("real_gpu_v34", False)),
                            row_count=int(st.session_state.get("real_row_count", 18) or 18),
                            ocr_scale=int(st.session_state.get("real_ocr_scale", 7) or 7),
                        )
                    st.session_state["real_perf_zip_path"] = str(perf_zip)
                    st.success("성능 타임라인 ZIP을 생성했습니다.")
        perf_zip_path = st.session_state.get("real_perf_zip_path")
        if perf_zip_path and Path(perf_zip_path).exists():
            st.download_button(
                "실전 성능 타임라인 ZIP 다운로드",
                data=Path(perf_zip_path).read_bytes(),
                file_name="real_perf_timeline.zip",
                mime="application/zip",
                key="sidebar_real_perf_timeline_download_v89_clean",
                use_container_width=True,
            )

        if fetch:
            if not token or not character_name:
                st.warning("API Key와 캐릭터명을 먼저 입력해줘.")
            else:
                with st.spinner("캐릭터 셋팅을 불러오는 중..."):
                    client = LostArkApiClient(token, CONFIG_DIR / "api_endpoints.yaml")
                    bundle = client.fetch_armory_bundle(character_name)
                    st.session_state.api_bundle = serializable_bundle(bundle)
                    st.session_state.api_summary = enrich_summary_with_identity(summarize_all(bundle), st.session_state.api_bundle)
                failed = {k: v for k, v in st.session_state.api_bundle.items() if not v.get("ok")}
                if failed:
                    st.warning(f"일부 API 호출 실패: {', '.join(failed.keys())}")
                else:
                    st.success("캐릭터 셋팅을 불러왔어.")

        summary = st.session_state.get("api_summary") or {}
        profile = summary.get("profile_summary", {}) if isinstance(summary, dict) else {}
        if profile:
            st.divider()
            st.caption("현재 분석 캐릭터")
            st.markdown(f"**{profile.get('캐릭터명') or character_name}**")
            meta = " · ".join(str(x) for x in [profile.get("서버"), profile.get("클래스"), profile.get("직업"), profile.get("아이템레벨")] if x)
            if meta:
                st.caption(meta)


def show_horizontal_damage_chart(top: pd.DataFrame) -> None:
    chart = top[["name", "damage", "share_rate"]].dropna(subset=["name", "damage"]).copy()
    if chart.empty:
        st.info("차트로 표시할 스킬 피해량이 없습니다.")
        return
    chart["damage_eok"] = chart["damage"] / 100_000_000
    chart["피해량"] = chart["damage"].apply(format_korean_number)
    chart["피해 지분"] = chart["share_rate"].apply(fmt_percent)
    chart["라벨"] = chart.apply(lambda r: f"{r['피해량']} ({r['피해 지분']})", axis=1)

    if alt is None:
        st.bar_chart(chart.set_index("name")[["damage_eok"]], use_container_width=True)
        return

    base = alt.Chart(chart).encode(
        y=alt.Y("name:N", sort="-x", title="스킬명", axis=alt.Axis(labelAngle=0)),
        x=alt.X("damage_eok:Q", title="피해량(억)", axis=alt.Axis(format=",.0f")),
        tooltip=[
            alt.Tooltip("name:N", title="스킬명"),
            alt.Tooltip("피해량:N"),
            alt.Tooltip("피해 지분:N"),
        ],
    )
    bars = base.mark_bar()
    text = base.mark_text(align="left", baseline="middle", dx=4).encode(text="라벨:N")
    st.altair_chart((bars + text).properties(height=max(320, len(chart) * 34)), use_container_width=True)
st.set_page_config(
    page_title="LOA 실전 효율 분석기",
    page_icon="⚔️",
    layout="wide",
)


@st.cache_data(show_spinner=False)
def load_yaml(path: str) -> Dict[str, Any]:
    with open(path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


def init_state() -> None:
    defaults = {
        "api_bundle": None,
        "api_summary": None,
        "real_table": empty_battle_table(),
        "dummy_table": empty_battle_table(),
        "real_elapsed_sec": 0.0,
        "dummy_elapsed_sec": 0.0,
        "manual_real_dps": None,
        "manual_real_dps_input": 0.0,
        "manual_dummy_dps": None,
        "manual_dummy_dps_input": 0.0,
        "real_meta": {},
        "dummy_meta": {},
        "last_ocr_lines": [],
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v


def metric_card(label: str, value: Any, help_text: str | None = None) -> None:
    st.metric(label=label, value=value if value is not None else "-", help=help_text)


def show_dataframe(df: pd.DataFrame, title: str, height: int = 320) -> None:
    st.subheader(title)
    if df is None or df.empty:
        st.info("표시할 데이터가 아직 없습니다.")
        return
    st.dataframe(df, use_container_width=True, height=height)


def _df_csv_bytes(df: pd.DataFrame) -> bytes:
    """엑셀에서 한글이 깨지지 않도록 UTF-8 BOM CSV로 변환합니다."""
    if df is None or df.empty:
        return b""
    return df.to_csv(index=False).encode("utf-8-sig")


def _json_safe_for_export(value: Any) -> Any:
    """Streamlit 세션/요약값을 JSON으로 안전하게 내보내기 위한 변환기입니다."""
    if isinstance(value, pd.DataFrame):
        return value.fillna("").to_dict(orient="records")
    if isinstance(value, dict):
        return {str(k): _json_safe_for_export(v) for k, v in value.items()}
    if isinstance(value, (list, tuple)):
        return [_json_safe_for_export(v) for v in value]
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            pass
    if value is None:
        return None
    try:
        if isinstance(value, float) and math.isnan(value):
            return None
    except Exception:
        pass
    return value


def _api_source_tables(summary: dict[str, Any]) -> dict[str, pd.DataFrame]:
    """치적/치피/피해군 이상값 검수에 필요한 API 출처표 묶음입니다."""
    if not isinstance(summary, dict):
        return {}
    keys = [
        ("damage_sources", "피해군_감지_출처"),
        ("crit_sources", "치명치피_출처"),
        ("base_damage_sources", "기준계산_출처_아크그리드제외"),
        ("arkgrid_damage_sources", "아크그리드_계산출처"),
        ("unresolved_sources", "자동매칭실패_검수필요"),
        ("loawa_like_base_breakdown", "로아와식_합산_아크그리드제외"),
        ("loawa_like_breakdown", "로아와식_합산_아크그리드포함"),
        ("combat_overview", "전투수치요약"),
        ("skill_crit_estimates", "최종계산표"),
        ("lostbuilds_base_skill_estimates", "기준계산표"),
        ("arkgrid_delta_estimates", "아크그리드추가분"),
        ("engravings", "각인"),
        ("arkpassive", "아크패시브"),
        ("arkgrid", "아크그리드"),
    ]
    tables: dict[str, pd.DataFrame] = {}
    for key, label in keys:
        df = summary.get(key)
        if isinstance(df, pd.DataFrame) and not df.empty:
            tables[label] = df.copy()
    return tables


def _api_source_export_zip(summary: dict[str, Any], bundle: dict[str, Any] | None = None) -> bytes:
    """피해군/치명 출처와 API 원본을 하나의 ZIP으로 묶습니다."""
    mem = io.BytesIO()
    with zipfile.ZipFile(mem, "w", zipfile.ZIP_DEFLATED) as zf:
        tables = _api_source_tables(summary)
        for name, df in tables.items():
            zf.writestr(f"{name}.csv", _df_csv_bytes(df))
        compact_summary = {
            "profile_summary": summary.get("profile_summary"),
            "base_crit_stat": summary.get("base_crit_stat"),
            "base_crit_percent": summary.get("base_crit_percent"),
            "avg_back_skill_crit_rate_percent": summary.get("avg_back_skill_crit_rate_percent"),
            "avg_back_basis_crit_rate_percent": summary.get("avg_back_basis_crit_rate_percent"),
            "global_crit_damage_no_skill_percent": summary.get("global_crit_damage_no_skill_percent"),
            "avg_conditional_crit_damage_percent": summary.get("avg_conditional_crit_damage_percent"),
            "avg_evolution_damage_percent": summary.get("avg_evolution_damage_percent"),
            "avg_enemy_damage_percent": summary.get("avg_enemy_damage_percent"),
            "avg_enemy_damage_unweighted_percent": summary.get("avg_enemy_damage_unweighted_percent"),
            "global_enemy_damage_percent": summary.get("global_enemy_damage_percent"),
            "avg_gem_damage_percent": summary.get("avg_gem_damage_percent"),
            "avg_gem_damage_unweighted_percent": summary.get("avg_gem_damage_unweighted_percent"),
            "avg_gem_damage_on_gem_skills_percent": summary.get("avg_gem_damage_on_gem_skills_percent"),
            "avg_final_multiplier": summary.get("avg_final_multiplier"),
            "base_crit_raw": summary.get("base_crit_raw"),
            "estimator_error": summary.get("estimator_error"),
            "_calc_version": summary.get("_calc_version"),
        }
        zf.writestr("summary_compact.json", json.dumps(_json_safe_for_export(compact_summary), ensure_ascii=False, indent=2))
        if bundle:
            zf.writestr("api_raw_bundle.json", json.dumps(_json_safe_for_export(bundle), ensure_ascii=False, indent=2))
    return mem.getvalue()



def ensure_api_summary_current() -> None:
    """코드 교체 후에도 session_state에 남은 오래된 api_summary를 자동 재계산합니다."""
    bundle = st.session_state.get("api_bundle")
    if not bundle:
        return
    summary = st.session_state.get("api_summary")
    needs = not isinstance(summary, dict)
    if not needs:
        if summary.get("_fast_setting_preview_v155"):
            needs = True
    if not needs:
        final_df = summary.get("arkgrid_final_skill_estimates")
        if not isinstance(final_df, pd.DataFrame) or final_df.empty:
            final_df = summary.get("skill_crit_estimates")
        skills_df = summary.get("skills")
        has_skills = isinstance(skills_df, pd.DataFrame) and not filter_adopted_skills_df(skills_df).empty
        missing_calc = not isinstance(final_df, pd.DataFrame) or final_df.empty
        needs = has_skills and missing_calc
        if not needs and summary.get("_calc_version") != APP_CALC_VERSION:
            summary["_calc_version"] = APP_CALC_VERSION
    if needs:
        with st.spinner("API 원본 기준으로 계산표를 다시 생성하는 중..."):
            new_summary = enrich_summary_with_identity(summarize_all(bundle), bundle)
            if isinstance(new_summary, dict):
                new_summary["_calc_version"] = APP_CALC_VERSION
            st.session_state.api_summary = new_summary


def _html_escape(value: Any) -> str:
    return html.escape(str(value if value is not None and value != "" else "-"))


def _summary_breakdown_value(summary: dict[str, Any], group_name: str) -> float | None:
    """로아와식 합산표에서 특정 피해군 합계를 꺼냅니다."""
    df = summary.get("loawa_like_breakdown")
    if not isinstance(df, pd.DataFrame) or df.empty:
        return None
    if "피해군" not in df.columns or "합계(%)" not in df.columns:
        return None
    hit = df[df["피해군"].astype(str).eq(group_name)]
    if hit.empty:
        return None
    val = pd.to_numeric(hit.iloc[0].get("합계(%)"), errors="coerce")
    try:
        return None if pd.isna(val) else float(val)
    except Exception:
        return None


def _avg_positive_metric(summary: dict[str, Any], col: str) -> float | None:
    df = summary.get("arkgrid_final_skill_estimates")
    if not isinstance(df, pd.DataFrame) or df.empty or col not in df.columns:
        return None
    vals = pd.to_numeric(df[col], errors="coerce")
    vals = vals[vals > 0]
    if vals.empty:
        return None
    return float(vals.mean())


def render_character_profile_card(profile: dict[str, Any]) -> None:
    """캐릭터 이미지와 기본 정보를 한 카드에 모아 보여줍니다."""
    name = _html_escape(profile.get("캐릭터명"))
    cls = _html_escape(profile.get("클래스"))
    job = _html_escape(profile.get("직업") or "추정 실패")
    server = _html_escape(profile.get("서버"))
    item_lv = _html_escape(profile.get("아이템레벨"))
    battle_lv = _html_escape(profile.get("전투레벨"))
    title = _html_escape(profile.get("칭호") or "-")
    guild = _html_escape(profile.get("길드") or "-")
    reason = _html_escape(profile.get("직업 추정 근거") or "아크패시브 깨달음/각인 원문 기준")

    img_col, info_col = st.columns([0.16, 0.84])
    with img_col:
        if profile.get("이미지"):
            st.image(profile.get("이미지"), width=210)
        else:
            st.info("캐릭터 이미지 없음")
    with info_col:
        st.markdown(
            f"""
            <div class="loa-character-card">
                <div class="loa-character-name">{name}</div>
                <div class="loa-character-sub">{cls} · {job} · {server}</div>
                <div class="loa-character-grid">
                    <div class="loa-character-pill"><span>클래스</span><b>{cls}</b></div>
                    <div class="loa-character-pill"><span>직업</span><b>{job}</b></div>
                    <div class="loa-character-pill"><span>아이템 레벨</span><b>{item_lv}</b></div>
                    <div class="loa-character-pill"><span>전투 레벨</span><b>{battle_lv}</b></div>
                    <div class="loa-character-pill"><span>서버</span><b>{server}</b></div>
                    <div class="loa-character-pill"><span>칭호</span><b>{title}</b></div>
                    <div class="loa-character-pill"><span>길드</span><b>{guild}</b></div>
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )



def show_summary_meta(meta: Dict[str, Any], title: str) -> None:
    if not meta:
        st.info(f"{title} 종합정보 이미지가 아직 인식되지 않았어.")
        return
    cols = st.columns(3)
    elapsed = meta.get("elapsed_seconds")
    elapsed_label = f"{int(elapsed // 60):02d}:{int(elapsed % 60):02d}" if elapsed else "-"
    elapsed_source = meta.get("elapsed_source")
    help_text = "이미지 OCR로 읽은 전투 시간입니다. v26부터 총피해량/DPS 역산은 하지 않습니다."
    if elapsed_source == "ocr_failed":
        help_text = "전투 시간 OCR 실패. 종합정보 수동 보정에서 직접 입력하세요."
    cols[0].metric("전투 시간", elapsed_label, help=help_text)
    cols[1].metric("총 피해량", meta.get("total_damage_text") or "-")
    cols[2].metric("DPS", meta.get("dps_text") or "-")
    if st.session_state.get("show_debug_tools", False):
        with st.expander(f"{title} 종합정보 OCR 원문"):
            st.json(meta.get("raw", {}))



def _sync_meta_to_manual_inputs(kind: str, meta: dict[str, Any]) -> None:
    """종합정보 OCR 결과를 수동 입력값에도 반영합니다.

    결과 계산은 meta를 우선 사용하지만, 사용자가 한눈에 확인/수정할 수 있도록
    전투 시간과 허수 DPS 입력창도 OCR 결과로 자동 채웁니다.
    """
    if not isinstance(meta, dict):
        return
    elapsed = meta.get("elapsed_seconds")
    if elapsed is not None:
        try:
            st.session_state[f"{kind}_elapsed_sec"] = float(elapsed)
        except Exception:
            pass
    if meta.get("dps"):
        try:
            if kind == "real":
                st.session_state.manual_real_dps = float(meta.get("dps"))
                st.session_state.manual_real_dps_input = float(meta.get("dps"))
            else:
                st.session_state.manual_dummy_dps = float(meta.get("dps"))
                st.session_state.manual_dummy_dps_input = float(meta.get("dps"))
        except Exception:
            pass
    if meta.get("total_damage"):
        try:
            st.session_state[f"manual_{kind}_total_damage_input"] = float(meta.get("total_damage"))
        except Exception:
            pass


def _compact_battle_column_config(skill_options: list[str] | None = None) -> dict[str, Any]:
    """전투분석기 검수표를 한 화면에 최대한 들어오게 하는 컬럼 설정.

    v25: 스킬명은 Selectbox가 아니라 TextColumn으로 둡니다.
    SelectboxColumn은 후보 밖의 글자를 직접 입력하면 Enter 후 값이 원복되는 경우가 있어,
    사용자가 표 안에서 직접 타이핑한 이름이 그대로 저장되도록 TextColumn을 사용합니다.
    """
    return {
        "이름": st.column_config.TextColumn("이름", width=150, help="직접 입력 가능합니다. Enter 또는 다음 칸 이동 후 값이 저장됩니다."),
        "피해량": st.column_config.TextColumn("피해량", width=96),
        "치명타 적중률": st.column_config.NumberColumn("치명타 적중률", format="%.2f%%", min_value=0.0, max_value=100.0, width=104),
        "치명타 비중": st.column_config.NumberColumn("치명타 비중", format="%.2f%%", min_value=0.0, max_value=100.0, width=100),
        "백어택 적중률": st.column_config.NumberColumn("백어택 적중률", format="%.2f%%", min_value=0.0, max_value=100.0, width=104),
        "백어택 비중": st.column_config.NumberColumn("백어택 비중", format="%.2f%%", min_value=0.0, max_value=100.0, width=100),
        "헤드어택 적중률": st.column_config.NumberColumn("헤드어택 적중률", format="%.2f%%", min_value=0.0, max_value=100.0, width=110),
        "헤드어택 비중": st.column_config.NumberColumn("헤드어택 비중", format="%.2f%%", min_value=0.0, max_value=100.0, width=104),
        "사용 횟수": st.column_config.NumberColumn("사용 횟수", min_value=0, step=1, width=80),
        "쿨타임 비율": st.column_config.NumberColumn("쿨타임 비율", format="%.2f%%", min_value=0.0, max_value=100.0, width=96),
        "피해량 지분": st.column_config.NumberColumn("피해량 지분", format="%.2f%%", min_value=0.0, max_value=100.0, width=96),
        "초당 피해량": st.column_config.TextColumn("초당 피해량", width=100),
    }




def _repair_battle_values_from_summary(df: pd.DataFrame, meta: dict[str, Any] | None, *, replace_damage_from_share: bool = False, elapsed_fallback: float | None = None) -> pd.DataFrame:
    """종합정보 총피해/전투시간으로 DPS와 피해량 지분을 역산합니다.

    v36 핵심:
    - 신규 전투분석기 공격정보는 `초당 피해량`과 `피해량 지분`이 빠지고
      `백어택 비중`, `치명타 비중`이 들어옵니다.
    - 스킬 DPS = 스킬 피해량 / 전투 시간
    - 스킬 피해량 지분 = 스킬 피해량 / 종합정보 총 피해량
    - 종합정보 총 피해량 - 표에 보이는 피해량 합계가 양수이면 `기타 추가딜` 행으로 더합니다.

    elapsed_fallback: 종합정보 OCR 이 전투시간을 못 읽었을 때(meta 에 elapsed_seconds 없음)
        수동 입력한 전투시간(예: st.session_state[f"{kind}_elapsed_sec"])으로 폴백해
        스킬별 DPS 역산을 계속 가능하게 합니다.
    """
    if df is None or df.empty:
        return df if isinstance(df, pd.DataFrame) else pd.DataFrame()
    out = df.copy()
    meta = meta or {}
    total_damage = meta.get("total_damage")
    elapsed = meta.get("elapsed_seconds")
    try:
        total_damage = float(total_damage) if total_damage is not None else None
    except Exception:
        total_damage = None
    try:
        elapsed = float(elapsed) if elapsed is not None else None
    except Exception:
        elapsed = None
    # 종합정보 OCR 전투시간이 없으면 수동 입력 전투시간으로 폴백합니다.
    if (elapsed is None or elapsed <= 0) and elapsed_fallback is not None:
        try:
            fb = float(elapsed_fallback)
            if fb > 0:
                elapsed = fb
        except Exception:
            pass

    from modules.calculators import parse_percent

    # 필요한 역산 컬럼이 없으면 생성합니다.
    if "초당 피해량" not in out.columns:
        out["초당 피해량"] = ""
    if "피해량 지분" not in out.columns:
        out["피해량 지분"] = ""

    # 이전에 자동 생성한 기타 추가딜 행은 다시 계산하기 위해 제거합니다.
    if "이름" in out.columns:
        out = out[out["이름"].astype(str).str.strip().ne("기타 추가딜")].reset_index(drop=True)

    for idx, row in out.iterrows():
        current_damage = parse_korean_number(row.get("피해량")) if "피해량" in out.columns else None
        share = parse_percent(row.get("피해량 지분")) if "피해량 지분" in out.columns else None

        # 구버전 표처럼 지분이 있고 피해량이 비어 있을 때만 지분으로 피해량을 복구합니다.
        if replace_damage_from_share and (current_damage is None or current_damage <= 0) and total_damage and share is not None and share > 0:
            current_damage = total_damage * share / 100.0
            out.at[idx, "피해량"] = format_korean_number(current_damage)

        if current_damage and current_damage > 0:
            if elapsed and elapsed > 0:
                out.at[idx, "초당 피해량"] = format_korean_number(current_damage / elapsed)
            if total_damage and total_damage > 0:
                out.at[idx, "피해량 지분"] = round(current_damage / total_damage * 100.0, 2)

    # 표에 보이는 행들의 피해량 합계와 종합정보 총 피해량 차이를 `기타 추가딜`로 보정합니다.
    if total_damage and total_damage > 0 and "피해량" in out.columns:
        visible_sum = 0.0
        for _, row in out.iterrows():
            dmg = parse_korean_number(row.get("피해량"))
            if dmg and dmg > 0:
                visible_sum += float(dmg)
        diff = float(total_damage) - visible_sum
        tolerance = max(100_000_000.0, float(total_damage) * 0.001)  # 1억 또는 0.1% 이하는 반올림 오차로 봄
        if diff > tolerance:
            extra = {c: "" for c in out.columns}
            extra["이름"] = "기타 추가딜"
            extra["피해량"] = format_korean_number(diff)
            if elapsed and elapsed > 0:
                extra["초당 피해량"] = format_korean_number(diff / elapsed)
            extra["피해량 지분"] = round(diff / float(total_damage) * 100.0, 2)
            # 기타 추가딜은 세부 적중/비중을 모르기 때문에 빈 값으로 둡니다.
            out = pd.concat([out, pd.DataFrame([extra])], ignore_index=True)
        elif diff < -tolerance:
            out.attrs["damage_sum_warning"] = f"표 피해량 합계가 종합정보보다 {format_korean_number(abs(diff))} 큽니다. OCR 피해량을 검수하세요."

    return out


def _drop_obvious_ocr_noise_rows(df: pd.DataFrame) -> pd.DataFrame:
    """스킬룬/채팅 찌꺼기처럼 계산에 거의 의미 없는 OCR 행을 정리합니다.

    피해량 지분이 없고, 사용 횟수도 없거나 0이며, 피해량이 1억 미만이면 잡음으로 봅니다.
    """
    if df is None or df.empty:
        return df if isinstance(df, pd.DataFrame) else pd.DataFrame()
    out = df.copy()
    keep = []
    from modules.calculators import parse_percent, parse_int
    for _, row in out.iterrows():
        share = parse_percent(row.get("피해량 지분"))
        casts = parse_int(row.get("사용 횟수"))
        damage = parse_korean_number(row.get("피해량"))
        name = str(row.get("이름") or "").strip()
        has_meaningful_share = share is not None and share > 0
        has_meaningful_damage = damage is not None and damage >= 100_000_000
        has_casts = casts is not None and casts > 0
        # 스킬명이 있고 피해량 지분/피해량/횟수 중 하나라도 의미 있으면 유지합니다.
        keep.append(bool(name and (has_meaningful_share or has_meaningful_damage or has_casts)))
    return out.loc[keep].reset_index(drop=True)





# ==============================
# v44: API 스킬 아이콘으로 확인 안 된 행은 기타 추가딜로 합산
# ==============================




def _aggregate_unmatched_rows_as_extra(df: pd.DataFrame, *, label: str = "인식 안된 기타 추가딜") -> pd.DataFrame:
    """API 스킬로 확인되지 않은 OCR 행을 하나의 기타 딜 행으로 합산합니다.

    스킬룬/보주/기타/기본공격처럼 전투분석기에는 보이지만 현재 API 스킬 아이콘으로
    확정할 수 없는 행은 개별 스킬 계산에서 제외하고, 딜 합계만 보존합니다.
    """
    if df is None or df.empty or "이름" not in df.columns:
        return df if isinstance(df, pd.DataFrame) else pd.DataFrame()

    from modules.calculators import parse_percent, parse_int

    keep_rows: list[pd.Series] = []
    extra_damage = 0.0
    extra_casts = 0
    extra_cool_weighted: list[tuple[float, float]] = []
    extra_names: list[str] = []

    for _, row in df.iterrows():
        name = str(row.get("이름") or "").strip()
        if not name or name in {"이름", "피해량", "DPS", "지분%", "백적중%", "백비중%", "치적%", "치비중%", "횟수", "쿨%"}:
            continue

        # 이미 자동 생성된 행은 다시 만들지 않습니다.
        if name in {label, "기타 추가딜"}:
            continue

        matched = _is_api_skill_matched_row(row)
        dmg = parse_korean_number(row.get("피해량")) if "피해량" in df.columns else None
        casts = parse_int(row.get("사용 횟수")) if "사용 횟수" in df.columns else None
        cool = parse_percent(row.get("쿨타임 비율")) if "쿨타임 비율" in df.columns else None

        if matched:
            keep_rows.append(row)
            continue

        # API 스킬로 확정되지 않았더라도 피해량이 없으면 계산에 의미가 없으므로 버립니다.
        if dmg is None or dmg <= 0:
            continue

        extra_damage += float(dmg)
        if casts is not None and casts > 0:
            extra_casts += int(casts)
        if cool is not None and cool >= 0:
            extra_cool_weighted.append((float(cool), float(dmg)))
        if name not in extra_names:
            extra_names.append(name)

    out = pd.DataFrame(keep_rows).reset_index(drop=True) if keep_rows else pd.DataFrame(columns=df.columns)

    if extra_damage > 0:
        extra = {c: "" for c in df.columns}
        extra["이름"] = label
        extra["피해량"] = format_korean_number(extra_damage)
        if "사용 횟수" in df.columns and extra_casts > 0:
            extra["사용 횟수"] = extra_casts
        if "쿨타임 비율" in df.columns and extra_cool_weighted:
            total_w = sum(w for _, w in extra_cool_weighted)
            if total_w > 0:
                extra["쿨타임 비율"] = round(sum(v * w for v, w in extra_cool_weighted) / total_w, 2)
        extra["_unmatched_aggregated"] = ", ".join(extra_names[:12])
        out = pd.concat([out, pd.DataFrame([extra])], ignore_index=True)

    # 원래 컬럼 순서 유지
    cols = [c for c in df.columns if c in out.columns]
    rest = [c for c in out.columns if c not in cols]
    return out[cols + rest].reset_index(drop=True)


def _bump_table_version(kind: str) -> None:
    st.session_state[f"{kind}_table_version"] = int(st.session_state.get(f"{kind}_table_version", 0) or 0) + 1


def _png_bytes(img: Any) -> bytes:
    buf = io.BytesIO()
    try:
        img.save(buf, format="PNG")
        return buf.getvalue()
    except Exception:
        return b""


def _save_ocr_debug_zip(kind: str, title: str, summary_image: Any, attack_image: Any, summary_debug: list[dict], attack_debug: list[dict]) -> Path:
    """OCR 디버그 자료를 zip으로 저장합니다.

    사용자가 이 파일을 보내주면 어떤 crop이 어느 텍스트로 읽혔는지 바로 확인할 수 있습니다.
    """
    EXPORT_DIR.mkdir(exist_ok=True)
    path = EXPORT_DIR / f"{kind}_ocr_debug.zip"
    summary_rows = []
    attack_rows = []
    with zipfile.ZipFile(path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        info = {
            "title": title,
            "kind": kind,
            "created_at": datetime.now().isoformat(timespec="seconds"),
            "note": "v96: Damage/DPS guide 유지 + 아이콘 crop 원복 + 공격정보 피해량/DPS 위험 글리프 차단.",
        }
        zf.writestr(
            "README.txt",
            "이 zip을 그대로 보내주시면 OCR 좌표/인식 문제를 확인할 수 있습니다.\n"
            "\n[v44] icons/ 폴더 설명:\n"
            " - row_XX_core.png  : 전투분석기에서 잘라 실제 매칭에 들어가는 '정규화된 아이콘'(가운데 정렬)\n"
            " - row_XX_candN_<점수>_<이름>.png : 그 행에 대해 아이콘 점수 상위 후보 API 아이콘(파일명에 점수/이름 포함)\n"
            " core 이미지와 cand0 이미지가 눈으로 봐도 다른 아이콘이면 매칭이 틀린 것입니다.\n",
        )
        zf.writestr("debug_info.json", json.dumps(info, ensure_ascii=False, indent=2))
        if summary_image is not None:
            zf.writestr("original_summary.png", _png_bytes(summary_image))
        if attack_image is not None:
            zf.writestr("original_attack.png", _png_bytes(attack_image))

        for item in summary_debug or []:
            key = str(item.get("key") or "")
            label = str(item.get("label") or key)
            text_value = str(item.get("text") or "")
            summary_rows.append({"key": key, "label": label, "ocr_text": text_value, "preprocess": str(item.get("preprocess") or ""), "score": item.get("score"), "window_box": str(item.get("window_box"))})
            if item.get("raw_crop") is not None:
                zf.writestr(f"summary/raw_{key}.png", _png_bytes(item["raw_crop"]))
            if item.get("processed_crop") is not None:
                zf.writestr(f"summary/processed_{key}.png", _png_bytes(item["processed_crop"]))

        for row in attack_debug or []:
            ri = int(row.get("row_index", 0) or 0)
            if row.get("row_crop") is not None:
                zf.writestr(f"attack/row_{ri:02d}_full.png", _png_bytes(row["row_crop"]))
            if row.get("icon_crop") is not None:
                zf.writestr(f"attack/row_{ri:02d}_icon.png", _png_bytes(row["icon_crop"]))
            # v44: 실제 비교에 들어가는 정규화 아이콘 코어 + 상위 후보 API 아이콘 이미지
            if row.get("icon_core_crop") is not None:
                zf.writestr(f"icons/row_{ri:02d}_core.png", _png_bytes(row["icon_core_crop"]))
            for ci, cand in enumerate(row.get("icon_candidate_images") or []):
                if cand.get("img") is not None:
                    safe_name = re.sub(r"[^0-9A-Za-z가-힣]+", "_", str(cand.get("name") or "")).strip("_")[:24]
                    score = cand.get("icon_score") or 0.0
                    zf.writestr(f"icons/row_{ri:02d}_cand{ci}_{score:0>5.1f}_{safe_name}.png", _png_bytes(cand["img"]))
            icon_match = row.get("icon_match") or {}
            if icon_match:
                attack_rows.append({
                    "row": ri,
                    "key": "icon_match",
                    "label": "아이콘 매칭",
                    "ocr_text": str(icon_match.get("name") or ""),
                    "preprocess": str(icon_match.get("reason") or ""),
                    "score": icon_match.get("icon_score"),
                    "window_box": str(row.get("window_box")),
                    "top_candidates": json.dumps(icon_match.get("top") or [], ensure_ascii=False),
                })
            for cell in row.get("cells", []) or []:
                key = str(cell.get("key") or "")
                label = str(cell.get("label") or key)
                text_value = str(cell.get("text") or "")
                attack_rows.append({"row": ri, "key": key, "label": label, "ocr_text": text_value, "preprocess": str(cell.get("preprocess") or ""), "score": cell.get("score"), "window_box": str(row.get("window_box"))})
                if cell.get("raw_crop") is not None:
                    zf.writestr(f"attack/row_{ri:02d}_{key}_raw.png", _png_bytes(cell["raw_crop"]))
                if cell.get("processed_crop") is not None:
                    zf.writestr(f"attack/row_{ri:02d}_{key}_processed.png", _png_bytes(cell["processed_crop"]))

        if summary_rows:
            zf.writestr("summary_ocr.csv", pd.DataFrame(summary_rows).to_csv(index=False).encode("utf-8-sig"))
        if attack_rows:
            zf.writestr("attack_ocr.csv", pd.DataFrame(attack_rows).to_csv(index=False).encode("utf-8-sig"))
    return path


def _text_input_value(key: str, default: Any) -> str:
    if key not in st.session_state:
        if default is None:
            st.session_state[key] = ""
        else:
            try:
                if pd.isna(default):
                    st.session_state[key] = ""
                else:
                    st.session_state[key] = str(default)
            except Exception:
                st.session_state[key] = str(default)

    # v87: 기존 세션에 남아 있는 Bleed rune / Poison rune text_input 값도 즉시 교체합니다.
    # Streamlit은 widget key가 같으면 default가 바뀌어도 기존 session_state 값을 우선하므로 여기서 직접 보정해야 합니다.
    try:
        if ("_이름" in str(key)) or str(key).endswith("_name") or str(key).endswith("_스킬명"):
            fixed = _normalize_special_battle_name_v86(st.session_state.get(key))
            if fixed != st.session_state.get(key):
                st.session_state[key] = fixed
    except Exception:
        pass
    return str(st.session_state.get(key) or "")


def render_compact_battle_editor(kind: str, df: pd.DataFrame) -> pd.DataFrame:
    """data_editor 대신 일반 입력칸으로 만든 검수표.

    Streamlit data_editor는 Enter가 셀 확정이 아니라 다음 칸 이동으로 처리되면서 값이 늦게 반영되는 경우가 있어,
    각 셀을 text_input으로 직접 렌더링합니다. Enter/Tab/마우스 이동 후 값이 session_state에 바로 남습니다.
    """
    if df is None or df.empty:
        df = empty_battle_table()
    df = prepare_battle_editor_df(df).copy()
    # 검수표 표시 순서: 이름 → 피해량 → 치명타 → 백어택 → 헤드어택 → 사용횟수 → 쿨타임 → 피해량지분 → 초당피해량
    _display_order = [
        "이름", "피해량",
        "치명타 적중률", "치명타 비중",
        "백어택 적중률", "백어택 비중",
        "헤드어택 적중률", "헤드어택 비중",
        "사용 횟수", "쿨타임 비율",
        "피해량 지분", "초당 피해량",
    ]
    cols = [c for c in _display_order if c in df.columns]
    df = df[cols].copy()
    # 사용자가 누락 행을 직접 추가할 수 있게 빈 행 2개를 항상 제공합니다.
    extra = pd.DataFrame([{c: "" for c in cols} for _ in range(2)])
    render_df = pd.concat([df, extra], ignore_index=True)
    version = int(st.session_state.get(f"{kind}_table_version", 0) or 0)

    widths = {
        "이름": 2.20,
        "피해량": 1.10,
        "치명타 적중률": 1.10,
        "치명타 비중": 1.05,
        "백어택 적중률": 1.10,
        "백어택 비중": 1.05,
        "헤드어택 적중률": 1.15,
        "헤드어택 비중": 1.10,
        "사용 횟수": 0.85,
        "쿨타임 비율": 1.00,
        "피해량 지분": 1.00,
        "초당 피해량": 1.08,
    }
    labels = {
        "이름": "이름",
        "피해량": "피해량",
        "치명타 적중률": "치명타 적중률",
        "치명타 비중": "치명타 비중",
        "백어택 적중률": "백어택 적중률",
        "백어택 비중": "백어택 비중",
        "헤드어택 적중률": "헤드어택 적중률",
        "헤드어택 비중": "헤드어택 비중",
        "사용 횟수": "사용 횟수",
        "쿨타임 비율": "쿨타임 비율",
        "피해량 지분": "피해량 지분",
        "초당 피해량": "초당 피해량",
    }

    st.markdown("<style>div[data-testid='stTextInput'] input{padding:0.22rem 0.40rem;font-size:0.88rem;min-width:0;} .loa-row-gap{height:0.10rem;}</style>", unsafe_allow_html=True)
    header_cols = st.columns([widths.get(c, 1.0) for c in cols], gap="small")
    for c, col in zip(cols, header_cols):
        col.caption(labels.get(c, c))

    output_rows = []
    for i, row in render_df.iterrows():
        row_cols = st.columns([widths.get(c, 1.0) for c in cols], gap="small")
        out_row = {}
        for c, col in zip(cols, row_cols):
            widget_key = f"{kind}_cell_{APP_CALC_VERSION}_{version}_{i}_{c}"
            default = row.get(c, "")
            if c in {"이름", "name", "스킬명"}:
                default = _normalize_special_battle_name_v86(default)
            _text_input_value(widget_key, default)
            with col:
                out_row[c] = st.text_input(
                    labels.get(c, c),
                    key=widget_key,
                    label_visibility="collapsed",
                    placeholder=labels.get(c, c),
                )
        # 완전히 빈 행은 반환하지 않습니다.
        if any(str(v or "").strip() for v in out_row.values()):
            output_rows.append(out_row)
    return pd.DataFrame(output_rows, columns=cols)




def _get_uploaded_image_from_session_v88(kind: str, slot: str) -> Any:
    """v88: 사이드바 디버그 버튼에서 전투분석기 업로드 이미지를 안전하게 꺼냅니다."""
    raw_key = f"{kind}_{slot}_image_bytes"
    raw = st.session_state.get(raw_key)
    if raw:
        try:
            return Image.open(io.BytesIO(raw)).convert("RGB")
        except Exception:
            pass

    # 아직 battle_input_panel에서 raw bytes를 저장하기 전이어도 file_uploader widget state에 남아 있으면 사용합니다.
    upload_key = f"{kind}_{slot}_img"
    uploaded = st.session_state.get(upload_key)
    if uploaded is not None:
        try:
            return image_from_upload(uploaded).convert("RGB")
        except Exception:
            try:
                return Image.open(io.BytesIO(uploaded.getvalue())).convert("RGB")
            except Exception:
                return None
    return None


def _render_sidebar_perf_timeline_debug_button() -> None:
    """v88: 검색 버튼 바로 아래에 항상 보이는 실전 전투분석기 성능 로그 버튼."""
    st.markdown("---")
    st.caption("⏱ OCR 성능 디버그")
    if st.button("실전 성능 타임라인 ZIP 생성", key="sidebar_real_perf_timeline_v88", use_container_width=True):
        summary_image = _get_uploaded_image_from_session_v88("real", "summary")
        attack_image = _get_uploaded_image_from_session_v88("real", "attack")
        if not easyocr_available():
            st.error("EasyOCR/RapidOCR이 설치되어 있지 않아 성능 로그를 만들 수 없습니다.")
        elif summary_image is None and attack_image is None:
            st.warning("실전 전투분석기 종합정보/공격정보 이미지를 먼저 업로드해 주세요.")
        else:
            with st.spinner("실전 OCR + 아이콘 매칭 성능 타임라인 수집 중..."):
                report, perf_zip, perf_parsed = _run_ocr_perf_timeline_debug(
                    "real",
                    "실전 전투분석기",
                    summary_image,
                    attack_image,
                    gpu=bool(st.session_state.get("real_gpu_v34", True)),
                    row_count=int(st.session_state.get("real_row_count", 18) or 18),
                    ocr_scale=int(st.session_state.get("real_ocr_scale", 7) or 7),
                    name_match_threshold=float(st.session_state.get("real_name_match_threshold", 52) or 52) / 100.0,
                    icon_match_threshold=float(st.session_state.get("real_icon_match_threshold", 74) or 74) / 100.0,
                )
            st.session_state["sidebar_real_perf_report_v88"] = report
            st.session_state["sidebar_real_perf_zip_path_v88"] = str(perf_zip)
            if isinstance(perf_parsed, pd.DataFrame):
                st.session_state["sidebar_real_perf_parsed_v88"] = perf_parsed
            st.success("성능 타임라인 ZIP을 생성했습니다.")

    perf_path = st.session_state.get("sidebar_real_perf_zip_path_v88")
    if perf_path and Path(perf_path).exists():
        st.download_button(
            "실전 성능 타임라인 ZIP 다운로드",
            data=Path(perf_path).read_bytes(),
            file_name=Path(perf_path).name,
            mime="application/zip",
            key="sidebar_real_perf_timeline_download_v88",
            use_container_width=True,
        )
    report = st.session_state.get("sidebar_real_perf_report_v88") or {}
    if report:
        try:
            st.caption(
                f"총 {float(report.get('total_elapsed_sec') or 0):.2f}s · "
                f"readtext {int(report.get('easyocr_calls') or 0)}회 · "
                f"readtext {float(report.get('easyocr_total_sec') or 0):.2f}s"
            )
        except Exception:
            pass

def _save_perf_timeline_zip(kind: str, title: str, report: dict[str, Any], parsed_df: pd.DataFrame | None = None) -> Path:
    """성능 타임라인/프로파일 결과를 zip으로 저장합니다."""
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_kind = re.sub(r"[^0-9A-Za-z가-힣_-]+", "_", str(kind or "ocr")) or "ocr"
    path = EXPORT_DIR / f"{safe_kind}_ocr_perf_timeline_{ts}.zip"
    events = report.get("events") or []
    phases = report.get("phases_sec") or {}
    with zipfile.ZipFile(path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("perf_report.json", json.dumps(report, ensure_ascii=False, indent=2))
        if events:
            zf.writestr("perf_events.csv", pd.DataFrame(events).to_csv(index=False).encode("utf-8-sig"))
        if phases:
            phase_df = pd.DataFrame([{"stage": k, "sec": v, "ms": round(float(v) * 1000.0, 3)} for k, v in phases.items()])
            phase_df = phase_df.sort_values("sec", ascending=False)
            zf.writestr("perf_phases.csv", phase_df.to_csv(index=False).encode("utf-8-sig"))
        if isinstance(parsed_df, pd.DataFrame) and not parsed_df.empty:
            zf.writestr("parsed_attack_after_icon_match.csv", parsed_df.to_csv(index=False).encode("utf-8-sig"))
        zf.writestr("README.txt", (
            f"{title} OCR 성능 타임라인 로그\n"
            "\n"
            "perf_report.json: 전체 요약/환경/단계별 시간\n"
            "perf_events.csv: timestamp가 찍힌 순차 이벤트 로그\n"
            "perf_phases.csv: 단계별 누적 시간 내림차순\n"
            "parsed_attack_after_icon_match.csv: 아이콘 보정까지 끝난 공격정보 표\n"
        ).encode("utf-8"))
    return path


def _run_ocr_perf_timeline_debug(
    kind: str,
    title: str,
    summary_image: Any,
    attack_image: Any,
    *,
    gpu: bool,
    row_count: int,
    ocr_scale: int,
    name_match_threshold: float = 0.52,
    icon_match_threshold: float = 0.74,
) -> tuple[dict[str, Any], Path, pd.DataFrame | None]:
    """실제 OCR+아이콘 보정 경로를 한 번 돌리며 timestamp 타임라인을 수집합니다."""
    import modules.fixed_grid_ocr as fgo

    old_profile = fgo.perf_profile_begin(label=f"{kind}:{title}")
    parsed: pd.DataFrame | None = None
    app_events: list[dict[str, Any]] = []
    app_t0 = time.perf_counter()

    def log(stage: str, detail: str = "", **data: Any) -> None:
        now = time.perf_counter()
        row: dict[str, Any] = {
            "timestamp": datetime.now().isoformat(timespec="milliseconds"),
            "elapsed_ms": round((now - app_t0) * 1000.0, 3),
            "stage": stage,
            "detail": detail,
        }
        for k, v in data.items():
            try:
                row[k] = v if isinstance(v, (str, int, float, bool)) or v is None else str(v)
            except Exception:
                pass
        app_events.append(row)
        try:
            fgo.perf_trace_event("app:" + stage, detail, **data)
        except Exception:
            pass

    try:
        log("reader_start", gpu=bool(gpu))
        t = time.perf_counter()
        reader_raw = get_easyocr_reader(gpu=gpu)
        reader = fgo.perf_wrap_reader(reader_raw)
        log("reader_done", elapsed_ms=round((time.perf_counter() - t) * 1000.0, 3), reader=type(reader_raw).__name__)

        if summary_image is not None:
            log("summary_parse_start", size=f"{summary_image.width}x{summary_image.height}")
            t = time.perf_counter()
            try:
                meta = parse_summary_fixed_grid(summary_image, reader)
                log(
                    "summary_parse_done",
                    elapsed_ms=round((time.perf_counter() - t) * 1000.0, 3),
                    elapsed_seconds=meta.get("elapsed_seconds"),
                    total_damage_text=meta.get("total_damage_text"),
                    dps_text=meta.get("dps_text"),
                )
            except Exception as e:  # noqa: BLE001
                log("summary_parse_error", error=repr(e))

        if attack_image is not None:
            log("attack_parse_start", size=f"{attack_image.width}x{attack_image.height}", row_count=int(row_count), scale=int(ocr_scale))
            t = time.perf_counter()
            try:
                parsed = parse_attack_fixed_grid(attack_image, reader, row_count=int(row_count), scale=int(ocr_scale))
                log("attack_parse_done", elapsed_ms=round((time.perf_counter() - t) * 1000.0, 3), rows=0 if parsed is None else len(parsed))
            except Exception as e:  # noqa: BLE001
                log("attack_parse_error", error=repr(e))
                parsed = pd.DataFrame()

            if parsed is not None and not parsed.empty:
                log("candidate_load_start")
                t = time.perf_counter()
                candidates_full = get_ocr_skill_candidates_full()
                skill_icon_candidates = _skill_only_ocr_candidates(candidates_full)
                log(
                    "candidate_load_done",
                    elapsed_ms=round((time.perf_counter() - t) * 1000.0, 3),
                    candidates_full=len(candidates_full),
                    skill_icon_candidates=len(skill_icon_candidates),
                )

                if skill_icon_candidates:
                    log("icon_correction_start", rows=len(parsed), candidates=len(skill_icon_candidates))
                    t = time.perf_counter()
                    parsed = correct_battle_skill_names_with_icons(
                        parsed,
                        attack_image,
                        skill_icon_candidates,
                        threshold=float(name_match_threshold),
                        icon_threshold=float(icon_match_threshold),
                        drop_unmatched=False,
                    )
                    log("icon_correction_done", elapsed_ms=round((time.perf_counter() - t) * 1000.0, 3), rows=len(parsed))
                else:
                    log("icon_correction_skip", reason="no_candidates")

                log("postprocess_start")
                t = time.perf_counter()
                try:
                    parsed = sanitize_battle_table(parsed)
                    parsed = _aggregate_unmatched_rows_as_extra(parsed)
                    parsed = _drop_obvious_ocr_noise_rows(parsed)
                except Exception as e:  # noqa: BLE001
                    log("postprocess_error", error=repr(e))
                log("postprocess_done", elapsed_ms=round((time.perf_counter() - t) * 1000.0, 3), rows=len(parsed))

    finally:
        report = fgo.perf_profile_end(old_profile)

    report["app_events"] = app_events
    # app_events도 events에 합쳐서 시간순으로 보기 쉽게 둡니다.
    merged_events = []
    merged_events.extend(report.get("events") or [])
    for i, ev in enumerate(app_events, start=len(merged_events) + 1):
        merged = dict(ev)
        merged["idx"] = i
        merged["stage"] = "app:" + str(merged.get("stage", ""))
        merged_events.append(merged)
    report["events"] = merged_events
    report["total_elapsed_sec"] = round(time.perf_counter() - app_t0, 3)
    # 상위 병목 후보를 간단 계산합니다.
    phase_rows = []
    for k, v in (report.get("phases_sec") or {}).items():
        try:
            phase_rows.append({"stage": k, "sec": float(v)})
        except Exception:
            pass
    app_phase_rows = []
    for ev in app_events:
        if str(ev.get("stage", "")).endswith("_done") and "elapsed_ms" in ev:
            try:
                app_phase_rows.append({"stage": "app:" + str(ev.get("stage")), "sec": float(ev.get("elapsed_ms")) / 1000.0})
            except Exception:
                pass
    bottlenecks = sorted(phase_rows + app_phase_rows, key=lambda x: x.get("sec", 0.0), reverse=True)[:12]
    report["bottlenecks_top"] = bottlenecks
    report["verdict"] = (
        f"총 {report.get('total_elapsed_sec', 0):.2f}s / readtext {report.get('easyocr_calls', 0)}회 "
        f"{report.get('easyocr_total_sec', 0):.2f}s / 후보 병목 1위: "
        f"{bottlenecks[0]['stage']} {bottlenecks[0]['sec']:.2f}s" if bottlenecks else "프로파일 이벤트가 부족합니다."
    )
    zip_path = _save_perf_timeline_zip(kind, title, report, parsed)
    return report, zip_path, parsed

def _show_ocr_debug_view(
    kind: str,
    title: str,
    summary_image: Any,
    attack_image: Any,
    *,
    gpu: bool,
    row_count: int,
    ocr_scale: int,
) -> None:
    """OCR이 실제로 어떤 영역을 잘라 읽는지 확인하는 디버그 UI."""
    with st.expander(f"{title} OCR 인식 과정 보기", expanded=False):
        st.caption(
            "v97 적용됨: FAST 모드 유지 + 아이콘은 행 높이 기준 정사각형 crop을 먼저 시도하고, 점수가 낮으면 기존 넓은 crop으로 자동 fallback합니다. "
            "여기서 보이는 crop이 실제 OCR에 들어가는 영역입니다. "
            "전투시간/스킬명이 누락되면 먼저 crop 위치가 맞는지 확인하세요."
        )

        # v85: 사용자가 OCR 디버그 버튼 바로 위에서 찾을 수 있도록 성능 타임라인 버튼을 고정 노출합니다.
        st.markdown("**⏱ 성능 타임라인 로그**")
        st.caption(
            "느린 구간을 찾기 위해 실제 OCR + 숫자 인식 + 아이콘 매칭 경로를 한 번 더 실행하고, "
            "timestamp가 찍힌 CSV/JSON ZIP을 만듭니다. 이 버튼으로 만든 ZIP을 보내주면 병목을 바로 분석할 수 있습니다."
        )
        if st.button(f"{title} 성능 타임라인 ZIP 생성", key=f"{kind}_make_perf_timeline_v85_fixed_visible", use_container_width=True):
            if not easyocr_available():
                st.error("EasyOCR/RapidOCR이 설치되어 있지 않아 성능 로그를 만들 수 없습니다.")
            elif summary_image is None and attack_image is None:
                st.warning("성능 로그를 만들 이미지가 없습니다. 종합정보 또는 공격정보 이미지를 올려주세요.")
            else:
                with st.spinner("OCR + 아이콘 매칭 성능 타임라인 수집 중..."):
                    report, perf_zip, perf_parsed = _run_ocr_perf_timeline_debug(
                        kind,
                        title,
                        summary_image,
                        attack_image,
                        gpu=bool(gpu),
                        row_count=int(row_count),
                        ocr_scale=int(ocr_scale),
                        name_match_threshold=float(st.session_state.get(f"{kind}_name_match_threshold", 52)) / 100.0,
                        icon_match_threshold=float(st.session_state.get(f"{kind}_icon_match_threshold", 74)) / 100.0,
                    )
                st.session_state[f"{kind}_perf_report"] = report
                st.session_state[f"{kind}_perf_zip_path"] = str(perf_zip)
                if isinstance(perf_parsed, pd.DataFrame):
                    st.session_state[f"{kind}_perf_parsed"] = perf_parsed
                st.success("성능 타임라인 ZIP을 생성했습니다. 아래 다운로드 버튼으로 저장해서 보내주세요.")

        perf_report = st.session_state.get(f"{kind}_perf_report") or {}
        perf_zip_path = st.session_state.get(f"{kind}_perf_zip_path")
        if perf_report:
            st.markdown("**성능 요약**")
            m1, m2, m3, m4 = st.columns(4)
            m1.metric("총 시간", f"{float(perf_report.get('total_elapsed_sec') or 0):.2f}s")
            m2.metric("readtext 호출", int(perf_report.get("easyocr_calls") or 0))
            m3.metric("readtext 시간", f"{float(perf_report.get('easyocr_total_sec') or 0):.2f}s")
            glyph = perf_report.get("glyph", {}).get("attack", {}) if isinstance(perf_report.get("glyph"), dict) else {}
            m4.metric("글리프 채택", f"{int(glyph.get('accept') or 0)}/{int(glyph.get('attempt') or 0)}")
            if perf_report.get("verdict"):
                st.info(str(perf_report.get("verdict")))

            bottlenecks = perf_report.get("bottlenecks_top") or []
            if bottlenecks:
                st.markdown("**병목 후보 TOP**")
                st.dataframe(pd.DataFrame(bottlenecks), use_container_width=True, hide_index=True, height=260)

            events = perf_report.get("events") or []
            if events:
                with st.expander("timestamp 이벤트 로그 미리보기", expanded=False):
                    st.dataframe(pd.DataFrame(events).tail(200), use_container_width=True, hide_index=True, height=360)

        if perf_zip_path and Path(perf_zip_path).exists():
            st.download_button(
                f"{title} 성능 타임라인 ZIP 다운로드",
                data=Path(perf_zip_path).read_bytes(),
                file_name=Path(perf_zip_path).name,
                mime="application/zip",
                key=f"{kind}_perf_zip_download_v85_fixed_visible",
                use_container_width=True,
            )

        st.divider()
        st.markdown("**🧩 OCR crop 디버그**")

        if st.button(f"{title} OCR 디버그 생성", key=f"{kind}_make_ocr_debug"):
            if not easyocr_available():
                st.error("EasyOCR/RapidOCR이 설치되어 있지 않아 디버그를 만들 수 없습니다.")
            else:
                # v86: 사용자가 실제로 보이는 이 버튼 하나로 OCR crop 디버그와 성능 타임라인을 같이 생성합니다.
                # 별도 성능 버튼이 화면에 안 보이는 환경에서도 이 버튼만 누르면 병목 분석 ZIP까지 만들어집니다.
                if summary_image is not None or attack_image is not None:
                    with st.spinner("성능 타임라인 수집 중..."):
                        try:
                            report, perf_zip, perf_parsed = _run_ocr_perf_timeline_debug(
                                kind,
                                title,
                                summary_image,
                                attack_image,
                                gpu=bool(gpu),
                                row_count=int(row_count),
                                ocr_scale=int(ocr_scale),
                                name_match_threshold=float(st.session_state.get(f"{kind}_name_match_threshold", 52)) / 100.0,
                                icon_match_threshold=float(st.session_state.get(f"{kind}_icon_match_threshold", 74)) / 100.0,
                            )
                            st.session_state[f"{kind}_perf_report"] = report
                            st.session_state[f"{kind}_perf_zip_path"] = str(perf_zip)
                            if isinstance(perf_parsed, pd.DataFrame):
                                st.session_state[f"{kind}_perf_parsed"] = perf_parsed
                        except Exception as e:  # noqa: BLE001
                            st.warning(f"성능 타임라인 생성 중 오류가 났습니다: {e}")
                reader = get_easyocr_reader(gpu=gpu)
                if summary_image is not None:
                    st.session_state[f"{kind}_summary_debug"] = make_summary_ocr_debug(summary_image, reader, scale=int(ocr_scale))
                if attack_image is not None:
                    st.session_state[f"{kind}_attack_debug"] = make_attack_ocr_debug(attack_image, reader, row_count=int(row_count), scale=int(ocr_scale), skill_candidates=_skill_only_ocr_candidates(get_ocr_skill_candidates_full()))
                debug_path = _save_ocr_debug_zip(
                    kind,
                    title,
                    summary_image,
                    attack_image,
                    st.session_state.get(f"{kind}_summary_debug") or [],
                    st.session_state.get(f"{kind}_attack_debug") or [],
                )
                st.session_state[f"{kind}_debug_zip_path"] = str(debug_path)
                st.success("OCR 디버그 ZIP과 성능 타임라인 ZIP을 함께 생성했습니다. 아래 다운로드 버튼으로 둘 다 저장해서 보내주세요.")

        debug_zip_path = st.session_state.get(f"{kind}_debug_zip_path")
        if debug_zip_path and Path(debug_zip_path).exists():
            st.download_button(
                f"{title} OCR 디버그 ZIP 다운로드",
                data=Path(debug_zip_path).read_bytes(),
                file_name=f"{kind}_ocr_debug.zip",
                mime="application/zip",
                key=f"{kind}_debug_zip_download",
                use_container_width=True,
            )

        # v87: 성능 ZIP은 별도 버튼이 안 보여도 OCR 디버그 버튼 하나로 생성되고, 바로 아래에서 받을 수 있게 고정 노출합니다.
        _perf_zip_path_v87 = st.session_state.get(f"{kind}_perf_zip_path")
        if _perf_zip_path_v87 and Path(_perf_zip_path_v87).exists():
            st.download_button(
                f"{title} 성능 타임라인 ZIP 다운로드",
                data=Path(_perf_zip_path_v87).read_bytes(),
                file_name=Path(_perf_zip_path_v87).name,
                mime="application/zip",
                key=f"{kind}_perf_zip_download_under_debug_v87",
                use_container_width=True,
            )

        summary_debug = st.session_state.get(f"{kind}_summary_debug") or []
        if summary_debug:
            st.markdown("**종합정보 OCR crop**")
            table_rows = []
            for item in summary_debug:
                table_rows.append({"영역": item.get("label"), "OCR 원문": item.get("text")})
            st.dataframe(pd.DataFrame(table_rows), use_container_width=True, hide_index=True, height=180)
            cols = st.columns(4)
            for i, item in enumerate(summary_debug[:8]):
                with cols[i % 4]:
                    st.caption(f"{item.get('label')} → {item.get('text') or '-'}")
                    st.image(item.get("raw_crop"), caption="원본 crop", use_container_width=True)

        attack_debug = st.session_state.get(f"{kind}_attack_debug") or []
        if attack_debug:
            st.markdown("**공격정보 행/셀 OCR crop**")
            max_rows = st.slider(
                "표시할 디버그 행 수",
                min_value=1,
                max_value=max(1, len(attack_debug)),
                value=min(8, len(attack_debug)),
                key=f"{kind}_debug_row_limit",
            )
            for row in attack_debug[:max_rows]:
                with st.container(border=True):
                    st.caption(f"행 {row.get('row_index')}")
                    st.image(row.get("row_crop"), caption="행 전체 crop", use_container_width=True)
                    cell_rows = [{"열": c.get("label"), "OCR 원문": c.get("text")} for c in row.get("cells", [])]
                    st.dataframe(pd.DataFrame(cell_rows), use_container_width=True, hide_index=True, height=180)
                    cell_cols = st.columns(4)
                    for i, c in enumerate(row.get("cells", [])[:8]):
                        with cell_cols[i % 4]:
                            st.caption(f"{c.get('label')} → {c.get('text') or '-'}")
                            st.image(c.get("raw_crop"), use_container_width=True)


def battle_input_panel(kind: str, title: str) -> pd.DataFrame:
    key = f"{kind}_table"
    meta_key = f"{kind}_meta"
    st.subheader(title)

    source = st.radio(
        f"{title} 입력 방식",
        ["이미지 OCR(창/표 자동 탐지)", "CSV 업로드", "화면에서 직접 입력/수정"],
        horizontal=True,
        key=f"{kind}_source",
    )

    if source == "이미지 OCR(창/표 자동 탐지)":
        st.caption("이미지를 먼저 인식하고, 아래 표에서 스킬명/수치만 바로 검수하는 흐름입니다.")
        summary_up = st.file_uploader(f"{title} 종합정보 이미지", type=["png", "jpg", "jpeg", "webp"], key=f"{kind}_summary_img")
        attack_up = st.file_uploader(f"{title} 공격정보 이미지", type=["png", "jpg", "jpeg", "webp"], key=f"{kind}_attack_img")

        p1, p2 = st.columns(2)
        summary_image = None
        attack_image = None
        if summary_up is not None:
            summary_image = image_from_upload(summary_up)
            # v88: 사이드바 성능 디버그 버튼에서도 현재 업로드 이미지를 쓸 수 있게 raw bytes를 보관합니다.
            try:
                st.session_state[f"{kind}_summary_image_bytes"] = summary_up.getvalue()
                st.session_state[f"{kind}_summary_image_name"] = getattr(summary_up, "name", "summary.png")
            except Exception:
                pass
            with p1:
                st.image(summary_image, caption="종합정보 이미지", use_container_width=True)
        if attack_up is not None:
            attack_image = image_from_upload(attack_up)
            # v88: 사이드바 성능 디버그 버튼에서도 현재 업로드 이미지를 쓸 수 있게 raw bytes를 보관합니다.
            try:
                st.session_state[f"{kind}_attack_image_bytes"] = attack_up.getvalue()
                st.session_state[f"{kind}_attack_image_name"] = getattr(attack_up, "name", "attack.png")
            except Exception:
                pass
            with p2:
                st.image(attack_image, caption="공격정보 이미지", use_container_width=True)

        with st.expander("OCR 세부 설정", expanded=False):
            gpu = st.checkbox("EasyOCR GPU 사용", value=True, key=f"{kind}_gpu_v34", help="기본 ON입니다. CUDA PyTorch가 설치되어 있으면 GPU로 실행하고, 실패하면 자동으로 CPU로 fallback합니다.")
            row_count = st.number_input("공격정보 인식 행 수", min_value=5, max_value=24, value=18, step=1, key=f"{kind}_row_count")
            ocr_scale = st.number_input(
                "OCR 확대 배율",
                min_value=2,
                max_value=10,
                value=7,
                step=1,
                key=f"{kind}_ocr_scale",
                help="글자가 작게 깨지면 7~8을 테스트하세요. 너무 높이면 느려질 수 있습니다.",
            )
            auto_name_correct = st.checkbox(
                "API 스킬/룬/보주 후보로 이름 자동 보정",
                value=True,
                key=f"{kind}_auto_name_correct",
            )
            name_match_threshold = st.slider(
                "스킬명 보정 민감도",
                min_value=40,
                max_value=90,
                value=52,
                step=1,
                key=f"{kind}_name_match_threshold",
            )
            icon_name_correct = st.checkbox(
                "API 스킬 아이콘 우선으로 이름 보정",
                value=True,
                key=f"{kind}_icon_name_correct",
            )
            icon_match_threshold = st.slider(
                "아이콘 보정 신뢰도",
                min_value=45,
                max_value=95,
                value=74,
                step=1,
                key=f"{kind}_icon_match_threshold",
                help="아이콘이 맞는데 보정이 안 되면 낮추고, 틀린 아이콘으로 바뀌면 올리세요.",
            )
            aggregate_unmatched_rows = st.checkbox(
                "API 스킬로 확인 안 된 행은 인식 안된 기타 추가딜로 합산",
                value=False,
                key=f"{kind}_aggregate_unmatched_rows",
                help="스킬룬/보주/기타/기본공격처럼 API 스킬 아이콘으로 확정되지 않는 행은 개별 스킬에서 제외하고 하나의 기타 딜로 합산합니다.",
            )

        if st.button(f"{title} 이미지 OCR 실행", key=f"{kind}_grid_ocr_btn", type="primary", use_container_width=True):
            if not easyocr_available():
                st.error("EasyOCR이 설치되어 있지 않아. pip install easyocr 또는 run_streamlit.bat 실행이 필요해.")
            elif summary_image is None and attack_image is None:
                st.warning("종합정보 또는 공격정보 이미지 중 최소 1개는 올려줘.")
            else:
                with st.spinner("창/표 자동 탐지 OCR 인식 중... 처음 실행은 모델 로딩 때문에 시간이 걸릴 수 있어."):
                    reader = get_easyocr_reader(gpu=gpu)
                    if summary_image is not None:
                        meta = parse_summary_fixed_grid(summary_image, reader)
                        st.session_state[meta_key] = meta
                        _sync_meta_to_manual_inputs(kind, meta)
                    if attack_image is not None:
                        parsed = parse_attack_fixed_grid(attack_image, reader, row_count=int(row_count), scale=int(ocr_scale))
                        if auto_name_correct and parsed is not None and not parsed.empty:
                            candidates_full = get_ocr_skill_candidates_full()
                            skill_icon_candidates = _skill_only_ocr_candidates(candidates_full)
                            candidates = [c["name"] for c in skill_icon_candidates]
                            if skill_icon_candidates and icon_name_correct:
                                parsed = correct_battle_skill_names_with_icons(
                                    parsed,
                                    attack_image,
                                    skill_icon_candidates,
                                    threshold=float(name_match_threshold) / 100.0,
                                    icon_threshold=float(icon_match_threshold) / 100.0,
                                    drop_unmatched=False,
                                )
                                if aggregate_unmatched_rows:
                                    parsed = _aggregate_unmatched_rows_as_extra(parsed)
                            elif candidates:
                                parsed = correct_battle_skill_names(
                                    parsed,
                                    candidates,
                                    threshold=float(name_match_threshold) / 100.0,
                                )
                            else:
                                st.info("API 스킬 아이콘 후보가 없어 스킬명 자동 보정은 건너뛰었어. API 셋팅을 먼저 불러오면 보정 정확도가 올라가.")
                        if parsed.empty:
                            st.warning("공격정보 표를 못 읽었어. OCR 디버그 ZIP을 생성해서 창/헤더/아이콘 감지 상태를 확인해줘.")
                        else:
                            parsed = sanitize_battle_table(parsed)
                            parsed = _repair_battle_values_from_summary(parsed, st.session_state.get(meta_key, {}), elapsed_fallback=st.session_state.get(f"{kind}_elapsed_sec"))
                            parsed = _drop_obvious_ocr_noise_rows(parsed)
                            st.session_state[key] = prepare_battle_editor_df(sanitize_battle_table(parsed))
                            _bump_table_version(kind)
                # 좌우 대칭 UI를 위해 완료 알림 문구는 표시하지 않습니다.
                pass

        show_summary_meta(st.session_state.get(meta_key, {}), title)
        _manual_meta_editor(kind, title)
        if st.session_state.get("show_debug_tools", False):
            _show_ocr_debug_view(
                kind,
                title,
                summary_image,
                attack_image,
                gpu=bool(locals().get("gpu", False)),
                row_count=int(locals().get("row_count", 18)),
                ocr_scale=int(locals().get("ocr_scale", 7)),
            )
        st.caption("행이 부족하면 OCR 세부 설정에서 공격정보 인식 행 수를 18~20으로 올려보세요.")

    elif source == "CSV 업로드":
        csv = st.file_uploader(f"{title} CSV", type=["csv"], key=f"{kind}_csv")
        if csv is not None:
            try:
                st.session_state[key] = prepare_battle_editor_df(sanitize_battle_table(pd.read_csv(csv)))
                _bump_table_version(kind)
                st.success("CSV를 불러왔어. 아래 표에서 다시 검수할 수 있어.")
            except Exception as e:  # noqa: BLE001
                st.error(f"CSV 읽기 실패: {e}")
        _manual_meta_editor(kind, title)
    else:
        _manual_meta_editor(kind, title)

    st.session_state[key] = prepare_battle_editor_df(sanitize_battle_table(st.session_state[key]))
    st.session_state[key] = prepare_battle_editor_df(sanitize_battle_table(_repair_battle_values_from_summary(st.session_state[key], st.session_state.get(meta_key, {}), elapsed_fallback=st.session_state.get(f"{kind}_elapsed_sec"))))
    st.caption("아래 표가 계산에 실제로 들어가는 값입니다. 표 안의 입력칸에서 직접 수정하세요. Enter/Tab/다음 칸 이동 후 값이 저장됩니다.")
    editor_df_full = prepare_battle_editor_df(st.session_state[key])
    hidden_internal_cols = [c for c in editor_df_full.columns if str(c).startswith("_")]
    display_cols = [c for c in editor_df_full.columns if c not in hidden_internal_cols]
    editor_df = editor_df_full[display_cols].copy()
    if not editor_df.empty:
        editor_df = editor_df.dropna(how="all")
    edited = render_compact_battle_editor(kind, editor_df)

    # 입력칸 검수표에서 보이는 값만 수정하더라도, OCR 아이콘 좌표 같은 내부 컬럼은 유지합니다.
    # 그래야 사용자가 이름을 고친 뒤 그 아이콘을 로컬 캐시에 학습할 수 있습니다.
    merged = edited.copy()
    if hidden_internal_cols and not editor_df_full.empty:
        for c in hidden_internal_cols:
            vals = list(editor_df_full[c].values)
            merged[c] = [vals[i] if i < len(vals) else None for i in range(len(merged))]
    st.session_state[key] = prepare_battle_editor_df(sanitize_battle_table(merged))

    # v43: 전분 아이콘 템플릿 라이브러리는 사용하지 않습니다.
    # 이름 보정은 현재 API에서 불러온 스킬/룬/보주 후보 + 더 좁게 자른 행 아이콘 crop만 사용합니다.

    with st.expander("계산용 숫자 변환 확인", expanded=False):
        st.caption("표에는 `억/만/%` 단위가 보이지만, 계산에는 아래 숫자로 변환됩니다.")
        try:
            from modules.calculators import normalize_battle_df
            norm_preview = normalize_battle_df(st.session_state[key]).copy()
            if not norm_preview.empty:
                show_cols = [c for c in ["name", "damage", "dps", "back_attack_rate", "crit_rate", "share_rate", "casts"] if c in norm_preview.columns]
                display_norm = norm_preview[show_cols].copy()
                display_norm = display_norm.rename(columns={
                    "name": "이름",
                    "damage": "피해량_숫자",
                    "dps": "DPS_숫자",
                    "back_attack_rate": "백어택%_숫자",
                    "crit_rate": "치명%_숫자",
                    "share_rate": "지분%_숫자",
                    "casts": "사용횟수_숫자",
                })
                st.dataframe(display_norm, use_container_width=True, height=180, hide_index=True)
            else:
                st.info("변환할 전투분석기 행이 없습니다.")
        except Exception as e:
            st.warning(f"숫자 변환 확인 중 오류: {e}")
    return st.session_state[key]


def ocr_tab() -> None:
    st.header("전투분석기 입력")
    st.caption("실전과 허수 전투분석기 이미지를 넣고 OCR 결과를 검수합니다. 표 값은 바로 결과 계산에 반영됩니다.")

    _render_step_guide(2, "전투분석기 이미지 넣고 검수하기", [
        "실전/허수 전투분석기 <b>스크린샷</b>을 올리고 <b>OCR 실행</b>을 누르세요.",
        "스킬명·전투시간·총피해량·DPS를 확인하고, 틀렸으면 <b>종합정보 수동 보정</b>에서 직접 고치세요.",
        "고쳤으면 <b>✅ 다시 계산</b> 버튼을 누르고 <b>③ 실력 분석 결과</b> 탭으로 이동하세요.",
    ])
    st.markdown('<div class="loa-wide-note">표가 잘리지 않도록 실전/허수를 각각 넓은 화면에서 검수합니다. OCR 후 값은 바로 결과 계산에 반영됩니다.</div>', unsafe_allow_html=True)
    real_tab, dummy_tab = st.tabs(["실전 전투분석기", "허수/기준 전투분석기"])
    with real_tab:
        battle_input_panel("real", "실전 전투분석기")
    with dummy_tab:
        battle_input_panel("dummy", "허수/기준 전투분석기")


RESULT_TABLE_HELP = {
    "스킬명": "전투분석기 행을 API 채용 스킬 아이콘과 매칭한 이름입니다.",
    "공격타입": "API 스킬 툴팁 기준 백어택/헤드어택/없음입니다.",
    "스킬 구조 타입": "스킬이 구조적으로 낼 수 있는 방향입니다. 백어택 전용/헤드어택 전용/백·헤드 가능/무방향. API 공격타입을 우선하고, 없으면 실전 관측으로 보조 판정합니다.",
    "착용 방향 각인": "캐릭터가 착용한 방향 각인입니다. 기습의 대가=백어택, 결투의 대가=헤드어택. 둘은 동시에 착용할 수 없습니다.",
    "목표 방향": "구조 타입과 착용 각인으로 정해지는 평가 목표 방향입니다. 백/헤드 가능 스킬은 각인이 없으면 목표 방향이 없습니다(그래도 기본 보너스는 계산).",
    "기본 방향 보너스 활용률": "기습/결투 각인을 뺀 백/헤드 자체 보너스(백 +5%·치적 +10%, 헤드 +20%)를 얼마나 살렸는지입니다. 각인이 없어도 표시됩니다.",
    "방향 각인 효율": "기습/결투 각인의 조건부 보너스를 실제 딜로 전환한 비율입니다. 각인이 없으면 표시되지 않습니다.",
    "방향 손실 피해": "방향만 최적으로 들어갔다면 더 나왔을 피해입니다. 무방향 스킬은 0, 백/헤드 가능인데 각인이 없으면 참고용입니다.",
    "방향 각인 손실 피해": "착용한 방향 각인 기준의 손실입니다. 각인이 없으면 0입니다.",
    "백어택 raw 비중": "전투분석기 피해 비중에서 백/헤드 보너스를 제거해 역산한 실제 백어택 적중 비중입니다.",
    "헤드어택 raw 비중": "헤드는 기본 20% 보너스가 있어 결과 피해 비중이 과대평가됩니다. 보너스를 제거한 실제 헤드 적중 비중입니다.",
    "무방향 raw 비중": "백/헤드 어느 쪽도 아닌(정면 등) 상태로 들어간 실제 비중입니다.",
    "무력화 기대 보너스": "헤드어택 무력화 +10% 기대치입니다. DPS 점수에는 넣지 않고 별도로만 표시합니다.",
    "방향 데이터 신뢰도": "이 스킬의 방향 평가에 필요한 데이터가 실제로 있는지입니다. '없음'이면 관측 데이터가 부족해 방향 점수 해석에 주의해야 합니다.",
    "기습/결투 효율": "조건부 최대 보너스 중 실제로 살린 비율입니다. 예: 기습 유물 15%에서 효율 78.67%면 실질 추가 딜증은 15% × 78.67% = 약 11.80%입니다.",
    "기습/결투 실질 딜증": "각인 조건 성공으로 실제로 받은 조건부 추가 딜증입니다. 조건부 보너스 × 기습/결투 효율로 표시합니다.",
    "실전 관측": "실전 전투분석기에서 읽은 실제 피해량입니다.",
    "허수 페이스": "실전 스킬 1회 보정 피해에 허수 전분의 사용횟수 페이스를 곱한 값입니다. 허수 피해량을 단순 시간 환산하지 않습니다.",
    "치명타 보정 값": "실전 관측 피해를 API 기대 치명률 기준으로 보정한 값입니다. 관측 치명타 비중이 기대보다 높으면 낮아질 수 있고, 낮으면 올라갑니다.",
    "데미지 보정값 / 허수페이스": "치명운 보정 후 실전 피해가 허수 페이스 대비 어느 정도인지 보는 핵심 수행률입니다.",
    "쿨소화율": "실전 쿨타임 비율을 허수 쿨타임 비율과 비교한 값입니다.",
    "사용횟수율": "실전 CPM/사용횟수 페이스를 허수 CPM/사용횟수 페이스와 비교한 값입니다.",
    "실전 포지션률": "해당 스킬의 백어택 또는 헤드어택 적중률입니다.",
    "실전 포지션 비중": "해당 스킬 피해 중 백어택 또는 헤드어택으로 들어간 피해 비중입니다.",
    "백어택/헤드어택 비중": "공격타입에 따라 백어택 스킬은 백어택 비중, 헤드어택 스킬은 헤드어택 비중을 표시합니다.",
    "방향 비중": "백어택 스킬은 백어택 비중, 헤드어택 스킬은 헤드어택 비중을 표시합니다.",
    "치명타 비중": "해당 스킬 피해 중 치명타로 들어간 피해 비중입니다. 단순 치명타 적중률보다 피해 보정에 더 적합합니다.",
    "제외사유": "허수 기준이 없거나 기타/보조 항목이면 비교 계산에서 제외된 이유입니다.",
    "우선순위": "손실피해가 큰 순서입니다. 1순위일수록 개선했을 때 기대 상승량이 큽니다.",
    "기존 피해량": "치명운을 보정한 실전 평균 피해입니다.",
    "허수페이스 기대": "실전 1회 보정 피해를 허수 사용횟수 페이스로 굴렸을 때의 기대 피해입니다.",
    "손실피해": "허수페이스 기대 피해 - 기존 피해량입니다. 이 값이 클수록 성장 여지가 큽니다.",
    "현재 대비 상승여지": "손실피해를 현재 실전 평균 피해로 나눈 값입니다.",
    "치명운": "실전 관측 피해가 치명운 보정 평균보다 얼마나 높거나 낮았는지입니다. 양수면 운이 좋았고 음수면 운이 나빴습니다.",
    "개선포인트": "사용횟수, 쿨소화, 포지션, 사이클 중 우선 확인할 항목입니다.",
    "추천 액션": "개선포인트에 맞춘 간단한 실전 체크 방향입니다.",
    "목표 달성 필요 개선률": "목표 상승률을 달성하려면 해당 손실피해 중 몇 퍼센트를 회수해야 하는지입니다.",
    "예상 최소 DPS": "치명/예둔 랜덤성이 최저로 굴렀다고 가정한 이론 DPS입니다. 실제 전투에서는 타수 때문에 이보다 좁게 흔들릴 수 있습니다.",
    "예상 평균 DPS": "현재 치명운 보정값을 기준으로 한 평균 기대 DPS입니다.",
    "예상 최대 DPS": "치명/예둔 랜덤성이 최고로 굴렀다고 가정한 이론 DPS입니다.",
    "예상 최소 피해": "전투 전체 기준 최소 피해 추정값입니다.",
    "예상 평균 피해": "전투 전체 기준 평균 피해 추정값입니다.",
    "예상 최대 피해": "전투 전체 기준 최대 피해 추정값입니다.",
    "고려 효과": "DPS 범위 계산에 포함한 랜덤성/특수 효과입니다.",
    "효과": "API Tooltip 또는 피해군 출처에서 감지한 각인/진화 효과입니다.",
    "등급/단계": "API Tooltip 또는 정리표에서 읽은 각인 등급/단계입니다.",
    "치명타 적중률 증가(%)": "해당 효과가 올려주는 치명타 적중률입니다. 아드레날린은 최대 중첩 기준을 우선합니다.",
    "공격력 증가(%)": "해당 효과가 올려주는 공격력입니다. 아드레날린은 1스택 수치 × 최대 중첩으로 검수합니다.",
    "치명타 피해 증가(%)": "해당 효과가 올려주는 치명타 피해량입니다. 예리한 둔기 감지에 사용합니다.",
    "랜덤 페널티 확률(%)": "예리한 둔기 같은 랜덤 피해 감소 페널티가 발동할 확률입니다. Tooltip에서 읽은 경우만 표시됩니다.",
    "랜덤 페널티 감소율(%)": "랜덤 페널티 발동 시 감소하는 피해량입니다. Tooltip에서 읽은 경우만 표시됩니다.",
    "대상 스킬 기준 각인 효율": "조건부 최대 보너스 중 대상 백/헤드 스킬에서 실제로 살린 비율입니다.",
    "대상 스킬 기준 실질 딜증": "대상 백/헤드 스킬 기준으로 실제 받은 조건부 추가 딜증입니다. 조건부 보너스 × 각인 효율입니다.",
    "각인 실질 딜증": "전체 실전 피해 기준으로 역산한 조건부 추가 딜증입니다.",
    "핵심 수행률": "치명운 보정 실전 피해 ÷ 허수페이스 기대 피해입니다. 100%면 허수 사용횟수 페이스와 동일합니다.",
    "허수 대비 실전 효율": "결과 화면의 대표 수행률입니다. 치명운 보정 실전 피해와 허수페이스 기대 피해를 비교합니다.",
}


def result_table_column_config(df: pd.DataFrame) -> dict[str, Any]:
    cfg: dict[str, Any] = {}
    if df is None or df.empty:
        return cfg
    width_map = {
        "스킬명": 170,
        "공격타입": 92,
        "기습/결투 효율": 120,
        "실전 관측": 118,
        "허수 페이스": 128,
        "치명타 보정 값": 150,
        "데미지 보정값 / 허수페이스": 150,
        "쿨소화율": 92,
        "사용횟수율": 104,
        "실전 포지션률": 120,
        "실전 포지션 비중": 130,
        "백어택/헤드어택 비중": 145,
        "치명타 비중": 104,
        "제외사유": 150,
        "우선순위": 82,
        "기존 피해량": 118,
        "허수페이스 기대": 136,
        "손실피해": 118,
        "현재 대비 상승여지": 132,
        "치명운": 92,
        "개선포인트": 128,
        "추천 액션": 330,
        "목표 달성 필요 개선률": 160,
        "각인": 120,
        "각인 등급": 100,
        "조건": 90,
        "대상 스킬 기준 활용률": 145,
        "대상 스킬 기준 각인 효율": 160,
        "전체 피해 기준 조건 비중": 160,
        "조건부 방향 피해(%)": 140,
        "전역 피해(%)": 110,
        "대상 스킬 기준 실질 딜증": 160,
        "각인 실질 딜증": 130,
        "조건부 보너스 활용 효율": 160,
        "기습/결투 실질 딜증": 130,
        "대상 스킬 피해": 130,
        "감지 출처": 140,
        "고려 효과": 180,
        "예상 최소 피해": 130,
        "예상 평균 피해": 130,
        "예상 최대 피해": 130,
        "예상 최소 DPS": 130,
        "예상 평균 DPS": 130,
        "예상 최대 DPS": 130,
        "주의": 360,
        "효과": 120,
        "등급/단계": 110,
        "계산 메모": 360,
    }
    for col in df.columns:
        help_text = RESULT_TABLE_HELP.get(str(col))
        cfg[col] = st.column_config.TextColumn(str(col), help=help_text, width=width_map.get(str(col), 110))
    return cfg


def _result_tables_by_synergy(
    real_table: pd.DataFrame,
    dummy_table: pd.DataFrame,
    real_elapsed: float | None,
    dummy_elapsed: float | None,
    target_gain_rate: float,
) -> dict[int, tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, dict[str, Any]]]:
    """치적 시너지 0/1/2명 결과를 한 번에 계산해 세션에 캐시합니다.

    입력(표·전투시간·목표·특수효과·각인·캐릭터)이 바뀌면 다시 계산하고, 그 외에는
    캐시를 재사용하므로 사용자가 '치적 시너지 인원 수'만 바꾸면 즉시 전환됩니다.
    """
    import hashlib as _hl
    sig = _hl.md5()
    for _df in (real_table, dummy_table):
        try:
            sig.update(pd.util.hash_pandas_object(_df.fillna(0), index=True).values.tobytes())
        except Exception:
            sig.update(repr(_df).encode("utf-8", "ignore"))
    for _fn in (_detect_special_combat_effects, _detect_directional_engravings):
        try:
            sig.update(repr(_fn()).encode("utf-8", "ignore"))
        except Exception:
            pass
    _api_marker = ""
    try:
        _b = st.session_state.get("api_bundle") or {}
        _api_marker = str(_b.get("character_name") or _b.get("name") or "")
    except Exception:
        pass
    try:
        _ver = APP_CALC_VERSION
    except Exception:
        _ver = ""
    # 아르카나 카드 사용 횟수 등 확률 보정에 쓰이는 세션 값이 바뀌면 캐시를 무효화합니다.
    _extra = "|".join(
        f"{_k}={st.session_state.get(_k)}"
        for _k in ("arcana_card_uses", "arcana_card_uses_ocr")
    )
    sig.update(f"|{real_elapsed}|{dummy_elapsed}|{target_gain_rate}|{_api_marker}|{_ver}|{_extra}".encode("utf-8", "ignore"))
    key = sig.hexdigest()

    cache = st.session_state.get("_crit_synergy_result_cache")
    if isinstance(cache, dict) and cache.get("key") == key and isinstance(cache.get("results"), dict):
        return cache["results"]

    results: dict[int, tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, dict[str, Any]]] = {}
    for _cnt in (0, 1, 2):
        results[_cnt] = build_result_and_improvement_tables(
            real_table, dummy_table, real_elapsed, dummy_elapsed, target_gain_rate,
            crit_synergy_count=_cnt,
        )
    st.session_state["_crit_synergy_result_cache"] = {"key": key, "results": results}
    return results


def _result_meta_times() -> tuple[dict[str, Any], dict[str, Any], float | None, float | None]:
    real_meta = st.session_state.get("real_meta", {}) or {}
    dummy_meta = st.session_state.get("dummy_meta", {}) or {}
    real_elapsed = real_meta.get("elapsed_seconds") or st.session_state.get("real_elapsed_sec")
    dummy_elapsed = dummy_meta.get("elapsed_seconds") or (st.session_state.get("dummy_elapsed_sec") if st.session_state.get("dummy_elapsed_sec", 0) > 0 else None)
    return real_meta, dummy_meta, real_elapsed, dummy_elapsed


def result_tab() -> None:
    st.header("실력 분석 결과")
    st.caption("허수 대비 실전 수행률, 확률 보정값, 성장 여지를 객관적으로 보는 화면입니다.")

    _render_step_guide(3, "결과 읽는 법", [
        "<b>허수 대비 실전 효율</b>·<b>평가</b>가 핵심 지표예요. 치명타·뭉툭한 가시 같은 확률 운을 보정해 비교합니다.",
        "전투시간·총피해량·DPS가 이상하면 <b>② 탭 → 수동 보정 → 다시 계산</b>으로 갱신하세요.",
        "<b>개선분석</b> 탭에서 다음에 무엇을 올리면 좋은지 우선순위를 확인하세요.",
    ])

    target_crit_percent = float(st.session_state.get("target_crit_percent", 100.0))
    target_back_percent = float(st.session_state.get("target_back_percent", 100.0))
    back_bonus_percent = float(st.session_state.get("back_bonus_percent", 5.0))
    target_gain_rate = float(st.session_state.get("target_gain_percent", 5.0)) / 100.0

    # v152: 파티 치명타 적중률 시너지 인원 수(최대 2명). 실전 기대 치명률에만 +10%p/명 반영.
    st.session_state.setdefault("crit_synergy_count", 0)
    _syn_c1, _syn_c2 = st.columns([1, 2])
    with _syn_c1:
        st.number_input(
            "치적 시너지 인원 수",
            min_value=0,
            max_value=2,
            step=1,
            key="crit_synergy_count",
            help=(
                "파티에서 치명타 적중률 시너지를 주는 인원 수입니다(최대 2명, 1명당 +10%p). "
                "실전은 시너지로 치적이 올라가지만(예: 80→90) 허수/전분은 시너지 없이 측정된 값이라, "
                "이 값을 실전 기대 치명률에만 더해 시너지로 오른 치명 피해가 '치명운'으로 깎이지 않게 보정합니다. "
                "0·1·2명 결과를 미리 계산해 두므로 숫자만 바꾸면 즉시 반영됩니다."
            ),
        )
    _synergy_count = max(0, min(2, int(st.session_state.get("crit_synergy_count", 0) or 0)))
    with _syn_c2:
        if _synergy_count > 0:
            st.caption(f"✅ 실전 기대 치명률에 **+{_synergy_count * 10}%p** 반영 중 (허수/전분은 시너지 없이 유지)")
        else:
            st.caption("치적 시너지가 없으면 0으로 두세요. 실전에서 치적 시너지를 받는다면 인원 수를 입력하세요.")

    real_meta, dummy_meta, real_elapsed, dummy_elapsed = _result_meta_times()
    real_summary = summarize_battle(st.session_state.real_table, real_elapsed)
    dummy_summary = summarize_battle(st.session_state.dummy_table, dummy_elapsed)
    real_summary = apply_summary_overrides(real_summary, real_meta)
    dummy_summary = apply_summary_overrides(dummy_summary, dummy_meta)

    _syn_results = _result_tables_by_synergy(
        st.session_state.real_table,
        st.session_state.dummy_table,
        real_elapsed,
        dummy_elapsed,
        target_gain_rate,
    )
    result_skill_df, improvement_df, summary_df, analysis_meta = _syn_results[_synergy_count]

    api_crit_rate_percent = api_estimated_crit_rate_for_battle(real_summary.get("df"))
    api_crit_damage_percent = api_estimated_crit_damage_for_battle(real_summary.get("df"))
    manual_dummy_dps = st.session_state.manual_dummy_dps or dummy_meta.get("dps")
    eff = compute_efficiency(real_summary, dummy_summary, manual_dummy_dps=manual_dummy_dps)
    corr = corrected_dps(
        eff.get("real_dps"),
        real_summary.get("weighted_crit_rate"),
        real_summary.get("weighted_back_attack_rate"),
        target_crit_percent,
        target_back_percent,
        api_crit_damage_percent,
        back_bonus_percent,
    )

    if result_skill_df.empty:
        st.info("실전/허수 전투분석기 표를 먼저 입력하면 결과와 개선분석이 표시됩니다.")
        return

    # v136: 입력값(OCR) 이상 감지 — 값은 바꾸지 않고 경고만 표시합니다.
    _data_warnings = analysis_meta.get("data_warnings") or []
    if _data_warnings:
        _lines = []
        for _w in _data_warnings:
            _lines.append(f"- **{_w.get('스킬명')}**: " + " · ".join(_w.get("사유") or []))
        st.warning(
            "다음 스킬은 입력값이 꼬였을 가능성이 있습니다(OCR 오독 등). "
            "값은 자동으로 바꾸지 않았으니, 위쪽 **편집 표에서 직접 수정**하면 결과가 정확해집니다.\n\n"
            + "\n".join(_lines)
        )

    tab_result, tab_improve = st.tabs(["결과", "개선분석"])

    with tab_result:
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("실전 총 피해량", format_korean_number(real_summary.get("total_damage")))
        c2.metric("실전 DPS", format_korean_number(eff.get("real_dps")))
        dummy_like_damage = analysis_meta.get("total_dummy_pace")
        if dummy_like_damage is not None and real_elapsed:
            c3.metric("허수처럼 칠 경우 피해량 / DPS", f"{format_korean_number(dummy_like_damage)} / {format_korean_number(dummy_like_damage / real_elapsed)}")
        else:
            c3.metric("허수처럼 칠 경우 피해량 / DPS", format_korean_number(eff.get("base_dps")))
        efficiency = analysis_meta.get("efficiency") or eff.get("efficiency_percent")
        c4.metric(
            "허수 대비 실전 효율",
            f"{efficiency:.2f}%" if efficiency is not None else "-",
            help="확률(치명타·뭉툭한 가시·예리한 둔기 등) 운을 보정한 실전 피해를 허수 사용횟수 페이스 기대 피해와 비교한 대표 수행률입니다. 100%면 허수 페이스와 동일합니다.",
        )

        c5, c6, c7, c8 = st.columns(4)
        grade_text = grade_efficiency(efficiency)
        grade_display = f"{grade_text} / {efficiency:.2f}%" if efficiency is not None else grade_text
        _prob_total = analysis_meta.get("total_real_avg_all") or analysis_meta.get("total_real_avg") or 0
        # c5: 실전 총 피해량(c1) 바로 아래 → 확률 보정 총 피해량
        c5.metric(
            "확률 보정 총 피해량",
            format_korean_number(_prob_total),
            help="실전 총 피해를 치명타 적중률·뭉툭한 가시·예리한 둔기 같은 확률 효과의 기대값으로 보정한 총 피해입니다. 위의 실전 총 피해량과 비교하면 이번 판의 확률 운을 알 수 있습니다.",
        )
        # c6: 실전 DPS(c2) 바로 아래 → 확률 보정 DPS (위아래로 바로 비교)
        c6.metric(
            "확률 보정 DPS",
            format_korean_number(_prob_total / real_elapsed) if real_elapsed else "-",
            help="치명타 적중률, 뭉툭한 가시, 예리한 둔기 등 확률 효과를 기대값으로 보정한 DPS입니다. 바로 위 실전 DPS와 비교하세요. 보조/기타 피해는 관측값 그대로 포함합니다.",
        )
        c7.metric("치명타 비중", fmt_percent(real_summary.get("weighted_crit_share") or real_summary.get("weighted_crit_rate")), help="신규 전분의 치명타 비중이 있으면 그 값을 우선 사용합니다.")
        # c8: 기존 치명운 보정 DPS 위치 → 평가
        c8.metric(
            "평가",
            grade_display,
            help="평가 기준: 90% 이상 매우 높음, 80~90% 높음, 70~80% 보통 이상, 60~70% 보통, 60% 미만 낮음입니다.",
        )

        # v137: 방향 원인 분석(최종 점수와 분리) — 착용 각인 / 기본 보너스 활용률 / 각인 효율(딜증%) / 백·헤드 비중
        _dir_warn = analysis_meta.get("direction_conflict_warning")
        if _dir_warn:
            st.error(f"방향 각인 감지 오류: {_dir_warn} 방향 각인을 '없음'으로 두고 분석했습니다.")
        d1, d2, d3, d4 = st.columns(4)
        _basic_util = analysis_meta.get("direction_weighted_basic_util")
        _eng_eff = analysis_meta.get("direction_weighted_engraving_eff")
        _eng_state_val = analysis_meta.get("global_engraving_state")
        _eng_state_label = {"AMBUSH_MASTER": "기습의 대가", "MASTER_BRAWLER": "결투의 대가", "NONE": "없음"}.get(_eng_state_val, "없음")
        # 방향 각인 효율 옆에 '실질 딜증(%)' = 조건부 보너스(%) × 효율 을 함께 표시합니다.
        _dir_rows_meta = analysis_meta.get("directional_engraving_efficiency") or []
        _eng_bonus_pct = _finite_float(_dir_rows_meta[0].get("조건부 방향 피해(%)"), None) if _dir_rows_meta else None
        d1.metric("착용 방향 각인", _eng_state_label, help="API에서 감지한 방향 각인입니다. 기습의 대가=백어택, 결투의 대가=헤드어택 조건부 보너스.")
        d2.metric("기본 방향 보너스 활용률", f"{_basic_util*100:.1f}%" if _basic_util is not None else "-", help="기습/결투 각인을 제외한 백/헤드 자체 보너스(백 +5%·치적, 헤드 +20%)를 피해량 가중으로 얼마나 살렸는지입니다. 최종 점수와는 분리된 원인 분석 지표입니다.")
        if _eng_eff is not None:
            _deal_up = (_eng_bonus_pct or 0.0) * _eng_eff
            _eng_disp = f"{_eng_eff*100:.1f}% (딜증 +{_deal_up:.2f}%)"
        else:
            _eng_disp = "해당 없음"
        d3.metric("방향 각인 효율", _eng_disp, help="기습/결투 각인의 조건부 보너스를 실제 딜로 전환한 비율(피해량 가중)입니다. 괄호의 딜증은 조건부 보너스(%) × 효율로 계산한 실제 추가 딜증입니다. 예: 15% × 75% = 11.25%.")
        d4.metric("백/헤드 비중", fmt_percent(real_summary.get("weighted_back_attack_share") or real_summary.get("weighted_head_attack_share") or real_summary.get("weighted_back_attack_rate") or real_summary.get("weighted_head_attack_rate")), help="신규 전분의 백/헤드 비중이 있으면 그 값을 우선 사용합니다.")

        directional_rows = analysis_meta.get("directional_engraving_efficiency") or []
        if directional_rows:
            st.subheader("기습/결투 각인 실전 활용률")
            eff_df = pd.DataFrame(directional_rows)
            for _, erow in eff_df.iterrows():
                eng_name = erow.get("각인") or "방향 각인"
                # 기습의 대가=백어택, 결투의 대가=헤드어택. 라벨을 방향 기준으로 표시합니다.
                _dir_word = "백어택" if "기습" in str(eng_name) else "헤드어택" if "결투" in str(eng_name) else "백/헤드"
                bonus_percent = _finite_float(erow.get("조건부 방향 피해(%)"), 0.0) or 0.0
                eff_val = erow.get("대상 스킬 기준 각인 효율")
                real_gain = erow.get("대상 스킬 기준 실질 딜증")
                if real_gain is None or (isinstance(real_gain, float) and math.isnan(real_gain)):
                    real_gain = (bonus_percent / 100.0) * float(eff_val) if eff_val is not None and not (isinstance(eff_val, float) and math.isnan(eff_val)) else None
                mcols = st.columns(5)
                mcols[0].metric(
                    f"{eng_name} 등급",
                    str(erow.get("각인 등급") or "-"),
                    help="API 각인/피해군 출처에서 읽은 장착 각인 등급과 단계입니다.",
                )
                mcols[1].metric(
                    f"{_dir_word} 조건 성공률",
                    fmt_percent(erow.get("대상 스킬 기준 활용률")),
                    help="대상 백/헤드 스킬 피해 안에서 실제 조건을 만족한 피해 비중입니다. 전분의 백어택/헤드어택 비중을 피해량 가중 평균으로 계산합니다.",
                )
                mcols[2].metric(
                    f"{_dir_word} 각인 효율",
                    fmt_ratio_percent(eff_val),
                    help="조건부 최대 보너스 중 실제로 살린 비율입니다. 예: 조건부 보너스 15%, 각인 효율 78.67%면 실질 추가 딜증은 15% × 78.67% = 약 11.80%입니다.",
                )
                mcols[3].metric(
                    f"{_dir_word} 성공으로 인한 추가 딜",
                    fmt_ratio_percent(real_gain),
                    help="백/헤드 성공으로 실제로 받은 조건부 추가 딜증입니다. 계산식은 조건부 보너스 × 각인 효율입니다. 기습의 대가라면 백어택 성공분, 결투의 대가라면 헤드어택 성공분만 반영합니다.",
                )
                mcols[4].metric(
                    f"{_dir_word} 최대 추가 딜",
                    fmt_percent(bonus_percent),
                    help="API 각인 Tooltip에서 읽은 백어택/헤드어택 성공 시 추가 피해입니다. 상시 피해 증가와 분리해서 계산합니다.",
                )
            st.dataframe(format_analysis_dataframe(eff_df), use_container_width=True, hide_index=True, height=190, column_config=result_table_column_config(format_analysis_dataframe(eff_df)))

        # v143: 예리한 둔기 + 뭉툭한 가시를 '하나로 합쳐' 확률 보정을 요약합니다.
        # 아드레날린은 확률 보정이 아니므로 제외합니다. 채용한 효과가 없으면 섹션 자체를 숨깁니다.
        _pc = analysis_meta.get("probability_correction")
        if _pc:
            st.subheader("확률 효과 보정 (치명타 확률 변동)")
            _names = _pc.get('채용') or []
            _name_txt = (', '.join(_names)) if _names else "치명타 확률"
            st.caption(f"기준: {_name_txt} · 치명타가 실제로 얼마나 터지느냐에 따른 실전 DPS 변동 범위입니다. "
                       f"예리한 둔기는 허수·실전 양쪽에 동일 확률로 적용돼 효율 비율에 영향이 없으므로 제외했습니다.")
            p1, p2, p3 = st.columns(3)
            p1.metric("최소 DPS (확률상 낮은)", format_korean_number(_pc.get("최소 DPS")),
                      help="치명타가 불운하게 적게 터졌을 때의 DPS입니다. 스킬별 타수를 반영한 치명 분포에서 "
                           "누적확률 0.1% 지점(약 −3.09σ)까지만 집계했습니다. 그보다 더 불운한 경우(모든 타격 비치명 등)는 "
                           "현실적으로 거의 불가능하므로 버렸습니다.")
            p2.metric("평균 DPS (기대값)", format_korean_number(_pc.get("평균 DPS")),
                      help="치명타가 기대 확률대로 터졌을 때의 DPS입니다. 실전 관측 DPS의 기준선입니다.")
            p3.metric("최대 DPS (확률상 높은)", format_korean_number(_pc.get("최대 DPS")),
                      help="치명타가 최대로 터졌을 때의 이론상 상한 DPS입니다.")

            _cardc = _pc.get("카드 보정")
            if _cardc and _cardc.get("항목"):
                st.markdown(f"**🃏 아르카나 카드 확률 보정 · {_cardc.get('각인') or ''}** "
                            f"(카드 사용 횟수 {int(_cardc.get('카드 사용 횟수') or 0):,}회 기준)")
                st.caption("황제·황후의 기사(직접피해 카드)의 사용 횟수를 '카드 뜰 확률 × 카드 사용 횟수'로 정규화했습니다. "
                           "기대(=최대) 딜은 확률 기대 횟수, 최소 딜은 이항분포 0.1% 하한 횟수로 계산합니다. "
                           "이 카드들은 매 판 뽑기 운에 따라 사용 횟수가 크게 흔들려서 확률 보정 대상입니다.")
                _card_rows = []
                for _it in _cardc.get("항목", []):
                    _card_rows.append({
                        "카드": _it.get("카드"),
                        "카드 뜰 확률": f"{_finite_float(_it.get('확률(%)'), 0.0):.1f}%",
                        "이번 판 관측 횟수": f"{int(_it.get('관측 횟수') or 0):,}회",
                        "기대 횟수(N×p)": f"{_finite_float(_it.get('기대 횟수'), 0.0):.1f}회",
                        "최소 횟수(0.1%)": f"{int(_it.get('최소 횟수') or 0):,}회",
                        "1회 평균 딜": format_korean_number(_it.get("1회 평균 딜")),
                        "기대 딜": format_korean_number(_it.get("기대 딜")),
                        "최소 딜": format_korean_number(_it.get("최소 딜")),
                    })
                st.dataframe(pd.DataFrame(_card_rows), use_container_width=True, hide_index=True)

        st.divider()
        left, right = st.columns([1.1, 0.9])
        with left:
            st.subheader("스킬별 피해 지분")
            top = real_summary.get("top_skills")
            if top is not None and not top.empty:
                show_horizontal_damage_chart(top)
                st.dataframe(display_battle_table(top), use_container_width=True, height=320, hide_index=True)
            else:
                st.info("실전 전투분석기 표가 비어 있습니다.")
        with right:
            st.subheader("핵심 지표")
            show_summary = summary_df.copy()
            if not show_summary.empty:
                # 피해량/비율을 사람이 읽기 좋게 표시
                def fmt_summary_value(row: pd.Series) -> Any:
                    label = str(row.get("핵심 지표", ""))
                    v = row.get("값")
                    if isinstance(v, str):
                        return v
                    if v is None or (isinstance(v, float) and math.isnan(v)):
                        return "-"
                    money_like_labels = {"총 상승 여지", "목표 상승에 필요한 피해", "목표 필요 피해", "치명타 보정 값", "비교 포함 실전 보정 피해"}
                    if label in money_like_labels or ((("피해" in label) or ("보정" in label)) and "비율" not in label and "영향" not in label and "상승률" not in label and label != "현재 대비 상승 여지"):
                        return format_korean_number(v)
                    if any(x in label for x in ["/", "영향", "소화율", "사용횟수율", "상승률", "상승 여지", "활용률", "비중", "실질 딜증", "효율"]):
                        # summary 내부에서 0~1 비율로 저장된 값과 0~100 퍼센트 값이 섞일 수 있어 보정합니다.
                        fv = float(v)
                        return f"{(fv * 100 if abs(fv) <= 1.5 else fv):.2f}%"
                    return v
                show_summary["값"] = show_summary.apply(fmt_summary_value, axis=1)
                st.dataframe(show_summary, use_container_width=True, hide_index=True, height=420)

        st.subheader("스킬별 결과표")
        st.caption("확률 보정 값은 치명타 적중률·뭉툭한 가시·예리한 둔기 같은 확률 효과를 기대값으로 보정한 값입니다. 허수 페이스는 `실전 1회 보정 피해 × 허수 사용횟수 페이스`로 계산하므로 허수 피해량 단순 시간환산이 아닙니다.")
        # v137: 표시용 컬럼명 변경(치명→확률 개념). 내부 계산 컬럼명은 유지합니다.
        _RESULT_DISPLAY_RENAME = {
            "실전 관측": "실전 피해량",
            "치명타 보정 값": "확률 보정 값",
            "데미지 보정값 / 허수페이스": "확률 보정 값 / 허수페이스",
            "기습/결투 실질 딜증": "기습/결대 실질 딜증",
            "치명운": "확률운",
        }
        formatted_result_df = format_analysis_dataframe(result_skill_df).rename(columns=_RESULT_DISPLAY_RENAME)
        st.dataframe(formatted_result_df, use_container_width=True, hide_index=True, height=460, column_config=result_table_column_config(formatted_result_df))
        st.download_button(
            "결과표 CSV 다운로드",
            data=result_skill_df.to_csv(index=False).encode("utf-8-sig"),
            file_name="loa_result_table.csv",
            mime="text/csv",
        )

    with tab_improve:
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("확률 보정 피해", format_korean_number(analysis_meta.get("total_real_avg")))
        c2.metric("허수페이스 기대 피해", format_korean_number(analysis_meta.get("total_dummy_pace")))
        c3.metric("총 상승 여지", format_korean_number(analysis_meta.get("total_loss")))
        c4.metric("현재 대비 상승 여지", f"{analysis_meta.get('current_gap_rate', 0) * 100:.2f}%")

        c5, c6, c7, c8 = st.columns(4)
        c5.metric("목표 상승률", f"{target_gain_rate * 100:.2f}%")
        c6.metric("목표 필요 피해", format_korean_number(analysis_meta.get("target_needed")))
        c7.metric("목표 가능 여부", analysis_meta.get("possible") or "-")
        c8.metric(
            "핵심 수행률",
            f"{analysis_meta.get('efficiency'):.2f}%" if analysis_meta.get("efficiency") is not None else "-",
            help="확률 보정 실전 피해 ÷ 허수페이스 기대 피해입니다. 개선분석에서 보는 가장 중요한 수행률입니다.",
        )

        st.subheader("상위 개선 후보")
        # v137: 치명운 → 확률 운 표시, '목표 달성 필요 개선률' 컬럼 제거.
        show_improve = format_analysis_dataframe(improvement_df)
        show_improve = show_improve.drop(columns=[c for c in ["목표 달성 필요 개선률"] if c in show_improve.columns])
        show_improve = show_improve.rename(columns={"치명운": "확률 운"})
        st.caption("표 제목의 ? 도움말에 각 항목 계산 의미를 넣어두었습니다. 손실피해가 큰 순서대로 우선순위를 봐주세요.")
        st.dataframe(show_improve, use_container_width=True, hide_index=True, height=560, column_config=result_table_column_config(show_improve))
        st.download_button(
            "개선분석 CSV 다운로드",
            data=improvement_df.to_csv(index=False).encode("utf-8-sig"),
            file_name="loa_improvement_analysis.csv",
            mime="text/csv",
        )

    # API 보정 기준 탭은 사용자 화면에서 제거했습니다. 내부 계산값은 JSON expander에서 확인할 수 있습니다.

    with st.expander("결과 JSON 저장/확인"):
        result_json = {
            "result_summary": analysis_meta,
            "efficiency": eff,
            "correction": corr,
            "api_estimated_crit_rate_percent": api_crit_rate_percent,
            "api_estimated_crit_damage_percent": api_crit_damage_percent,
            "api_crit_basis": st.session_state.get("api_crit_basis"),
        }
        st.json(result_json)
        st.download_button(
            "계산 결과 JSON 다운로드",
            data=json.dumps(result_json, ensure_ascii=False, indent=2),
            file_name="loa_analysis_result.json",
            mime="application/json",
        )



# ==============================
# v40: OCR 후보 생성 정리
# ==============================
def get_ocr_skill_candidates_full() -> list[dict[str, str]]:  # type: ignore[override]
    """전투분석기 이름 보정 후보 v40.

    핵심 변경:
    - 스킬 후보는 채용/계산에 실제 들어간 스킬만 사용합니다.
    - API 원문 전체 Name+Icon 스캔은 오보정 원인이 커서 기본 후보에서 제외합니다.
    - 룬은 combat_skills의 Rune에서만 `스킬룬: 이름` 형태로 추가합니다.
    - 보주는 Type=보주 Tooltip 안의 실제 전투분석기 스킬명을 추출합니다.
    """
    summary = st.session_state.get("api_summary") or {}
    bundle = st.session_state.get("api_bundle") or {}
    candidates: list[dict[str, str]] = []
    seen: set[tuple[str, str, str]] = set()

    def add_candidate(name: Any, icon: Any = "", source: str = "") -> None:
        name_s = str(name or "").strip()
        icon_s = str(icon or "").strip()
        src_s = str(source or "").strip()
        if not name_s or len(name_s) > 35:
            return
        key = (name_s, icon_s, src_s)
        if key in seen:
            return
        seen.add(key)
        candidates.append({"name": name_s, "icon": icon_s, "source": src_s})

    def col_of(df: pd.DataFrame, names: list[str]) -> str | None:
        for n in names:
            if n in df.columns:
                return n
        return None

    # 전체 스킬명/아이콘, 그리고 채용 스킬명 집합을 분리합니다.
    skill_df = summary.get("skills")
    icon_by_name: dict[str, str] = {}
    all_skill_names: set[str] = set()
    adopted_skill_names: set[str] = set()
    if isinstance(skill_df, pd.DataFrame) and not skill_df.empty:
        name_col = col_of(skill_df, ["스킬명", "name"])
        icon_col = col_of(skill_df, ["아이콘", "icon"])
        if name_col:
            for _, row in skill_df.iterrows():
                n = str(row.get(name_col) or "").strip()
                if not n:
                    continue
                all_skill_names.add(n)
                if icon_col:
                    icon_by_name[n] = str(row.get(icon_col) or "").strip()
            adopted = filter_adopted_skills_df(skill_df)
            if isinstance(adopted, pd.DataFrame) and not adopted.empty:
                adopted_col = col_of(adopted, ["스킬명", "name"])
                if adopted_col:
                    adopted_skill_names.update(str(x).strip() for x in adopted[adopted_col].dropna().tolist() if str(x).strip())

    # 계산표에 들어간 스킬은 레벨 1 초각성 스킬이어도 실제 사용 후보로 봅니다.
    for key in ["skill_crit_estimates", "lostbuilds_base_skill_estimates"]:
        df = summary.get(key)
        if isinstance(df, pd.DataFrame) and not df.empty:
            ncol = col_of(df, ["스킬명", "name"])
            if ncol:
                adopted_skill_names.update(str(x).strip() for x in df[ncol].dropna().tolist() if str(x).strip())

    # 스킬 후보는 실제 채용/계산 스킬만 추가합니다. 여기서 퓨리 블레이드 같은 미채용 스킬 오보정을 막습니다.
    for name in sorted(adopted_skill_names):
        add_candidate(name, icon_by_name.get(name, ""), "skill")

    def icon_from_obj(obj: Any) -> str:
        if not isinstance(obj, dict):
            return ""
        for k in ["Icon", "icon", "IconPath", "iconPath", "imagePath", "ImagePath"]:
            v = obj.get(k)
            if isinstance(v, str) and v.strip():
                return v.strip()
        slot = obj.get("slotData") or obj.get("SlotData")
        if isinstance(slot, dict):
            return icon_from_obj(slot)
        return ""

    def strip_api_text(value: Any) -> str:
        text = str(value or "")
        text = re.sub(r"<BR\s*/?>", "\n", text, flags=re.I)
        text = re.sub(r"<[^>]+>", "", text)
        text = text.replace("&nbsp;", " ").replace("&lt;", "<").replace("&gt;", ">")
        text = re.sub(r"\s+", " ", text)
        return text.strip()

    def flatten_tooltip_strings(value: Any, out: list[str] | None = None) -> list[str]:
        if out is None:
            out = []
        if value is None:
            return out
        if isinstance(value, str):
            raw = value
            try:
                if raw.strip().startswith("{") or raw.strip().startswith("["):
                    return flatten_tooltip_strings(json.loads(raw), out)
            except Exception:
                pass
            cleaned = strip_api_text(raw)
            if cleaned:
                out.append(cleaned)
            return out
        if isinstance(value, dict):
            for v in value.values():
                flatten_tooltip_strings(v, out)
        elif isinstance(value, list):
            for v in value:
                flatten_tooltip_strings(v, out)
        return out

    def extract_orb_skill_names_from_tooltip(item: dict[str, Any]) -> list[str]:
        """보주 Tooltip에서 전투분석기에 표시되는 실제 보주 스킬명을 최대한 범용 추출합니다.

        예: `눈부신 비전의 보주`의 Tooltip 안에 있는
        `<FONT COLOR='#FE43FC'>맥스웰 맥시마</FONT><BR>...` 같은 짧은 제목 라인을 후보로 잡습니다.
        특정 보주 이름을 하드코딩하지 않고, Tooltip 구조에서 짧은 한글 스킬명 라인만 추출합니다.
        """
        tooltip_raw = item.get("Tooltip") or item.get("tooltip") or ""
        raw_text = str(tooltip_raw or "")
        found: list[str] = []

        def add_if_skill_like(value: Any) -> None:
            cand = strip_api_text(value)
            cand = re.sub(r"^[|:·\-\s]+|[|:·\-\s]+$", "", cand)
            cand = re.sub(r"\s+", " ", cand)
            if not cand:
                return
            # 보주 스킬명은 보통 2~20자 내외의 짧은 제목입니다.
            if not (2 <= len(cand) <= 20 and re.search(r"[가-힣]", cand)):
                return
            bad_words = [
                "보주", "유물", "전설", "고대", "효과", "피해", "재사용", "장착", "특수",
                "등급", "몬스터", "초", "m", "%", "스킬 룬", "귀속", "거래"
            ]
            if any(b in cand for b in bad_words):
                return
            if cand not in found:
                found.append(cand)

        # 1) FONT 태그 바로 뒤 BR이 오는 제목 라인. 보주 스킬명은 이 형태가 가장 흔합니다.
        for m in re.finditer(r"<FONT[^>]*>([^<>]{2,30})</FONT>\s*(?:<BR\s*/?>|\\n|\n)", raw_text, flags=re.I):
            add_if_skill_like(m.group(1))

        # 2) 강조 색상 안의 짧은 한글 문자열. 색상은 보주마다 다를 수 있어 특정 색에 묶지 않습니다.
        for m in re.finditer(r"<FONT[^>]*COLOR=['\"]?#[0-9A-Fa-f]{6}['\"]?[^>]*>([^<>]{2,30})</FONT>", raw_text, flags=re.I):
            add_if_skill_like(m.group(1))

        joined = re.sub(r"\s+", " ", " ".join(flatten_tooltip_strings(tooltip_raw)))

        # 3) `<스킬명> 보스 등급 이상의...` 형태
        for m in re.finditer(r"([가-힣A-Za-z0-9][가-힣A-Za-z0-9 '\-]{1,28}?)\s*보스\s*등급", joined):
            cand = re.sub(r"^(보주|효과|스킬|특수 효과)\s*", "", strip_api_text(m.group(1))).strip()
            add_if_skill_like(cand)

        # 4) `<스킬명>에게 ... 피해` 형태
        for m in re.finditer(r"([가-힣A-Za-z0-9][가-힣A-Za-z0-9 '\-]{1,28}?)\s*에게\s+[^.]{0,50}피해", joined):
            cand = re.sub(r".*?(?:효과|스킬|특수 효과)\s*", "", strip_api_text(m.group(1))).strip()
            add_if_skill_like(cand)

        return found

    # API 상 채용 스킬 추정 보강: 레벨/룬/트포가 있는 스킬은 우선 후보에 추가합니다.
    # 단순 Level 1 무채용 스킬은 제외하고, 도약/초각성처럼 레벨 1이지만 다른 툴팁에 언급된 스킬은 아래에서 추가합니다.
    combat_data = (bundle.get("combat_skills") or {}).get("data") if isinstance(bundle, dict) else None
    if isinstance(combat_data, list):
        for s in combat_data:
            if not isinstance(s, dict):
                continue
            sname = str(s.get("Name") or s.get("name") or "").strip()
            try:
                slevel = int(float(s.get("Level") or s.get("level") or 0))
            except Exception:
                slevel = 0
            tripod_text = " ".join(flatten_tooltip_strings(s.get("Tripods") or s.get("tripods") or s.get("Tooltip") or ""))
            rune = s.get("Rune") or {}
            has_rune = isinstance(rune, dict) and bool(str(rune.get("Name") or rune.get("name") or "").strip())
            has_tripod = bool(tripod_text and "None" not in tripod_text)
            if sname and (slevel >= 7 or has_rune or has_tripod):
                add_candidate(sname, icon_from_obj(s) or icon_by_name.get(sname, ""), "skill_combat_used")
            if isinstance(rune, dict):
                rname = str(rune.get("Name") or rune.get("name") or "").strip()
                ricon = icon_from_obj(rune)
                if rname:
                    add_candidate(f"스킬룬: {rname}", ricon, "rune")

    # 도약/아크패시브/아크그리드/장비 툴팁에 직접 언급된 스킬명은 초각성 스킬 후보로 추가합니다.
    # combat_skills 원문 자체를 검색하면 모든 스킬명이 다 들어가므로 제외합니다.
    def collect_non_combat_text() -> str:
        parts: list[str] = []
        if isinstance(bundle, dict):
            for k, payload in bundle.items():
                if str(k) == "combat_skills":
                    continue
                if isinstance(payload, dict):
                    parts.extend(flatten_tooltip_strings(payload.get("data")))
        return " ".join(parts)

    non_combat_text = collect_non_combat_text()
    for sname in sorted(all_skill_names):
        if sname and sname in non_combat_text:
            add_candidate(sname, icon_by_name.get(sname, ""), "skill_tooltip_mention")

    def walk_orbs(obj: Any, depth: int = 0) -> None:
        if depth > 10:
            return
        if isinstance(obj, dict):
            typ = str(obj.get("Type") or obj.get("type") or "").strip()
            name_text = str(obj.get("Name") or obj.get("name") or "")
            if typ == "보주" or "보주" in name_text:
                icon = icon_from_obj(obj)
                for skill_name in extract_orb_skill_names_from_tooltip(obj):
                    add_candidate(skill_name, icon, "orb_tooltip_skill")
            for v in obj.values():
                walk_orbs(v, depth + 1)
        elif isinstance(obj, list):
            for v in obj:
                walk_orbs(v, depth + 1)

    if isinstance(bundle, dict):
        for payload in bundle.values():
            if isinstance(payload, dict):
                walk_orbs(payload.get("data"), 0)

    # v80: 로컬 lostark_skill_db_output 스킬 아이콘 추가 (API CDN 대신 로컬 파일 사용).
    # API에서 읽은 클래스명으로 해당 클래스 폴더의 PNG를 찾아 후보에 추가합니다.
    _prof = summary.get("profile_summary", {}) if isinstance(summary, dict) else {}
    _char_class = str(_prof.get("클래스") or "").strip()
    if _char_class:
        try:
            for _item in _load_local_skill_db_candidates_v80(_char_class):
                add_candidate(_item.get("name", ""), _item.get("icon", ""), _item.get("source", "skill_local_db"))
        except Exception as _e:
            print(f"[skill_db] 로컬 스킬 DB 후보 추가 실패: {_e}")

    return candidates

def settings_tab() -> None:
    st.header("4) 수정 포인트")
    st.markdown(
        """
이 프로젝트는 엑셀을 쓰지 않고 Streamlit 화면에서 결과를 바로 보는 구조입니다.

- 공격정보는 스킬 아이콘 위치를 먼저 감지해 행을 나눕니다. 그래도 밀리면 OCR 디버그 ZIP을 보내 좌표/전처리를 확인하세요.
- API endpoint를 끄고 싶으면 `configs/api_endpoints.yaml`에서 enabled 수정
- 백어택/개선분석 목표값은 왼쪽 사이드바에서 수정
- 계산식을 바꾸고 싶으면 `modules/calculators.py` 수정
- API 파싱 결과를 바꾸고 싶으면 `modules/api_parser.py` 수정
- 스킬별 치명/치피/진피/피해군 추정식을 바꾸고 싶으면 `modules/api_skill_estimator.py` 수정
- 화면 구성을 바꾸고 싶으면 `app.py` 수정
        """
    )
    with st.expander("현재 계산 설정 YAML"):
        st.code((CONFIG_DIR / "calculation_presets.yaml").read_text(encoding="utf-8"), language="yaml")
    with st.expander("현재 OCR 헤더 YAML"):
        st.code((CONFIG_DIR / "ocr_columns.yaml").read_text(encoding="utf-8"), language="yaml")


# v70 -------------------------------------------------------------------------
# 검수 단계용 API 탭 단순화.
# - 일반 화면에서는 스탯과 피해군 출처만 노출합니다.
# - 디스트로이어 검증용으로 진화/깨달음/도약 수동 시뮬레이터를 임시 제공합니다.

def _v70_pct(value: Any, default: float = 0.0) -> float:
    try:
        if value is None:
            return default
        if isinstance(value, str):
            value = value.replace("%", "").replace(",", "").strip()
        return float(value)
    except Exception:
        return default


def _v70_damage_source_sum(df: Any, col: str, names: list[str] | None = None, scopes: list[str] | None = None) -> float:
    if not isinstance(df, pd.DataFrame) or df.empty or col not in df.columns:
        return 0.0
    t = df.copy()
    if names:
        nset = {_norm_name(x) for x in names if x}
        t = t[t.get("이름", pd.Series([""] * len(t))).map(lambda x: _norm_name(x) in nset)]
    if scopes:
        sset = set(scopes)
        t = t[t.get("적용범위", pd.Series([""] * len(t))).astype(str).isin(sset)]
    vals = pd.to_numeric(t[col], errors="coerce").fillna(0.0)
    return float(vals.sum())


def _v70_render_stat_cards(summary: dict[str, Any]) -> None:
    damage_sources = summary.get("damage_sources")
    global_enemy = summary.get("global_enemy_damage_percent")
    if global_enemy is None:
        global_enemy = _summary_breakdown_value(summary, "적에게 주는 피해")
    avg_enemy = summary.get("avg_enemy_damage_unweighted_percent", summary.get("avg_enemy_damage_percent", 0.0))
    avg_gem_all = summary.get("avg_gem_damage_unweighted_percent", summary.get("avg_gem_damage_percent", 0.0))
    avg_gem_active = summary.get("avg_gem_damage_on_gem_skills_percent")
    if avg_gem_active is None:
        avg_gem_active = _avg_positive_metric(summary, "보석 피해(조건부)(%)")

    cols = st.columns(4)
    cols[0].metric("치명 스탯", f"{summary.get('base_crit_stat', 0):,.0f}")
    cols[1].metric("치명 스탯 치적", f"{summary.get('base_crit_percent', 0):.2f}%")
    cols[2].metric(
        "기본 기준 치명",
        f"{summary.get('global_basis_crit_rate_percent', summary.get('avg_back_basis_crit_rate_percent', 0)):.2f}%",
        help="스킬 자체 트라이포드 치적과 방향성 치적(기습의 대가, 결투의 대가, 백어택 기본 보너스)을 제외한 기준 치명입니다. 스탯/전체범위 각인/장비/아크패시브만 합산합니다.",
    )
    cols[3].metric(
        "최종 전역 치피",
        f"{summary.get('global_crit_damage_no_skill_percent', summary.get('avg_conditional_crit_damage_percent', 200)):.2f}%",
        help="기본 200% + 전역 치피 + 방향성 치피입니다. 특정 스킬 전용 치피는 제외합니다.",
    )

    cols = st.columns(4)
    cols[0].metric("평균 진화형 피해", f"{summary.get('avg_evolution_damage_percent', 0):.2f}%")
    cols[1].metric("전역 적피 합계", f"{(global_enemy if global_enemy is not None else 0):.2f}%")
    cols[2].metric("스킬 포함 적피 평균", f"{avg_enemy:.2f}%")
    cols[3].metric("평균 보석 피해", f"{avg_gem_all:.2f}%")

    cols = st.columns(2)
    cols[0].metric("보석 장착 스킬 평균", f"{(avg_gem_active if avg_gem_active is not None else 0):.2f}%")
    cols[1].metric("평균 최종 배율", f"{summary.get('avg_final_multiplier', 1):.4f}x")


def _v70_destroyer_review_simulator(summary: dict[str, Any]) -> None:
    st.subheader("검수용 노드 시뮬레이터")
    st.caption("임시 검수용입니다. 진화/깨달음/도약 값을 바꿔 보면서 치명/피해군 합산이 기대대로 움직이는지 확인합니다.")

    src = summary.get("damage_sources")
    if not isinstance(src, pd.DataFrame):
        src = pd.DataFrame()

    base_stat = _v70_pct(summary.get("base_crit_percent", 0))
    current_basis = _v70_pct(summary.get("global_basis_crit_rate_percent", 0))
    current_evo = _v70_pct(summary.get("avg_evolution_damage_percent", 0))
    current_enemy = _v70_pct(summary.get("global_enemy_damage_percent", _summary_breakdown_value(summary, "적에게 주는 피해") or 0))
    current_skill = _v70_pct(summary.get("avg_skill_damage_percent", 0))
    current_gem = _v70_pct(summary.get("avg_gem_damage_unweighted_percent", summary.get("avg_gem_damage_percent", 0)))

    with st.expander("디스트로이어 / 공용 노드 수동 조정", expanded=False):
        c1, c2, c3, c4 = st.columns(4)
        ilgyeok_lv = c1.number_input("일격 Lv", 0, 5, 2, 1, help="Lv당 치명 +10%, 백/헤드 치피 +16%")
        sharp_hammer = c2.number_input("날카로운 해머 치적", 0.0, 30.0, 6.0, 0.5, help="디스트로이어 해방 스킬 3코어 기준 치적")
        area_lv = c3.number_input("영역 강화 Lv", 0, 5, 1, 1, help="Lv 1~5 = 치적 2/4/6/8/10")
        back_head_bonus = c4.number_input("백/헤드 기본 치적", 0.0, 10.0, 0.0, 1.0, help="백어택 기준 캐릭터 검수 때만 10 입력")

        c5, c6, c7, c8 = st.columns(4)
        extra_crit = c5.number_input("추가 치적", -100.0, 200.0, 0.0, 0.5)
        extra_evo = c6.number_input("추가 진화형 피해", -100.0, 300.0, 0.0, 0.5)
        extra_enemy = c7.number_input("추가 적피", -100.0, 300.0, 0.0, 0.5)
        extra_skill = c8.number_input("추가 스킬 피해", -100.0, 300.0, 0.0, 0.5)

        area_table = {0: 0.0, 1: 2.0, 2: 4.0, 3: 6.0, 4: 8.0, 5: 10.0}
        sim_crit = (
            base_stat
            + _v70_damage_source_sum(src, "치명타 적중률 증가(%)", scopes=["전체", "백/헤드 스킬", "백어택 스킬", "헤드어택 스킬"])
            + float(ilgyeok_lv) * 10.0
            + float(sharp_hammer)
            + area_table.get(int(area_lv), 0.0)
            + float(back_head_bonus)
            + float(extra_crit)
        )
        sim_evo = current_evo + float(extra_evo)
        sim_enemy = current_enemy + float(extra_enemy)
        sim_skill = current_skill + float(extra_skill)

        p1, p2, p3, p4 = st.columns(4)
        p1.metric("시뮬 치명", f"{sim_crit:.2f}%", delta=f"{sim_crit - current_basis:+.2f}%p")
        p2.metric("시뮬 진화형 피해", f"{sim_evo:.2f}%", delta=f"{sim_evo - current_evo:+.2f}%p")
        p3.metric("시뮬 전역 적피", f"{sim_enemy:.2f}%", delta=f"{sim_enemy - current_enemy:+.2f}%p")
        p4.metric("시뮬 스킬 피해", f"{sim_skill:.2f}%", delta=f"{sim_skill - current_skill:+.2f}%p")

        st.caption("주의: 이 시뮬레이터는 검수용 표시값입니다. 실제 계산표에 반영하려면 class_rules.yaml/API 파서 룰로 확정해야 합니다.")






def _inject_ux_polish_css_v48() -> None:
    """v48: 기존 테마 위에 얹는 추가 폴리시(버튼/익스팬더/경고/가이드/표 헤더)."""
    st.markdown(
        """
        <style>
        div.stButton > button {
            border-radius: 12px; font-weight: 700;
            border: 1px solid rgba(255,255,255,0.12); transition: all .12s ease;
        }
        div.stButton > button:hover { border-color: rgba(255,120,90,0.65); transform: translateY(-1px); }
        div[data-testid="stExpander"] {
            border: 1px solid rgba(255,255,255,0.08); border-radius: 12px;
            margin-bottom: 10px; background: rgba(255,255,255,0.02);
        }
        div[data-testid="stAlertContainer"] { border-radius: 12px; }
        div[data-testid="stMetric"]:hover { border-color: rgba(255,120,90,0.35); }
        div[data-testid="stDataFrame"] div[role="columnheader"] { font-weight: 700; }
        .stTabs [data-baseweb="tab"][aria-selected="true"] {
            background: rgba(255,120,90,0.12);
            border-bottom: 2px solid rgba(255,120,90,0.9);
        }
        .loa-guide {
            display:flex; gap:14px; align-items:flex-start;
            border:1px solid rgba(120,170,255,0.25);
            background:linear-gradient(135deg, rgba(67,139,255,0.12), rgba(67,139,255,0.03));
            border-radius:14px; padding:14px 18px; margin:4px 0 18px 0;
        }
        .loa-guide .loa-guide-badge {
            background:rgba(67,139,255,0.9); color:#fff; font-weight:800;
            border-radius:8px; padding:3px 12px; white-space:nowrap; font-size:0.92rem; margin-top:2px;
        }
        .loa-guide ol { margin:0; padding-left:1.15rem; line-height:1.75; color:rgba(255,255,255,0.88); }
        .loa-guide b { color:#ffd479; }
        </style>
        """,
        unsafe_allow_html=True,
    )


def _render_step_guide(step: int, title: str, items: list[str]) -> None:
    """탭 상단 '이 단계에서 할 일' 안내 배너."""
    lis = "".join(f"<li>{x}</li>" for x in items)
    st.markdown(
        f'<div class="loa-guide"><span class="loa-guide-badge">STEP {step}</span>'
        f'<div><div style="font-weight:800;font-size:1.02rem;margin-bottom:4px;">{title}</div>'
        f'<ol>{lis}</ol></div></div>',
        unsafe_allow_html=True,
    )




# ==============================================================================
# v81: 공용 로컬 아이콘 후보 추가
# ==============================================================================
# add_icon/class 아이덴티티 아이콘은 lostark_skill_db_output/<class> 로 정리되어
# 기존 _load_local_skill_db_candidates_v80(class) 경로로 클래스별 후보에 들어갑니다.
# 여기서는 기본공격/기타딜/룬 같은 공용 아이콘만 추가합니다.
_get_ocr_skill_candidates_full_prev_v81 = get_ocr_skill_candidates_full


def get_ocr_skill_candidates_full() -> list[dict[str, str]]:  # type: ignore[override]
    candidates = list(_get_ocr_skill_candidates_full_prev_v81() or [])
    seen: set[tuple[str, str, str]] = set()
    deduped: list[dict[str, str]] = []
    for c in candidates:
        key = (str(c.get("name") or ""), str(c.get("icon") or ""), str(c.get("source") or ""))
        if key in seen:
            continue
        seen.add(key)
        deduped.append(c)
    try:
        from modules.fixed_grid_ocr import _local_icon_candidates_v32
        for item in _local_icon_candidates_v32():
            name = str(item.get("name") or "").strip()
            icon = str(item.get("icon") or "").strip()
            source = str(item.get("source") or "local_common").strip()
            if not name:
                continue
            key = (name, icon, source)
            if key in seen:
                continue
            seen.add(key)
            deduped.append({"name": name, "icon": icon, "source": source})
    except Exception as e:  # noqa: BLE001
        print(f"[add_icon:v81] 공용 로컬 아이콘 후보 추가 실패: {e}")
    return deduped




# ==============================================================================
# v90: 공용 룬 행 보존 + 성능 병목 기준 후보 확장
# ==============================================================================
APP_CALC_VERSION = "v97_square_icon_fallback"

# v91: v90 override에서 EDIT_COLUMNS를 참조했지만 상수가 정의되지 않아 NameError가 발생했습니다.
# 검수표 표준 편집 컬럼은 기존 STANDARD_BATTLE_COLUMN_ORDER를 그대로 사용합니다.
EDIT_COLUMNS = STANDARD_BATTLE_COLUMN_ORDER


def _skill_only_ocr_candidates(candidates: list[dict[str, str]] | None) -> list[dict[str, str]]:  # type: ignore[override]
    """v90: 이름 확정 후보는 스킬 + 공용 룬/기본공격/기타 아이콘까지 포함합니다.

    이전 v89는 source=skill만 남겨서 Bleed/Poison rune 행이 API 스킬로 확정되지 못하고
    `인식 안된 기타 추가딜`로 합산되었습니다. 전투분석기 검수표에는 룬 행도 보여야 하므로
    local_common/local_rune/rune 후보를 함께 둡니다.
    """
    out: list[dict[str, str]] = []
    seen: set[tuple[str, str, str]] = set()
    for c in candidates or []:
        if not isinstance(c, dict):
            continue
        source = str(c.get("source") or "").strip()
        source_l = source.lower()
        name = _normalize_special_battle_name_v86(str(c.get("name") or "").strip())
        icon = str(c.get("icon") or "").strip()
        if not name or not icon:
            continue
        keep = False
        if source_l.startswith("skill"):
            keep = True
        if "rune" in source_l or "local_common" in source_l or "local_rune" in source_l:
            keep = True
        if name.startswith("스킬룬") or name in {"기본 공격", "기타"}:
            keep = True
        if not keep:
            continue
        key = (name, icon, source)
        if key in seen:
            continue
        seen.add(key)
        out.append({"name": name, "icon": icon, "source": source})
    return out


def _is_api_skill_matched_row(row: pd.Series | dict[str, Any]) -> bool:  # type: ignore[override]
    """v90: API 스킬뿐 아니라 공용 룬/기본공격/기타 아이콘 확정 행도 보존합니다."""
    name = _normalize_special_battle_name_v86(str(row.get("이름") or "").strip())
    reason = str(row.get("_name_match_reason", "") or "")
    source = str(row.get("_icon_match_source", "") or "").lower()
    icon_score = row.get("_name_match_icon_score", row.get("_icon_match_score", 0))
    try:
        icon_score_f = float(icon_score or 0)
    except Exception:
        icon_score_f = 0.0

    if name.startswith("스킬룬") or name in {"기본 공격", "기타"}:
        return True
    if reason.startswith("icon_common") and icon_score_f > 0:
        return True
    if "rune" in source or "local_common" in source or "local_rune" in source:
        return icon_score_f > 0 or reason != "unmatched"
    if reason.startswith("icon_first_skill") or reason.startswith("text_fallback:skill") or reason.startswith("icon_first:skill"):
        return True
    if source.startswith("skill") and icon_score_f > 0:
        return reason != "unmatched"
    return False


def prepare_battle_editor_df(df: pd.DataFrame) -> pd.DataFrame:  # type: ignore[override]
    """v91: 검수표 진입 직전에 룬 이름을 한글명으로 정규화하고 표준 편집 컬럼만 유지합니다."""
    if df is None or df.empty:
        return pd.DataFrame(columns=EDIT_COLUMNS)
    out = strip_battle_helper_columns(df).copy()
    if "이름" in out.columns:
        out["이름"] = out["이름"].map(_normalize_special_battle_name_v86)
    cols = [c for c in EDIT_COLUMNS if c in out.columns]
    out = out[cols].copy() if cols else pd.DataFrame(columns=EDIT_COLUMNS)
    for c in EDIT_COLUMNS:
        if c not in out.columns:
            out[c] = ""
    return out[EDIT_COLUMNS].fillna("")



# ==============================================================================
# v96: 정확도 복구용 후처리
# ==============================================================================
# v95에서 종합정보 총피해가 OCR 실패로 실제보다 훨씬 작게 들어오면,
# 피해량 지분이 전부 100%가 되는 문제가 있었습니다.
# 표의 피해량 합계가 종합정보 총피해보다 말이 안 되게 크면 표 합계를 기준으로 복구합니다.
_repair_battle_values_from_summary_prev_v96 = _repair_battle_values_from_summary


def _repair_battle_values_from_summary(df: pd.DataFrame, meta: dict[str, Any] | None, *, replace_damage_from_share: bool = False, elapsed_fallback: float | None = None) -> pd.DataFrame:  # type: ignore[override]
    meta2 = dict(meta or {})
    try:
        if df is not None and not df.empty and "피해량" in df.columns:
            from modules.calculators import parse_korean_number
            visible_sum = 0.0
            max_damage = 0.0
            for _, row in df.iterrows():
                dmg = parse_korean_number(row.get("피해량"))
                if dmg and dmg > 0:
                    visible_sum += float(dmg)
                    max_damage = max(max_damage, float(dmg))
            total = meta2.get("total_damage")
            try:
                total_f = float(total) if total is not None else None
            except Exception:
                total_f = None
            # 종합정보 OCR이 `4,641.26억`/원시숫자를 잘못 읽어 4천만 같은 작은 값으로 들어온 경우.
            if visible_sum > 0 and (total_f is None or total_f <= 0 or total_f < max(max_damage * 1.05, visible_sum * 0.25)):
                meta2["total_damage"] = visible_sum
                try:
                    meta2["total_damage_text"] = format_korean_number(visible_sum)
                except Exception:
                    pass
                meta2["total_damage_source"] = "attack_visible_sum_fallback_v96"
    except Exception:
        pass
    return _repair_battle_values_from_summary_prev_v96(df, meta2, replace_damage_from_share=replace_damage_from_share, elapsed_fallback=elapsed_fallback)


# ==============================================================================
# v98: 클래스 전용 아이콘 후보 강제 + 후보 구조 점검
# ==============================================================================
APP_CALC_VERSION = "v98_class_icon_strict"


def _current_api_class_name_v98() -> str:
    """현재 API 프로필에서 클래스명을 가져옵니다."""
    try:
        summary = st.session_state.get("api_summary") or {}
        prof = summary.get("profile_summary") if isinstance(summary, dict) else {}
        cls = str((prof or {}).get("클래스") or "").strip()
        if cls:
            return cls
    except Exception:
        pass
    try:
        bundle = st.session_state.get("api_bundle") or {}
        prof = ((bundle.get("profile") or {}).get("data") or {}) if isinstance(bundle, dict) else {}
        cls = str(prof.get("CharacterClassName") or prof.get("ClassName") or "").strip()
        if cls:
            return cls
    except Exception:
        pass
    return ""


def _dedupe_icon_candidates_v98(items: list[dict[str, str]]) -> list[dict[str, str]]:
    out: list[dict[str, str]] = []
    seen: set[tuple[str, str]] = set()
    for c in items or []:
        if not isinstance(c, dict):
            continue
        name = _normalize_special_battle_name_v86(str(c.get("name") or "").strip())
        icon = str(c.get("icon") or "").strip()
        source = str(c.get("source") or "").strip()
        if not name or not icon:
            continue
        key = (name, icon)
        if key in seen:
            continue
        seen.add(key)
        out.append({"name": name, "icon": icon, "source": source})
    return out


def _build_strict_class_icon_candidates_v98(fallback_candidates: list[dict[str, str]] | None = None) -> list[dict[str, str]]:
    """API 클래스 폴더의 아이콘 + 공용 룬/기타만 후보로 사용합니다.

    목적:
    - 600개 전체 또는 다른 직업 아이콘이 섞이지 않게 합니다.
    - API CDN 후보와 로컬 DB 후보가 중복되어 top-k가 흔들리는 문제를 줄입니다.
    - 클래스 폴더가 없을 때만 기존 후보로 fallback합니다.
    """
    items: list[dict[str, str]] = []
    cls = _current_api_class_name_v98()
    try:
        from modules.fixed_grid_ocr import _load_local_skill_db_candidates_v80, _local_icon_candidates_v32
        if cls:
            for it in _load_local_skill_db_candidates_v80(cls) or []:
                name = str(it.get("name") or "").strip()
                icon = str(it.get("icon") or "").strip()
                if name and icon:
                    items.append({"name": name, "icon": icon, "source": "skill_local_db_class_strict"})
        # 룬/기본공격/기타는 클래스와 무관한 공용 후보로 유지합니다.
        for it in _local_icon_candidates_v32() or []:
            name = _normalize_special_battle_name_v86(str(it.get("name") or "").strip())
            icon = str(it.get("icon") or "").strip()
            src = str(it.get("source") or "local_common").strip()
            if not name or not icon:
                continue
            if name.startswith("스킬룬") or name in {"기본 공격", "기타"} or "rune" in src.lower() or "local_common" in src.lower():
                items.append({"name": name, "icon": icon, "source": src})
    except Exception as e:  # noqa: BLE001
        print(f"[v98] 클래스 전용 아이콘 후보 생성 실패: {e}")

    strict = _dedupe_icon_candidates_v98(items)
    if strict:
        try:
            st.session_state["last_icon_candidate_mode"] = "class_strict"
            st.session_state["last_icon_candidate_class"] = cls or "-"
            st.session_state["last_icon_candidate_count"] = len(strict)
        except Exception:
            pass
        return strict

    # 클래스 폴더를 못 찾은 경우에만 기존 후보로 fallback합니다.
    fallback = _dedupe_icon_candidates_v98(list(fallback_candidates or []))
    try:
        st.session_state["last_icon_candidate_mode"] = "fallback_all"
        st.session_state["last_icon_candidate_class"] = cls or "-"
        st.session_state["last_icon_candidate_count"] = len(fallback)
    except Exception:
        pass
    return fallback


_skill_only_ocr_candidates_prev_v98 = _skill_only_ocr_candidates


def _skill_only_ocr_candidates(candidates: list[dict[str, str]] | None) -> list[dict[str, str]]:  # type: ignore[override]
    """v98: 아이콘 매칭 후보를 API 클래스 폴더 + 공용 아이콘으로 강제 제한합니다."""
    fallback = _skill_only_ocr_candidates_prev_v98(candidates)
    return _build_strict_class_icon_candidates_v98(fallback)


# v98 기본값: 클래스 후보가 25~70개 수준이라 TopK를 12로 올려 정확도를 우선합니다.
# 속도가 필요하면 run.bat에서 LOA_ICON_HEAVY_TOPK 값을 낮추면 됩니다.
try:
    import os as _os_v98
    _os_v98.environ.setdefault("LOA_ICON_HEAVY_TOPK", "12")
except Exception:
    pass


# __main__ moved below v99 overrides


# ==============================================================================
# v99: 검색 버튼 아래 통합 디버그 ZIP + 버전 표시 고정
# ==============================================================================
APP_CALC_VERSION = "v99_integrated_debug"


def _zip_copy_with_prefix_v99(dst: zipfile.ZipFile, src_zip_path: Path | str, prefix: str) -> None:
    """기존 OCR 디버그 ZIP/성능 ZIP을 하나의 통합 ZIP 안에 prefix로 넣습니다."""
    try:
        src_path = Path(src_zip_path)
        if not src_path.exists():
            return
        with zipfile.ZipFile(src_path, "r") as src:
            for name in src.namelist():
                if name.endswith("/"):
                    continue
                try:
                    dst.writestr(f"{prefix}/{name}", src.read(name))
                except Exception:
                    pass
    except Exception as e:  # noqa: BLE001
        try:
            dst.writestr(f"{prefix}/COPY_ERROR.txt", repr(e))
        except Exception:
            pass


def _safe_session_json_v99() -> dict[str, Any]:
    """디버그용으로 안전한 session_state 일부만 저장합니다. API 키 원문은 저장하지 않습니다."""
    keys = [
        "character_name", "target_gain_percent", "real_row_count", "real_ocr_scale", "real_gpu_v34",
        "real_icon_match_threshold", "real_name_match_threshold", "last_icon_candidate_mode",
        "last_icon_candidate_class", "last_icon_candidate_count", "api_summary",
    ]
    out: dict[str, Any] = {}
    for k in keys:
        try:
            v = st.session_state.get(k)
            json.dumps(v, ensure_ascii=False, default=str)
            out[k] = v
        except Exception:
            out[k] = str(st.session_state.get(k))
    if "api_token" in st.session_state:
        out["api_token"] = "***masked***"
    return out


def _make_real_integrated_debug_zip_v99(summary_image: Any, attack_image: Any) -> tuple[Path, dict[str, Any]]:
    """OCR crop 디버그 + 성능 타임라인 + 아이콘 후보 정보를 한 ZIP으로 묶습니다."""
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    title = "실전 전투분석기"
    kind = "real"
    row_count = int(st.session_state.get("real_row_count", 18) or 18)
    ocr_scale = int(st.session_state.get("real_ocr_scale", 7) or 7)
    gpu = bool(st.session_state.get("real_gpu_v34", False))
    icon_threshold = float(st.session_state.get("real_icon_match_threshold", 74) or 74) / 100.0
    name_threshold = float(st.session_state.get("real_name_match_threshold", 52) or 52) / 100.0

    context: dict[str, Any] = {
        "version": APP_CALC_VERSION,
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "row_count": row_count,
        "ocr_scale": ocr_scale,
        "gpu": gpu,
        "icon_threshold": icon_threshold,
        "name_threshold": name_threshold,
        "api_class": _current_api_class_name_v98() if "_current_api_class_name_v98" in globals() else "",
        "env": {
            "LOA_ATTACK_RECHECK": __import__("os").environ.get("LOA_ATTACK_RECHECK"),
            "LOA_ICON_HEAVY_TOPK": __import__("os").environ.get("LOA_ICON_HEAVY_TOPK"),
            "LOA_ICON_SQUARE_CROP": __import__("os").environ.get("LOA_ICON_SQUARE_CROP"),
            "LOA_ICON_SQUARE_SIDE_RATIO": __import__("os").environ.get("LOA_ICON_SQUARE_SIDE_RATIO"),
            "LOA_ICON_SQUARE_FALLBACK_MARGIN": __import__("os").environ.get("LOA_ICON_SQUARE_FALLBACK_MARGIN"),
        },
        "errors": [],
    }

    candidates: list[dict[str, str]] = []
    try:
        candidates = _skill_only_ocr_candidates(get_ocr_skill_candidates_full())
        context["candidate_count"] = len(candidates)
        context["candidate_mode"] = st.session_state.get("last_icon_candidate_mode")
        context["candidate_class"] = st.session_state.get("last_icon_candidate_class")
    except Exception as e:  # noqa: BLE001
        context["errors"].append({"stage": "candidate_build", "error": repr(e)})

    reader = None
    ocr_debug_zip: Path | None = None
    perf_zip: Path | None = None
    perf_report: dict[str, Any] = {}
    perf_parsed: pd.DataFrame | None = None

    try:
        reader = get_easyocr_reader(gpu=gpu)
    except Exception as e:  # noqa: BLE001
        context["errors"].append({"stage": "reader", "error": repr(e)})

    if reader is not None:
        # 1) OCR crop 디버그
        try:
            summary_debug = make_summary_ocr_debug(summary_image, reader, scale=ocr_scale) if summary_image is not None else []
            attack_debug = make_attack_ocr_debug(
                attack_image,
                reader,
                row_count=row_count,
                scale=ocr_scale,
                skill_candidates=candidates,
            ) if attack_image is not None else []
            ocr_debug_zip = _save_ocr_debug_zip(kind, title, summary_image, attack_image, summary_debug, attack_debug)
        except Exception as e:  # noqa: BLE001
            context["errors"].append({"stage": "ocr_debug_zip", "error": repr(e)})

        # 2) 실제 분석 경로 성능 타임라인
        try:
            perf_report, perf_zip, perf_parsed = _run_ocr_perf_timeline_debug(
                kind,
                title,
                summary_image,
                attack_image,
                gpu=gpu,
                row_count=row_count,
                ocr_scale=ocr_scale,
                name_match_threshold=name_threshold,
                icon_match_threshold=icon_threshold,
            )
        except Exception as e:  # noqa: BLE001
            context["errors"].append({"stage": "perf_timeline", "error": repr(e)})

    out_zip = EXPORT_DIR / f"real_integrated_debug_{ts}.zip"
    with zipfile.ZipFile(out_zip, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        readme = (
            "LOA 실전 전투분석기 통합 디버그 ZIP\n"
            "\n"
            "이 ZIP 하나만 보내면 OCR crop, 성능 타임라인, 아이콘 후보/매칭 결과를 같이 볼 수 있습니다.\n"
            "\n"
            "폴더 설명:\n"
            "- ocr_debug/: 화면 crop, 행별 아이콘 crop, 셀별 OCR 원본 이미지\n"
            "- perf/: timestamp 기반 성능 로그와 icon_match 후 결과표\n"
            "- context/: 현재 클래스, 후보 목록, 세션 설정, 환경변수\n"
            "\n"
            "아이콘 매칭 문제를 볼 때 우선 확인할 파일:\n"
            "1) context/candidates_used.csv\n"
            "2) perf/parsed_attack_after_icon_match.csv\n"
            "3) ocr_debug/icons/row_XX_core.png 와 ocr_debug/icons/row_XX_cand*.png\n"
        )
        zf.writestr("README.txt", readme)
        zf.writestr("context/context.json", json.dumps(context, ensure_ascii=False, indent=2, default=str))
        zf.writestr("context/session_state_safe.json", json.dumps(_safe_session_json_v99(), ensure_ascii=False, indent=2, default=str))
        try:
            zf.writestr("context/candidates_used.csv", pd.DataFrame(candidates).to_csv(index=False).encode("utf-8-sig"))
        except Exception as e:
            zf.writestr("context/candidates_used_ERROR.txt", repr(e))
        if perf_report:
            zf.writestr("context/perf_report_copy.json", json.dumps(perf_report, ensure_ascii=False, indent=2, default=str))
        if isinstance(perf_parsed, pd.DataFrame):
            try:
                zf.writestr("context/parsed_attack_after_icon_match_copy.csv", perf_parsed.to_csv(index=False).encode("utf-8-sig"))
                icon_cols = [c for c in perf_parsed.columns if c in {
                    "이름", "피해량", "초당 피해량", "피해량 지분", "사용 횟수",
                    "_ocr_row_index", "_icon_match_name", "_icon_match_score", "_icon_match_source",
                    "_icon_match_top3", "_icon_crop_mode", "_icon_square_score", "_icon_legacy_score",
                    "_name_match_reason", "_name_match_icon_score",
                }]
                if icon_cols:
                    zf.writestr("context/icon_match_rows.csv", perf_parsed[icon_cols].to_csv(index=False).encode("utf-8-sig"))
            except Exception as e:
                zf.writestr("context/parsed_copy_ERROR.txt", repr(e))
        if ocr_debug_zip:
            _zip_copy_with_prefix_v99(zf, ocr_debug_zip, "ocr_debug")
        if perf_zip:
            _zip_copy_with_prefix_v99(zf, perf_zip, "perf")
    return out_zip, perf_report or {}


def _render_integrated_debug_button_v99() -> None:
    st.markdown("**🧪 통합 디버그**")
    st.caption("OCR crop + 성능 타임라인 + 아이콘 후보/Top3를 ZIP 하나로 묶습니다.")
    if st.button("실전 통합 디버그 ZIP 생성", key="sidebar_real_integrated_debug_v99", use_container_width=True):
        summary_image = _get_uploaded_image_from_session_v88("real", "summary")
        attack_image = _get_uploaded_image_from_session_v88("real", "attack")
        if summary_image is None and attack_image is None:
            st.warning("먼저 실전 전투분석기 종합정보/공격정보 이미지를 업로드하고 OCR을 한 번 실행해 주세요.")
        elif not easyocr_available():
            st.error("OCR 엔진이 설치되어 있지 않아 통합 디버그를 만들 수 없습니다.")
        else:
            with st.spinner("통합 디버그 ZIP 생성 중... OCR crop과 성능 타임라인을 함께 수집합니다."):
                zpath, report = _make_real_integrated_debug_zip_v99(summary_image, attack_image)
            st.session_state["real_integrated_debug_zip_v99"] = str(zpath)
            st.session_state["real_integrated_debug_report_v99"] = report
            st.success("통합 디버그 ZIP을 생성했습니다.")
    zpath = st.session_state.get("real_integrated_debug_zip_v99")
    if zpath and Path(zpath).exists():
        st.download_button(
            "실전 통합 디버그 ZIP 다운로드",
            data=Path(zpath).read_bytes(),
            file_name=Path(zpath).name,
            mime="application/zip",
            key="sidebar_real_integrated_debug_download_v99",
            use_container_width=True,
        )
    report = st.session_state.get("real_integrated_debug_report_v99") or {}
    if report:
        try:
            st.caption(
                f"최근 로그: 총 {float(report.get('total_elapsed_sec') or 0):.2f}s · "
                f"readtext {int(report.get('easyocr_calls') or 0)}회 · "
                f"OCR {float(report.get('easyocr_total_sec') or 0):.2f}s"
            )
        except Exception:
            pass


def _build_speed_debug_download_v154() -> tuple[bytes, str, bool]:
    timing = st.session_state.get("api_last_timing_v120") or st.session_state.get("api_last_timing_v119") or st.session_state.get("api_last_timing_v117") or {}
    summary = st.session_state.get("api_summary") or {}
    bundle = st.session_state.get("api_bundle") or {}
    profile = summary.get("profile_summary", {}) if isinstance(summary, dict) else {}

    endpoint_status: dict[str, Any] = {}
    if isinstance(bundle, dict):
        for key, value in bundle.items():
            if isinstance(value, dict):
                data = value.get("data")
                endpoint_status[str(key)] = {
                    "ok": value.get("ok"),
                    "status_code": value.get("status_code"),
                    "error": value.get("error"),
                    "data_type": type(data).__name__,
                    "data_size_chars": len(str(data)) if data is not None else 0,
                }

    recommendations: list[str] = []
    if isinstance(timing, dict):
        fetch_ms = float(timing.get("fetch_armory_bundle_ms") or 0)
        summarize_ms = float(timing.get("summarize_all_ms") or 0)
        enrich_ms = float(timing.get("enrich_summary_with_identity_ms") or 0)
        wall_ms = float(timing.get("search_button_wall_before_render_ms_v121") or timing.get("total_ms") or 0)
        endpoint_timings = timing.get("endpoint_timings_ms") or {}
        if fetch_ms > 5000:
            recommendations.append("API 응답 시간이 5초를 넘었습니다. endpoint_timings_ms에서 느린 Lost Ark API 요청을 확인하세요.")
        if summarize_ms > 3000:
            recommendations.append("API 응답 이후 계산표 생성이 3초를 넘었습니다. summarize_detail_v117 안의 단계별 시간을 확인하세요.")
        if enrich_ms > 2000:
            recommendations.append("직업/아이덴티티 보정 계산이 2초를 넘었습니다. enrich_summary_with_identity_ms가 병목일 수 있습니다.")
        if wall_ms > fetch_ms + summarize_ms + enrich_ms + 2000:
            recommendations.append("측정된 API/계산 시간보다 화면 대기 시간이 큽니다. Streamlit 재실행 또는 화면 렌더링 비용을 의심할 수 있습니다.")
        if isinstance(endpoint_timings, dict) and endpoint_timings:
            slowest = sorted(endpoint_timings.items(), key=lambda item: float(item[1] or 0), reverse=True)[:3]
            recommendations.append(f"가장 느린 API 요청 TOP3: {slowest}")
        estimator_detail = timing.get("estimate_skill_crit_tables_detail_v157") if isinstance(timing, dict) else {}
        if isinstance(estimator_detail, dict) and estimator_detail.get("top_stages"):
            recommendations.append(f"estimator internal TOP: {estimator_detail.get('top_stages')[:5]}")
        if not recommendations:
            recommendations.append("큰 단일 병목은 기록되지 않았습니다. total_ms, endpoint_timings_ms, summarize_detail_v117을 함께 비교하세요.")

    payload = {
        "debug_version": "v157_search_speed_estimator_detail",
        "created_at_epoch": _time_v120_app.time(),
        "character_name": st.session_state.get("character_name", ""),
        "profile_summary": profile,
        "timing": timing,
        "estimator_timing": (
            (timing.get("estimate_skill_crit_tables_detail_v157") if isinstance(timing, dict) else None)
            or (summary.get("_estimator_timing_v157") if isinstance(summary, dict) else None)
            or {}
        ),
        "endpoint_status": endpoint_status,
        "summary_keys": list(summary.keys()) if isinstance(summary, dict) else [],
        "recommendations": recommendations,
        "note": "API 키 원문은 포함하지 않습니다.",
    }
    raw_name = str(st.session_state.get("character_name") or "loa").strip()
    safe_name = re.sub(r"[^0-9A-Za-z가-힣_-]+", "_", raw_name)[:40] or "loa"
    data = _json_v117_app.dumps(payload, ensure_ascii=False, indent=2, default=str).encode("utf-8")
    return data, f"loa_search_speed_debug_{safe_name}.json", bool(timing)


def sidebar_controls() -> None:  # type: ignore[override]
    """v99: 좌측 검색 아래에 통합 디버그 버튼을 고정 노출합니다."""
    calc_config = load_yaml(str(CONFIG_DIR / "calculation_presets.yaml"))
    defaults = calc_config.get("defaults", {})

    st.session_state.setdefault("api_crit_basis", "백어택 기준(조건부 포함)")
    st.session_state.setdefault("target_crit_percent", float(defaults.get("target_crit_percent", 100.0)))
    st.session_state.setdefault("target_back_percent", float(defaults.get("target_back_attack_percent", 100.0)))
    st.session_state.setdefault("back_bonus_percent", float(defaults.get("back_attack_damage_bonus_percent", 5.0)))

    with st.sidebar:
        st.markdown("### LOA 실전 분석")
        st.info("통합 디버그 ZIP + 클래스 전용 아이콘 후보")
        st.caption("캐릭터를 불러오고, 목표 상승률만 정한 뒤 전투분석기 이미지를 넣으면 됩니다.")

        st.session_state["show_debug_tools"] = st.checkbox(
            "🔧 디버그 도구 표시",
            value=bool(st.session_state.get("show_debug_tools", False)),
            help="본문의 OCR 세부 디버그 영역을 표시합니다. 통합 ZIP 버튼은 검색 아래에 항상 있습니다.",
        )

        token = st.text_input(
            "Lost Ark API Key",
            type="password",
            key="api_token",
            help="developer-lostark에서 발급받은 JWT/API KEY를 넣으세요.",
        )
        character_name = st.text_input(
            "캐릭터명",
            value=st.session_state.get("character_name", ""),
            key="character_name",
            placeholder="캐릭터명을 입력하세요",
        )
        st.number_input(
            "목표 상승률(%)",
            value=float(st.session_state.get("target_gain_percent", defaults.get("target_gain_percent", 5.0))),
            min_value=0.0,
            max_value=100.0,
            step=0.5,
            key="target_gain_percent",
            help="개선분석에서 목표로 삼을 총 딜 상승률입니다.",
        )

        fetch = st.button("검색", type="primary", use_container_width=True, key="sidebar_fetch_api_v99")

        # 사용자가 찾기 쉽게 검색 바로 아래에 통합 디버그 버튼을 고정합니다.
        _render_integrated_debug_button_v99()

        if fetch:
            if not token or not character_name:
                st.warning("API Key와 캐릭터명을 먼저 입력해줘.")
            else:
                with st.spinner("캐릭터 셋팅을 불러오는 중..."):
                    client = LostArkApiClient(token, CONFIG_DIR / "api_endpoints.yaml")
                    bundle = client.fetch_armory_bundle(character_name)
                    st.session_state.api_bundle = serializable_bundle(bundle)
                    st.session_state.api_summary = enrich_summary_with_identity(summarize_all(bundle), st.session_state.api_bundle)
                failed = {k: v for k, v in st.session_state.api_bundle.items() if not v.get("ok")}
                if failed:
                    st.warning(f"일부 API 호출 실패: {', '.join(failed.keys())}")
                else:
                    st.success("캐릭터 셋팅을 불러왔어.")

        summary = st.session_state.get("api_summary") or {}
        profile = summary.get("profile_summary", {}) if isinstance(summary, dict) else {}
        if profile:
            st.divider()
            st.caption("현재 분석 캐릭터")
            st.markdown(f"**{profile.get('캐릭터명') or character_name}**")
            meta = " · ".join(str(x) for x in [profile.get("서버"), profile.get("클래스"), profile.get("직업"), profile.get("아이템레벨")] if x)
            if meta:
                st.caption(meta)
            try:
                st.caption(
                    f"아이콘 후보: {st.session_state.get('last_icon_candidate_mode', '-')} · "
                    f"{st.session_state.get('last_icon_candidate_class', '-')} · "
                    f"{st.session_state.get('last_icon_candidate_count', '-')}개"
                )
            except Exception:
                pass


# __main__ moved below v100 overrides

# ==============================================================================
# v100: OCR 숫자 단위 복구 + 통합 디버그 유지
# ==============================================================================
APP_CALC_VERSION = "v100_damage_unit_repair"


def _repair_damage_text_unit_v100(value: Any) -> Any:
    """공격정보 피해량 OCR에서 `억`이 9/91/94처럼 붙거나 누락된 값을 복구합니다.

    예:
    - `2,083.449`  -> `2,083.44억`
    - `544.219`   -> `544.21억`
    - `395.4091`  -> `395.40억`

    전투분석기 공격정보 피해량 칸은 대부분 `xx.xx억` 형태인데, RapidOCR이 한글 단위 `억`을
    숫자 9/91/94로 읽는 경우가 많습니다. 이 함수는 피해량 컬럼에만 적용합니다.
    """
    try:
        if value is None:
            return value
        if isinstance(value, float) and math.isnan(value):
            return value
        text = str(value).strip()
        if not text or text in {"-", "None", "nan", "피해량", "DPS", "지분%"}:
            return value
        compact = text.replace(" ", "")
        # 이미 한글 단위가 있으면 그대로 둡니다.
        if any(unit in compact for unit in ["조", "억", "만"]):
            return text
        # 원시 숫자 형태 464,126,236,306 처럼 큰 정수면 그대로 둡니다.
        raw_num = re.sub(r"[^0-9.]", "", compact)
        if not raw_num or raw_num.count(".") > 1:
            return value
        if "." not in raw_num:
            try:
                as_int = int(raw_num)
                if as_int >= 10_000_000:
                    return text
            except Exception:
                pass
            return value
        int_part, dec_part = raw_num.split(".", 1)
        if not int_part:
            return value
        try:
            base = float(int_part + "." + re.sub(r"[^0-9]", "", dec_part))
        except Exception:
            return value
        # 공격정보 피해량으로 보기에는 너무 큰 원시값이면 건드리지 않습니다.
        if base >= 10_000_000:
            return text
        # 소수부 뒤쪽의 9/91/94는 `억` 오인식인 경우가 많으므로 소수 2자리만 사용합니다.
        dec_digits = re.sub(r"[^0-9]", "", dec_part)
        if len(dec_digits) >= 2:
            repaired_num = f"{int(int_part):,}.{dec_digits[:2]}"
        elif len(dec_digits) == 1:
            repaired_num = f"{int(int_part):,}.{dec_digits}0"
        else:
            repaired_num = f"{int(int_part):,}"
        return f"{repaired_num}억"
    except Exception:
        return value


def _repair_attack_damage_units_v100(df: pd.DataFrame | None) -> pd.DataFrame:
    if df is None or not isinstance(df, pd.DataFrame) or df.empty:
        return df if isinstance(df, pd.DataFrame) else pd.DataFrame()
    out = df.copy()
    for col in ["피해량", "damage_text"]:
        if col in out.columns:
            out[col] = out[col].apply(_repair_damage_text_unit_v100)
    return out


_sanitize_battle_table_prev_v100 = sanitize_battle_table


def sanitize_battle_table(df: pd.DataFrame) -> pd.DataFrame:  # type: ignore[override]
    df2 = _repair_attack_damage_units_v100(df)
    return _sanitize_battle_table_prev_v100(df2)


_aggregate_unmatched_rows_as_extra_prev_v100 = _aggregate_unmatched_rows_as_extra


def _aggregate_unmatched_rows_as_extra(df: pd.DataFrame, *, label: str = "인식 안된 기타 추가딜") -> pd.DataFrame:  # type: ignore[override]
    return _aggregate_unmatched_rows_as_extra_prev_v100(_repair_attack_damage_units_v100(df), label=label)


_repair_battle_values_from_summary_prev_v100 = _repair_battle_values_from_summary


def _repair_battle_values_from_summary(df: pd.DataFrame, meta: dict[str, Any] | None, *, replace_damage_from_share: bool = False, elapsed_fallback: float | None = None) -> pd.DataFrame:  # type: ignore[override]
    return _repair_battle_values_from_summary_prev_v100(
        _repair_attack_damage_units_v100(df),
        meta,
        replace_damage_from_share=replace_damage_from_share,
        elapsed_fallback=elapsed_fallback,
    )


_prepare_battle_editor_df_prev_v100 = prepare_battle_editor_df


def prepare_battle_editor_df(df: pd.DataFrame) -> pd.DataFrame:  # type: ignore[override]
    return _prepare_battle_editor_df_prev_v100(_repair_attack_damage_units_v100(df))


_sidebar_controls_prev_v100 = sidebar_controls


def sidebar_controls() -> None:  # type: ignore[override]
    pass  # 버전 변경 안내 제거
    return _sidebar_controls_prev_v100()


# ==============================================================================
# v101: glyph_templates.json 숫자 전용 강제 실험 모드
# ==============================================================================
APP_CALC_VERSION = "v101_glyph_only_numbers"

_sidebar_controls_prev_v101 = sidebar_controls


def sidebar_controls() -> None:  # type: ignore[override]
    
    import os as _os_v103_msg
    if str(_os_v103_msg.environ.get("LOA_GLYPH_ONLY_NUMBERS", "0")).strip().lower() in {"1", "true", "yes", "on"}:
        st.sidebar.success("v101 glyph-only 숫자 실험 ON")
        st.sidebar.caption("피해량/DPS/퍼센트/횟수 숫자를 OCR 없이 글리프 결과로 강제 표시합니다. 정확도 검증용입니다.")
    else:
        st.sidebar.caption("glyph-only 숫자 실험 OFF · 숫자는 OCR + 단위보정 사용")
    return _sidebar_controls_prev_v101()


# __main__ moved below v102 overrides


# ==============================================================================
# v102: 아이콘 후보 판정 수정 + 통합 디버그 crop 생성 오류 수정
# ==============================================================================
APP_CALC_VERSION = "v102_icon_common_fix"

_make_real_integrated_debug_zip_prev_v102 = _make_real_integrated_debug_zip_v99

def _make_real_integrated_debug_zip_v99(summary_image: Any, attack_image: Any) -> tuple[Path, dict[str, Any]]:  # type: ignore[override]
    """v102: make_attack_ocr_debug 인자 불일치로 crop 디버그가 빠지던 문제를 복구합니다.

    v99는 make_attack_ocr_debug(..., skill_candidates=...)를 호출했지만 설치된 함수 시그니처가
    이를 받지 않아 OCR crop 이미지가 통합 ZIP에 들어가지 않았습니다. 여기서는 원래 함수 호출로
    돌아가고, 후보/아이콘 매칭 결과는 context CSV로 별도 저장합니다.
    """
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    title = "실전 전투분석기"
    kind = "real"
    row_count = int(st.session_state.get("real_row_count", 18) or 18)
    ocr_scale = int(st.session_state.get("real_ocr_scale", 7) or 7)
    gpu = bool(st.session_state.get("real_gpu_v34", False))
    icon_threshold = float(st.session_state.get("real_icon_match_threshold", 74) or 74) / 100.0
    name_threshold = float(st.session_state.get("real_name_match_threshold", 52) or 52) / 100.0

    context: dict[str, Any] = {
        "version": APP_CALC_VERSION,
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "row_count": row_count,
        "ocr_scale": ocr_scale,
        "gpu": gpu,
        "icon_threshold": icon_threshold,
        "name_threshold": name_threshold,
        "api_class": _current_api_class_name_v98() if "_current_api_class_name_v98" in globals() else "",
        "env": {
            "LOA_GLYPH_ONLY_NUMBERS": __import__("os").environ.get("LOA_GLYPH_ONLY_NUMBERS"),
            "LOA_ATTACK_RECHECK": __import__("os").environ.get("LOA_ATTACK_RECHECK"),
            "LOA_ICON_HEAVY_TOPK": __import__("os").environ.get("LOA_ICON_HEAVY_TOPK"),
            "LOA_ICON_SQUARE_CROP": __import__("os").environ.get("LOA_ICON_SQUARE_CROP"),
            "LOA_ICON_SQUARE_SIDE_RATIO": __import__("os").environ.get("LOA_ICON_SQUARE_SIDE_RATIO"),
            "LOA_ICON_SQUARE_FALLBACK_MARGIN": __import__("os").environ.get("LOA_ICON_SQUARE_FALLBACK_MARGIN"),
        },
        "errors": [],
    }

    candidates: list[dict[str, str]] = []
    try:
        candidates = _skill_only_ocr_candidates(get_ocr_skill_candidates_full())
        context["candidate_count"] = len(candidates)
        context["candidate_mode"] = st.session_state.get("last_icon_candidate_mode")
        context["candidate_class"] = st.session_state.get("last_icon_candidate_class")
    except Exception as e:  # noqa: BLE001
        context["errors"].append({"stage": "candidate_build", "error": repr(e)})

    reader = None
    ocr_debug_zip: Path | None = None
    perf_zip: Path | None = None
    perf_report: dict[str, Any] = {}
    perf_parsed: pd.DataFrame | None = None

    try:
        reader = get_easyocr_reader(gpu=gpu)
    except Exception as e:  # noqa: BLE001
        context["errors"].append({"stage": "reader", "error": repr(e)})

    if reader is not None:
        try:
            summary_debug = make_summary_ocr_debug(summary_image, reader, scale=ocr_scale) if summary_image is not None else []
            try:
                attack_debug = make_attack_ocr_debug(attack_image, reader, row_count=row_count, scale=ocr_scale) if attack_image is not None else []
            except TypeError:
                # 혹시 다른 버전의 함수 시그니처가 섞여도 최소한 디버그 생성은 살립니다.
                attack_debug = make_attack_ocr_debug(attack_image, reader, row_count=row_count) if attack_image is not None else []
            ocr_debug_zip = _save_ocr_debug_zip(kind, title, summary_image, attack_image, summary_debug, attack_debug)
        except Exception as e:  # noqa: BLE001
            context["errors"].append({"stage": "ocr_debug_zip", "error": repr(e)})

        try:
            perf_report, perf_zip, perf_parsed = _run_ocr_perf_timeline_debug(
                kind,
                title,
                summary_image,
                attack_image,
                reader,
                row_count=row_count,
                scale=ocr_scale,
                icon_match_threshold=icon_threshold,
                name_match_threshold=name_threshold,
            )
        except Exception as e:  # noqa: BLE001
            context["errors"].append({"stage": "perf_zip", "error": repr(e)})

    out_path = Path(f"real_integrated_debug_{ts}.zip")
    try:
        with zipfile.ZipFile(out_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
            zf.writestr("README.txt", "LOA v102 통합 디버그 ZIP\n- ocr_debug/: OCR crop 원본\n- perf/: 성능 타임라인\n- context/: 후보/세션/아이콘 매칭 결과\n")
            zf.writestr("context/context.json", json.dumps(context, ensure_ascii=False, indent=2, default=str))
            zf.writestr("context/session_state_safe.json", json.dumps(_safe_session_json_v99(), ensure_ascii=False, indent=2, default=str))
            if candidates:
                zf.writestr("context/candidates_used.csv", pd.DataFrame(candidates).to_csv(index=False, encoding="utf-8-sig"))
            if perf_report:
                zf.writestr("context/perf_report_copy.json", json.dumps(perf_report, ensure_ascii=False, indent=2, default=str))
            if isinstance(perf_parsed, pd.DataFrame) and not perf_parsed.empty:
                zf.writestr("context/parsed_attack_after_icon_match_copy.csv", perf_parsed.to_csv(index=False, encoding="utf-8-sig"))
                icon_cols = [c for c in perf_parsed.columns if c.startswith("_icon") or c.startswith("_name_match") or c in ["이름", "피해량", "사용 횟수", "초당 피해량", "피해량 지분", "_ocr_row_index"]]
                if icon_cols:
                    zf.writestr("context/icon_match_rows.csv", perf_parsed[icon_cols].to_csv(index=False, encoding="utf-8-sig"))
            if ocr_debug_zip:
                _zip_copy_with_prefix_v99(zf, ocr_debug_zip, "ocr_debug")
            if perf_zip:
                _zip_copy_with_prefix_v99(zf, perf_zip, "perf")
    except Exception as e:  # noqa: BLE001
        context["errors"].append({"stage": "integrated_zip_write", "error": repr(e)})
        with zipfile.ZipFile(out_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
            zf.writestr("context/context.json", json.dumps(context, ensure_ascii=False, indent=2, default=str))
    return out_path, context

_sidebar_controls_prev_v102 = sidebar_controls

def sidebar_controls() -> None:  # type: ignore[override]
    pass  # 버전 변경 안내 제거
    st.sidebar.caption("아이콘은 파일 크기와 무관하게 내부에서 64×64 정규화 후 비교합니다.")
    return _sidebar_controls_prev_v102()


# ==============================================================================
# v103: OCR + unit repair restore after glyph-only experiment
# ==============================================================================
APP_CALC_VERSION = "v103_ocr_unit_repair_icon_common"

# glyph-only 실험은 기본 OFF로 고정합니다. run.bat이 누락되어도 OCR+단위보정 경로를 사용합니다.
try:
    import os as _os_v103
    _os_v103.environ.setdefault("LOA_GLYPH_ONLY_NUMBERS", "0")
except Exception:
    pass

_sidebar_controls_prev_v103 = sidebar_controls

def sidebar_controls() -> None:  # type: ignore[override]
    pass  # 버전 변경 안내 제거
    st.sidebar.caption("피해량/DPS/지분 등 숫자는 glyph-only가 아니라 OCR 결과를 읽은 뒤 억/만/조 단위를 보정합니다.")
    return _sidebar_controls_prev_v103()

# __main__ moved below v104 overrides

# ==============================================================================
# v104: comma-formatted manual summary inputs + preserve unmatched rows for review
# ==============================================================================
APP_CALC_VERSION = "v104_comma_inputs_ocr_unit_icon_review"


def _parse_big_number_text_v104(value: Any) -> float:
    """수동 입력 문자열을 숫자로 변환합니다.

    허용 예시:
    - 479,320,187,168
    - 479320187168
    - 4,793.20억
    - 15.77억
    """
    try:
        if value is None:
            return 0.0
        if isinstance(value, (int, float)):
            if isinstance(value, float) and math.isnan(value):
                return 0.0
            return float(value)
        text = str(value).strip()
        if not text or text in {"-", "None", "nan"}:
            return 0.0
        if any(unit in text for unit in ["조", "억", "만"]):
            parsed = parse_korean_number(text)
            return float(parsed or 0.0)
        cleaned = re.sub(r"[^0-9.\-]", "", text.replace(",", ""))
        if not cleaned or cleaned in {".", "-", "-."}:
            return 0.0
        return float(cleaned)
    except Exception:
        return 0.0


def _format_commas_v104(value: Any) -> str:
    try:
        num = _parse_big_number_text_v104(value)
        if not num or num <= 0:
            return "0"
        if abs(num - round(num)) < 0.0001:
            return f"{int(round(num)):,}"
        return f"{num:,.2f}".rstrip("0").rstrip(".")
    except Exception:
        return str(value or "0")


_sync_meta_to_manual_inputs_prev_v104 = _sync_meta_to_manual_inputs


def _sync_meta_to_manual_inputs(kind: str, meta: dict[str, Any]) -> None:  # type: ignore[override]
    """v104: OCR 결과를 숫자 키와 콤마 표시용 text 키에 동시에 반영합니다."""
    _sync_meta_to_manual_inputs_prev_v104(kind, meta)
    if not isinstance(meta, dict):
        return
    try:
        total = meta.get("total_damage")
        if total:
            st.session_state[f"manual_{kind}_total_damage_text_v104"] = _format_commas_v104(total)
    except Exception:
        pass
    try:
        dps = meta.get("dps")
        if dps:
            st.session_state[f"manual_{kind}_dps_text_v104"] = _format_commas_v104(dps)
    except Exception:
        pass




_aggregate_unmatched_rows_as_extra_prev_v104 = _aggregate_unmatched_rows_as_extra


def _aggregate_unmatched_rows_as_extra(df: pd.DataFrame, *, label: str = "인식 안된 기타 추가딜") -> pd.DataFrame:  # type: ignore[override]
    """v104: 아이콘 매칭이 애매해도 행을 접지 않고 검수표에 남깁니다.

    이전 방식은 매칭 실패 행을 하나의 `기타 추가딜`로 합산했기 때문에, 아이콘 매칭 문제가 생기면
    스킬별 피해량이 모두 사라져 원인 확인이 어려웠습니다. 이제는 개별 행을 그대로 남기고,
    이름만 `기타` 또는 OCR/아이콘 후보 결과로 보여 사용자가 직접 고칠 수 있게 합니다.
    """
    if df is None or not isinstance(df, pd.DataFrame) or df.empty:
        return df if isinstance(df, pd.DataFrame) else pd.DataFrame()
    out = _repair_attack_damage_units_v100(df.copy()) if "_repair_attack_damage_units_v100" in globals() else df.copy()
    if "이름" in out.columns:
        # 헤더/완전 빈 행만 제거하고, 실제 피해량/횟수가 있는 행은 보존합니다.
        bad_names = {"", "이름", "피해량", "DPS", "지분%", "백적중%", "백비중%", "치적%", "치비중%", "횟수", "쿨%"}
        keep = []
        for _, row in out.iterrows():
            name = str(row.get("이름") or "").strip()
            dmg = parse_korean_number(row.get("피해량")) if "피해량" in out.columns else None
            casts = safe_int_value(row.get("사용 횟수"), 0) if "사용 횟수" in out.columns else 0
            if name in bad_names and not (dmg and dmg > 0) and not (casts and casts > 0):
                keep.append(False)
            else:
                keep.append(True)
        out = out.loc[keep].reset_index(drop=True)
    return out


_sidebar_controls_prev_v104 = sidebar_controls


def sidebar_controls() -> None:  # type: ignore[override]
    pass  # 버전 변경 안내 제거
    st.sidebar.caption("숫자는 glyph-only가 아니라 OCR+단위보정이며, 아이콘 매칭 실패 행도 기타추가딜로 접지 않고 표에 남깁니다.")
    return _sidebar_controls_prev_v104()

# v103의 __main__ 호출 뒤에 붙은 경우를 대비해, 파일 끝 실행을 위해 별도 호출하지 않습니다.



# ==============================================================================
# v106: clean sidebar + repaired manual comma inputs + integrated debug refresh
# ==============================================================================
APP_CALC_VERSION = "v107_accuracy_mode"

# glyph-only 실험은 기본 OFF로 고정합니다. 숫자는 OCR + 억/만/조 단위보정 경로를 사용합니다.
try:
    import os as _os_v106
    _os_v106.environ["LOA_GLYPH_ONLY_NUMBERS"] = "0"
except Exception:
    pass


def _v106_now_stamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def _v106_text_key(kind: str, name: str) -> str:
    return f"manual_{kind}_{name}_text_v106"


def _seed_manual_text_inputs_v106(kind: str, meta: dict[str, Any] | None = None, *, force: bool = False) -> None:
    """OCR/meta 결과를 수동 입력 text key에 안전하게 반영합니다.

    Streamlit widget key는 해당 run에서 widget 생성 전에만 수정해야 합니다.
    이 함수는 OCR 버튼 처리 직후, 수동 입력 위젯이 그려지기 전에 호출됩니다.
    """
    meta = dict(meta or st.session_state.get(f"{kind}_meta", {}) or {})
    total_key = _v106_text_key(kind, "total_damage")
    dps_key = _v106_text_key(kind, "dps")
    total_input_key = f"manual_{kind}_total_damage_input"
    dps_input_key = f"manual_{kind}_dps_input"
    dps_value_key = f"manual_{kind}_dps"

    elapsed = meta.get("elapsed_seconds")
    if elapsed is not None:
        try:
            st.session_state[f"{kind}_elapsed_sec"] = float(elapsed)
        except Exception:
            pass

    total = meta.get("total_damage") or st.session_state.get(total_input_key)
    dps = meta.get("dps") or st.session_state.get(dps_input_key) or st.session_state.get(dps_value_key)

    try:
        if force or total_key not in st.session_state or str(st.session_state.get(total_key) or "").strip() in {"", "0", "0.0"}:
            st.session_state[total_key] = _format_commas_v104(total or 0)
        st.session_state[total_input_key] = float(_parse_big_number_text_v104(st.session_state.get(total_key, "0")) or 0.0)
    except Exception:
        st.session_state[total_key] = "0"
        st.session_state[total_input_key] = 0.0

    try:
        if force or dps_key not in st.session_state or str(st.session_state.get(dps_key) or "").strip() in {"", "0", "0.0"}:
            st.session_state[dps_key] = _format_commas_v104(dps or 0)
        dps_value = float(_parse_big_number_text_v104(st.session_state.get(dps_key, "0")) or 0.0)
        st.session_state[dps_input_key] = dps_value
        st.session_state[dps_value_key] = dps_value if dps_value > 0 else st.session_state.get(dps_value_key)
    except Exception:
        st.session_state[dps_key] = "0"
        st.session_state[dps_input_key] = 0.0


def _sync_meta_to_manual_inputs(kind: str, meta: dict[str, Any]) -> None:  # type: ignore[override]
    """v106: OCR 종합정보를 수동 입력칸에도 즉시 반영합니다."""
    if not isinstance(meta, dict):
        return
    try:
        if meta.get("total_damage") or meta.get("dps") or meta.get("elapsed_seconds") is not None:
            _seed_manual_text_inputs_v106(kind, meta, force=True)
    except Exception:
        pass


def _normalize_manual_big_number_text_v106(text_key: str, number_key: str) -> None:
    """text_input on_change 콜백: 1000단위 콤마 자동 정리."""
    raw = st.session_state.get(text_key, "")
    num = _parse_big_number_text_v104(raw)
    st.session_state[text_key] = _format_commas_v104(num)
    st.session_state[number_key] = float(num or 0.0)


def _manual_meta_editor(kind: str, title: str) -> None:  # type: ignore[override]
    """v106: 직접 입력칸을 단일 source of truth로 정리한 버전."""
    st.markdown(f"#### {title} 종합정보")
    with st.container(border=True):
        st.caption("전투 시간 · 총 피해량 · DPS가 잘못 읽혔으면 여기서 직접 고치세요. 콤마는 Enter/포커스 이동 시 자동 정리됩니다.")
        elapsed_key = f"{kind}_elapsed_sec"
        meta_key = f"{kind}_meta"
        total_input_key = f"manual_{kind}_total_damage_input"
        dps_input_key = f"manual_{kind}_dps_input"
        dps_value_key = f"manual_{kind}_dps"
        total_text_key = _v106_text_key(kind, "total_damage")
        dps_text_key = _v106_text_key(kind, "dps")

        # widget 생성 전에만 seed합니다. 이미 사용자가 입력 중인 값은 덮지 않습니다.
        _seed_manual_text_inputs_v106(kind, st.session_state.get(meta_key, {}), force=False)

        mc1, mc2 = st.columns(2)
        with mc1:
            st.number_input(
                f"{title} 전투 시간(초)",
                min_value=0.0,
                value=float(st.session_state.get(elapsed_key, 0.0) or 0.0),
                step=1.0,
                key=elapsed_key,
                help="예: 5분 4초 = 304초.",
            )
            try:
                elapsed_i = int(float(st.session_state.get(elapsed_key, 0.0) or 0.0))
                st.caption(f"= {elapsed_i // 60:02d}:{elapsed_i % 60:02d}")
            except Exception:
                pass
        with mc2:
            st.text_input(
                f"{title} 총 피해량(숫자)",
                key=total_text_key,
                help="콤마/한글 단위 입력 가능. 예: 479,320,187,168 또는 4,793.20억",
                on_change=_normalize_manual_big_number_text_v106,
                args=(total_text_key, total_input_key),
            )
            tv = _parse_big_number_text_v104(st.session_state.get(total_text_key, "0"))
            st.session_state[total_input_key] = float(tv or 0.0)
            if tv > 0:
                st.caption(f"= {_format_commas_v104(tv)} / {format_korean_number(tv)}")

        st.text_input(
            f"{title} DPS 직접 입력",
            key=dps_text_key,
            help="콤마/한글 단위 입력 가능. 예: 1,576,711,142 또는 15.77억",
            on_change=_normalize_manual_big_number_text_v106,
            args=(dps_text_key, dps_input_key),
        )
        dv = _parse_big_number_text_v104(st.session_state.get(dps_text_key, "0"))
        st.session_state[dps_input_key] = float(dv or 0.0)
        if dv > 0:
            st.caption(f"DPS = {_format_commas_v104(dv)} / {format_korean_number(dv)}")

        # 수동 입력값을 meta에 반영합니다. 0은 OCR 값을 지워버리지 않도록 반영하지 않습니다.
        meta = dict(st.session_state.get(meta_key, {}) or {})
        changed = False
        if tv > 0:
            meta["total_damage"] = float(tv)
            meta["total_damage_text"] = format_korean_number(float(tv))
            changed = True
        if dv > 0:
            st.session_state[dps_value_key] = float(dv)
            meta["dps"] = float(dv)
            meta["dps_text"] = format_korean_number(float(dv))
            changed = True
        if changed:
            st.session_state[meta_key] = meta

        if st.button(f"✅ 수정값으로 ③ 실력 분석 결과 다시 계산", key=f"{kind}_recompute_btn_v106", use_container_width=True):
            try:
                table_key = f"{kind}_table"
                st.session_state[table_key] = prepare_battle_editor_df(sanitize_battle_table(
                    _repair_battle_values_from_summary(
                        st.session_state.get(table_key),
                        st.session_state.get(meta_key, {}),
                        elapsed_fallback=st.session_state.get(elapsed_key),
                    )
                ))
                _bump_table_version(kind)
            except Exception as e:
                st.warning(f"재계산 중 일부 오류가 있었어요: {e}")
            st.rerun()


def _v106_icon_debug_columns(df: pd.DataFrame | None) -> pd.DataFrame:
    if df is None or not isinstance(df, pd.DataFrame) or df.empty:
        return pd.DataFrame()
    wanted = [
        "_ocr_row_index", "이름", "피해량", "초당 피해량", "피해량 지분", "사용 횟수", "쿨타임 비율",
        "_icon_match_name", "_icon_match_score", "_icon_match_source", "_icon_match_top3",
        "_name_match_reason", "_name_match_icon_score", "_icon_crop_mode", "_icon_square_score", "_icon_legacy_score",
    ]
    cols = [c for c in wanted if c in df.columns]
    extra = [c for c in df.columns if c.startswith("_icon") or c.startswith("_name_match")]
    cols += [c for c in extra if c not in cols]
    return df[cols].copy() if cols else df.copy()


def _make_real_integrated_debug_zip_v106(summary_image: Any, attack_image: Any) -> tuple[Path, dict[str, Any]]:
    """OCR crop + 성능 타임라인 + 후보/매칭 결과를 하나로 묶는 최신 통합 디버그."""
    EXPORT_DIR.mkdir(exist_ok=True)
    ts = _v106_now_stamp()
    kind = "real"
    title = "실전 전투분석기"
    row_count = int(st.session_state.get("real_row_count", 18) or 18)
    ocr_scale = int(st.session_state.get("real_ocr_scale", 7) or 7)
    gpu = bool(st.session_state.get("real_gpu_v34", False))
    icon_threshold = float(st.session_state.get("real_icon_match_threshold", 74) or 74) / 100.0
    name_threshold = float(st.session_state.get("real_name_match_threshold", 52) or 52) / 100.0

    context: dict[str, Any] = {
        "version": APP_CALC_VERSION,
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "row_count": row_count,
        "ocr_scale": ocr_scale,
        "gpu": gpu,
        "icon_threshold": icon_threshold,
        "name_threshold": name_threshold,
        "api_class": _current_api_class_name_v98() if "_current_api_class_name_v98" in globals() else "",
        "candidate_mode": st.session_state.get("last_icon_candidate_mode"),
        "candidate_class": st.session_state.get("last_icon_candidate_class"),
        "candidate_count_session": st.session_state.get("last_icon_candidate_count"),
        "env": {},
        "errors": [],
    }
    try:
        import os as _os
        for k in ["LOA_GLYPH_ONLY_NUMBERS", "LOA_ATTACK_RECHECK", "LOA_ICON_HEAVY_TOPK", "LOA_ICON_SQUARE_CROP", "LOA_ICON_SQUARE_SIDE_RATIO"]:
            context["env"][k] = _os.environ.get(k)
    except Exception:
        pass

    candidates: list[dict[str, str]] = []
    try:
        candidates = _skill_only_ocr_candidates(get_ocr_skill_candidates_full())
        context["candidate_count_actual"] = len(candidates)
        context["candidate_mode_after_build"] = st.session_state.get("last_icon_candidate_mode")
        context["candidate_class_after_build"] = st.session_state.get("last_icon_candidate_class")
    except Exception as e:
        context["errors"].append({"stage": "candidate_build", "error": repr(e)})

    ocr_debug_zip: Path | None = None
    perf_zip: Path | None = None
    perf_report: dict[str, Any] = {}
    perf_parsed: pd.DataFrame | None = None
    summary_debug: list[dict[str, Any]] = []
    attack_debug: list[dict[str, Any]] = []

    if easyocr_available():
        try:
            reader = get_easyocr_reader(gpu=gpu)
            if summary_image is not None:
                summary_debug = make_summary_ocr_debug(summary_image, reader, scale=ocr_scale)
            if attack_image is not None:
                try:
                    attack_debug = make_attack_ocr_debug(attack_image, reader, row_count=row_count, scale=ocr_scale)
                except TypeError:
                    attack_debug = make_attack_ocr_debug(attack_image, reader, row_count=row_count)
            ocr_debug_zip = _save_ocr_debug_zip(kind, title, summary_image, attack_image, summary_debug, attack_debug)
        except Exception as e:
            context["errors"].append({"stage": "ocr_debug", "error": repr(e)})

        try:
            perf_report, perf_zip, perf_parsed = _run_ocr_perf_timeline_debug(
                kind,
                title,
                summary_image,
                attack_image,
                gpu=gpu,
                row_count=row_count,
                ocr_scale=ocr_scale,
                name_match_threshold=name_threshold,
                icon_match_threshold=icon_threshold,
            )
        except Exception as e:
            context["errors"].append({"stage": "perf_timeline", "error": repr(e)})
    else:
        context["errors"].append({"stage": "reader", "error": "OCR engine unavailable"})

    out_zip = EXPORT_DIR / f"real_integrated_debug_{ts}.zip"
    with zipfile.ZipFile(out_zip, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("README.txt", (
            "LOA v107 통합 디버그 ZIP\n"
            "- context/: API 클래스, 후보 목록, 세션, 최종 표\n"
            "- ocr_debug/: 종합/공격정보 crop 원본과 OCR 원문\n"
            "- perf/: timestamp 성능 로그와 아이콘 매칭 후 표\n"
            "아이콘 문제는 context/icon_match_rows.csv와 ocr_debug/icons/row_XX_core.png를 먼저 확인하세요.\n"
        ))
        zf.writestr("context/context.json", json.dumps(context, ensure_ascii=False, indent=2, default=str))
        try:
            zf.writestr("context/session_state_safe.json", json.dumps(_safe_session_json_v99(), ensure_ascii=False, indent=2, default=str))
        except Exception as e:
            zf.writestr("context/session_state_error.txt", repr(e))
        if candidates:
            zf.writestr("context/candidates_used.csv", pd.DataFrame(candidates).to_csv(index=False).encode("utf-8-sig"))
        try:
            current_table = st.session_state.get("real_table")
            if isinstance(current_table, pd.DataFrame) and not current_table.empty:
                zf.writestr("context/current_real_table.csv", current_table.to_csv(index=False).encode("utf-8-sig"))
        except Exception:
            pass
        try:
            zf.writestr("context/current_real_meta.json", json.dumps(st.session_state.get("real_meta", {}), ensure_ascii=False, indent=2, default=str))
        except Exception:
            pass
        if perf_report:
            zf.writestr("context/perf_report_copy.json", json.dumps(perf_report, ensure_ascii=False, indent=2, default=str))
        if isinstance(perf_parsed, pd.DataFrame) and not perf_parsed.empty:
            zf.writestr("context/parsed_attack_after_icon_match_copy.csv", perf_parsed.to_csv(index=False).encode("utf-8-sig"))
            icon_df = _v106_icon_debug_columns(perf_parsed)
            if not icon_df.empty:
                zf.writestr("context/icon_match_rows.csv", icon_df.to_csv(index=False).encode("utf-8-sig"))
        if summary_debug:
            rows = [{k: str(v) for k, v in item.items() if k not in {"raw_crop", "processed_crop"}} for item in summary_debug]
            zf.writestr("context/summary_debug_rows.csv", pd.DataFrame(rows).to_csv(index=False).encode("utf-8-sig"))
        if attack_debug:
            rows = []
            for item in attack_debug:
                d = {k: str(v) for k, v in item.items() if k not in {"row_crop", "icon_crop", "icon_core_crop", "icon_candidate_images"}}
                rows.append(d)
            zf.writestr("context/attack_debug_rows.csv", pd.DataFrame(rows).to_csv(index=False).encode("utf-8-sig"))
        if ocr_debug_zip and Path(ocr_debug_zip).exists():
            _zip_copy_with_prefix_v99(zf, Path(ocr_debug_zip), "ocr_debug")
        if perf_zip and Path(perf_zip).exists():
            _zip_copy_with_prefix_v99(zf, Path(perf_zip), "perf")
    return out_zip, context


def _safe_class_job_slug_v149() -> tuple[str, str]:
    """현재 API 프로필의 클래스/직업을 파일명용 문자열로 반환합니다."""
    prof = (st.session_state.get("api_summary") or {}).get("profile_summary", {}) if isinstance(st.session_state.get("api_summary"), dict) else {}
    cls = str((prof or {}).get("클래스") or "").strip() or "클래스미상"
    job = str((prof or {}).get("직업") or "").strip() or "직업미상"
    def _slug(s: str) -> str:
        return re.sub(r"[^0-9A-Za-z가-힣]+", "", s) or "미상"
    return _slug(cls), _slug(job)


def _make_icon_match_debug_zip_v149(attack_image: Any) -> Path | None:
    """아이콘 매칭 전용 디버그 ZIP.

    - 파일명에 클래스명_직업을 붙입니다.
    - 표(이름/아이콘 매칭명/점수/매칭 방식/상위후보)는 전투분석기 위→아래 순서(_ocr_row_index)로 정렬합니다.
    - 행별 아이콘 crop과 원본 공격정보 이미지를 함께 담아 오매칭 원인을 눈으로 볼 수 있게 합니다.
    """
    if attack_image is None:
        return None
    cls_slug, job_slug = _safe_class_job_slug_v149()
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    # 1) 아이콘 보정까지 끝난 결과표
    parsed = None
    try:
        _rep, _zip, parsed = _run_ocr_perf_timeline_debug(
            "icon_match", "아이콘 매칭 디버그", None, attack_image,
            gpu=False, row_count=14, ocr_scale=7,
        )
    except Exception:
        parsed = None
    # 2) 행별 아이콘 crop(디버그)
    debug_rows = []
    try:
        reader = get_easyocr_reader(gpu=False)
        debug_rows = make_attack_ocr_debug(attack_image, reader, row_count=14, scale=7) or []
    except Exception:
        debug_rows = []

    path = EXPORT_DIR / f"{cls_slug}_{job_slug}_아이콘매칭디버그_{ts}.zip"
    with zipfile.ZipFile(path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("context.json", json.dumps({
            "클래스": cls_slug, "직업": job_slug,
            "created_at": datetime.now().isoformat(timespec="seconds"),
            "row_count": 14,
            "note": "아이콘 매칭 전용 디버그. icon_match_table.csv를 위→아래 순서로 확인하세요.",
        }, ensure_ascii=False, indent=2))
        if attack_image is not None:
            try:
                zf.writestr("original_attack.png", _png_bytes(attack_image))
            except Exception:
                pass
        # 결과표 CSV (위→아래 순서 정렬)
        if isinstance(parsed, pd.DataFrame) and not parsed.empty:
            df = parsed.copy()
            if "_ocr_row_index" in df.columns:
                df["_sort_idx"] = pd.to_numeric(df["_ocr_row_index"], errors="coerce")
                df = df.sort_values("_sort_idx", na_position="last").drop(columns=["_sort_idx"])
            pref = [
                "_ocr_row_index", "이름", "OCR 원본 이름", "_icon_match_name", "_icon_match_score",
                "_name_match_reason", "_icon_match_source", "_icon_match_top3",
                "피해량", "치명타 적중률", "사용 횟수",
            ]
            cols = [c for c in pref if c in df.columns] or list(df.columns)
            out_df = df[cols].rename(columns={
                "_ocr_row_index": "순번(위→아래)",
                "_icon_match_name": "아이콘 매칭명",
                "_icon_match_score": "아이콘 점수",
                "_name_match_reason": "매칭 방식",
                "_icon_match_source": "아이콘 소스",
                "_icon_match_top3": "상위 후보(top3)",
            })
            zf.writestr("icon_match_table.csv", out_df.to_csv(index=False).encode("utf-8-sig"))
        # 행별 아이콘 crop (위→아래 순서 = row_index 순)
        for row in sorted(debug_rows, key=lambda r: int(r.get("row_index", 0) or 0)):
            ri = int(row.get("row_index", 0) or 0)
            if row.get("icon_crop") is not None:
                try:
                    zf.writestr(f"icons/row_{ri:02d}_icon.png", _png_bytes(row["icon_crop"]))
                except Exception:
                    pass
            if row.get("row_crop") is not None:
                try:
                    zf.writestr(f"rows/row_{ri:02d}_full.png", _png_bytes(row["row_crop"]))
                except Exception:
                    pass
        zf.writestr("README.txt", (
            "아이콘 매칭 전용 디버그 ZIP\n"
            f"클래스/직업: {cls_slug} / {job_slug}\n\n"
            "icon_match_table.csv : 전투분석기 위→아래 순서로 정렬된 아이콘 매칭 결과\n"
            "  · 이름=최종 매칭 스킬명, 아이콘 매칭명/점수=아이콘 인식 결과, 상위 후보(top3)=경합 후보\n"
            "icons/row_XX_icon.png : 각 행에서 실제 잘라 매칭에 쓴 아이콘 crop (row_XX = 위에서부터 순서)\n"
            "rows/row_XX_full.png  : 각 행 전체 crop\n"
            "original_attack.png   : 업로드한 공격정보 원본\n"
        ).encode("utf-8"))

    # 웹페이지에서 다운로드 없이 바로 확인할 수 있도록 인라인 표시용 데이터를 세션에 저장합니다.
    try:
        inline_tbl = None
        match_by_idx: dict[int, dict[str, Any]] = {}
        if isinstance(parsed, pd.DataFrame) and not parsed.empty:
            df2 = parsed.copy()
            if "_ocr_row_index" in df2.columns:
                df2["_sort_idx"] = pd.to_numeric(df2["_ocr_row_index"], errors="coerce")
                df2 = df2.sort_values("_sort_idx", na_position="last").drop(columns=["_sort_idx"])
            pref2 = ["_ocr_row_index", "이름", "_icon_match_name", "_icon_match_score",
                     "_name_match_reason", "_icon_match_top3", "피해량", "사용 횟수"]
            cols2 = [c for c in pref2 if c in df2.columns] or list(df2.columns)
            inline_tbl = df2[cols2].rename(columns={
                "_ocr_row_index": "순번", "_icon_match_name": "아이콘 매칭명",
                "_icon_match_score": "점수", "_name_match_reason": "매칭 방식",
                "_icon_match_top3": "상위 후보(top3)",
            })
            for _, pr in df2.iterrows():
                try:
                    ridx = int(pd.to_numeric(pr.get("_ocr_row_index"), errors="coerce"))
                except Exception:
                    continue
                match_by_idx[ridx] = {
                    "ocr": str(pr.get("이름") or ""),
                    "matched": str(pr.get("_icon_match_name") or ""),
                    "score": pr.get("_icon_match_score"),
                    "top3": str(pr.get("_icon_match_top3") or ""),
                    "casts": pr.get("사용 횟수"),
                }
        inline_rows = []
        for row in sorted(debug_rows, key=lambda r: int(r.get("row_index", 0) or 0)):
            ri = int(row.get("row_index", 0) or 0)
            m = match_by_idx.get(ri, {})
            inline_rows.append({
                "idx": ri, "crop": row.get("icon_crop"),
                "ocr": m.get("ocr", ""), "matched": m.get("matched", ""),
                "score": m.get("score"), "top3": m.get("top3", ""), "casts": m.get("casts"),
            })
        st.session_state["icon_match_inline_v150"] = {
            "class": cls_slug, "job": job_slug, "ts": ts,
            "table_df": inline_tbl, "rows": inline_rows,
        }
    except Exception:
        pass
    return path


def _render_icon_match_inline_result_v150() -> None:
    """아이콘 매칭 디버그 결과를 다운로드 없이 웹페이지에서 바로 보여줍니다.

    - 위→아래 순서 결과표 + 행별 아이콘 crop 썸네일 + 매칭명/점수/상위후보.
    - crop 그림과 '아이콘 매칭명'이 다르면 오매칭이므로 눈으로 바로 확인됩니다.
    """
    data = st.session_state.get("icon_match_inline_v150")
    if not data:
        return
    rows = data.get("rows") or []
    tbl = data.get("table_df")
    if (tbl is None or (isinstance(tbl, pd.DataFrame) and tbl.empty)) and not rows:
        return
    with st.expander(f"🎯 아이콘 매칭 결과 (바로 확인) · {data.get('class','')} {data.get('job','')}", expanded=True):
        cc1, cc2 = st.columns([4, 1])
        with cc1:
            st.caption("아래 crop 그림과 ‘아이콘 매칭명’이 다르면 오매칭입니다. 표는 전투분석기 위→아래 순서예요.")
        with cc2:
            if st.button("결과 지우기", key="icon_match_inline_clear_v150", use_container_width=True):
                st.session_state.pop("icon_match_inline_v150", None)
                st.rerun()
        if isinstance(tbl, pd.DataFrame) and not tbl.empty:
            st.dataframe(tbl, use_container_width=True, hide_index=True)
        if rows:
            st.markdown("**행별 아이콘 crop → 매칭 결과**")
            ncol = 6
            grid = st.columns(ncol)
            for i, r in enumerate(rows):
                with grid[i % ncol]:
                    if r.get("crop") is not None:
                        try:
                            st.image(r["crop"], use_container_width=True)
                        except Exception:
                            pass
                    _sc = r.get("score")
                    _sc_txt = f"{float(_sc):.1f}" if isinstance(_sc, (int, float)) else (str(_sc) if _sc not in (None, "") else "-")
                    _matched = r.get("matched") or "-"
                    _cast = r.get("casts")
                    _cast_txt = f" · {int(_cast)}회" if isinstance(_cast, (int, float)) and not pd.isna(_cast) else ""
                    st.markdown(f"<div style='font-size:0.78rem;line-height:1.25'>"
                                f"<b>row {r.get('idx')}</b>{_cast_txt}<br>→ <b>{_html_escape(_matched)}</b> ({_sc_txt})</div>",
                                unsafe_allow_html=True)


def _render_icon_match_debug_button_v149() -> None:
    """독립 실행 아이콘 매칭 디버그. 자체 업로더가 있어 캐릭터 검색/화면 공유 없이 바로 됩니다."""
    st.markdown("**🎯 아이콘 매칭 디버그 (독립 실행)**")
    st.caption("공격정보 이미지만 올리면 캐릭터 검색·화면 공유 없이 바로 아이콘 매칭 디버그 ZIP을 만들어요. "
               "파일명에 클래스·직업이 붙고, 표는 전투분석기 위→아래 순서입니다.")
    up = st.file_uploader(
        "공격정보 이미지 업로드 (백/헤드/치명 열이 보이는 화면)",
        type=["png", "jpg", "jpeg", "webp"],
        key="icon_match_debug_upload_v149",
    )
    img = None
    if up is not None:
        try:
            img = Image.open(io.BytesIO(up.getvalue())).convert("RGB")
        except Exception:
            img = None
    if img is None:
        # 업로드가 없으면, 화면 공유로 저장해 둔 실전 공격 이미지를 폴백으로 사용합니다.
        img = _get_uploaded_image_from_session_v88("real", "attack")

    if st.button("아이콘 매칭 디버그 ZIP 생성", key="sidebar_icon_match_debug_v149",
                 use_container_width=True, disabled=(img is None)):
        with st.spinner("아이콘 매칭 디버그 ZIP 생성 중..."):
            try:
                zp = _make_icon_match_debug_zip_v149(img)
            except Exception as e:  # noqa: BLE001
                zp = None
                st.error(f"아이콘 매칭 디버그 생성 중 오류: {e}")
        if zp is not None:
            st.session_state["icon_match_debug_zip_v149"] = str(zp)
            st.success("아이콘 매칭 디버그 ZIP을 생성했습니다.")
    if img is None:
        st.caption("↑ 이미지를 올리면 생성 버튼이 활성화됩니다. (캐릭터를 불러온 상태면 그 직업 기준으로 매칭까지 확인돼요.)")
    _zp = st.session_state.get("icon_match_debug_zip_v149")
    if _zp and Path(_zp).exists():
        st.download_button(
            "아이콘 매칭 디버그 ZIP 다운로드",
            data=Path(_zp).read_bytes(),
            file_name=Path(_zp).name,
            mime="application/zip",
            key="sidebar_icon_match_debug_download_v149",
            use_container_width=True,
        )


def _render_integrated_debug_button_v106() -> None:
    st.markdown("**🧪 통합 디버그**")
    if st.button("실전 통합 디버그 ZIP 생성", key="sidebar_real_integrated_debug_v106", use_container_width=True):
        summary_image = _get_uploaded_image_from_session_v88("real", "summary")
        attack_image = _get_uploaded_image_from_session_v88("real", "attack")
        if summary_image is None and attack_image is None:
            st.warning("먼저 실전 전투분석기 종합정보/공격정보 이미지를 업로드하고 OCR을 한 번 실행해 주세요.")
        else:
            with st.spinner("통합 디버그 ZIP 생성 중..."):
                zpath, report = _make_real_integrated_debug_zip_v106(summary_image, attack_image)
            st.session_state["real_integrated_debug_zip_v106"] = str(zpath)
            st.session_state["real_integrated_debug_report_v106"] = report
            st.success("통합 디버그 ZIP을 생성했습니다.")
    zpath = st.session_state.get("real_integrated_debug_zip_v106")
    if zpath and Path(zpath).exists():
        st.download_button(
            "실전 통합 디버그 ZIP 다운로드",
            data=Path(zpath).read_bytes(),
            file_name=Path(zpath).name,
            mime="application/zip",
            key="sidebar_real_integrated_debug_download_v106",
            use_container_width=True,
        )
        report = st.session_state.get("real_integrated_debug_report_v106") or {}
        if report:
            st.caption(
                f"후보 {report.get('candidate_count_actual', report.get('candidate_count_session', '-'))}개 · "
                f"클래스 {report.get('candidate_class_after_build') or report.get('candidate_class') or '-'}"
            )
    # 통합 디버그 바로 아래에 아이콘 매칭 전용 디버그 버튼을 붙입니다.
    # 기본적으로 숨김. 신규 캐릭터 작업 시 '🔧 디버그 도구 표시'를 켜면 다시 나타납니다.
    if st.session_state.get("show_debug_tools", False):
        st.divider()
        _render_icon_match_debug_button_v149()


def sidebar_controls() -> None:  # type: ignore[override]
    """v106: 이전 버전의 초록/파랑 누적 안내를 모두 제거한 깨끗한 사이드바."""
    calc_config = load_yaml(str(CONFIG_DIR / "calculation_presets.yaml"))
    defaults = calc_config.get("defaults", {})
    st.session_state.setdefault("api_crit_basis", "백어택 기준(조건부 포함)")
    st.session_state.setdefault("target_crit_percent", float(defaults.get("target_crit_percent", 100.0)))
    st.session_state.setdefault("target_back_percent", float(defaults.get("target_back_attack_percent", 100.0)))
    st.session_state.setdefault("back_bonus_percent", float(defaults.get("back_attack_damage_bonus_percent", 5.0)))

    with st.sidebar:
        st.markdown("### LOA 실전 분석")
        st.caption("정확도 우선 · 고정 행 + 셀별 OCR 재검수 + 통합 디버그")
        st.caption("캐릭터를 불러오고, 목표 상승률만 정한 뒤 전투분석기 이미지를 넣으면 됩니다.")

        st.session_state["show_debug_tools"] = st.checkbox(
            "🔧 디버그 도구 표시",
            value=bool(st.session_state.get("show_debug_tools", False)),
            help="본문의 OCR 원문/crop 디버그를 표시합니다. 통합 디버그 ZIP은 아래 버튼에서 생성합니다.",
        )

        token = st.text_input("Lost Ark API Key", type="password", key="api_token")
        character_name = st.text_input("캐릭터명", value=st.session_state.get("character_name", ""), key="character_name")
        st.number_input(
            "목표 상승률(%)",
            value=float(st.session_state.get("target_gain_percent", defaults.get("target_gain_percent", 5.0))),
            min_value=0.0,
            max_value=100.0,
            step=0.5,
            key="target_gain_percent",
        )
        fetch = st.button("검색", type="primary", use_container_width=True, key="sidebar_fetch_api_v106")

        _render_integrated_debug_button_v106()

        if fetch:
            if not token or not character_name:
                st.warning("API Key와 캐릭터명을 먼저 입력해줘.")
            else:
                with st.spinner("캐릭터 셋팅을 불러오는 중..."):
                    client = LostArkApiClient(token, CONFIG_DIR / "api_endpoints.yaml")
                    bundle = client.fetch_armory_bundle(character_name)
                    st.session_state.api_bundle = serializable_bundle(bundle)
                    st.session_state.api_summary = enrich_summary_with_identity(summarize_all(bundle), st.session_state.api_bundle)
                failed = {k: v for k, v in st.session_state.api_bundle.items() if not v.get("ok")}
                if failed:
                    st.warning(f"일부 API 호출 실패: {', '.join(failed.keys())}")
                else:
                    st.success("캐릭터 셋팅을 불러왔어.")

        summary = st.session_state.get("api_summary") or {}
        profile = summary.get("profile_summary", {}) if isinstance(summary, dict) else {}
        if profile:
            st.divider()
            st.caption("현재 분석 캐릭터")
            st.markdown(f"**{profile.get('캐릭터명') or character_name}**")
            meta = " · ".join(str(x) for x in [profile.get("서버"), profile.get("클래스"), profile.get("직업"), profile.get("아이템레벨")] if x)
            if meta:
                st.caption(meta)




# ==============================================================================
# v110 sidebar marker
# ==============================================================================
try:
    _sidebar_controls_prev_v110 = sidebar_controls
    def sidebar_controls():  # type: ignore[override]
        pass  # 버전 변경 안내 제거
        return _sidebar_controls_prev_v110()
except Exception:
    pass



# ==============================================================================
# v112: low-confidence icon rows -> 기타 추가딜 + clean marker
# ==============================================================================
APP_CALC_VERSION = "v113_fast_icon_api_debug"

try:
    _repair_battle_values_from_summary_prev_v112 = _repair_battle_values_from_summary
except Exception:
    _repair_battle_values_from_summary_prev_v112 = None


def _is_low_conf_icon_row_v112(row: Any) -> bool:
    try:
        name = str(row.get("이름") or "").strip()
        score = row.get("_name_match_icon_score")
        if score is None or str(score).strip() == "":
            score = row.get("_icon_match_score")
        try:
            score_f = float(score)
        except Exception:
            score_f = 0.0
        source = str(row.get("_icon_match_source") or "")
        reason = str(row.get("_name_match_reason") or "")
        match_name = str(row.get("_icon_match_name") or "")
        # 진짜 파란 기타 아이콘은 local_common + 높은 점수면 유지합니다.
        if name == "기타" and source == "local_common" and score_f >= 90:
            return False
        # 낮은 점수/low_conf fallback/스킬 후보로 억지 매칭된 기타는 기타추가딜로 보냅니다.
        if score_f and score_f < 70:
            return True
        if "icon_low_conf_etc" in reason or "low_conf" in reason:
            return True
        if name in {"기타", "인식 안된 기타 추가딜"} and source != "local_common":
            return True
        if name == "기타" and match_name and match_name != "기타" and score_f < 90:
            return True
    except Exception:
        return False
    return False


def _aggregate_low_conf_icon_rows_v112(df: pd.DataFrame, meta: dict[str, Any] | None = None) -> pd.DataFrame:
    if df is None or not isinstance(df, pd.DataFrame) or df.empty or "이름" not in df.columns:
        return df if isinstance(df, pd.DataFrame) else pd.DataFrame()
    meta = meta or {}
    try:
        elapsed = float(meta.get("elapsed_seconds") or 0)
    except Exception:
        elapsed = 0.0
    try:
        total_damage = float(meta.get("total_damage") or 0)
    except Exception:
        total_damage = 0.0

    keep_rows = []
    extra_damage = 0.0
    extra_names = []
    for _, row in df.iterrows():
        name = str(row.get("이름") or "").strip()
        dmg = parse_korean_number(row.get("피해량")) if "피해량" in df.columns else None
        if name == "기타 추가딜":
            if dmg and dmg > 0:
                extra_damage += float(dmg)
            continue
        if _is_low_conf_icon_row_v112(row) and dmg and dmg > 0:
            extra_damage += float(dmg)
            label = str(row.get("_icon_match_name") or name or "미확정")
            if label not in extra_names:
                extra_names.append(label)
            continue
        keep_rows.append(row)

    out = pd.DataFrame(keep_rows).reset_index(drop=True) if keep_rows else pd.DataFrame(columns=df.columns)
    if extra_damage > 0:
        extra = {c: "" for c in df.columns}
        extra["이름"] = "기타 추가딜"
        extra["피해량"] = format_korean_number(extra_damage)
        if elapsed > 0 and "초당 피해량" in df.columns:
            extra["초당 피해량"] = format_korean_number(extra_damage / elapsed)
        if total_damage > 0 and "피해량 지분" in df.columns:
            extra["피해량 지분"] = round(extra_damage / total_damage * 100.0, 2)
        extra["_low_conf_aggregated_v112"] = ", ".join(extra_names[:12])
        out = pd.concat([out, pd.DataFrame([extra])], ignore_index=True)
    cols = [c for c in df.columns if c in out.columns]
    rest = [c for c in out.columns if c not in cols]
    return out[cols + rest].reset_index(drop=True)


def _repair_battle_values_from_summary(df: pd.DataFrame, meta: dict[str, Any] | None, *, replace_damage_from_share: bool = False, elapsed_fallback: float | None = None) -> pd.DataFrame:  # type: ignore[override]
    if callable(_repair_battle_values_from_summary_prev_v112):
        out = _repair_battle_values_from_summary_prev_v112(df, meta, replace_damage_from_share=replace_damage_from_share, elapsed_fallback=elapsed_fallback)
    else:
        out = df
    return _aggregate_low_conf_icon_rows_v112(out, meta or {})

try:
    _sidebar_controls_prev_v112 = sidebar_controls
    def sidebar_controls():  # type: ignore[override]
        pass  # 버전 변경 안내 제거
        return _sidebar_controls_prev_v112()
except Exception:
    pass



# ==============================================================================
# v113: speed tuning + API timing/debug + cleaner sidebar
# ==============================================================================
APP_CALC_VERSION = "v113_fast_icon_api_debug"

import hashlib as _hashlib_v113
import time as _time_v113


def _api_cache_key_v113(token: str, character_name: str) -> str:
    h = _hashlib_v113.sha1((token or '').encode('utf-8')).hexdigest()[:10]
    return f"{h}:{str(character_name or '').strip().lower()}"


def _fetch_api_bundle_timed_v113(token: str, character_name: str) -> tuple[dict[str, Any], dict[str, Any]]:
    """API 호출/요약 계산 시간을 측정하고 같은 캐릭터 재검색은 세션 캐시를 사용합니다."""
    cache = st.session_state.setdefault('api_cache_v113', {})
    key = _api_cache_key_v113(token, character_name)
    t0 = _time_v113.perf_counter()
    timing: dict[str, Any] = {'cache_key': key, 'character_name': character_name, 'cache_hit': False}
    if key in cache:
        cached = cache[key]
        timing['cache_hit'] = True
        timing['total_ms'] = round((_time_v113.perf_counter() - t0) * 1000.0, 3)
        return cached, timing

    t = _time_v113.perf_counter()
    client = LostArkApiClient(token, CONFIG_DIR / 'api_endpoints.yaml')
    timing['client_init_ms'] = round((_time_v113.perf_counter() - t) * 1000.0, 3)

    t = _time_v113.perf_counter()
    bundle = client.fetch_armory_bundle(character_name)
    timing['fetch_armory_bundle_ms'] = round((_time_v113.perf_counter() - t) * 1000.0, 3)

    t = _time_v113.perf_counter()
    serial = serializable_bundle(bundle)
    timing['serializable_bundle_ms'] = round((_time_v113.perf_counter() - t) * 1000.0, 3)

    t = _time_v113.perf_counter()
    summary0 = summarize_all(bundle)
    timing['summarize_all_ms'] = round((_time_v113.perf_counter() - t) * 1000.0, 3)

    t = _time_v113.perf_counter()
    summary = enrich_summary_with_identity(summary0, serial)
    if isinstance(summary, dict):
        summary["_calc_version"] = APP_CALC_VERSION
    timing['enrich_summary_with_identity_ms'] = round((_time_v113.perf_counter() - t) * 1000.0, 3)

    failed = {k: {'status_code': getattr(v, 'status_code', None), 'error': getattr(v, 'error', '')} for k, v in bundle.items() if not getattr(v, 'ok', False)}
    timing['failed'] = failed
    timing['endpoint_count'] = len(bundle)
    timing['total_ms'] = round((_time_v113.perf_counter() - t0) * 1000.0, 3)
    result = {'api_bundle': serial, 'api_summary': summary}
    cache[key] = result
    return result, timing


try:
    _make_real_integrated_debug_zip_prev_v113 = _make_real_integrated_debug_zip_v106
    def _make_real_integrated_debug_zip_v106(summary_image: Any, attack_image: Any) -> tuple[Path, dict[str, Any]]:  # type: ignore[override]
        zpath, context = _make_real_integrated_debug_zip_prev_v113(summary_image, attack_image)
        try:
            api_timing = st.session_state.get('api_last_timing_v113') or {}
            if api_timing:
                import tempfile, zipfile, shutil
                tmp = Path(str(zpath) + '.tmp')
                with zipfile.ZipFile(zpath, 'r') as zin, zipfile.ZipFile(tmp, 'w', compression=zipfile.ZIP_DEFLATED) as zout:
                    for item in zin.infolist():
                        zout.writestr(item, zin.read(item.filename))
                    zout.writestr('context/api_timing_v113.json', json.dumps(api_timing, ensure_ascii=False, indent=2, default=str))
                tmp.replace(zpath)
                context['api_timing_v113'] = api_timing
        except Exception as e:
            try:
                context.setdefault('errors', []).append({'stage': 'api_timing_append_v113', 'error': repr(e)})
            except Exception:
                pass
        return zpath, context
except Exception:
    pass


def sidebar_controls() -> None:  # type: ignore[override]
    """v113: 사이드바 정리 + API 캐시/타이밍."""
    calc_config = load_yaml(str(CONFIG_DIR / 'calculation_presets.yaml'))
    defaults = calc_config.get('defaults', {})
    st.session_state.setdefault('api_crit_basis', '백어택 기준(조건부 포함)')
    st.session_state.setdefault('target_crit_percent', float(defaults.get('target_crit_percent', 100.0)))
    st.session_state.setdefault('target_back_percent', float(defaults.get('target_back_attack_percent', 100.0)))
    st.session_state.setdefault('back_bonus_percent', float(defaults.get('back_attack_damage_bonus_percent', 5.0)))

    with st.sidebar:
        st.markdown('### LOA 실전 분석')
        st.caption('전투분석기 이미지에서 값을 자동으로 읽어옵니다.')
        st.caption('캐릭터를 불러오고, 목표 상승률만 정한 뒤 전투분석기 이미지를 넣으면 됩니다.')
        st.session_state['show_debug_tools'] = st.checkbox(
            '🔧 디버그 도구 표시',
            value=bool(st.session_state.get('show_debug_tools', False)),
            help='본문의 OCR 원문/crop 디버그를 표시합니다. 통합 디버그 ZIP은 아래 버튼에서 생성합니다.',
        )
        token = st.text_input('Lost Ark API Key', type='password', key='api_token')
        character_name = st.text_input('캐릭터명', value=st.session_state.get('character_name', ''), key='character_name')
        st.number_input(
            '목표 상승률(%)',
            value=float(st.session_state.get('target_gain_percent', defaults.get('target_gain_percent', 5.0))),
            min_value=0.0,
            max_value=100.0,
            step=0.5,
            key='target_gain_percent',
        )
        fetch = st.button('검색', type='primary', use_container_width=True, key='sidebar_fetch_api_v113')
        _render_integrated_debug_button_v106()
        if fetch:
            if not token or not character_name:
                st.warning('API Key와 캐릭터명을 먼저 입력해줘.')
            else:
                with st.spinner('캐릭터 셋팅을 불러오는 중...'):
                    result, timing = _fetch_api_bundle_timed_v113(token, character_name)
                    st.session_state.api_bundle = result['api_bundle']
                    st.session_state.api_summary = result['api_summary']
                    st.session_state.api_last_timing_v113 = timing
                failed = timing.get('failed') or {}
                if failed:
                    st.warning(f"일부 API 호출 실패: {', '.join(failed.keys())}")
                else:
                    msg = '캐릭터 셋팅을 불러왔어.'
                    if timing.get('cache_hit'):
                        msg += ' (캐시 사용)'
                    st.success(msg)
                st.caption(f"API/계산 시간: {float(timing.get('total_ms') or 0):,.0f} ms · fetch {float(timing.get('fetch_armory_bundle_ms') or 0):,.0f} ms · 요약 {float(timing.get('summarize_all_ms') or 0):,.0f} ms")

        summary = st.session_state.get('api_summary') or {}
        profile = summary.get('profile_summary', {}) if isinstance(summary, dict) else {}
        if profile:
            st.divider()
            st.caption('현재 분석 캐릭터')
            st.markdown(f"**{profile.get('캐릭터명') or character_name}**")
            meta = ' · '.join(str(x) for x in [profile.get('서버'), profile.get('클래스'), profile.get('직업'), profile.get('아이템레벨')] if x)
            if meta:
                st.caption(meta)




# v111 sidebar marker
try:
    _sidebar_controls_prev_v111 = sidebar_controls
    def sidebar_controls():  # type: ignore[override]
        st.sidebar.success('v111 적용됨 · 아이콘 y감지 그리드 + summary raw 전처리 완화')
        return _sidebar_controls_prev_v111()
except Exception:
    pass


# ==============================================================================
# v114: speed mode + persistent API/recalculation cache + clean final sidebar
# ==============================================================================
APP_CALC_VERSION = "v114_speed_api_cache"

import os as _os_v114
import json as _json_v114
import time as _time_v114
import hashlib as _hashlib_v114
from pathlib import Path as _Path_v114


def _cache_dir_v114() -> _Path_v114:
    p = _Path_v114("cache") / "api_v114"
    try:
        p.mkdir(parents=True, exist_ok=True)
    except Exception:
        pass
    return p


def _api_cache_key_v114(token: str, character_name: str) -> str:
    token_hash = _hashlib_v114.sha1((token or "").encode("utf-8")).hexdigest()[:12]
    name_key = re.sub(r"[^0-9A-Za-z가-힣_-]+", "_", str(character_name or "").strip().lower())[:80]
    return f"{token_hash}_{name_key}"


def _api_cache_path_v114(token: str, character_name: str) -> _Path_v114:
    return _cache_dir_v114() / f"{_api_cache_key_v114(token, character_name)}.json"


def _api_cache_ttl_sec_v114() -> int:
    try:
        return int(float(_os_v114.environ.get("LOA_API_CACHE_TTL_SEC", "3600")))
    except Exception:
        return 3600


def _read_api_cache_v114(token: str, character_name: str) -> tuple[dict[str, Any] | None, dict[str, Any]]:
    path = _api_cache_path_v114(token, character_name)
    t0 = _time_v114.perf_counter()
    info: dict[str, Any] = {"cache_path": str(path), "disk_cache_hit": False, "cache_read_ms": 0.0}
    try:
        if not path.exists():
            info["cache_read_ms"] = round((_time_v114.perf_counter() - t0) * 1000.0, 3)
            return None, info
        data = _json_v114.loads(path.read_text(encoding="utf-8"))
        created = float(data.get("created_at") or 0.0)
        age = _time_v114.time() - created
        info["cache_age_sec"] = round(age, 3)
        info["cache_ttl_sec"] = _api_cache_ttl_sec_v114()
        if age > _api_cache_ttl_sec_v114():
            info["cache_expired"] = True
            info["cache_read_ms"] = round((_time_v114.perf_counter() - t0) * 1000.0, 3)
            return None, info
        result = data.get("result")
        if not isinstance(result, dict) or "api_bundle" not in result or "api_summary" not in result:
            info["cache_invalid"] = True
            info["cache_read_ms"] = round((_time_v114.perf_counter() - t0) * 1000.0, 3)
            return None, info
        info["disk_cache_hit"] = True
        info["cache_read_ms"] = round((_time_v114.perf_counter() - t0) * 1000.0, 3)
        return result, info
    except Exception as e:
        info["cache_error"] = repr(e)
        info["cache_read_ms"] = round((_time_v114.perf_counter() - t0) * 1000.0, 3)
        return None, info


def _write_api_cache_v114(token: str, character_name: str, result: dict[str, Any]) -> dict[str, Any]:
    path = _api_cache_path_v114(token, character_name)
    t0 = _time_v114.perf_counter()
    info: dict[str, Any] = {"cache_path": str(path), "cache_write_ms": 0.0}
    try:
        payload = {"created_at": _time_v114.time(), "character_name": character_name, "result": result}
        tmp = path.with_suffix(".tmp")
        tmp.write_text(_json_v114.dumps(payload, ensure_ascii=False, default=str), encoding="utf-8")
        tmp.replace(path)
    except Exception as e:
        info["cache_write_error"] = repr(e)
    info["cache_write_ms"] = round((_time_v114.perf_counter() - t0) * 1000.0, 3)
    return info




try:
    _make_real_integrated_debug_zip_prev_v114 = _make_real_integrated_debug_zip_v106
    def _make_real_integrated_debug_zip_v106(summary_image: Any, attack_image: Any) -> tuple[Path, dict[str, Any]]:  # type: ignore[override]
        zpath, context = _make_real_integrated_debug_zip_prev_v114(summary_image, attack_image)
        try:
            import zipfile
            api_timing = st.session_state.get("api_last_timing_v114") or st.session_state.get("api_last_timing_v113") or {}
            speed_env = {
                "LOA_ATTACK_FAST_GRID": _os_v114.environ.get("LOA_ATTACK_FAST_GRID"),
                "LOA_ATTACK_ROW_PASS_SCALE": _os_v114.environ.get("LOA_ATTACK_ROW_PASS_SCALE"),
                "LOA_ATTACK_DAMAGE_RECHECK": _os_v114.environ.get("LOA_ATTACK_DAMAGE_RECHECK"),
                "LOA_ATTACK_CRITICAL_CELL_SCALE": _os_v114.environ.get("LOA_ATTACK_CRITICAL_CELL_SCALE"),
                "LOA_ATTACK_RECHECK": _os_v114.environ.get("LOA_ATTACK_RECHECK"),
                "LOA_FORCE_DAMAGE_CELL_OCR": _os_v114.environ.get("LOA_FORCE_DAMAGE_CELL_OCR"),
                "LOA_ICON_HEAVY_TOPK": _os_v114.environ.get("LOA_ICON_HEAVY_TOPK"),
                "LOA_ICON_VARIANT_RETRY": _os_v114.environ.get("LOA_ICON_VARIANT_RETRY"),
                "LOA_API_CACHE_TTL_SEC": _os_v114.environ.get("LOA_API_CACHE_TTL_SEC"),
            }
            tmp = Path(str(zpath) + ".tmp")
            with zipfile.ZipFile(zpath, "r") as zin, zipfile.ZipFile(tmp, "w", compression=zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    zout.writestr(item, zin.read(item.filename))
                zout.writestr("context/api_timing_v114.json", _json_v114.dumps(api_timing, ensure_ascii=False, indent=2, default=str))
                zout.writestr("context/speed_env_v114.json", _json_v114.dumps(speed_env, ensure_ascii=False, indent=2, default=str))
            tmp.replace(zpath)
            context["api_timing_v114"] = api_timing
            context["speed_env_v114"] = speed_env
        except Exception as e:
            try:
                context.setdefault("errors", []).append({"stage": "v114_debug_append", "error": repr(e)})
            except Exception:
                pass
        return zpath, context
except Exception:
    pass


def _clear_api_cache_v114(token: str = "", character_name: str = "") -> None:
    try:
        st.session_state["api_cache_v114"] = {}
        if token and character_name:
            p = _api_cache_path_v114(token, character_name)
            if p.exists():
                p.unlink()
    except Exception:
        pass


def sidebar_controls() -> None:  # type: ignore[override]
    """v114: 최종 사이드바. API/계산 캐시와 속도 디버그를 노출합니다."""
    calc_config = load_yaml(str(CONFIG_DIR / "calculation_presets.yaml"))
    defaults = calc_config.get("defaults", {})
    st.session_state.setdefault("api_crit_basis", "백어택 기준(조건부 포함)")
    st.session_state.setdefault("target_crit_percent", float(defaults.get("target_crit_percent", 100.0)))
    st.session_state.setdefault("target_back_percent", float(defaults.get("target_back_attack_percent", 100.0)))
    st.session_state.setdefault("back_bonus_percent", float(defaults.get("back_attack_damage_bonus_percent", 5.0)))

    with st.sidebar:
        st.markdown("### LOA 실전 분석")
        st.caption("속도 진단 · 통합 업로드 + API/계산 상세 타이밍")
        st.session_state["show_debug_tools"] = st.checkbox(
            "🔧 디버그 도구 표시",
            value=bool(st.session_state.get("show_debug_tools", False)),
            help="본문의 OCR 원문/crop 디버그를 표시합니다. 통합 디버그 ZIP은 검색 버튼 아래에서 생성합니다.",
        )
        token = st.text_input("Lost Ark API Key", type="password", key="api_token")
        character_name = st.text_input("캐릭터명", value=st.session_state.get("character_name", ""), key="character_name")
        st.number_input(
            "목표 상승률(%)",
            value=float(st.session_state.get("target_gain_percent", defaults.get("target_gain_percent", 5.0))),
            min_value=0.0,
            max_value=100.0,
            step=0.5,
            key="target_gain_percent",
        )
        force_refresh = st.checkbox("API 캐시 무시하고 새로고침", value=False, key="api_force_refresh_v114")
        fetch = st.button("검색", type="primary", use_container_width=True, key="sidebar_fetch_api_v114")
        _render_integrated_debug_button_v106()

        if fetch:
            if not token or not character_name:
                st.warning("API Key와 캐릭터명을 먼저 입력해줘.")
            else:
                with st.spinner("캐릭터 셋팅/계산표를 불러오는 중..."):
                    result, timing = _fetch_api_bundle_timed_v114(token, character_name, force_refresh=bool(force_refresh))
                    st.session_state.api_bundle = result["api_bundle"]
                    st.session_state.api_summary = result["api_summary"]
                    st.session_state.api_last_timing_v114 = timing
                failed = timing.get("failed") or {}
                if failed:
                    st.warning(f"일부 API 호출 실패: {', '.join(failed.keys())}")
                else:
                    if timing.get("session_cache_hit"):
                        st.success("캐릭터 셋팅을 불러왔어. (세션 캐시)")
                    elif timing.get("disk_cache_hit"):
                        st.success("캐릭터 셋팅을 불러왔어. (디스크 캐시)")
                    else:
                        st.success("캐릭터 셋팅을 불러왔어.")
                st.caption(
                    f"API/계산 시간: {float(timing.get('total_ms') or 0):,.0f} ms · "
                    f"fetch {float(timing.get('fetch_armory_bundle_ms') or 0):,.0f} ms · "
                    f"요약 {float(timing.get('summarize_all_ms') or 0):,.0f} ms · "
                    f"캐시읽기 {float(timing.get('cache_read_ms') or 0):,.0f} ms"
                )
        if st.button("현재 캐릭터 API 캐시 삭제", use_container_width=True, key="clear_api_cache_v114"):
            _clear_api_cache_v114(token, character_name)
            st.info("현재 캐릭터 API 캐시를 삭제했습니다.")

        summary = st.session_state.get("api_summary") or {}
        profile = summary.get("profile_summary", {}) if isinstance(summary, dict) else {}
        if profile:
            st.divider()
            st.caption("현재 분석 캐릭터")
            st.markdown(f"**{profile.get('캐릭터명') or character_name}**")
            meta = " · ".join(str(x) for x in [profile.get("서버"), profile.get("클래스"), profile.get("직업"), profile.get("아이템레벨")] if x)
            if meta:
                st.caption(meta)



# ==============================================================================
# v115: 통합 업로드 + API/계산 상세 타이밍 + 병목 진단 파일
# ==============================================================================
APP_CALC_VERSION = "v115_unified_speed_diagnostics"

import time as _time_v115
import json as _json_v115


def _summarize_all_detailed_v115(bundle: dict[str, Any]) -> tuple[dict[str, Any], dict[str, Any]]:
    """api_parser.summarize_all 내부를 단계별로 쪼개 시간을 기록합니다."""
    import modules.api_parser as ap
    timings: dict[str, Any] = {"version": "v115"}
    t_all = _time_v115.perf_counter()

    def timed(name: str, fn, *args, **kwargs):
        t = _time_v115.perf_counter()
        value = fn(*args, **kwargs)
        timings[f"{name}_ms"] = round((_time_v115.perf_counter() - t) * 1000.0, 3)
        return value

    try:
        profile_summary, stats = timed("parse_profile", ap.parse_profile, bundle)
    except Exception as e:
        profile_summary, stats = {}, pd.DataFrame()
        timings["parse_profile_error"] = repr(e)

    try:
        t = _time_v115.perf_counter()
        from modules.api_skill_estimator import estimate_skill_crit_tables
        crit_tables = estimate_skill_crit_tables(bundle)
        timings["estimate_skill_crit_tables_ms"] = round((_time_v115.perf_counter() - t) * 1000.0, 3)
        if isinstance(crit_tables, dict) and isinstance(crit_tables.get("_estimator_timing_v157"), dict):
            timings["estimate_skill_crit_tables_detail_v157"] = crit_tables.get("_estimator_timing_v157")
    except Exception as e:
        timings["estimate_skill_crit_tables_error"] = repr(e)
        crit_tables = {
            "skill_crit_estimates": pd.DataFrame(),
            "crit_sources": pd.DataFrame(),
            "damage_sources": pd.DataFrame(),
            "base_crit_percent": 0.0,
            "base_crit_raw": f"치명/치피/피해군 추정 실패: {e}",
            "global_static_crit_rate_percent": 0.0,
            "global_static_crit_damage_percent": 0.0,
            "avg_evolution_damage_percent": 0.0,
            "avg_final_multiplier": 1.0,
        }

    summary: dict[str, Any] = {
        "profile_summary": profile_summary,
        "stats": stats,
        "equipment": timed("parse_equipment", ap.parse_equipment, bundle),
        "skills": timed("parse_combat_skills", ap.parse_combat_skills, bundle),
        "engravings": timed("parse_engravings", ap.parse_engravings, bundle),
        "cards": timed("parse_cards", ap.parse_cards, bundle),
        "gems": timed("parse_gems", ap.parse_gems, bundle),
        "arkpassive": timed("parse_arkpassive", ap.parse_arkpassive, bundle),
        "arkgrid": timed("parse_arkgrid", ap.parse_arkgrid, bundle),
        **crit_tables,
    }
    timings["summarize_all_total_ms"] = round((_time_v115.perf_counter() - t_all) * 1000.0, 3)
    # 병목 상위 정렬용
    timing_rows = []
    for k, v in timings.items():
        if k.endswith("_ms"):
            try:
                timing_rows.append({"stage": k, "ms": float(v)})
            except Exception:
                pass
    timings["top_stages"] = sorted(timing_rows, key=lambda x: x["ms"], reverse=True)[:10]
    return summary, timings






def _run_ocr_for_kind_v115(kind: str, title: str, summary_image: Image.Image | None, attack_image: Image.Image | None, *, row_count: int, ocr_scale: int, icon_threshold: float, name_threshold: float, aggregate_unmatched: bool) -> dict[str, Any]:
    key = f"{kind}_table"
    meta_key = f"{kind}_meta"
    report: dict[str, Any] = {"kind": kind, "title": title, "summary": bool(summary_image), "attack": bool(attack_image)}
    t_all = _time_v115.perf_counter()
    reader = get_easyocr_reader(gpu=False)
    if summary_image is not None:
        t = _time_v115.perf_counter()
        meta = parse_summary_fixed_grid(summary_image, reader)
        st.session_state[meta_key] = meta
        _sync_meta_to_manual_inputs(kind, meta)
        report["summary_parse_ms"] = round((_time_v115.perf_counter() - t) * 1000.0, 3)
    if attack_image is not None:
        t = _time_v115.perf_counter()
        parsed = parse_attack_fixed_grid(attack_image, reader, row_count=int(row_count), scale=int(ocr_scale))
        report["attack_parse_ms"] = round((_time_v115.perf_counter() - t) * 1000.0, 3)
        report["attack_rows_before_icon"] = 0 if parsed is None else len(parsed)
        if parsed is not None and not parsed.empty:
            t = _time_v115.perf_counter()
            candidates_full = get_ocr_skill_candidates_full()
            skill_icon_candidates = _skill_only_ocr_candidates(candidates_full)
            report["candidate_load_ms"] = round((_time_v115.perf_counter() - t) * 1000.0, 3)
            report["candidate_count"] = len(skill_icon_candidates)
            if skill_icon_candidates:
                t = _time_v115.perf_counter()
                parsed = correct_battle_skill_names_with_icons(
                    parsed,
                    attack_image,
                    skill_icon_candidates,
                    threshold=float(name_threshold),
                    icon_threshold=float(icon_threshold),
                    drop_unmatched=False,
                )
                report["icon_match_ms"] = round((_time_v115.perf_counter() - t) * 1000.0, 3)
                if aggregate_unmatched:
                    parsed = _aggregate_unmatched_rows_as_extra(parsed)
            parsed = sanitize_battle_table(parsed)
            parsed = _repair_battle_values_from_summary(parsed, st.session_state.get(meta_key, {}), elapsed_fallback=st.session_state.get(f"{kind}_elapsed_sec"))
            parsed = _drop_obvious_ocr_noise_rows(parsed)
            st.session_state[key] = prepare_battle_editor_df(sanitize_battle_table(parsed))
            _bump_table_version(kind)
            report["attack_rows_after"] = len(st.session_state[key])
    report["total_ms"] = round((_time_v115.perf_counter() - t_all) * 1000.0, 3)
    st.session_state[f"{kind}_last_ocr_timing_v115"] = report
    return report




_ocr_tab_prev_v115 = ocr_tab

def ocr_tab() -> None:  # type: ignore[override]
    st.header("전투분석기 입력")
    st.caption("실전/허수 종합정보·공격정보 이미지를 한 화면에서 한 번에 업로드하고 OCR 실행합니다.")
    _render_step_guide(2, "전투분석기 이미지 넣고 검수하기", [
        "실전/허수 전투분석기 이미지를 한 번에 올리고 <b>통합 OCR 실행</b>을 누르세요.",
        "OCR 후 실전/허수 표를 각각 검수하세요.",
        "고쳤으면 <b>③ 실력 분석 결과</b> 탭으로 이동하세요.",
    ])

    with st.container(border=True):
        st.markdown("#### 통합 이미지 업로드")
        c1, c2 = st.columns(2)
        with c1:
            st.markdown("**실전 전투분석기**")
            real_summary_up = st.file_uploader("실전 종합정보 이미지", type=["png", "jpg", "jpeg", "webp"], key="unified_real_summary_v115")
            real_attack_up = st.file_uploader("실전 공격정보 이미지", type=["png", "jpg", "jpeg", "webp"], key="unified_real_attack_v115")
        with c2:
            st.markdown("**허수/기준 전투분석기**")
            dummy_summary_up = st.file_uploader("허수/기준 종합정보 이미지", type=["png", "jpg", "jpeg", "webp"], key="unified_dummy_summary_v115")
            dummy_attack_up = st.file_uploader("허수/기준 공격정보 이미지", type=["png", "jpg", "jpeg", "webp"], key="unified_dummy_attack_v115")

        real_summary_img = _image_from_uploader_or_session_v115(real_summary_up, "real", "summary")
        real_attack_img = _image_from_uploader_or_session_v115(real_attack_up, "real", "attack")
        dummy_summary_img = _image_from_uploader_or_session_v115(dummy_summary_up, "dummy", "summary")
        dummy_attack_img = _image_from_uploader_or_session_v115(dummy_attack_up, "dummy", "attack")

        with st.expander("통합 OCR 설정", expanded=False):
            row_count = st.number_input("공격정보 인식 행 수", min_value=5, max_value=24, value=int(st.session_state.get("unified_row_count_v115", 18) or 18), step=1, key="unified_row_count_v115")
            ocr_scale = st.number_input("OCR 확대 배율", min_value=2, max_value=10, value=int(st.session_state.get("unified_ocr_scale_v115", 7) or 7), step=1, key="unified_ocr_scale_v115")
            icon_match_threshold = st.slider("아이콘 보정 신뢰도", min_value=45, max_value=95, value=int(st.session_state.get("unified_icon_threshold_v115", 74) or 74), step=1, key="unified_icon_threshold_v115")
            name_match_threshold = st.slider("스킬명 보정 민감도", min_value=40, max_value=90, value=int(st.session_state.get("unified_name_threshold_v115", 52) or 52), step=1, key="unified_name_threshold_v115")
            aggregate_unmatched = st.checkbox("낮은 신뢰도 행은 기타 추가딜로 합산", value=bool(st.session_state.get("unified_aggregate_unmatched_v115", False)), key="unified_aggregate_unmatched_v115")

        if st.button("실전 + 허수/기준 통합 OCR 실행", type="primary", use_container_width=True, key="unified_run_ocr_v115"):
            if not easyocr_available():
                st.error("OCR 엔진이 설치되어 있지 않습니다.")
            elif not any([real_summary_img, real_attack_img, dummy_summary_img, dummy_attack_img]):
                st.warning("업로드된 이미지가 없습니다.")
            else:
                reports = []
                with st.spinner("실전/허수 전투분석기 OCR 실행 중..."):
                    if real_summary_img is not None or real_attack_img is not None:
                        reports.append(_run_ocr_for_kind_v115("real", "실전 전투분석기", real_summary_img, real_attack_img, row_count=int(row_count), ocr_scale=int(ocr_scale), icon_threshold=float(icon_match_threshold)/100.0, name_threshold=float(name_match_threshold)/100.0, aggregate_unmatched=bool(aggregate_unmatched)))
                    if dummy_summary_img is not None or dummy_attack_img is not None:
                        reports.append(_run_ocr_for_kind_v115("dummy", "허수/기준 전투분석기", dummy_summary_img, dummy_attack_img, row_count=int(row_count), ocr_scale=int(ocr_scale), icon_threshold=float(icon_match_threshold)/100.0, name_threshold=float(name_match_threshold)/100.0, aggregate_unmatched=bool(aggregate_unmatched)))
                st.session_state["unified_ocr_reports_v115"] = reports
                st.success("통합 OCR이 끝났습니다. 아래 실전/허수 표를 검수하세요.")

        reports = st.session_state.get("unified_ocr_reports_v115") or []
        if reports:
            rep_df = pd.DataFrame(reports)
            st.caption("최근 OCR 실행 시간")
            show_cols = [c for c in ["title", "summary_parse_ms", "attack_parse_ms", "candidate_load_ms", "icon_match_ms", "total_ms", "attack_rows_after"] if c in rep_df.columns]
            st.dataframe(rep_df[show_cols], use_container_width=True, hide_index=True)

    real_tab, dummy_tab = st.tabs(["실전 검수", "허수/기준 검수"])
    with real_tab:
        _render_battle_review_only_v115("real", "실전 전투분석기")
    with dummy_tab:
        _render_battle_review_only_v115("dummy", "허수/기준 전투분석기")

    with st.expander("기존 개별 업로드 UI 사용", expanded=False):
        st.caption("문제가 생기면 기존 실전/허수 개별 업로드 UI로도 사용할 수 있습니다.")
        if st.checkbox("기존 UI 표시", value=False, key="show_legacy_ocr_ui_v115"):
            _ocr_tab_prev_v115()




try:
    _make_real_integrated_debug_zip_prev_v115 = _make_real_integrated_debug_zip_v106
    def _make_real_integrated_debug_zip_v106(summary_image: Any, attack_image: Any) -> tuple[Path, dict[str, Any]]:  # type: ignore[override]
        zpath, context = _make_real_integrated_debug_zip_prev_v115(summary_image, attack_image)
        try:
            import zipfile
            perf_report = {}
            with zipfile.ZipFile(zpath, "r") as zin:
                try:
                    perf_report = _json_v115.loads(zin.read("perf/perf_report.json").decode("utf-8"))
                except Exception:
                    perf_report = {}
            api_timing = st.session_state.get("api_last_timing_v114") or st.session_state.get("api_last_timing_v113") or {}
            bottleneck = _build_bottleneck_summary_v115(perf_report, api_timing)
            tmp = Path(str(zpath) + ".tmp")
            with zipfile.ZipFile(zpath, "r") as zin, zipfile.ZipFile(tmp, "w", compression=zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    zout.writestr(item, zin.read(item.filename))
                zout.writestr("context/api_timing_v115.json", _json_v115.dumps(api_timing, ensure_ascii=False, indent=2, default=str))
                zout.writestr("context/bottleneck_summary_v115.json", _json_v115.dumps(bottleneck, ensure_ascii=False, indent=2, default=str))
            tmp.replace(zpath)
            context["api_timing_v115"] = api_timing
            context["bottleneck_summary_v115"] = bottleneck
        except Exception as e:
            try:
                context.setdefault("errors", []).append({"stage": "v115_bottleneck_append", "error": repr(e)})
            except Exception:
                pass
        return zpath, context
except Exception:
    pass



# ==============================================================================
# v116: 웹 배포용 no-cache API + 병목 진단 정리
# ==============================================================================
# 사용자가 지적한 것처럼 공개 웹사이트에서는 대부분 첫 방문/첫 검색이므로
# 디스크 캐시/세션 캐시 hit를 전제로 한 최적화는 병목 판단을 흐립니다.
# v116은 API 캐시를 완전히 끄고, 매 검색마다 실제 API fetch + 계산표 생성 시간을 기록합니다.

APP_CALC_VERSION = "v116_no_cache_diagnostics"
import time as _time_v116
import json as _json_v116


def _fetch_api_bundle_timed_v114(token: str, character_name: str, *, force_refresh: bool = False) -> tuple[dict[str, Any], dict[str, Any]]:  # type: ignore[override]
    """v116: no-cache 웹 배포 기준 API/계산 타이밍.

    디스크/세션 캐시는 사용하지 않습니다. force_refresh 인자는 하위 호환용으로만 받습니다.
    """
    t0 = _time_v116.perf_counter()
    timing: dict[str, Any] = {
        "version": "v116_no_cache",
        "character_name": character_name,
        "cache_enabled": False,
        "session_cache_hit": False,
        "disk_cache_hit": False,
        "force_refresh": True,
    }

    t = _time_v116.perf_counter()
    client = LostArkApiClient(token, CONFIG_DIR / "api_endpoints.yaml")
    timing["client_init_ms"] = round((_time_v116.perf_counter() - t) * 1000.0, 3)

    t = _time_v116.perf_counter()
    bundle = client.fetch_armory_bundle(character_name)
    timing["fetch_armory_bundle_ms"] = round((_time_v116.perf_counter() - t) * 1000.0, 3)

    t = _time_v116.perf_counter()
    serial = serializable_bundle(bundle)
    timing["serializable_bundle_ms"] = round((_time_v116.perf_counter() - t) * 1000.0, 3)

    t = _time_v116.perf_counter()
    try:
        summary0, detailed = _summarize_all_detailed_v115(bundle)
    except Exception as e:
        detailed = {"version": "v116", "summarize_all_detailed_error": repr(e)}
        summary0 = summarize_all(bundle)
    timing["summarize_all_ms"] = round((_time_v116.perf_counter() - t) * 1000.0, 3)
    timing["summarize_detail_v115"] = detailed
    timing["summarize_detail_v116"] = detailed

    t = _time_v116.perf_counter()
    summary = enrich_summary_with_identity(summary0, serial)
    if isinstance(summary, dict):
        summary["_calc_version"] = APP_CALC_VERSION
    timing["enrich_summary_with_identity_ms"] = round((_time_v116.perf_counter() - t) * 1000.0, 3)

    failed = {k: {"status_code": getattr(v, "status_code", None), "error": getattr(v, "error", "")} for k, v in bundle.items() if not getattr(v, "ok", False)}
    timing["failed"] = failed
    timing["endpoint_count"] = len(bundle)
    timing["total_ms"] = round((_time_v116.perf_counter() - t0) * 1000.0, 3)
    return {"api_bundle": serial, "api_summary": summary}, timing


def _render_api_timing_caption_v116(timing: dict[str, Any]) -> None:
    try:
        st.caption(
            f"API/계산 시간: {float(timing.get('total_ms') or 0):,.0f} ms · "
            f"API fetch {float(timing.get('fetch_armory_bundle_ms') or 0):,.0f} ms · "
            f"계산표 {float(timing.get('summarize_all_ms') or 0):,.0f} ms · "
            f"identity {float(timing.get('enrich_summary_with_identity_ms') or 0):,.0f} ms · "
            "캐시 OFF"
        )
    except Exception:
        pass


def sidebar_controls() -> None:  # type: ignore[override]
    """v116: 캐시 없는 웹 배포 기준 사이드바."""
    calc_config = load_yaml(str(CONFIG_DIR / "calculation_presets.yaml"))
    defaults = calc_config.get("defaults", {})
    st.session_state.setdefault("api_crit_basis", "백어택 기준(조건부 포함)")
    st.session_state.setdefault("target_crit_percent", float(defaults.get("target_crit_percent", 100.0)))
    st.session_state.setdefault("target_back_percent", float(defaults.get("target_back_attack_percent", 100.0)))
    st.session_state.setdefault("back_bonus_percent", float(defaults.get("back_attack_damage_bonus_percent", 5.0)))

    with st.sidebar:
        st.markdown("### LOA 실전 분석")
        st.caption("속도 진단 · 캐시 OFF · 통합 업로드")
        st.session_state["show_debug_tools"] = st.checkbox(
            "🔧 디버그 도구 표시",
            value=bool(st.session_state.get("show_debug_tools", False)),
            help="본문의 OCR 원문/crop 디버그를 표시합니다. 통합 디버그 ZIP은 검색 버튼 아래에서 생성합니다.",
        )
        token = st.text_input("Lost Ark API Key", type="password", key="api_token")
        character_name = st.text_input("캐릭터명", value=st.session_state.get("character_name", ""), key="character_name")
        st.number_input(
            "목표 상승률(%)",
            value=float(st.session_state.get("target_gain_percent", defaults.get("target_gain_percent", 5.0))),
            min_value=0.0,
            max_value=100.0,
            step=0.5,
            key="target_gain_percent",
        )
        fetch = st.button("검색", type="primary", use_container_width=True, key="sidebar_fetch_api_v116_no_cache")
        _render_integrated_debug_button_v106()

        if fetch:
            if not token or not character_name:
                st.warning("API Key와 캐릭터명을 먼저 입력해줘.")
            else:
                with st.spinner("캐릭터 셋팅/계산표를 불러오는 중... (캐시 OFF)"):
                    result, timing = _fetch_api_bundle_timed_v114(token, character_name, force_refresh=True)
                    st.session_state.api_bundle = result["api_bundle"]
                    st.session_state.api_summary = result["api_summary"]
                    st.session_state.api_last_timing_v114 = timing
                    st.session_state.api_last_timing_v116 = timing
                failed = timing.get("failed") or {}
                if failed:
                    st.warning(f"일부 API 호출 실패: {', '.join(failed.keys())}")
                else:
                    st.success("캐릭터 셋팅을 불러왔어. (캐시 OFF)")
                _render_api_timing_caption_v116(timing)

        # 마지막 검색 시간은 재실행 후에도 보이게 둡니다.
        timing = st.session_state.get("api_last_timing_v116") or st.session_state.get("api_last_timing_v114") or {}
        if timing:
            _render_api_timing_caption_v116(timing)

        summary = st.session_state.get("api_summary") or {}
        profile = summary.get("profile_summary", {}) if isinstance(summary, dict) else {}
        if profile:
            st.divider()
            st.caption("현재 분석 캐릭터")
            st.markdown(f"**{profile.get('캐릭터명') or character_name}**")
            meta = " · ".join(str(x) for x in [profile.get("서버"), profile.get("클래스"), profile.get("직업"), profile.get("아이템레벨")] if x)
            if meta:
                st.caption(meta)


def _build_bottleneck_summary_v115(perf_report: dict[str, Any], api_timing: dict[str, Any]) -> dict[str, Any]:  # type: ignore[override]
    """v116: 캐시 항목을 제거하고, 실제 첫 방문 기준 병목만 요약합니다."""
    out: dict[str, Any] = {"version": "v116_no_cache"}
    out["ocr_total_sec"] = perf_report.get("total_elapsed_sec")
    out["ocr_readtext_calls"] = perf_report.get("easyocr_calls")
    out["ocr_readtext_total_sec"] = perf_report.get("easyocr_total_sec")
    out["api_total_ms"] = api_timing.get("total_ms")
    out["api_fetch_ms"] = api_timing.get("fetch_armory_bundle_ms")
    out["api_summarize_ms"] = api_timing.get("summarize_all_ms")
    out["api_identity_ms"] = api_timing.get("enrich_summary_with_identity_ms")
    out["api_cache"] = {"enabled": False, "note": "v116은 웹 첫 방문 기준이라 API 캐시를 사용하지 않습니다."}
    detail = api_timing.get("summarize_detail_v116") or api_timing.get("summarize_detail_v115") or {}
    out["api_summarize_top"] = detail.get("top_stages") or []

    events = perf_report.get("events") or []
    app_done = []
    read_shapes = {}
    for e in events:
        stage = str(e.get("stage") or "")
        if stage.startswith("app:") and stage.endswith("_done") and e.get("elapsed_ms") is not None:
            try:
                app_done.append({"stage": stage, "ms": float(e.get("elapsed_ms") or 0)})
            except Exception:
                pass
        if stage == "readtext":
            sh = str(e.get("shape") or "")
            item = read_shapes.setdefault(sh, {"calls": 0, "ms": 0.0})
            item["calls"] += 1
            try:
                item["ms"] += float(e.get("elapsed_ms") or 0)
            except Exception:
                pass
    out["ocr_app_done_top"] = sorted(app_done, key=lambda x: x["ms"], reverse=True)[:10]
    out["readtext_shapes_top"] = sorted([{"shape": k, **v} for k, v in read_shapes.items()], key=lambda x: x["ms"], reverse=True)[:12]

    recommendations = []
    try:
        if float(api_timing.get("fetch_armory_bundle_ms") or 0) > 1500:
            recommendations.append("API fetch가 1.5초 이상입니다. Lost Ark API 응답/네트워크가 병목입니다. 병렬 endpoint 호출이 가능한지 확인하세요.")
        else:
            recommendations.append("이번 로그에서는 API fetch가 큰 병목이 아닙니다. 캐시보다 OCR/아이콘 매칭 최적화가 우선입니다.")
    except Exception:
        pass
    try:
        if float(api_timing.get("summarize_all_ms") or 0) > 500:
            recommendations.append("계산표 생성이 0.5초 이상입니다. api_parser.summarize_all 내부 단계별 최적화가 필요합니다.")
    except Exception:
        pass
    try:
        if float(perf_report.get("easyocr_total_sec") or 0) > 5:
            recommendations.append("OCR readtext 누적 시간이 큽니다. summary 총피해/DPS와 공격정보 피해량 OCR 호출 수를 줄이는 것이 우선입니다.")
    except Exception:
        pass
    out["recommendations"] = recommendations
    return out


try:
    _make_real_integrated_debug_zip_prev_v116 = _make_real_integrated_debug_zip_v106
    def _make_real_integrated_debug_zip_v106(summary_image: Any, attack_image: Any) -> tuple[Path, dict[str, Any]]:  # type: ignore[override]
        zpath, context = _make_real_integrated_debug_zip_prev_v116(summary_image, attack_image)
        try:
            import zipfile
            perf_report = {}
            with zipfile.ZipFile(zpath, "r") as zin:
                try:
                    perf_report = _json_v116.loads(zin.read("perf/perf_report.json").decode("utf-8"))
                except Exception:
                    perf_report = {}
            api_timing = st.session_state.get("api_last_timing_v116") or st.session_state.get("api_last_timing_v114") or {}
            bottleneck = _build_bottleneck_summary_v115(perf_report, api_timing)
            tmp = Path(str(zpath) + ".tmp")
            with zipfile.ZipFile(zpath, "r") as zin, zipfile.ZipFile(tmp, "w", compression=zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    # v115 파일이 있어도 유지하되, v116 파일을 추가합니다.
                    zout.writestr(item, zin.read(item.filename))
                zout.writestr("context/api_timing_v116.json", _json_v116.dumps(api_timing, ensure_ascii=False, indent=2, default=str))
                zout.writestr("context/bottleneck_summary_v116.json", _json_v116.dumps(bottleneck, ensure_ascii=False, indent=2, default=str))
            tmp.replace(zpath)
            context["api_timing_v116"] = api_timing
            context["bottleneck_summary_v116"] = bottleneck
        except Exception as e:
            try:
                context.setdefault("errors", []).append({"stage": "v116_bottleneck_append", "error": repr(e)})
            except Exception:
                pass
        return zpath, context
except Exception:
    pass




# ==============================================================================
# v117: 속도 최적화 + 검색 UI wall-time 진단
# ==============================================================================
import os as _os_v117_app
import time as _time_v117_app
import json as _json_v117_app

_os_v117_app.environ.setdefault("LOA_ATTACK_ROW_PASS_SCALE", "2")
_os_v117_app.environ.setdefault("LOA_ATTACK_DAMAGE_RECHECK", "smart")
_os_v117_app.environ.setdefault("LOA_ATTACK_CRITICAL_CELL_SCALE", "3")
_os_v117_app.environ.setdefault("LOA_ICON_HEAVY_TOPK", "5")
_os_v117_app.environ.setdefault("LOA_ICON_VARIANT_RETRY", "0")
_os_v117_app.environ.setdefault("LOA_SUMMARY_RAW_SCALE", "3")
_os_v117_app.environ.setdefault("LOA_API_CACHE_TTL_SEC", "0")

APP_CALC_VERSION = "v117_speed_optimized"


def _fetch_api_bundle_timed_v117(token: str, character_name: str) -> tuple[dict[str, Any], dict[str, Any]]:
    """캐시 없이 실제 첫 방문 기준 API+요약 시간을 측정합니다."""
    t0 = _time_v117_app.perf_counter()
    timing: dict[str, Any] = {
        "version": "v117_no_cache_wall_diagnostics",
        "character_name": character_name,
        "cache_enabled": False,
        "session_cache_hit": False,
        "disk_cache_hit": False,
    }
    t = _time_v117_app.perf_counter()
    client = LostArkApiClient(token, CONFIG_DIR / "api_endpoints.yaml")
    timing["client_init_ms"] = round((_time_v117_app.perf_counter() - t) * 1000.0, 3)

    t = _time_v117_app.perf_counter()
    bundle = client.fetch_armory_bundle(character_name)
    timing["fetch_armory_bundle_ms"] = round((_time_v117_app.perf_counter() - t) * 1000.0, 3)
    timing["endpoint_timings_ms"] = getattr(client, "last_timings", {})

    t = _time_v117_app.perf_counter()
    serial = serializable_bundle(bundle)
    timing["serializable_bundle_ms"] = round((_time_v117_app.perf_counter() - t) * 1000.0, 3)

    t = _time_v117_app.perf_counter()
    try:
        summary0, detailed = _summarize_all_detailed_v115(bundle)
    except Exception as e:
        detailed = {"version": "v117", "summarize_all_detailed_error": repr(e)}
        summary0 = summarize_all(bundle)
    timing["summarize_all_ms"] = round((_time_v117_app.perf_counter() - t) * 1000.0, 3)
    timing["summarize_detail_v117"] = detailed

    t = _time_v117_app.perf_counter()
    summary = enrich_summary_with_identity(summary0, serial)
    if isinstance(summary, dict):
        summary["_calc_version"] = APP_CALC_VERSION
    timing["enrich_summary_with_identity_ms"] = round((_time_v117_app.perf_counter() - t) * 1000.0, 3)

    failed = {k: {"status_code": getattr(v, "status_code", None), "error": getattr(v, "error", "")} for k, v in bundle.items() if not getattr(v, "ok", False)}
    timing["failed"] = failed
    timing["endpoint_count"] = len(bundle)
    timing["api_pipeline_ms"] = round((_time_v117_app.perf_counter() - t0) * 1000.0, 3)
    timing["total_ms"] = timing["api_pipeline_ms"]
    return {"api_bundle": serial, "api_summary": summary}, timing


def _render_api_timing_caption_v117(timing: dict[str, Any]) -> None:
    if not timing:
        return
    try:
        st.caption(
            f"검색/렌더 진단: 전체 {float(timing.get('search_button_wall_ms') or timing.get('total_ms') or 0):,.0f} ms · "
            f"API {float(timing.get('fetch_armory_bundle_ms') or 0):,.0f} ms · "
            f"계산표 {float(timing.get('summarize_all_ms') or 0):,.0f} ms · "
            f"identity {float(timing.get('enrich_summary_with_identity_ms') or 0):,.0f} ms · "
            f"세션대입 {float(timing.get('state_assign_ms') or 0):,.0f} ms · 캐시 OFF"
        )
    except Exception:
        pass


def sidebar_controls() -> None:  # type: ignore[override]
    calc_config = load_yaml(str(CONFIG_DIR / "calculation_presets.yaml"))
    defaults = calc_config.get("defaults", {})
    st.session_state.setdefault("api_crit_basis", "백어택 기준(조건부 포함)")
    st.session_state.setdefault("target_crit_percent", float(defaults.get("target_crit_percent", 100.0)))
    st.session_state.setdefault("target_back_percent", float(defaults.get("target_back_attack_percent", 100.0)))
    st.session_state.setdefault("back_bonus_percent", float(defaults.get("back_attack_damage_bonus_percent", 5.0)))

    sidebar_t0 = _time_v117_app.perf_counter()
    with st.sidebar:
        st.markdown("### LOA 실전 분석")
        st.caption("속도 최적화 · no-cache · OCR/아이콘 병목 축소")
        st.session_state["show_debug_tools"] = st.checkbox(
            "🔧 디버그 도구 표시",
            value=bool(st.session_state.get("show_debug_tools", False)),
            help="본문의 OCR 원문/crop 디버그를 표시합니다. 통합 디버그 ZIP은 검색 버튼 아래에서 생성합니다.",
        )
        token = st.text_input("Lost Ark API Key", type="password", key="api_token")
        character_name = st.text_input("캐릭터명", value=st.session_state.get("character_name", ""), key="character_name")
        st.number_input(
            "목표 상승률(%)",
            value=float(st.session_state.get("target_gain_percent", defaults.get("target_gain_percent", 5.0))),
            min_value=0.0,
            max_value=100.0,
            step=0.5,
            key="target_gain_percent",
        )
        fetch = st.button("검색", type="primary", use_container_width=True, key="sidebar_fetch_api_v117_no_cache")
        _render_integrated_debug_button_v106()

        if fetch:
            search_t0 = _time_v117_app.perf_counter()
            if not token or not character_name:
                st.warning("API Key와 캐릭터명을 먼저 입력해줘.")
            else:
                with st.spinner("캐릭터 셋팅/계산표를 불러오는 중... (캐시 OFF)"):
                    result, timing = _fetch_api_bundle_timed_v117(token, character_name)
                    t_assign = _time_v117_app.perf_counter()
                    st.session_state.api_bundle = result["api_bundle"]
                    st.session_state.api_summary = result["api_summary"]
                    st.session_state.api_last_timing_v117 = timing
                    st.session_state.api_last_timing_v116 = timing
                    st.session_state.api_last_timing_v114 = timing
                    timing["state_assign_ms"] = round((_time_v117_app.perf_counter() - t_assign) * 1000.0, 3)
                    timing["search_button_wall_ms"] = round((_time_v117_app.perf_counter() - search_t0) * 1000.0, 3)
                failed = timing.get("failed") or {}
                if failed:
                    st.warning(f"일부 API 호출 실패: {', '.join(failed.keys())}")
                else:
                    st.success("캐릭터 셋팅을 불러왔어. (캐시 OFF)")
                _render_api_timing_caption_v117(timing)

        timing = st.session_state.get("api_last_timing_v117") or st.session_state.get("api_last_timing_v116") or st.session_state.get("api_last_timing_v114") or {}
        if timing:
            # sidebar 자체 렌더 시간도 남겨서, 영상상의 긴 대기와 API 타이밍 차이를 확인합니다.
            try:
                timing["sidebar_render_ms_last"] = round((_time_v117_app.perf_counter() - sidebar_t0) * 1000.0, 3)
            except Exception:
                pass
            _render_api_timing_caption_v117(timing)

        summary = st.session_state.get("api_summary") or {}
        profile = summary.get("profile_summary", {}) if isinstance(summary, dict) else {}
        if profile:
            st.divider()
            st.markdown(f"**{profile.get('캐릭터명') or character_name}**")
            st.caption(f"{profile.get('클래스명', '-')} · Lv.{profile.get('아이템레벨', '-')}")


_make_real_integrated_debug_zip_v106_prev_v117 = _make_real_integrated_debug_zip_v106

def _make_real_integrated_debug_zip_v106(summary_image: Any, attack_image: Any) -> tuple[Path, dict[str, Any]]:  # type: ignore[override]
    """v117: 기존 통합 ZIP 생성 후 API wall-time/속도 설정을 추가합니다."""
    zpath, context = _make_real_integrated_debug_zip_v106_prev_v117(summary_image, attack_image) if '_make_real_integrated_debug_zip_v106_prev_v117' in globals() else (None, {})
    if zpath is None:
        raise RuntimeError('previous integrated debug function missing')
    try:
        api_timing = st.session_state.get('api_last_timing_v117') or st.session_state.get('api_last_timing_v116') or st.session_state.get('api_last_timing_v114') or {}
        speed_env = {k: _os_v117_app.environ.get(k) for k in [
            'LOA_ATTACK_ROW_PASS_SCALE','LOA_ATTACK_DAMAGE_RECHECK','LOA_ATTACK_CRITICAL_CELL_SCALE',
            'LOA_ICON_HEAVY_TOPK','LOA_ICON_VARIANT_RETRY','LOA_SUMMARY_RAW_SCALE','LOA_API_CACHE_TTL_SEC'
        ]}
        tmp = Path(str(zpath) + '.tmp')
        with zipfile.ZipFile(zpath, 'r') as zin, zipfile.ZipFile(tmp, 'w', compression=zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                zout.writestr(item, zin.read(item.filename))
            zout.writestr('context/api_timing_v117.json', _json_v117_app.dumps(api_timing, ensure_ascii=False, indent=2, default=str))
            zout.writestr('context/speed_env_v117.json', _json_v117_app.dumps(speed_env, ensure_ascii=False, indent=2, default=str))
        tmp.replace(zpath)
        context['api_timing_v117'] = api_timing
        context['speed_env_v117'] = speed_env
    except Exception as e:
        try:
            context.setdefault('errors', []).append({'stage': 'v117_append', 'error': repr(e)})
        except Exception:
            pass
    return zpath, context

# 기존 함수를 보관한 뒤 v117 wrapper를 사용합니다.

# ==============================================================================
# v118: speed/back-column fix UI helpers
# ==============================================================================
APP_CALC_VERSION = "v118_speed_backfix"

import os as _os_v118_app
_os_v118_app.environ.setdefault("LOA_ATTACK_ROW_PASS_SCALE", "3")
_os_v118_app.environ.setdefault("LOA_ATTACK_DAMAGE_RECHECK", "smart")
_os_v118_app.environ.setdefault("LOA_ATTACK_CRITICAL_CELL_SCALE", "3")
_os_v118_app.environ.setdefault("LOA_ATTACK_PERCENT_FALLBACK", "1")
_os_v118_app.environ.setdefault("LOA_ATTACK_CAST_COOLDOWN_FALLBACK", "1")
_os_v118_app.environ.setdefault("LOA_ICON_HEAVY_TOPK", "5")
_os_v118_app.environ.setdefault("LOA_ICON_VARIANT_RETRY", "0")
_os_v118_app.environ.setdefault("LOA_SUMMARY_RAW_SCALE", "3")
_os_v118_app.environ.setdefault("LOA_API_CACHE_TTL_SEC", "0")

_drop_obvious_ocr_noise_rows_prev_v118 = globals().get("_drop_obvious_ocr_noise_rows")

def _drop_obvious_ocr_noise_rows(df: pd.DataFrame) -> pd.DataFrame:  # type: ignore[override]
    """Remove header-like trailing rows produced when row_count is larger than actual table rows."""
    if df is None or not isinstance(df, pd.DataFrame) or df.empty:
        return df
    out = df.copy()
    keep = []
    header_words = {"이름", "피해량", "dps", "지분%", "백적중%", "백비중%", "헤드적중%", "헤드비중%", "치적%", "치비중%", "횟수", "쿨%"}
    for _, row in out.iterrows():
        name = str(row.get("이름") or "").strip()
        dmg = str(row.get("피해량") or "").strip()
        dps = str(row.get("초당 피해량") or row.get("DPS") or "").strip()
        # Header/placeholder rows from the editor or OCR should not enter calculations.
        if name.lower() in {"name"} or name in header_words:
            continue
        if dmg in {"", "피해량", "damage"} and dps in {"", "DPS", "dps"}:
            continue
        # Rows with no digit in both name/damage are noise.
        if not re.search(r"\d", dmg) and name in {"", "이름"}:
            continue
        keep.append(row)
    if not keep:
        return out.iloc[0:0].copy()
    cleaned = pd.DataFrame(keep).reset_index(drop=True)
    if callable(_drop_obvious_ocr_noise_rows_prev_v118):
        try:
            cleaned = _drop_obvious_ocr_noise_rows_prev_v118(cleaned)
        except Exception:
            pass
    return cleaned

_make_bottleneck_summary_prev_v118 = globals().get("_make_bottleneck_summary_v116") or globals().get("_make_bottleneck_summary_v115")

def _write_v118_speed_env_snapshot() -> dict[str, str]:
    keys = [
        "LOA_ATTACK_ROW_PASS_SCALE", "LOA_ATTACK_DAMAGE_RECHECK", "LOA_ATTACK_CRITICAL_CELL_SCALE",
        "LOA_ATTACK_PERCENT_FALLBACK", "LOA_ATTACK_CAST_COOLDOWN_FALLBACK",
        "LOA_ICON_HEAVY_TOPK", "LOA_ICON_VARIANT_RETRY", "LOA_SUMMARY_RAW_SCALE", "LOA_API_CACHE_TTL_SEC",
    ]
    return {k: str(_os_v118_app.environ.get(k, "")) for k in keys}

# If the integrated-debug collector asks for v117 env, also include v118 values.
try:
    _debug_context_extra_prev_v118 = globals().get("_debug_context_extra_v117")
except Exception:
    _debug_context_extra_prev_v118 = None

try:
    _sidebar_controls_prev_v118 = sidebar_controls
    def sidebar_controls() -> None:  # type: ignore[override]
        pass  # 버전 변경 안내 제거
        try:
            st.sidebar.caption("OCR: row-pass 3배율 + 필요한 셀만 빠른 재검수 / API 캐시 OFF")
        except Exception:
            pass
        return _sidebar_controls_prev_v118()
except Exception:
    pass


# ==============================================================================
# v119: Streamlit lazy UI / rerun reduction / v118 backfix retained
# ==============================================================================
# Streamlit tabs render every tab on each rerun.  The previous UI rendered API,
# OCR and result pages plus many nested API tables after pressing Search, so the
# visible wait could be much longer than the measured Lost Ark API time.  v119
# switches the top-level navigation and API details to lazy single-section
# rendering.  OCR accuracy/back-column v118 logic is kept.

APP_CALC_VERSION = "v119_lazy_ui_speed_backfix"

import os as _os_v119_app
import time as _time_v119_app

_os_v119_app.environ.setdefault("LOA_UI_FAST_MODE", "1")
_os_v119_app.environ.setdefault("LOA_ATTACK_FAST_GRID", "1")
_os_v119_app.environ.setdefault("LOA_ATTACK_ROW_PASS_SCALE", "3")
_os_v119_app.environ.setdefault("LOA_ATTACK_DAMAGE_RECHECK", "smart")
_os_v119_app.environ.setdefault("LOA_ATTACK_CRITICAL_CELL_SCALE", "3")
_os_v119_app.environ.setdefault("LOA_ATTACK_PERCENT_FALLBACK", "1")
_os_v119_app.environ.setdefault("LOA_ATTACK_CAST_COOLDOWN_FALLBACK", "1")
_os_v119_app.environ.setdefault("LOA_ICON_HEAVY_TOPK", "5")
_os_v119_app.environ.setdefault("LOA_ICON_VARIANT_RETRY", "0")
_os_v119_app.environ.setdefault("LOA_SUMMARY_RAW_SCALE", "3")
_os_v119_app.environ.setdefault("LOA_API_CACHE_TTL_SEC", "0")

try:
    @st.cache_data(show_spinner=False, max_entries=12)
    def _load_image_from_bytes_cached_v119(raw: bytes) -> bytes:
        # cache_data should store simple bytes reliably; PIL decode is done after
        # retrieving the cached normalized PNG bytes.
        import io as _io_v119
        from PIL import Image as _Image_v119
        im = _Image_v119.open(_io_v119.BytesIO(raw)).convert("RGB")
        buf = _io_v119.BytesIO()
        im.save(buf, format="PNG", optimize=False)
        return buf.getvalue()
except Exception:
    _load_image_from_bytes_cached_v119 = None  # type: ignore[assignment]


def _image_from_uploader_or_session_v115(uploaded: Any, kind: str, slot: str) -> Image.Image | None:  # type: ignore[override]
    """v119: avoid re-decoding uploaded images on every Streamlit rerun."""
    raw = None
    if uploaded is not None:
        try:
            raw = uploaded.getvalue()
            st.session_state[f"{kind}_{slot}_image_bytes"] = raw
            st.session_state[f"{kind}_{slot}_image_name"] = getattr(uploaded, "name", f"{kind}_{slot}.png")
        except Exception:
            raw = None
    else:
        raw = st.session_state.get(f"{kind}_{slot}_image_bytes")
    if not raw:
        return None
    try:
        import io as _io_v119
        if callable(_load_image_from_bytes_cached_v119):
            png_bytes = _load_image_from_bytes_cached_v119(bytes(raw))
            return Image.open(_io_v119.BytesIO(png_bytes)).convert("RGB")
        return Image.open(_io_v119.BytesIO(raw)).convert("RGB")
    except Exception:
        try:
            return _get_uploaded_image_from_session_v88(kind, slot)
        except Exception:
            return None


def _render_api_summary_fast_v119(summary: dict[str, Any], character_name: str = "") -> None:
    profile = summary.get("profile_summary", {}) if isinstance(summary, dict) else {}
    if profile:
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("캐릭터", str(profile.get("캐릭터명") or character_name or "-"))
        c2.metric("클래스", str(profile.get("클래스") or profile.get("클래스명") or "-"))
        c3.metric("직업", str(profile.get("직업") or "-"))
        c4.metric("아이템 레벨", str(profile.get("아이템레벨") or "-"))
    timing = st.session_state.get("api_last_timing_v117") or st.session_state.get("api_last_timing_v116") or st.session_state.get("api_last_timing_v114") or {}
    if timing:
        _render_api_timing_caption_v117(timing)
    overview = summary.get("combat_overview") if isinstance(summary, dict) else None
    if isinstance(overview, pd.DataFrame) and not overview.empty:
        st.dataframe(overview, use_container_width=True, hide_index=True, height=170)






def sidebar_controls() -> None:  # type: ignore[override]
    """v119: use a form to reduce reruns while entering API key/character."""
    calc_config = load_yaml(str(CONFIG_DIR / "calculation_presets.yaml"))
    defaults = calc_config.get("defaults", {})
    st.session_state.setdefault("api_crit_basis", "백어택 기준(조건부 포함)")
    st.session_state.setdefault("target_crit_percent", float(defaults.get("target_crit_percent", 100.0)))
    st.session_state.setdefault("target_back_percent", float(defaults.get("target_back_attack_percent", 100.0)))
    st.session_state.setdefault("back_bonus_percent", float(defaults.get("back_attack_damage_bonus_percent", 5.0)))

    sidebar_t0 = _time_v119_app.perf_counter()
    with st.sidebar:
        st.markdown("### LOA 실전 분석")
        pass  # 버전 변경 안내 제거
        st.session_state["show_debug_tools"] = st.checkbox(
            "🔧 디버그 도구 표시",
            value=bool(st.session_state.get("show_debug_tools", False)),
            help="본문 OCR 디버그와 통합 디버그 ZIP을 표시합니다.",
        )
        st.session_state["fast_page_v119"] = st.radio(
            "화면",
            ["① 캐릭터 세팅", "② 전투분석기 입력", "③ 실력 분석 결과"],
            index=["① 캐릭터 세팅", "② 전투분석기 입력", "③ 실력 분석 결과"].index(st.session_state.get("fast_page_v119", "① 캐릭터 세팅")) if st.session_state.get("fast_page_v119") in ["① 캐릭터 세팅", "② 전투분석기 입력", "③ 실력 분석 결과"] else 0,
            key="fast_page_radio_v119",
        )

        with st.form("api_search_form_v119", clear_on_submit=False):
            token = st.text_input("Lost Ark API Key", type="password", key="api_token")
            character_name = st.text_input("캐릭터명", value=st.session_state.get("character_name", ""), key="character_name")
            st.number_input(
                "목표 상승률(%)",
                value=float(st.session_state.get("target_gain_percent", defaults.get("target_gain_percent", 5.0))),
                min_value=0.0,
                max_value=100.0,
                step=0.5,
                key="target_gain_percent",
            )
            fetch = st.form_submit_button("검색", type="primary", use_container_width=True)

        _render_integrated_debug_button_v106()

        if fetch:
            search_t0 = _time_v119_app.perf_counter()
            if not token or not character_name:
                st.warning("API Key와 캐릭터명을 먼저 입력해줘.")
            else:
                with st.spinner("캐릭터 셋팅/계산표를 불러오는 중... (캐시 OFF)"):
                    result, timing = _fetch_api_bundle_timed_v117(token, character_name)
                    t_assign = _time_v119_app.perf_counter()
                    st.session_state.api_bundle = result["api_bundle"]
                    st.session_state.api_summary = result["api_summary"]
                    st.session_state.api_last_timing_v119 = timing
                    st.session_state.api_last_timing_v117 = timing
                    st.session_state.api_last_timing_v116 = timing
                    st.session_state.api_last_timing_v114 = timing
                    timing["state_assign_ms"] = round((_time_v119_app.perf_counter() - t_assign) * 1000.0, 3)
                    timing["search_button_wall_ms"] = round((_time_v119_app.perf_counter() - search_t0) * 1000.0, 3)
                    timing["ui_mode_v119"] = "lazy_single_page"
                    st.session_state["fast_page_radio_v119"] = "① 캐릭터 세팅"
            if token and character_name:
                failed = (st.session_state.get("api_last_timing_v119") or {}).get("failed") or {}
                if failed:
                    st.warning(f"일부 API 호출 실패: {', '.join(failed.keys())}")
                else:
                    st.success("캐릭터 셋팅을 불러왔어. 빠른 렌더 모드로 표시합니다.")

        timing = st.session_state.get("api_last_timing_v119") or st.session_state.get("api_last_timing_v117") or st.session_state.get("api_last_timing_v116") or st.session_state.get("api_last_timing_v114") or {}
        if timing:
            try:
                timing["sidebar_render_ms_last"] = round((_time_v119_app.perf_counter() - sidebar_t0) * 1000.0, 3)
            except Exception:
                pass
            _render_api_timing_caption_v117(timing)

        summary = st.session_state.get("api_summary") or {}
        profile = summary.get("profile_summary", {}) if isinstance(summary, dict) else {}
        if profile:
            st.divider()
            st.markdown(f"**{profile.get('캐릭터명') or st.session_state.get('character_name','')}**")
            st.caption(f"{profile.get('클래스명', profile.get('클래스', '-'))} · Lv.{profile.get('아이템레벨', '-')}")




try:
    _make_real_integrated_debug_zip_prev_v119 = _make_real_integrated_debug_zip_v106
    def _make_real_integrated_debug_zip_v106(summary_image: Any, attack_image: Any) -> tuple[Path, dict[str, Any]]:  # type: ignore[override]
        zpath, context = _make_real_integrated_debug_zip_prev_v119(summary_image, attack_image)
        try:
            import zipfile as _zipfile_v119
            import json as _json_v119
            api_timing = st.session_state.get("api_last_timing_v119") or st.session_state.get("api_last_timing_v117") or st.session_state.get("api_last_timing_v116") or {}
            ui_env = {
                "version": "v119_lazy_ui_speed_backfix",
                "LOA_UI_FAST_MODE": _os_v119_app.environ.get("LOA_UI_FAST_MODE"),
                "top_nav": "sidebar_radio_single_page",
                "api_detail": "radio_lazy_one_section",
                "battle_editor": "read_only_by_default",
                "uploaded_image_decode": "cache_data_bytes",
            }
            tmp = Path(str(zpath) + ".tmp")
            with _zipfile_v119.ZipFile(zpath, "r") as zin, _zipfile_v119.ZipFile(tmp, "w", compression=_zipfile_v119.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    zout.writestr(item, zin.read(item.filename))
                zout.writestr("context/api_timing_v119.json", _json_v119.dumps(api_timing, ensure_ascii=False, indent=2, default=str))
                zout.writestr("context/ui_speed_v119.json", _json_v119.dumps(ui_env, ensure_ascii=False, indent=2, default=str))
            tmp.replace(zpath)
            context["api_timing_v119"] = api_timing
            context["ui_speed_v119"] = ui_env
        except Exception as e:
            try:
                context.setdefault("errors", []).append({"stage": "v119_append", "error": repr(e)})
            except Exception:
                pass
        return zpath, context
except Exception:
    pass



# ==============================================================================
# v120: Static-first UI renderer + browser auto-open support
# ==============================================================================
# 목표:
# - 화면 선택식 lazy UI(v119)를 사용자가 불편해해서 상단 3탭 구조를 복구한다.
# - 대신 모든 표를 Streamlit dataframe/data_editor가 아니라 HTML 기반 정적 테이블로 먼저 렌더한다.
# - 수정/원본 dataframe은 버튼을 눌렀을 때만 렌더한다.
# - 검색 버튼 클릭부터 실제 API 탭 렌더 완료까지 wall time을 측정한다.

APP_CALC_VERSION = "v120_static_first_ui_browser"

import os as _os_v120_app
import time as _time_v120_app
import html as _html_v120_app
import re as _re_v120_app

_os_v120_app.environ.setdefault("LOA_UI_FAST_MODE", "static_first")
_os_v120_app.environ.setdefault("LOA_RENDER_STATIC_TABLE", "1")
_os_v120_app.environ.setdefault("LOA_ATTACK_FAST_GRID", "1")
_os_v120_app.environ.setdefault("LOA_ATTACK_ROW_PASS_SCALE", "3")
_os_v120_app.environ.setdefault("LOA_ATTACK_DAMAGE_RECHECK", "smart")
_os_v120_app.environ.setdefault("LOA_ATTACK_CRITICAL_CELL_SCALE", "3")
_os_v120_app.environ.setdefault("LOA_ATTACK_PERCENT_FALLBACK", "1")
_os_v120_app.environ.setdefault("LOA_ATTACK_CAST_COOLDOWN_FALLBACK", "1")
_os_v120_app.environ.setdefault("LOA_ICON_HEAVY_TOPK", "5")
_os_v120_app.environ.setdefault("LOA_ICON_VARIANT_RETRY", "0")
_os_v120_app.environ.setdefault("LOA_SUMMARY_RAW_SCALE", "3")
_os_v120_app.environ.setdefault("LOA_API_CACHE_TTL_SEC", "0")


def _v120_safe_key(s: str) -> str:
    return _re_v120_app.sub(r"[^0-9A-Za-z가-힣_]+", "_", str(s))[:80]


def _v120_fmt_cell(v: Any) -> str:
    try:
        if v is None:
            return ""
        # pandas NA
        if pd.isna(v):
            return ""
    except Exception:
        pass
    if isinstance(v, float):
        # 원본 표시가 이미 문자열이면 건드리지 않고, float만 짧게 표시
        if abs(v) >= 1000:
            return f"{v:,.2f}"
        return f"{v:.2f}".rstrip("0").rstrip(".")
    return str(v)


def _v120_fast_html_table(df: Any, max_rows: int = 80, max_cols: int = 14, class_name: str = "loa-v120-table") -> str:
    try:
        import pandas as _pd_v120
        if df is None:
            return '<div class="loa-v120-empty">표 데이터가 없습니다.</div>'
        if isinstance(df, list):
            df = _pd_v120.DataFrame(df)
        elif isinstance(df, dict):
            df = _pd_v120.DataFrame([df])
        elif not isinstance(df, _pd_v120.DataFrame):
            return f'<div class="loa-v120-empty">{_html_v120_app.escape(str(df))}</div>'
        if df.empty:
            return '<div class="loa-v120-empty">표 데이터가 없습니다.</div>'
        view = df.copy()
        hidden = [c for c in view.columns if str(c).startswith("_")]
        if hidden:
            view = view.drop(columns=hidden, errors="ignore")
        if len(view.columns) > max_cols:
            view = view.iloc[:, :max_cols]
        if len(view) > max_rows:
            view = view.head(max_rows)
        ths = "".join(f"<th>{_html_v120_app.escape(str(c))}</th>" for c in view.columns)
        rows = []
        for _, row in view.iterrows():
            tds = "".join(f"<td>{_html_v120_app.escape(_v120_fmt_cell(row.get(c)))}</td>" for c in view.columns)
            rows.append(f"<tr>{tds}</tr>")
        more = ""
        if len(df) > len(view):
            more = f'<div class="loa-v120-more">표시: {len(view):,} / 전체 {len(df):,}행</div>'
        return f'<div class="loa-v120-scroll"><table class="{class_name}"><thead><tr>{ths}</tr></thead><tbody>{"".join(rows)}</tbody></table></div>{more}'
    except Exception as e:
        return f'<div class="loa-v120-empty">표 렌더 오류: {_html_v120_app.escape(repr(e))}</div>'


def _inject_v120_static_css() -> None:
    st.markdown(
        """
        <style>
        .loa-v120-banner{padding:10px 14px;border-radius:12px;background:rgba(38,42,54,.72);border:1px solid rgba(255,255,255,.08);margin:8px 0 14px 0}
        .loa-v120-banner b{color:#fff}
        .loa-v120-grid{display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:10px;margin:10px 0 14px 0}
        .loa-v120-card{background:#242733;border:1px solid rgba(255,255,255,.08);border-radius:12px;padding:12px;min-height:70px}
        .loa-v120-card .k{color:#aab0c0;font-size:12px;margin-bottom:6px}.loa-v120-card .v{color:#fff;font-weight:700;font-size:18px;word-break:break-word}
        .loa-v120-scroll{overflow:auto;max-height:560px;border-radius:10px;border:1px solid rgba(255,255,255,.08);background:#151922;margin:6px 0 10px 0}
        table.loa-v120-table{border-collapse:separate;border-spacing:0;width:100%;font-size:13px;color:#f6f7fb;table-layout:auto}
        table.loa-v120-table th{position:sticky;top:0;background:#202431;color:#cbd1df;font-weight:700;padding:8px 10px;border-bottom:1px solid rgba(255,255,255,.10);white-space:nowrap;text-align:left;z-index:1}
        table.loa-v120-table td{background:#252833;padding:8px 10px;border-bottom:6px solid #11151d;white-space:nowrap;border-radius:6px}
        table.loa-v120-table tr:hover td{background:#2d3140}
        .loa-v120-empty{padding:12px;color:#aab0c0;background:#202431;border-radius:10px;border:1px solid rgba(255,255,255,.08)}
        .loa-v120-more{font-size:12px;color:#aab0c0;margin:4px 0 12px 2px}
        .loa-v120-placeholder{height:16px;border-radius:6px;background:linear-gradient(90deg,rgba(255,255,255,.06),rgba(255,255,255,.14),rgba(255,255,255,.06));margin:6px 0}
        @media (max-width: 900px){.loa-v120-grid{grid-template-columns:repeat(2,minmax(0,1fr));}}
        </style>
        """,
        unsafe_allow_html=True,
    )


def _v120_cards(items: list[tuple[str, Any]]) -> None:
    cards = []
    for k, v in items:
        cards.append(f'<div class="loa-v120-card"><div class="k">{_html_v120_app.escape(str(k))}</div><div class="v">{_html_v120_app.escape(_v120_fmt_cell(v))}</div></div>')
    st.markdown(f'<div class="loa-v120-grid">{"".join(cards)}</div>', unsafe_allow_html=True)


# show_dataframe 전체를 정적 HTML 우선으로 바꾼다. 원본 dataframe은 체크박스를 켰을 때만 그린다.
try:
    _show_dataframe_prev_v120 = show_dataframe
    def show_dataframe(df: Any, title: str, height: int = 320) -> None:  # type: ignore[override]
        st.markdown(f"#### {title}")
        st.markdown(_v120_fast_html_table(df, max_rows=80, max_cols=16), unsafe_allow_html=True)
        key = "v120_raw_df_" + _v120_safe_key(title)
        if st.checkbox(f"{title} 원본 Streamlit 표 열기", value=False, key=key):
            _show_dataframe_prev_v120(df, title + " · 원본", height)
except Exception:
    pass


def _render_api_summary_fast_v120(summary: dict[str, Any], character_name: str = "") -> None:
    profile = summary.get("profile_summary", {}) if isinstance(summary, dict) else {}
    overview = summary.get("combat_overview") if isinstance(summary, dict) else None
    _v120_cards([
        ("캐릭터", profile.get("캐릭터명") or character_name or "-"),
        ("클래스", profile.get("클래스") or profile.get("클래스명") or "-"),
        ("직업", profile.get("직업") or "-"),
        ("아이템 레벨", profile.get("아이템레벨") or "-"),
    ])
    timing = st.session_state.get("api_last_timing_v120") or st.session_state.get("api_last_timing_v119") or st.session_state.get("api_last_timing_v117") or {}
    if timing:
        _render_api_timing_caption_v117(timing)
    if isinstance(overview, pd.DataFrame) and not overview.empty:
        show_dataframe(overview, "전투/캐릭터 요약", 170)






def sidebar_controls() -> None:  # type: ignore[override]
    """v120: 화면 선택은 제거하고 상단 탭 구조를 사용한다. 검색은 form으로 rerun 최소화."""
    calc_config = load_yaml(str(CONFIG_DIR / "calculation_presets.yaml"))
    defaults = calc_config.get("defaults", {})
    st.session_state.setdefault("api_crit_basis", "백어택 기준(조건부 포함)")
    st.session_state.setdefault("target_crit_percent", float(defaults.get("target_crit_percent", 100.0)))
    st.session_state.setdefault("target_back_percent", float(defaults.get("target_back_attack_percent", 100.0)))
    st.session_state.setdefault("back_bonus_percent", float(defaults.get("back_attack_damage_bonus_percent", 5.0)))

    sidebar_t0 = _time_v120_app.perf_counter()
    with st.sidebar:
        st.markdown("### LOA 실전 분석")
        pass  # 버전 변경 안내 제거
        st.session_state["show_debug_tools"] = st.checkbox(
            "🔧 디버그 도구 표시",
            value=bool(st.session_state.get("show_debug_tools", False)),
            help="본문 OCR 디버그와 통합 디버그 ZIP을 표시합니다.",
        )
        with st.form("api_search_form_v120", clear_on_submit=False):
            token = st.text_input("Lost Ark API Key", type="password", key="api_token")
            character_name = st.text_input("캐릭터명", value=st.session_state.get("character_name", ""), key="character_name")
            st.number_input(
                "목표 상승률(%)",
                value=float(st.session_state.get("target_gain_percent", defaults.get("target_gain_percent", 5.0))),
                min_value=0.0,
                max_value=100.0,
                step=0.5,
                key="target_gain_percent",
            )
            fetch = st.form_submit_button("검색", type="primary", use_container_width=True)

        _render_integrated_debug_button_v106()

        if fetch:
            search_t0 = _time_v120_app.perf_counter()
            st.session_state["_search_click_t0_v120"] = search_t0
            if not token or not character_name:
                st.warning("API Key와 캐릭터명을 먼저 입력해줘.")
            else:
                # 여기서 보이는 시간은 API+계산+state assign까지. 실제 화면 렌더 완료는 api_tab에서 기록한다.
                with st.spinner("캐릭터 셋팅/계산표를 불러오는 중... (캐시 OFF)"):
                    result, timing = _fetch_api_bundle_timed_v117(token, character_name)
                    t_assign = _time_v120_app.perf_counter()
                    st.session_state.api_bundle = result["api_bundle"]
                    st.session_state.api_summary = result["api_summary"]
                    timing["state_assign_ms"] = round((_time_v120_app.perf_counter() - t_assign) * 1000.0, 3)
                    timing["search_button_wall_before_render_ms_v120"] = round((_time_v120_app.perf_counter() - search_t0) * 1000.0, 3)
                    timing["ui_mode_v120"] = "top_tabs_static_html_first"
                    st.session_state.api_last_timing_v120 = timing
                    st.session_state.api_last_timing_v119 = timing
                    st.session_state.api_last_timing_v117 = timing
                    st.session_state.api_last_timing_v116 = timing
                    st.session_state.api_last_timing_v114 = timing
            if token and character_name:
                failed = (st.session_state.get("api_last_timing_v120") or {}).get("failed") or {}
                if failed:
                    st.warning(f"일부 API 호출 실패: {', '.join(failed.keys())}")
                else:
                    st.success("캐릭터 셋팅을 불러왔어. 표는 빠른 HTML 렌더로 표시합니다.")

        timing = st.session_state.get("api_last_timing_v120") or st.session_state.get("api_last_timing_v119") or st.session_state.get("api_last_timing_v117") or {}
        if timing:
            try:
                timing["sidebar_render_ms_last_v120"] = round((_time_v120_app.perf_counter() - sidebar_t0) * 1000.0, 3)
            except Exception:
                pass
            _render_api_timing_caption_v117(timing)

        summary = st.session_state.get("api_summary") or {}
        profile = summary.get("profile_summary", {}) if isinstance(summary, dict) else {}
        if profile:
            st.divider()
            st.markdown(f"**{profile.get('캐릭터명') or st.session_state.get('character_name','')}**")
            st.caption(f"{profile.get('클래스명', profile.get('클래스', '-'))} · Lv.{profile.get('아이템레벨', '-')}")




try:
    _make_real_integrated_debug_zip_prev_v120 = _make_real_integrated_debug_zip_v106
    def _make_real_integrated_debug_zip_v106(summary_image: Any, attack_image: Any) -> tuple[Path, dict[str, Any]]:  # type: ignore[override]
        zpath, context = _make_real_integrated_debug_zip_prev_v120(summary_image, attack_image)
        try:
            import zipfile as _zipfile_v120
            import json as _json_v120
            api_timing = st.session_state.get("api_last_timing_v120") or st.session_state.get("api_last_timing_v119") or st.session_state.get("api_last_timing_v117") or {}
            ui_env = {
                "version": "v120_static_first_ui_browser",
                "LOA_UI_FAST_MODE": _os_v120_app.environ.get("LOA_UI_FAST_MODE"),
                "top_nav": "streamlit_tabs_restored",
                "table_render": "static_html_first",
                "data_editor": "only_when_checkbox_on",
                "uploaded_image_decode": "cache_data_bytes_from_v119",
                "api_cache": "off",
            }
            tmp = Path(str(zpath) + ".tmp")
            with _zipfile_v120.ZipFile(zpath, "r") as zin, _zipfile_v120.ZipFile(tmp, "w", compression=_zipfile_v120.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    zout.writestr(item, zin.read(item.filename))
                zout.writestr("context/api_timing_v120.json", _json_v120.dumps(api_timing, ensure_ascii=False, indent=2, default=str))
                zout.writestr("context/ui_speed_v120.json", _json_v120.dumps(ui_env, ensure_ascii=False, indent=2, default=str))
            tmp.replace(zpath)
            context["api_timing_v120"] = api_timing
            context["ui_speed_v120"] = ui_env
        except Exception as e:
            try:
                context.setdefault("errors", []).append({"stage": "v120_append", "error": repr(e)})
            except Exception:
                pass
        return zpath, context
except Exception:
    pass



# ==============================================================================
# v121 UI 정리 패치
# - v120의 정적 HTML 표 방향은 유지하되, 사용자 요청에 맞춰 표 컬럼을 줄인다.
# - 피해군 출처 NameError를 방지하고, 기본 화면에서는 무거운 출처 표를 렌더하지 않는다.
# - Streamlit dataframe과 비슷한 어두운 정적 CSS 테이블로 통일한다.

APP_CALC_VERSION = "v121_clean_static_css"


def _v121_inject_css() -> None:
    st.markdown(
        """
        <style>
        .loa-v121-section{margin:22px 0 10px 0;}
        .loa-v121-section h3{font-size:1.35rem;margin:0 0 10px 0;color:#f7f8ff;letter-spacing:-.02em;}
        .loa-v121-help{color:#9ca3b5;font-size:12px;margin:-4px 0 8px 0;}
        .loa-v121-scroll{overflow:auto;border-radius:10px;border:1px solid #303545;background:#0f1219;margin:4px 0 14px 0;max-height:520px;}
        .loa-v121-full{max-height:none;overflow:visible;}
        table.loa-v121-table{border-collapse:collapse;width:100%;font-size:13px;color:#f5f7fb;table-layout:auto;}
        table.loa-v121-table thead th{position:sticky;top:0;background:#1b1f2a;color:#c7cedd;font-weight:700;text-align:left;padding:10px 12px;border-bottom:1px solid #363b4b;white-space:nowrap;z-index:1;}
        table.loa-v121-table tbody td{background:#11151d;color:#f5f7fb;padding:10px 12px;border-bottom:1px solid #2a2f3d;border-right:1px solid #252a36;white-space:nowrap;vertical-align:middle;}
        table.loa-v121-table tbody tr:nth-child(even) td{background:#151a23;}
        table.loa-v121-table tbody tr:hover td{background:#222838;}
        table.loa-v121-table td:first-child,table.loa-v121-table th:first-child{font-weight:700;}
        .loa-v121-empty{padding:14px 16px;border-radius:10px;border:1px solid #303545;background:#11151d;color:#aab0c0;margin:6px 0 14px 0;}
        .loa-v121-grid{display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:10px;margin:8px 0 18px 0;}
        .loa-v121-card{background:#151a23;border:1px solid #303545;border-radius:12px;padding:14px 16px;min-height:76px;}
        .loa-v121-card .k{color:#aab0c0;font-size:12px;margin-bottom:6px;}
        .loa-v121-card .v{color:#ffffff;font-weight:800;font-size:22px;letter-spacing:-.01em;}
        .loa-v121-caption{font-size:12px;color:#9ca3b5;margin:4px 0 12px 2px;}
        @media (max-width: 900px){.loa-v121-grid{grid-template-columns:repeat(2,minmax(0,1fr));}}
        </style>
        """,
        unsafe_allow_html=True,
    )


def _v121_as_df(obj: Any) -> pd.DataFrame:
    try:
        if isinstance(obj, pd.DataFrame):
            return obj.copy()
        if isinstance(obj, list):
            return pd.DataFrame(obj)
        if isinstance(obj, dict):
            return pd.DataFrame(obj)
    except Exception:
        pass
    return pd.DataFrame()


def _v121_pick_col(df: pd.DataFrame, candidates: list[str]) -> str | None:
    for c in candidates:
        if c in df.columns:
            return c
    # 공백/괄호가 조금 달라도 최대한 잡기
    norm = {str(c).replace(" ", "").replace("％", "%"): c for c in df.columns}
    for c in candidates:
        key = c.replace(" ", "").replace("％", "%")
        if key in norm:
            return norm[key]
    return None


def _v121_fmt(v: Any) -> str:
    try:
        if v is None or pd.isna(v):
            return "-"
    except Exception:
        if v is None:
            return "-"
    if isinstance(v, float):
        return (f"{v:,.2f}" if abs(v) >= 1000 else f"{v:.2f}").rstrip("0").rstrip(".")
    return str(v) if str(v) != "" else "-"


def _v121_table_html(df: pd.DataFrame, max_rows: int = 80, class_name: str = "loa-v121-table", full_height: bool = False) -> str:
    if not isinstance(df, pd.DataFrame) or df.empty:
        return '<div class="loa-v121-empty">표 데이터가 없습니다.</div>'
    view = df.copy()
    view = view.drop(columns=[c for c in view.columns if str(c).startswith("_")], errors="ignore")
    if len(view) > max_rows:
        view = view.head(max_rows)
    th = "".join(f"<th>{_html_v120_app.escape(str(c))}</th>" for c in view.columns)
    rows = []
    for _, row in view.iterrows():
        rows.append("<tr>" + "".join(f"<td>{_html_v120_app.escape(_v121_fmt(row.get(c)))}</td>" for c in view.columns) + "</tr>")
    # full_height: 스크롤 없이 전체 행을 한 번에 표시합니다(최종 계산표 등).
    wrap_cls = "loa-v121-scroll loa-v121-full" if full_height else "loa-v121-scroll"
    return f'<div class="{wrap_cls}"><table class="{class_name}"><thead><tr>{th}</tr></thead><tbody>{"".join(rows)}</tbody></table></div>'


def _v121_render_table(title: str, df: pd.DataFrame, max_rows: int = 80, help_text: str = "", full_height: bool = False) -> None:
    st.markdown(f'<div class="loa-v121-section"><h3>{_html_v120_app.escape(title)}</h3></div>', unsafe_allow_html=True)
    if help_text:
        st.markdown(f'<div class="loa-v121-help">{_html_v120_app.escape(help_text)}</div>', unsafe_allow_html=True)
    st.markdown(_v121_table_html(df, max_rows=max_rows, full_height=full_height), unsafe_allow_html=True)


def _v155_source_table(summary: dict[str, Any], key: str, max_cols: int = 8) -> pd.DataFrame:
    df = _v121_as_df(summary.get(key))
    if df.empty:
        return df
    drop_words = ["아이콘", "Icon", "icon", "Tooltip", "tooltip", "툴팁", "설명"]
    out = df.drop(columns=[c for c in df.columns if any(w in str(c) for w in drop_words)], errors="ignore").copy()
    if len(out.columns) > max_cols:
        out = out.iloc[:, :max_cols]
    return out


def _v155_render_character_source_tables(summary: dict[str, Any]) -> None:
    sections = [
        ("스탯", "stats", 12, 6),
        ("장비", "equipment", 20, 7),
        ("전투 스킬", "skills", 40, 8),
        ("각인", "engravings", 20, 6),
        ("카드", "cards", 20, 6),
        ("보석", "gems", 20, 6),
        ("아크패시브", "arkpassive", 40, 6),
    ]
    rendered = False
    for title, key, max_rows, max_cols in sections:
        df = _v155_source_table(summary, key, max_cols=max_cols)
        if isinstance(df, pd.DataFrame) and not df.empty:
            _v121_render_table(title, df, max_rows)
            rendered = True
    if not rendered:
        _v121_render_table("전투/캐릭터 요약", _v121_combat_overview(summary), 80)


def _v155_pick_col(df: pd.DataFrame, names: list[str]) -> str | None:
    for name in names:
        if name in df.columns:
            return name
    lowered = {str(c).lower(): c for c in df.columns}
    for name in names:
        hit = lowered.get(str(name).lower())
        if hit is not None:
            return hit
    return None


def _v155_float(value: Any, default: float = 0.0) -> float:
    try:
        text = re.sub(r"[^0-9.\-]", "", str(value or ""))
        return float(text) if text else default
    except Exception:
        return default


def _v155_fast_preview_stats(summary: dict[str, Any]) -> tuple[float, float]:
    stats = _v121_as_df(summary.get("stats"))
    crit_stat = 0.0
    if not stats.empty:
        type_col = _v155_pick_col(stats, ["스탯", "Type", "type"])
        value_col = _v155_pick_col(stats, ["값", "Value", "value"])
        if (not type_col or not value_col) and len(stats.columns) >= 2:
            type_col = type_col or stats.columns[0]
            value_col = value_col or stats.columns[1]
        if type_col and value_col:
            for _, row in stats.iterrows():
                if "치명" in str(row.get(type_col, "")):
                    crit_stat = _v155_float(row.get(value_col), 0.0)
                    break
    stat_rate = crit_stat * 0.03579
    return crit_stat, stat_rate


def _v155_apply_fast_setting_preview(summary: dict[str, Any]) -> None:
    final_df = _v121_as_df(summary.get("arkgrid_final_skill_estimates"))
    if final_df.empty:
        final_df = _v121_as_df(summary.get("skill_crit_estimates"))
    if not final_df.empty:
        return

    crit_stat, stat_rate = _v155_fast_preview_stats(summary)
    skills = _v121_adopted_skills(summary)
    if skills.empty:
        skills = _v121_as_df(summary.get("skills"))

    name_col = _v155_pick_col(skills, ["스킬명", "이름", "Name", "name"])
    level_col = _v155_pick_col(skills, ["스킬레벨", "레벨", "Level", "level"])
    if not name_col and len(skills.columns) >= 1:
        name_col = skills.columns[0]
    if not level_col and len(skills.columns) >= 2:
        level_col = skills.columns[1]
    rows: list[dict[str, Any]] = []
    if name_col:
        for _, row in skills.head(24).iterrows():
            name = str(row.get(name_col) or "").strip()
            if not name:
                continue
            rows.append({
                "스킬명": name,
                "스킬레벨": row.get(level_col) if level_col else "",
                "예상 치명 확률": round(stat_rate, 2),
                "예상 치피": 200,
                "진화형 피해": 0,
            })

    summary["skill_crit_estimates"] = pd.DataFrame(rows)
    summary["combat_overview"] = pd.DataFrame([
        {"항목": "치명 스탯", "아크그리드 제외 기준": round(crit_stat, 2), "아크그리드 포함 최종": round(crit_stat, 2)},
        {"항목": "치명 스탯 치적", "아크그리드 제외 기준": f"{stat_rate:.2f}%", "아크그리드 포함 최종": f"{stat_rate:.2f}%"},
        {"항목": "백어택 스킬 기준 치명", "아크그리드 제외 기준": f"{stat_rate:.2f}%", "아크그리드 포함 최종": f"{stat_rate:.2f}%"},
        {"항목": "전역 치명타 피해량(스킬 전용 제외/방향성 기준)", "아크그리드 제외 기준": "200.00%", "아크그리드 포함 최종": "200.00%"},
        {"항목": "평균 치명타 피해량(스킬 전용 포함)", "아크그리드 제외 기준": "200.00%", "아크그리드 포함 최종": "200.00%"},
        {"항목": "평균 진화형 피해", "아크그리드 제외 기준": "-", "아크그리드 포함 최종": "-"},
        {"항목": "전역 적피 합계", "아크그리드 제외 기준": "-", "아크그리드 포함 최종": "-"},
        {"항목": "스킬 포함 적피 평균", "아크그리드 제외 기준": "-", "아크그리드 포함 최종": "-"},
        {"항목": "평균 스킬 피해", "아크그리드 제외 기준": "-", "아크그리드 포함 최종": "-"},
        {"항목": "평균 보석 피해(전체 스킬)", "아크그리드 제외 기준": "-", "아크그리드 포함 최종": "-"},
        {"항목": "평균 예상 최종 배율", "아크그리드 제외 기준": "-", "아크그리드 포함 최종": "-"},
        {"항목": "기습 각인 감지", "아크그리드 제외 기준": "-", "아크그리드 포함 최종": "-"},
        {"항목": "결투 각인 감지", "아크그리드 제외 기준": "-", "아크그리드 포함 최종": "-"},
    ])
    summary["_fast_setting_preview_v155"] = True


def _v121_cards(items: list[tuple[str, Any]]) -> None:
    html_cards = []
    for k, v in items:
        html_cards.append(f'<div class="loa-v121-card"><div class="k">{_html_v120_app.escape(str(k))}</div><div class="v">{_html_v120_app.escape(_v121_fmt(v))}</div></div>')
    st.markdown(f'<div class="loa-v121-grid">{"".join(html_cards)}</div>', unsafe_allow_html=True)


def _v121_combat_overview(summary: dict[str, Any]) -> pd.DataFrame:
    df = _v121_as_df(summary.get("combat_overview"))
    if df.empty:
        return df
    keep = [c for c in ["항목", "아크그리드 제외 기준", "아크그리드 포함 최종"] if c in df.columns]
    out = df[keep].copy() if keep else df
    if "항목" not in out.columns:
        return out
    wanted = [
        "치명 스탯",
        "치명 스탯 치적",
        "백어택 스킬 기준 치명",
        "전역 치명타 피해량(스킬 전용 제외/방향성 기준)",
        "평균 치명타 피해량(스킬 전용 포함)",
        "평균 진화형 피해",
        "전역 적피 합계",
        "스킬 포함 적피 평균",
        "평균 스킬 피해",
        "평균 보석 피해(전체 스킬)",
        "평균 예상 최종 배율",
        "기습 각인 감지",
        "결투 각인 감지",
    ]
    by_name = {str(row.get("항목")): row for _, row in out.iterrows()}
    rows = []
    for name in wanted:
        if name in by_name:
            rows.append(dict(by_name[name]))
        else:
            rows.append({"항목": name, "아크그리드 제외 기준": "-", "아크그리드 포함 최종": "-"})
    return pd.DataFrame(rows)


def _v121_breakdown(summary: dict[str, Any]) -> pd.DataFrame:
    df = _v121_as_df(summary.get("loawa_like_breakdown"))
    if df.empty:
        return df
    keep = [c for c in ["피해군", "합계(%)"] if c in df.columns]
    out = df[keep].copy() if keep else df.iloc[:, :2].copy()
    if "합계(%)" in out.columns:
        out = out.rename(columns={"합계(%)": "합계"})
    return out


def _v121_adopted_skills(summary: dict[str, Any]) -> pd.DataFrame:
    awakening = set(summary.get("awakening_skill_names") or [])
    try:
        df = filter_adopted_skills_df(summary.get("skills"), include_names=awakening)
    except Exception:
        df = _v121_as_df(summary.get("skills"))
    if df.empty:
        return df
    out = df.drop(columns=[c for c in df.columns if str(c) in {"아이콘", "icon", "Icon", "스킬 아이콘"}], errors="ignore").copy()
    # 기본타입(조작 타입) 대신 '공격 방향'(백어택/헤드어택/무방향)을 보여줍니다.
    # 방향은 API 판정 + 방향 보정 파일까지 반영한 최종값을 사용합니다.
    try:
        _dir_lookup = _build_api_skill_lookup()
    except Exception:
        _dir_lookup = {}
    _name_col = "스킬명" if "스킬명" in out.columns else ("name" if "name" in out.columns else None)
    if _name_col:
        out["공격 방향"] = out[_name_col].map(
            lambda nm: (_dir_lookup.get(normalize_name(str(nm)), {}) or {}).get("공격타입") or "없음"
        )
    # 채용 스킬 표는 중요한 기본 정보만 먼저 보여준다. (기본타입 → 공격 방향으로 대체)
    preferred = ["스킬명", "레벨", "스킬레벨", "공격 방향", "타입", "타입변경근거", "룬", "트포"]
    cols = [c for c in preferred if c in out.columns]
    if cols:
        out = out[cols]
    return out


def _v121_final_calc(summary: dict[str, Any]) -> pd.DataFrame:
    df = _v121_as_df(summary.get("arkgrid_final_skill_estimates"))
    if df.empty:
        df = _v121_as_df(summary.get("skill_crit_estimates"))
    if df.empty:
        return df
    name_col = _v121_pick_col(df, ["스킬명", "name"])
    level_col = _v121_pick_col(df, ["스킬레벨", "레벨"])
    crit_col = _v121_pick_col(df, [
        "예상 치명 확률",
        "예상 치명 확률(백어택 기준)(%)",
        "예상 치명 확률(조건부 포함)(%)",
        "치명 증가 합계(조건부)(%)",
        "기준 치명(백어택)(%)",
    ])
    crit_dmg_col = _v121_pick_col(df, [
        "예상 치피",
        "예상 치피(조건부 포함)(%)",
        "예상 치피(상시)(%)",
        "기준 치피(%)",
    ])
    evo_col = _v121_pick_col(df, [
        "진화형 피해",
        "진화형 피해(조건부)(%)",
        "기준 진화형 피해(%)",
        "진화형 피해(%)",
        "예상 진화형 피해(조건부)(%)",
    ])
    cols = []
    data = pd.DataFrame()
    if name_col:
        data["스킬명"] = df[name_col]
    if level_col:
        data["스킬레벨"] = df[level_col]
    if crit_col:
        data["예상 치명 확률"] = df[crit_col]
    if crit_dmg_col:
        data["예상 치피"] = df[crit_dmg_col]
    if evo_col:
        data["진화형 피해"] = df[evo_col]
    # 없으면 빈 DF가 아니라 원본 앞 컬럼이라도 보여준다.
    if data.empty:
        return df.iloc[:, :min(5, len(df.columns))].copy()
    return data


def show_damage_source_tab(summary: dict[str, Any]) -> None:  # type: ignore[override]
    """v121: 이전 버전에서 함수가 누락되어 NameError가 나던 피해군 출처 탭 안전 처리."""
    st.info("피해군 출처 상세 표는 기본 화면에서 제외했습니다.")


def _render_icon_recognition_test_v144() -> None:
    """직업별로 아이콘·이름 인식이 잘 되는지 확인하는 테스트 뷰.

    이미지 인식이 끝난 뒤, 실전/허수 검수표에 들어간 스킬명과 값을 그대로 보여줘서
    '글자(이름)와 값이 제대로 들어갔는지'를 눈으로 빠르게 확인할 수 있습니다.
    """
    with st.expander("🎯 아이콘·이름 인식 테스트 (직업별 확인용)", expanded=False):
        prof = (st.session_state.get("api_summary") or {}).get("profile_summary", {}) if isinstance(st.session_state.get("api_summary"), dict) else {}
        _cls = prof.get("클래스") or "-"
        _job = prof.get("직업") or "-"
        st.caption(f"현재 클래스: **{_cls}** · 직업: **{_job}** · 공격정보 이미지를 올리면 서버에서 아이콘·이름을 인식해 보여줍니다. "
                   f"직업별로 아이콘 매칭과 스킬명이 잘 들어가는지 여기서 바로 확인하세요.")

        # 1) 이미지 업로드 테스트
        st.markdown("**이미지 올려서 인식 테스트**")
        up = st.file_uploader("공격정보 이미지 (백/헤드/치명 열이 보이는 전투분석기 화면)",
                              type=["png", "jpg", "jpeg", "webp"], key="icon_test_upload_v144")
        if up is not None:
            try:
                test_img = Image.open(io.BytesIO(up.getvalue())).convert("RGB")
                st.image(test_img, caption="업로드한 이미지", use_container_width=True)
                with st.spinner("서버에서 아이콘·이름 인식 중... (첫 실행은 인식 모델 준비로 조금 걸릴 수 있어요)"):
                    _rep, _zip, parsed = _run_ocr_perf_timeline_debug(
                        "icon_test", "아이콘 인식 테스트", None, test_img,
                        gpu=False, row_count=14, ocr_scale=7,
                    )
                if isinstance(parsed, pd.DataFrame) and not parsed.empty:
                    # 전투분석기 위→아래 순서로 정렬
                    parsed_sorted = parsed.copy()
                    if "_ocr_row_index" in parsed_sorted.columns:
                        parsed_sorted["_sort_idx"] = pd.to_numeric(parsed_sorted["_ocr_row_index"], errors="coerce")
                        parsed_sorted = parsed_sorted.sort_values("_sort_idx", na_position="last").drop(columns=["_sort_idx"])
                    pref = ["_ocr_row_index", "이름", "_icon_match_name", "_icon_match_score", "_name_match_reason",
                            "OCR 원본 이름", "피해량", "치명타 적중률", "사용 횟수"]
                    cols = [c for c in pref if c in parsed_sorted.columns]
                    view = parsed_sorted[cols].copy() if cols else parsed_sorted.copy()
                    ren = {"_ocr_row_index": "순번(위→아래)", "_icon_match_name": "아이콘 매칭명", "_icon_match_score": "아이콘 점수", "_name_match_reason": "매칭 방식"}
                    view = view.rename(columns={k: v for k, v in ren.items() if k in view.columns})
                    st.dataframe(view, use_container_width=True, hide_index=True)
                    st.caption(f"인식된 행 {len(view)}개 · 위→아래 순서 · ‘이름’이 실제 매칭된 스킬명, ‘아이콘 매칭명/점수’로 아이콘 인식 품질을 확인하세요.")
                    # 이 업로드 이미지에 대한 아이콘 매칭 전용 디버그 ZIP(클래스명_직업 파일명)
                    if st.button("🎯 이 이미지 아이콘 매칭 디버그 ZIP 생성", key="icon_test_make_debug_v149", use_container_width=True):
                        with st.spinner("아이콘 매칭 디버그 ZIP 생성 중..."):
                            try:
                                _zpi = _make_icon_match_debug_zip_v149(test_img)
                            except Exception as _e:  # noqa: BLE001
                                _zpi = None
                                st.error(f"디버그 생성 오류: {_e}")
                        if _zpi is not None:
                            st.session_state["icon_test_debug_zip_v149"] = str(_zpi)
                            st.success("생성했습니다. 아래에서 받아 보내주세요.")
                    _zpi2 = st.session_state.get("icon_test_debug_zip_v149")
                    if _zpi2 and Path(_zpi2).exists():
                        st.download_button(
                            "아이콘 매칭 디버그 ZIP 다운로드",
                            data=Path(_zpi2).read_bytes(),
                            file_name=Path(_zpi2).name,
                            mime="application/zip",
                            key="icon_test_debug_zip_download_v149",
                            use_container_width=True,
                        )
                else:
                    st.warning("인식 결과가 비었어요. 이미지가 공격정보 화면인지(백/헤드/치명 열이 보이는지) 확인해주세요.")
            except Exception as e:  # noqa: BLE001
                st.error(f"서버 인식을 실행하지 못했어요: {e}\n"
                         f"이 환경에 서버 인식 라이브러리가 없을 수 있습니다. 그럴 땐 아래의 화면공유 인식 결과로 확인하세요.")

        st.divider()
        # 2) 마지막 화면공유 인식 결과(참고)
        st.markdown("**최근 화면공유 인식 결과(참고)**")
        show_cols = ["이름", "피해량", "치명타 적중률", "사용 횟수"]
        c1, c2 = st.columns(2)
        for col, kind, label in ((c1, "real", "실전"), (c2, "dummy", "허수/기준")):
            with col:
                st.markdown(f"*{label}*")
                df = st.session_state.get(f"{kind}_table")
                if isinstance(df, pd.DataFrame) and not df.empty:
                    cols = [c for c in show_cols if c in df.columns] or list(df.columns[:4])
                    st.dataframe(df[cols], use_container_width=True, hide_index=True)
                else:
                    st.info("아직 인식 결과가 없습니다.")


def _render_character_skill_info_v144(summary: dict[str, Any], bundle: dict[str, Any] | None = None) -> None:
    """캐릭터 검색 후, 스킬명·방향·타입·레벨·룬·선택 트라이포드를 한 표로 보여줍니다.

    방향 보정 파일(skill_direction_overrides.xlsx)을 만들 때 참고하기 좋습니다.
    """
    df = summary.get("skill_crit_estimates")
    if not isinstance(df, pd.DataFrame) or df.empty:
        df = summary.get("lostbuilds_base_skill_estimates")
    with st.expander("🗂️ 스킬 정보 보기 (이름 · 방향 · 타입 · 트라이포드)", expanded=False):
        if not isinstance(df, pd.DataFrame) or df.empty or "스킬명" not in df.columns:
            st.info("캐릭터를 검색하면 스킬 목록이 여기에 표시됩니다.")
            return
        st.caption("이 캐릭터의 스킬과 방향(백/헤드/무방향)입니다. ‘방향(최종)’은 수동 보정 파일까지 반영한 값이고, "
                   "‘API 방향’은 API 원본 판정입니다. 방향을 바꾸려면 data/skill_direction_overrides.xlsx를 수정하세요.")
        try:
            final = _build_api_skill_lookup()
        except Exception:
            final = {}
        try:
            trips = _selected_tripods_by_skill(bundle or {})
        except Exception:
            trips = {}
        base_type_lookup: dict[str, str] = {}
        skills_df = summary.get("skills")
        if isinstance(skills_df, pd.DataFrame) and not skills_df.empty and "스킬명" in skills_df.columns:
            _bt_col = next((c for c in ["기본타입", "타입", "SkillType"] if c in skills_df.columns), None)
            if _bt_col:
                for _, s in skills_df.iterrows():
                    base_type_lookup[normalize_name(str(s.get("스킬명")))] = str(s.get(_bt_col) or "")
        rows = []
        for _, r in df.iterrows():
            nm = str(r.get("스킬명") or "").strip()
            if not nm:
                continue
            nkey = normalize_name(nm)
            api_dir = _direction_type_from_attack_type(r.get("공격타입"))
            fin_dir = final.get(nkey, {}).get("공격타입") or api_dir
            rows.append({
                "스킬명": nm,
                "방향(최종)": fin_dir,
                "API 방향": api_dir,
                "기본타입": base_type_lookup.get(nkey, ""),
                "스킬레벨": r.get("스킬레벨") or r.get("레벨") or "",
                "룬": r.get("룬") or "",
                "선택 트라이포드": ", ".join(sorted(trips.get(nkey, set()))),
            })
        if rows:
            st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)
            st.caption(f"스킬 {len(rows)}개 · 표 우측 상단 아이콘으로 CSV로 받을 수 있어요.")
        else:
            st.info("표시할 스킬이 없습니다.")


def _render_api_skill_direction_debug_v144(summary: dict[str, Any]) -> None:
    """API가 각 스킬에 방향(백어택/헤드어택)을 주는지, 그리고 전투분석기 스킬 중
    API에 아예 없는(=매칭 실패) 스킬이 무엇인지 보여줍니다.

    아이덴티티/아덴 스킬(예: 슬레이어 블러드러스트)이 API combat-skills 목록에 없으면
    여기서 '전투분석기에만 있고 API에는 없음'으로 잡히고, 방향 보정이 안 들어갑니다.
    """
    with st.expander("🧭 스킬 방향 디버그 — 각 스킬 방향이 어디서 왔나(API/파일)", expanded=False):
        st.caption("전투분석기 스킬별로 최종 방향과 그 출처를 보여줍니다. "
                   "출처가 '수동 보정 파일'이면 data/skill_direction_overrides.xlsx에서 온 것이고, "
                   "'API'면 API 스킬 툴팁에서, '없음(미지정)'이면 API에도 없고 파일에도 없어 방향이 안 붙은 것입니다. "
                   "방향을 바꾸거나 추가하려면 그 파일의 ‘방향보정’ 시트를 수정하세요.")
        # API 스킬 방향(파일 보정 적용 전) 집합
        df = summary.get("skill_crit_estimates")
        if not isinstance(df, pd.DataFrame) or df.empty:
            df = summary.get("lostbuilds_base_skill_estimates")
        api_dir: dict[str, str] = {}
        if isinstance(df, pd.DataFrame) and not df.empty and "스킬명" in df.columns and "공격타입" in df.columns:
            for _, r in df.iterrows():
                api_dir[normalize_name(str(r.get("스킬명")))] = _direction_type_from_attack_type(r.get("공격타입"))
        overrides = _load_skill_direction_overrides()  # {nkey: {raw, name}}
        final_lookup = _build_api_skill_lookup()        # 파일 보정 반영된 최종

        rows = []
        seen: set[str] = set()
        for kind, label in (("real", "실전"), ("dummy", "허수")):
            bt = st.session_state.get(f"{kind}_table")
            if not isinstance(bt, pd.DataFrame) or bt.empty or "이름" not in bt.columns:
                continue
            for nm in bt["이름"].tolist():
                nm = str(nm).strip()
                if not nm or _is_rune_or_misc(nm):
                    continue
                nkey = normalize_name(nm)
                if nkey in seen:
                    continue
                seen.add(nkey)
                _api = api_dir.get(nkey, "—(API에 없음)")
                _ov = overrides.get(nkey)
                _final = final_lookup.get(nkey, {}).get("공격타입") or "없음"
                if _ov:
                    src = "수동 보정 파일"
                elif nkey in api_dir and _api not in ("없음", "—(API에 없음)"):
                    src = "API"
                elif nkey in api_dir:
                    src = "API(방향 없음)"
                else:
                    src = "없음(미지정)"
                rows.append({"스킬명": nm, "API 방향": _api, "파일 보정": (_ov or {}).get("raw", "-"), "최종 방향": _final, "출처": src})
        if rows:
            st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)
            st.caption("‘없음(미지정)’ 스킬을 방향 지정하려면 skill_direction_overrides.xlsx에 스킬명·공격타입을 추가하세요.")
        else:
            st.info("전투분석기 인식 결과가 없습니다. 먼저 이미지 인식을 실행하세요.")


def _render_crit_sum_debug_v144(summary: dict[str, Any]) -> None:
    """치명타 적중률이 어디서 얼마나 더해졌는지 출처별로 보여주는 디버그.

    동일 스펙 반지 2개처럼 같은 값이 여러 개여도 각각 집계되는지 눈으로 검증할 수 있습니다.
    """
    src = summary.get("damage_sources")
    with st.expander("🔎 치명 합산 디버그 — 치명타 적중률이 어디서 더해졌나", expanded=False):
        st.caption("각 출처(각인/장비/아크패시브 등)가 더해준 치명타 적중률을 나열합니다. "
                   "같은 이름이 (2),(3)으로 보이면 동일 스펙 장비가 각각 계산된 것입니다.")
        if not isinstance(src, pd.DataFrame) or src.empty:
            st.info("치명 출처 데이터가 아직 없습니다. 먼저 캐릭터를 검색하세요.")
            return
        cols = [c for c in ["출처구분", "이름", "적용범위", "치명타 적중률 증가(%)", "치명타 적중률 고정(%)", "조건부 여부", "설명"] if c in src.columns]
        view = src[cols].copy()
        inc = pd.to_numeric(view.get("치명타 적중률 증가(%)"), errors="coerce").fillna(0.0)
        fix = pd.to_numeric(view.get("치명타 적중률 고정(%)"), errors="coerce").fillna(0.0) if "치명타 적중률 고정(%)" in view.columns else 0.0
        contrib = view[(inc != 0) | (fix != 0)].copy()
        stat_rate = float(summary.get("base_crit_percent", 0) or 0)
        st.markdown(f"**치명 스탯 치적:** {stat_rate:.2f}%  ·  "
                    f"**전체 범위 치적 합:** {float(contrib[contrib['적용범위'].astype(str).isin(['전체','전체/범위 기준'])]['치명타 적중률 증가(%)'].pipe(pd.to_numeric, errors='coerce').fillna(0).sum()):.2f}%  ·  "
                    f"**기본 기준 치명:** {float(summary.get('global_basis_crit_rate_percent', 0) or 0):.2f}%")
        if contrib.empty:
            st.info("치명타 적중률을 더해주는 출처가 없습니다(스탯 치적만 사용).")
        else:
            st.dataframe(contrib, use_container_width=True, hide_index=True)


def _render_synergy_dupe_debug_v144(summary: dict[str, Any]) -> None:
    """시너지(치명타 적중률·피해 계수 등)가 중복 적용됐는지 점검하는 디버그.

    같은 (출처구분·이름·적용범위·설명) 시너지가 2번 이상 잡혀 값이 두 배로 들어가는지,
    치명타 적중률 합이 물리적 상한(100%)을 넘는지 등을 표시합니다.
    """
    src = summary.get("damage_sources")
    with st.expander("🔁 시너지 중복 점검 디버그", expanded=False):
        st.caption("같은 시너지가 두 번 잡혀 값이 중복 합산되는지 점검합니다. "
                   "장비처럼 (2),(3)로 구분된 동일 스펙은 정상 중복이며, 그 외에 완전히 동일한 출처가 여러 번 나오면 의심 대상입니다.")
        if not isinstance(src, pd.DataFrame) or src.empty:
            st.info("데이터가 아직 없습니다.")
            return
        # 감지된 시너지 트라이포드(급소노출 등)를 먼저 보여줍니다.
        # 이름이 '시너지:'로 시작하는 행이 v65 시너지 판정으로 잡힌 것들입니다.
        st.markdown("**감지된 시너지 (전체 적용)**")
        _syn = src[src["이름"].astype(str).str.startswith("시너지:")].copy() if "이름" in src.columns else pd.DataFrame()
        if _syn.empty:
            st.warning("이 캐릭터에서 감지된 시너지 트라이포드가 없습니다. "
                       "급소노출 같은 시너지 트포를 채용했는데도 비어 있으면, 트포가 '선택' 상태인지 확인하세요.")
        else:
            _syn_cols = [c for c in ["이름", "적용범위", "치명타 적중률 증가(%)", "치명타 피해량 증가(%)", "적에게 주는 피해(%)", "공격력 증가(%)", "설명"] if c in _syn.columns]
            st.dataframe(_syn[_syn_cols], use_container_width=True, hide_index=True)
        st.divider()
        key_cols = [c for c in ["출처구분", "이름", "적용범위", "조건부 여부", "설명"] if c in src.columns]
        eff_cols = [c for c in src.columns if c.endswith("(%)")]
        grp = src.groupby(key_cols, dropna=False).size().reset_index(name="중복수")
        dupes = grp[grp["중복수"] >= 2].copy()
        # 장비 동일 스펙((2),(3))은 정상이므로, 이름에 "(숫자)"가 붙은 케이스는 제외하고 표시합니다.
        if not dupes.empty and "이름" in dupes.columns:
            _is_slot = dupes["이름"].astype(str).str.contains(r"\(\d+\)$", regex=True)
            dupes = dupes[~_is_slot]
        if dupes.empty:
            st.success("완전히 동일한 시너지가 2번 이상 잡힌 케이스는 없습니다.")
        else:
            st.warning("아래 출처가 중복으로 잡혔습니다. 시너지 이중 적용일 수 있으니 확인하세요.")
            st.dataframe(dupes, use_container_width=True, hide_index=True)
        # 치명타 적중률 총합(전체 범위)이 100%를 넘는지 확인.
        if "치명타 적중률 증가(%)" in src.columns:
            all_scope = src[src["적용범위"].astype(str).isin(["전체", "전체/범위 기준"])]
            tot = float(pd.to_numeric(all_scope["치명타 적중률 증가(%)"], errors="coerce").fillna(0).sum())
            stat_rate = float(summary.get("base_crit_percent", 0) or 0)
            grand = stat_rate + tot
            st.markdown(f"**전체 범위 치명타 적중률 합(스탯 포함):** {grand:.2f}%")
            if grand > 100.0:
                st.warning("전체 치명타 적중률이 100%를 넘습니다. 시너지가 중복 적용됐을 가능성이 있어요.")


def api_tab() -> None:  # type: ignore[override]
    t0 = _time_v120_app.perf_counter()
    st.header("캐릭터 세팅 확인")
    st.caption("캐릭터 세팅과 계산에 사용된 값을 한눈에 확인할 수 있습니다.")
    summary = st.session_state.get("api_summary") or {}
    if not summary:
        st.info("왼쪽에서 API Key와 캐릭터명을 넣고 검색하면 여기에 표시됩니다.")
        st.markdown('<div class="loa-v120-placeholder"></div><div class="loa-v120-placeholder"></div><div class="loa-v120-placeholder"></div>', unsafe_allow_html=True)
        return
    if isinstance(summary, dict) and summary.get("estimator_error"):
        st.error(f"계산표 생성 오류: {summary.get('estimator_error')}")

    profile = summary.get("profile_summary", {}) if isinstance(summary, dict) else {}
    _v121_cards([
        ("캐릭터", profile.get("캐릭터명") or st.session_state.get("character_name", "-") or "-"),
        ("클래스", profile.get("클래스") or profile.get("클래스명") or "-"),
        ("직업", profile.get("직업") or "-"),
        ("아이템 레벨", profile.get("아이템레벨") or "-"),
    ])
    timing = st.session_state.get("api_last_timing_v120") or st.session_state.get("api_last_timing_v119") or st.session_state.get("api_last_timing_v117") or {}
    if timing:
        _render_api_timing_caption_v117(timing)

    # 빠른 핵심 카드
    cols = st.columns(4)
    cols[0].metric("치명 스탯", f"{float(summary.get('base_crit_stat', 0) or 0):,.0f}")
    cols[1].metric("치명 스탯 치적", f"{float(summary.get('base_crit_percent', 0) or 0):.2f}%")
    cols[2].metric("기본 기준 치명", f"{float(summary.get('global_basis_crit_rate_percent', summary.get('avg_back_basis_crit_rate_percent', 0)) or 0):.2f}%")
    cols[3].metric("전역 치피", f"{float(summary.get('global_crit_damage_no_skill_percent', summary.get('avg_conditional_crit_damage_percent', 200)) or 200):.2f}%")

    # 표시 순서: 최종 계산표 → 전투/캐릭터 요약 → 효과 합산표 → 전투 스킬(채용 스킬만)
    _v121_render_table(
        "최종 계산표",
        _v121_final_calc(summary),
        1000,  # 스크롤 없이 전체 스킬을 한 번에 표시합니다.
        "예상 치명 확률은 백어택 기준/시너지/조건부를 포함한 최종 계산표 기준값입니다.",
        full_height=True,
    )
    _v121_render_table("전투/캐릭터 요약", _v121_combat_overview(summary), 80)
    _v121_render_table("효과 합산표", _v121_breakdown(summary), 40)
    _v121_render_table("전투 스킬 - 채용 스킬만", _v121_adopted_skills(summary), 60)

    # 캐릭터 검색 시 항상 볼 수 있는 스킬 정보(이름·방향·타입·트라이포드) 뷰
    # 스킬 정보 바로 밑에 이미지 업로드 인식 테스트.
    # 기본적으로 숨김. 신규 캐릭터 작업 시 '🔧 디버그 도구 표시'를 켜면 다시 나타납니다.
    if st.session_state.get("show_debug_tools", False):
        _render_character_skill_info_v144(summary, st.session_state.get("api_bundle"))
        _render_icon_recognition_test_v144()

    if st.session_state.get("show_debug_tools"):
        _render_api_skill_direction_debug_v144(summary)
        _render_crit_sum_debug_v144(summary)
        _render_synergy_dupe_debug_v144(summary)

    timing = st.session_state.get("api_last_timing_v120") or st.session_state.get("api_last_timing_v119") or st.session_state.get("api_last_timing_v117") or {}
    if isinstance(timing, dict):
        timing["api_tab_render_ms_v121"] = round((_time_v120_app.perf_counter() - t0) * 1000.0, 3)
        start = st.session_state.get("_search_click_t0_v120")
        if start:
            try:
                timing["search_to_api_tab_render_done_ms_v121"] = round((_time_v120_app.perf_counter() - float(start)) * 1000.0, 3)
            except Exception:
                pass
            st.session_state.api_last_timing_v120 = timing


def sidebar_controls() -> None:  # type: ignore[override]
    calc_config = load_yaml(str(CONFIG_DIR / "calculation_presets.yaml"))
    defaults = calc_config.get("defaults", {})
    st.session_state.setdefault("api_crit_basis", "백어택 기준(조건부 포함)")
    st.session_state.setdefault("target_crit_percent", float(defaults.get("target_crit_percent", 100.0)))
    st.session_state.setdefault("target_back_percent", float(defaults.get("target_back_attack_percent", 100.0)))
    st.session_state.setdefault("back_bonus_percent", float(defaults.get("back_attack_damage_bonus_percent", 5.0)))

    sidebar_t0 = _time_v120_app.perf_counter()
    with st.sidebar:
        st.markdown("### LOA 실전 분석")
        pass  # 버전 변경 안내 제거
        st.session_state["show_debug_tools"] = st.checkbox(
            "🔧 디버그 도구 표시",
            value=bool(st.session_state.get("show_debug_tools", False)),
            help="본문 OCR 디버그와 통합 디버그 ZIP을 표시합니다.",
        )
        with st.form("api_search_form_v121", clear_on_submit=False):
            token = st.text_input("Lost Ark API Key", type="password", key="api_token")
            character_name = st.text_input("캐릭터명", value=st.session_state.get("character_name", ""), key="character_name")
            st.number_input(
                "목표 상승률(%)",
                value=float(st.session_state.get("target_gain_percent", defaults.get("target_gain_percent", 5.0))),
                min_value=0.0,
                max_value=100.0,
                step=0.5,
                key="target_gain_percent",
            )
            fetch = st.form_submit_button("검색", type="primary", use_container_width=True)

        _render_integrated_debug_button_v106()

        if fetch:
            search_t0 = _time_v120_app.perf_counter()
            st.session_state["_search_click_t0_v120"] = search_t0
            if not token or not character_name:
                st.warning("API Key와 캐릭터명을 먼저 입력해줘.")
            else:
                with st.spinner("캐릭터 셋팅/계산표를 불러오는 중... (캐시 OFF)"):
                    result, timing = _fetch_api_bundle_timed_v117(token, character_name)
                    t_assign = _time_v120_app.perf_counter()
                    st.session_state.api_bundle = result["api_bundle"]
                    st.session_state.api_summary = result["api_summary"]
                    timing["state_assign_ms"] = round((_time_v120_app.perf_counter() - t_assign) * 1000.0, 3)
                    timing["search_button_wall_before_render_ms_v121"] = round((_time_v120_app.perf_counter() - search_t0) * 1000.0, 3)
                    timing["ui_mode_v121"] = "static_css_clean_tables"
                    st.session_state.api_last_timing_v120 = timing
                    st.session_state.api_last_timing_v119 = timing
                    st.session_state.api_last_timing_v117 = timing
                    st.session_state.api_last_timing_v116 = timing
                    st.session_state.api_last_timing_v114 = timing
            if token and character_name:
                failed = (st.session_state.get("api_last_timing_v120") or {}).get("failed") or {}
                if failed:
                    st.warning(f"일부 API 호출 실패: {', '.join(failed.keys())}")
                else:
                    st.success("캐릭터 셋팅을 불러왔어. CSS 정적표로 표시합니다.")

        timing = st.session_state.get("api_last_timing_v120") or st.session_state.get("api_last_timing_v119") or st.session_state.get("api_last_timing_v117") or {}
        if timing:
            try:
                timing["sidebar_render_ms_last_v121"] = round((_time_v120_app.perf_counter() - sidebar_t0) * 1000.0, 3)
            except Exception:
                pass
            _render_api_timing_caption_v117(timing)

        summary = st.session_state.get("api_summary") or {}
        profile = summary.get("profile_summary", {}) if isinstance(summary, dict) else {}
        if profile:
            st.divider()
            st.markdown(f"**{profile.get('캐릭터명') or st.session_state.get('character_name','')}**")
            st.caption(f"{profile.get('클래스명', profile.get('클래스', '-'))} · Lv.{profile.get('아이템레벨', '-')}")



try:
    _make_real_integrated_debug_zip_prev_v121 = _make_real_integrated_debug_zip_v106
    def _make_real_integrated_debug_zip_v106(summary_image: Any, attack_image: Any) -> tuple[Path, dict[str, Any]]:  # type: ignore[override]
        zpath, context = _make_real_integrated_debug_zip_prev_v121(summary_image, attack_image)
        try:
            import zipfile as _zipfile_v121
            import json as _json_v121
            api_timing = st.session_state.get("api_last_timing_v120") or st.session_state.get("api_last_timing_v119") or st.session_state.get("api_last_timing_v117") or {}
            ui_env = {
                "version": "v121_clean_static_css",
                "table_render": "static_css_clean_tables",
                "default_columns": {
                    "combat_overview": ["항목", "아크그리드 제외 기준", "아크그리드 포함 최종"],
                    "breakdown": ["피해군", "합계"],
                    "skills": "아이콘 열 제외",
                    "final_calc": ["스킬명", "스킬레벨", "예상 치명 확률", "예상 치피", "진화형 피해"],
                },
                "damage_source_tab": "safe_stub_no_name_error",
                "api_cache": "off",
            }
            tmp = Path(str(zpath) + ".tmp")
            with _zipfile_v121.ZipFile(zpath, "r") as zin, _zipfile_v121.ZipFile(tmp, "w", compression=_zipfile_v121.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    zout.writestr(item, zin.read(item.filename))
                zout.writestr("context/api_timing_v121.json", _json_v121.dumps(api_timing, ensure_ascii=False, indent=2, default=str))
                zout.writestr("context/ui_speed_v121.json", _json_v121.dumps(ui_env, ensure_ascii=False, indent=2, default=str))
            tmp.replace(zpath)
            context["api_timing_v121"] = api_timing
            context["ui_speed_v121"] = ui_env
        except Exception as e:
            try:
                context.setdefault("errors", []).append({"stage": "v121_append", "error": repr(e)})
            except Exception:
                pass
        return zpath, context
except Exception:
    pass


# if __name__ block moved to file end by v122

# ==============================================================================
# v122: analysis-only share mode UI
# ==============================================================================
APP_CALC_VERSION = "v122_analysis_share_fast"
import os as _os_v122_app
_os_v122_app.environ.setdefault("LOA_ANALYSIS_SHARE_ONLY", "1")
_os_v122_app.environ.setdefault("LOA_ATTACK_PERCENT_FALLBACK", "1")
_os_v122_app.environ.setdefault("LOA_ATTACK_CAST_COOLDOWN_FALLBACK", "1")
_os_v122_app.environ.setdefault("LOA_ATTACK_ROW_PASS_SCALE", "3")
_os_v122_app.environ.setdefault("LOA_ATTACK_DAMAGE_RECHECK", "smart")
_os_v122_app.environ.setdefault("LOA_ATTACK_CRITICAL_CELL_SCALE", "3")

# v134f: 백/헤드/치명 '적중률'도 검수표에 표시·편집하도록 포함합니다.
# (이전엔 비중만 표시 컬럼에 있어서, 에디터가 적중률을 빈 컬럼으로 되돌려 OCR로 읽은 값을 덮어썼습니다.)
_BATTLE_ANALYSIS_DISPLAY_COLS_V122 = [
    "이름", "피해량", "초당 피해량", "피해량 지분",
    "백어택 적중률", "백어택 비중", "헤드어택 적중률", "헤드어택 비중",
    "치명타 적중률", "치명타 비중", "사용 횟수", "쿨타임 비율",
]



_sidebar_controls_prev_v122 = globals().get("sidebar_controls")
def sidebar_controls() -> None:  # type: ignore[override]
    if callable(_sidebar_controls_prev_v122):
        _sidebar_controls_prev_v122()
    try:
        pass  # 버전 변경 안내 제거
        st.sidebar.caption("백적중/치적 셀 재OCR 생략 · 백비중/치비중/피해량 중심")
    except Exception:
        pass


try:
    _make_real_integrated_debug_zip_prev_v122 = _make_real_integrated_debug_zip_v106
    def _make_real_integrated_debug_zip_v106(summary_image: Any, attack_image: Any) -> tuple[Path, dict[str, Any]]:  # type: ignore[override]
        zpath, context = _make_real_integrated_debug_zip_prev_v122(summary_image, attack_image)
        try:
            import zipfile as _zipfile_v122
            import json as _json_v122
            api_timing = st.session_state.get("api_last_timing_v120") or st.session_state.get("api_last_timing_v119") or st.session_state.get("api_last_timing_v117") or {}
            speed_env = {
                "version": "v122_analysis_share_fast",
                "analysis_share_only": str(_os_v122_app.environ.get("LOA_ANALYSIS_SHARE_ONLY", "1")),
                "skipped_default_columns": ["백어택 적중률", "치명타 적중률"],
                "required_ocr_columns": ["피해량", "백어택 비중", "헤드어택 비중", "치명타 비중", "사용 횟수", "쿨타임 비율"],
                "reason": "analysis-only: shares reflect actual damage contribution; hit-rate columns are not required for current-result analysis",
            }
            tmp = Path(str(zpath) + ".tmp")
            with _zipfile_v122.ZipFile(zpath, "r") as zin, _zipfile_v122.ZipFile(tmp, "w", compression=_zipfile_v122.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    zout.writestr(item, zin.read(item.filename))
                zout.writestr("context/api_timing_v122.json", _json_v122.dumps(api_timing, ensure_ascii=False, indent=2, default=str))
                zout.writestr("context/speed_env_v122.json", _json_v122.dumps(speed_env, ensure_ascii=False, indent=2, default=str))
            tmp.replace(zpath)
            context["api_timing_v122"] = api_timing
            context["speed_env_v122"] = speed_env
        except Exception as e:
            try:
                context.setdefault("errors", []).append({"stage": "v122_append", "error": repr(e)})
            except Exception:
                pass
        return zpath, context
except Exception:
    pass


# ==============================================================================
# v123: CPU profiler + column-pass OCR diagnostics
# ==============================================================================
APP_CALC_VERSION = "v123_column_cpu_debug"
import os as _os_v123_app
_os_v123_app.environ.setdefault("LOA_ATTACK_COLUMN_PASS", "1")
_os_v123_app.environ.setdefault("LOA_ATTACK_COLUMN_PASS_SCALE", "2")
_os_v123_app.environ.setdefault("LOA_ATTACK_COLUMN_CELL_FALLBACK", "0")
_os_v123_app.environ.setdefault("LOA_WEB_LOW_CPU", "1")
_os_v123_app.environ.setdefault("LOA_CPU_PROFILE", "1")

class _CpuProfileV123:
    def __init__(self, label: str, interval: float = 0.20):
        self.label = label
        self.interval = max(0.05, float(interval or 0.20))
        self.samples = []
        self.error = None
        self._stop = None
        self._thread = None
        self._t0 = 0.0
        self._proc = None

    def __enter__(self):
        self._t0 = time.perf_counter()
        if str(_os_v123_app.environ.get("LOA_CPU_PROFILE", "1")).strip().lower() in {"0","false","no","off"}:
            return self
        try:
            import psutil  # type: ignore
            import threading
            self._proc = psutil.Process()
            self._proc.cpu_percent(interval=None)
            self._stop = threading.Event()
            def _loop():
                while self._stop is not None and not self._stop.wait(self.interval):
                    try:
                        with self._proc.oneshot():
                            self.samples.append({
                                "t_ms": round((time.perf_counter() - self._t0) * 1000.0, 3),
                                "process_cpu_pct": float(self._proc.cpu_percent(interval=None)),
                                "num_threads": int(self._proc.num_threads()),
                                "rss_mb": round(float(self._proc.memory_info().rss) / (1024 * 1024), 2),
                            })
                    except Exception:
                        pass
            self._thread = threading.Thread(target=_loop, name="loa_cpu_profiler_v123", daemon=True)
            self._thread.start()
        except Exception as e:
            self.error = repr(e)
        return self

    def __exit__(self, exc_type, exc, tb):
        try:
            if self._stop is not None:
                self._stop.set()
            if self._thread is not None:
                self._thread.join(timeout=1.0)
        except Exception:
            pass
        return False

    def summary(self) -> dict[str, Any]:
        vals = [float(x.get("process_cpu_pct") or 0.0) for x in self.samples]
        ths = [int(x.get("num_threads") or 0) for x in self.samples]
        rss = [float(x.get("rss_mb") or 0.0) for x in self.samples]
        wall = round((time.perf_counter() - self._t0) * 1000.0, 3) if self._t0 else 0.0
        return {
            "version": "v123_column_cpu_debug",
            "label": self.label,
            "wall_ms": wall,
            "sample_count": len(self.samples),
            "sample_interval_sec": self.interval,
            "process_cpu_pct_avg": round(sum(vals) / len(vals), 2) if vals else None,
            "process_cpu_pct_max": round(max(vals), 2) if vals else None,
            "num_threads_max": max(ths) if ths else None,
            "rss_mb_max": round(max(rss), 2) if rss else None,
            "error": self.error,
            "samples": self.samples[-300:],
        }


def _make_speed_env_v123() -> dict[str, Any]:
    return {
        "version": "v123_column_cpu_debug",
        "column_pass": str(_os_v123_app.environ.get("LOA_ATTACK_COLUMN_PASS", "1")),
        "column_pass_scale": str(_os_v123_app.environ.get("LOA_ATTACK_COLUMN_PASS_SCALE", "2")),
        "column_cell_fallback": str(_os_v123_app.environ.get("LOA_ATTACK_COLUMN_CELL_FALLBACK", "0")),
        "web_low_cpu": str(_os_v123_app.environ.get("LOA_WEB_LOW_CPU", "1")),
        "cpu_profile": str(_os_v123_app.environ.get("LOA_CPU_PROFILE", "1")),
        "rapidocr_threads": str(_os_v123_app.environ.get("RAPIDOCR_THREADS", "")),
        "omp_num_threads": str(_os_v123_app.environ.get("OMP_NUM_THREADS", "")),
        "ort_num_threads": str(_os_v123_app.environ.get("ORT_NUM_THREADS", "")),
        "openblas_num_threads": str(_os_v123_app.environ.get("OPENBLAS_NUM_THREADS", "")),
        "mkl_num_threads": str(_os_v123_app.environ.get("MKL_NUM_THREADS", "")),
        "expected_effect": "공격정보 셀별 OCR을 열 단위 OCR로 바꿔 readtext 호출 수와 CPU 점유율을 낮춥니다.",
    }

_run_ocr_for_kind_v115_prev_v123 = globals().get("_run_ocr_for_kind_v115")
def _run_ocr_for_kind_v115(kind: str, title: str, summary_image: Any, attack_image: Any, *, row_count: int, ocr_scale: int, icon_threshold: float, name_threshold: float, aggregate_unmatched: bool) -> dict[str, Any]:  # type: ignore[override]
    if not callable(_run_ocr_for_kind_v115_prev_v123):
        return {}
    with _CpuProfileV123(f"ocr_{kind}:{title}") as prof:
        result = _run_ocr_for_kind_v115_prev_v123(kind, title, summary_image, attack_image, row_count=row_count, ocr_scale=ocr_scale, icon_threshold=icon_threshold, name_threshold=name_threshold, aggregate_unmatched=aggregate_unmatched)
    cpu = prof.summary()
    try:
        result["cpu_profile_v123"] = {k: v for k, v in cpu.items() if k != "samples"}
        profiles = st.session_state.setdefault("cpu_profiles_v123", [])
        profiles.append(cpu)
        # 너무 커지는 것 방지
        st.session_state["cpu_profiles_v123"] = profiles[-8:]
    except Exception:
        pass
    return result

_sidebar_controls_prev_v123 = globals().get("sidebar_controls")
def sidebar_controls() -> None:  # type: ignore[override]
    if callable(_sidebar_controls_prev_v123):
        _sidebar_controls_prev_v123()
    try:
        pass  # 버전 변경 안내 제거
        st.sidebar.caption("readtext 호출 수 절감 · RAPIDOCR_THREADS 기본 2")
    except Exception:
        pass


try:
    _make_real_integrated_debug_zip_prev_v123 = _make_real_integrated_debug_zip_v106
    def _make_real_integrated_debug_zip_v106(summary_image: Any, attack_image: Any) -> tuple[Path, dict[str, Any]]:  # type: ignore[override]
        zpath, context = _make_real_integrated_debug_zip_prev_v123(summary_image, attack_image)
        try:
            import zipfile as _zipfile_v123
            import json as _json_v123
            profiles = st.session_state.get("cpu_profiles_v123") or []
            speed_env = _make_speed_env_v123()
            tmp = Path(str(zpath) + ".tmp")
            with _zipfile_v123.ZipFile(zpath, "r") as zin, _zipfile_v123.ZipFile(tmp, "w", compression=_zipfile_v123.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    zout.writestr(item, zin.read(item.filename))
                zout.writestr("context/cpu_profile_v123.json", _json_v123.dumps({"profiles": profiles, "speed_env": speed_env}, ensure_ascii=False, indent=2, default=str))
                zout.writestr("context/speed_env_v123.json", _json_v123.dumps(speed_env, ensure_ascii=False, indent=2, default=str))
            tmp.replace(zpath)
            context["cpu_profile_v123"] = {"profiles": profiles, "speed_env": speed_env}
            context["speed_env_v123"] = speed_env
        except Exception as e:
            try:
                context.setdefault("errors", []).append({"stage": "v123_append", "error": repr(e)})
            except Exception:
                pass
        return zpath, context
except Exception:
    pass



# ==============================================================================
# v124: web CPU limited mode + real CPU/thread diagnostics
# ==============================================================================
APP_CALC_VERSION = "v124_web_cpu_limited"
import os as _os_v124_app

# 직접 `streamlit run app.py`로 실행해도 웹 저CPU 모드가 기본이 되도록 보장합니다.
try:
    _loa_apply_cpu_limits_early()
except Exception:
    pass

def _apply_runtime_thread_limits_v124() -> dict[str, Any]:
    info: dict[str, Any] = {}
    try:
        threads = int(str(_os_v124_app.environ.get("LOA_EFFECTIVE_OCR_THREADS") or _os_v124_app.environ.get("RAPIDOCR_THREADS") or "1"))
    except Exception:
        threads = 1
    threads = max(1, min(8, threads))
    info["effective_threads"] = threads
    try:
        import cv2 as _cv2_v124  # type: ignore
        _cv2_v124.setNumThreads(threads)
        try:
            _cv2_v124.ocl.setUseOpenCL(False)
        except Exception:
            pass
        info["cv2_num_threads"] = int(_cv2_v124.getNumThreads())
    except Exception as e:
        info["cv2_error"] = repr(e)
    try:
        import onnxruntime as _ort_v124  # type: ignore
        info["onnxruntime_version"] = str(getattr(_ort_v124, "__version__", ""))
        info["onnxruntime_available"] = True
    except Exception as e:
        info["onnxruntime_available"] = False
        info["onnxruntime_error"] = repr(e)
    return info

class _CpuProfileV124:
    def __init__(self, label: str, interval: float | None = None):
        self.label = label
        try:
            interval = float(interval if interval is not None else _os_v124_app.environ.get("LOA_CPU_PROFILE_INTERVAL", "0.10"))
        except Exception:
            interval = 0.10
        self.interval = max(0.05, min(1.0, interval))
        self.samples: list[dict[str, Any]] = []
        self.error: str | None = None
        self._stop = None
        self._thread = None
        self._t0 = 0.0
        self._proc = None
        self.logical_cpu = _os_v124_app.cpu_count() or 1

    def __enter__(self):
        self._t0 = time.perf_counter()
        if str(_os_v124_app.environ.get("LOA_CPU_PROFILE", "1")).strip().lower() in {"0", "false", "no", "off"}:
            return self
        try:
            import psutil  # type: ignore
            import threading
            self._proc = psutil.Process()
            self._proc.cpu_percent(interval=None)
            psutil.cpu_percent(interval=None)
            self._stop = threading.Event()
            def _loop():
                while self._stop is not None and not self._stop.wait(self.interval):
                    try:
                        with self._proc.oneshot():
                            raw = float(self._proc.cpu_percent(interval=None))
                            # psutil process cpu는 멀티코어 기준 100%를 넘을 수 있습니다.
                            # 작업 관리자 전체 CPU%와 비교하기 쉽게 논리 CPU 수로 나눈 값도 같이 저장합니다.
                            pct_total = raw / max(1, self.logical_cpu)
                            try:
                                io_c = self._proc.io_counters()
                                io_read = int(getattr(io_c, "read_bytes", 0))
                                io_write = int(getattr(io_c, "write_bytes", 0))
                            except Exception:
                                io_read = io_write = 0
                            self.samples.append({
                                "t_ms": round((time.perf_counter() - self._t0) * 1000.0, 3),
                                "process_cpu_pct_raw": round(raw, 2),
                                "process_cpu_pct_of_total": round(pct_total, 2),
                                "system_cpu_pct": float(psutil.cpu_percent(interval=None)),
                                "num_threads": int(self._proc.num_threads()),
                                "rss_mb": round(float(self._proc.memory_info().rss) / (1024 * 1024), 2),
                                "io_read_mb": round(io_read / (1024 * 1024), 2),
                                "io_write_mb": round(io_write / (1024 * 1024), 2),
                            })
                    except Exception:
                        pass
            self._thread = threading.Thread(target=_loop, name="loa_cpu_profiler_v124", daemon=True)
            self._thread.start()
        except Exception as e:
            self.error = repr(e)
        return self

    def __exit__(self, exc_type, exc, tb):
        try:
            if self._stop is not None:
                self._stop.set()
            if self._thread is not None:
                self._thread.join(timeout=1.0)
        except Exception:
            pass
        return False

    def summary(self) -> dict[str, Any]:
        def vals(key: str) -> list[float]:
            return [float(x.get(key) or 0.0) for x in self.samples]
        def avg(xs: list[float]) -> float | None:
            return round(sum(xs) / len(xs), 2) if xs else None
        def mx(xs: list[float]) -> float | None:
            return round(max(xs), 2) if xs else None
        ths = [int(x.get("num_threads") or 0) for x in self.samples]
        rss = vals("rss_mb")
        wall = round((time.perf_counter() - self._t0) * 1000.0, 3) if self._t0 else 0.0
        return {
            "version": "v124_web_cpu_limited",
            "label": self.label,
            "wall_ms": wall,
            "sample_count": len(self.samples),
            "sample_interval_sec": self.interval,
            "logical_cpu": self.logical_cpu,
            "process_cpu_pct_raw_avg": avg(vals("process_cpu_pct_raw")),
            "process_cpu_pct_raw_max": mx(vals("process_cpu_pct_raw")),
            "process_cpu_pct_of_total_avg": avg(vals("process_cpu_pct_of_total")),
            "process_cpu_pct_of_total_max": mx(vals("process_cpu_pct_of_total")),
            "system_cpu_pct_avg": avg(vals("system_cpu_pct")),
            "system_cpu_pct_max": mx(vals("system_cpu_pct")),
            "num_threads_max": max(ths) if ths else None,
            "rss_mb_max": mx(rss),
            "error": self.error,
            "samples": self.samples[-500:],
        }

def _make_speed_env_v124() -> dict[str, Any]:
    runtime = _apply_runtime_thread_limits_v124()
    keys = [
        "LOA_OCR_CPU_MODE", "LOA_OCR_THREADS", "LOA_EFFECTIVE_OCR_THREADS",
        "LOA_FORCE_CPU_LIMIT", "LOA_WEB_LOW_CPU", "LOA_CPU_PROFILE",
        "RAPIDOCR_THREADS", "OMP_NUM_THREADS", "ORT_NUM_THREADS",
        "OPENBLAS_NUM_THREADS", "MKL_NUM_THREADS", "VECLIB_MAXIMUM_THREADS",
        "NUMEXPR_NUM_THREADS", "NUMEXPR_MAX_THREADS", "OMP_WAIT_POLICY", "KMP_BLOCKTIME",
        "LOA_ATTACK_COLUMN_PASS", "LOA_ATTACK_COLUMN_PASS_SCALE", "LOA_ATTACK_COLUMN_CELL_FALLBACK",
    ]
    return {
        "version": "v124_web_cpu_limited",
        "mode_note": "web_low=CPU 제한 우선, balanced=중간, local_fast=로컬 속도 우선",
        "env": {k: str(_os_v124_app.environ.get(k, "")) for k in keys},
        "runtime_thread_info": runtime,
        "expected_effect": "ONNX/RapidOCR/OpenCV 스레드를 실제 1~2개로 제한하고, psutil로 프로세스 CPU/스레드 수를 디버그에 기록합니다.",
    }

_run_ocr_for_kind_v115_prev_v124 = globals().get("_run_ocr_for_kind_v115")
def _run_ocr_for_kind_v115(kind: str, title: str, summary_image: Any, attack_image: Any, *, row_count: int, ocr_scale: int, icon_threshold: float, name_threshold: float, aggregate_unmatched: bool) -> dict[str, Any]:  # type: ignore[override]
    if not callable(_run_ocr_for_kind_v115_prev_v124):
        return {}
    _apply_runtime_thread_limits_v124()
    with _CpuProfileV124(f"ocr_{kind}:{title}") as prof:
        result = _run_ocr_for_kind_v115_prev_v124(kind, title, summary_image, attack_image, row_count=row_count, ocr_scale=ocr_scale, icon_threshold=icon_threshold, name_threshold=name_threshold, aggregate_unmatched=aggregate_unmatched)
    cpu = prof.summary()
    try:
        result["cpu_profile_v124"] = {k: v for k, v in cpu.items() if k != "samples"}
        profiles = st.session_state.setdefault("cpu_profiles_v124", [])
        profiles.append(cpu)
        st.session_state["cpu_profiles_v124"] = profiles[-10:]
    except Exception:
        pass
    return result

_sidebar_controls_prev_v124 = globals().get("sidebar_controls")
def sidebar_controls() -> None:  # type: ignore[override]
    if callable(_sidebar_controls_prev_v124):
        _sidebar_controls_prev_v124()
    try:
        env = _make_speed_env_v124()
        mode = env.get("env", {}).get("LOA_OCR_CPU_MODE", "web_low")
        th = env.get("env", {}).get("LOA_EFFECTIVE_OCR_THREADS", "1")
        pass  # 버전 변경 안내 제거
        st.sidebar.caption("psutil CPU/스레드 프로파일 + ONNX/OpenCV 스레드 제한")
    except Exception:
        pass


try:
    _make_real_integrated_debug_zip_prev_v124 = _make_real_integrated_debug_zip_v106
    def _make_real_integrated_debug_zip_v106(summary_image: Any, attack_image: Any) -> tuple[Path, dict[str, Any]]:  # type: ignore[override]
        zpath, context = _make_real_integrated_debug_zip_prev_v124(summary_image, attack_image)
        try:
            import zipfile as _zipfile_v124
            import json as _json_v124
            profiles = st.session_state.get("cpu_profiles_v124") or []
            speed_env = _make_speed_env_v124()
            tmp = Path(str(zpath) + ".tmp")
            with _zipfile_v124.ZipFile(zpath, "r") as zin, _zipfile_v124.ZipFile(tmp, "w", compression=_zipfile_v124.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    zout.writestr(item, zin.read(item.filename))
                zout.writestr("context/cpu_profile_v124.json", _json_v124.dumps({"profiles": profiles, "speed_env": speed_env}, ensure_ascii=False, indent=2, default=str))
                zout.writestr("context/speed_env_v124.json", _json_v124.dumps(speed_env, ensure_ascii=False, indent=2, default=str))
            tmp.replace(zpath)
            context["cpu_profile_v124"] = {"profiles": profiles, "speed_env": speed_env}
            context["speed_env_v124"] = speed_env
        except Exception as e:
            try:
                context.setdefault("errors", []).append({"stage": "v124_append", "error": repr(e)})
            except Exception:
                pass
        return zpath, context
except Exception:
    pass


# ==============================================================================
# v127: 브라우저 OCR 전용 모드
# - OCR을 Python/RapidOCR 서버가 아니라 방문자 브라우저(Tesseract.js/WASM)에서 실행합니다.
# - 서버는 결과 JSON 적용, API 계산, 아이콘 매칭만 담당합니다.
# - 완전 프론트 전환 전 단계이므로 서버 OCR은 폴백으로 남겨둡니다.
# ==============================================================================
APP_CALC_VERSION = "v127_browser_summary_fix"

try:
    import base64 as _base64_v125
    import streamlit.components.v1 as _components_v125
    _CLIENT_OCR_COMPONENT_V125 = _components_v125.declare_component(
        "loa_client_ocr_v127",
        path=str((Path(__file__).parent / "components" / "loa_client_ocr_v127").resolve()),
    )
except Exception:  # noqa: BLE001
    _base64_v125 = None  # type: ignore[assignment]
    _CLIENT_OCR_COMPONENT_V125 = None  # type: ignore[assignment]


def _upload_to_data_url_v125(upload: Any) -> str:
    if upload is None or _base64_v125 is None:
        return ""
    try:
        data = upload.getvalue()
        name = str(getattr(upload, "name", "") or "").lower()
        mime = "image/png"
        if name.endswith(".jpg") or name.endswith(".jpeg"):
            mime = "image/jpeg"
        elif name.endswith(".webp"):
            mime = "image/webp"
        return f"data:{mime};base64," + _base64_v125.b64encode(data).decode("ascii")
    except Exception:
        return ""


def _meta_from_browser_summary_v125(summary: dict[str, Any] | None) -> dict[str, Any]:
    summary = summary or {}
    raw = {
        "elapsed_text": str(summary.get("elapsed_text") or ""),
        "total_damage_text": str(summary.get("total_damage_text") or ""),
        "dps_text": str(summary.get("dps_text") or ""),
        "total_damage_raw_text": str(summary.get("total_damage_raw_text") or ""),
        "total_damage_yellow_text": str(summary.get("total_damage_yellow_text") or ""),
        "dps_raw_text": str(summary.get("dps_raw_text") or ""),
        "dps_yellow_text": str(summary.get("dps_yellow_text") or ""),
    }
    try:
        from modules.fixed_grid_ocr import extract_elapsed_seconds as _extract_elapsed_seconds_v125
        from modules.fixed_grid_ocr import full_raw_number_from_summary as _full_raw_number_from_summary_v125
    except Exception:
        _extract_elapsed_seconds_v125 = None
        _full_raw_number_from_summary_v125 = None
    def _num(text: str) -> float | None:
        val = None
        try:
            if callable(_full_raw_number_from_summary_v125):
                val = _full_raw_number_from_summary_v125(text)
        except Exception:
            val = None
        if val is None:
            try:
                val = parse_korean_number(text)
            except Exception:
                val = None
        return val
    elapsed_seconds = None
    try:
        if callable(_extract_elapsed_seconds_v125):
            elapsed_seconds = _extract_elapsed_seconds_v125(raw["elapsed_text"])
    except Exception:
        elapsed_seconds = None
    total_value = _num(raw["total_damage_text"])
    dps_value = _num(raw["dps_text"])
    # v127: 브라우저 OCR이 하단 원시 숫자를 놓치면 큰 노란 억 단위 숫자를 폴백으로 사용합니다.
    try:
        if not total_value and raw.get("total_damage_yellow_text"):
            total_value = _num(str(raw.get("total_damage_yellow_text")) + "억")
        if not dps_value and raw.get("dps_yellow_text"):
            dps_value = _num(str(raw.get("dps_yellow_text")) + "억")
    except Exception:
        pass
    meta = {
        "raw": raw,
        "browser_ocr_v125": True,
        "elapsed_seconds": elapsed_seconds,
        "elapsed_source": "browser_ocr_v125" if elapsed_seconds is not None else "browser_ocr_failed",
        "total_damage": total_value,
        "dps": dps_value,
        "total_damage_text": format_korean_number(total_value) if total_value else raw["total_damage_text"],
        "dps_text": format_korean_number(dps_value) if dps_value else raw["dps_text"],
        "raw_display": {
            "elapsed_text": raw["elapsed_text"],
            "total_damage_text": format_korean_number(total_value) if total_value else raw["total_damage_text"],
            "dps_text": format_korean_number(dps_value) if dps_value else raw["dps_text"],
        },
        "browser_ms": summary.get("browser_ms"),
    }
    return meta


def _browser_attack_metadata_rows_v125(attack_image: Image.Image | None, row_count: int) -> list[dict[str, Any]]:
    if attack_image is None:
        return []
    rows: list[dict[str, Any]] = []
    try:
        import modules.fixed_grid_ocr as _fgo_v125
        window_box = _fgo_v125._detect_window_no_ocr_v101(attack_image, None)
        iterator = _fgo_v125._iter_fixed_grid_rows_v111(attack_image, window_box, int(row_count or 14))
        for i, row_box, icon_box, source in iterator:
            rows.append({
                "_ocr_row_index": str(i),
                "_row_source": f"browser_ocr_v125:{source}",
                "_direction_kind": "back",
                "_window_x1": window_box[0], "_window_y1": window_box[1], "_window_x2": window_box[2], "_window_y2": window_box[3],
                "_icon_x1": icon_box[0], "_icon_y1": icon_box[1], "_icon_x2": icon_box[2], "_icon_y2": icon_box[3],
                "_row_x1": row_box[0], "_row_y1": row_box[1], "_row_x2": row_box[2], "_row_y2": row_box[3],
                "_grid_mode_v125": "browser_client_ocr",
            })
    except Exception:
        rows = []
    return rows


def _browser_attack_to_df_v125(attack_payload: dict[str, Any] | None, attack_image: Image.Image | None, row_count: int) -> pd.DataFrame:
    attack_payload = attack_payload or {}
    source_rows = attack_payload.get("rows") or []
    if not isinstance(source_rows, list) or not source_rows:
        return pd.DataFrame()
    meta_rows = _browser_attack_metadata_rows_v125(attack_image, row_count)
    out_rows: list[dict[str, Any]] = []
    for idx, r in enumerate(source_rows):
        if not isinstance(r, dict):
            continue
        row_index = int(r.get("row_index", idx) or idx)
        item: dict[str, Any] = {
            "이름": f"행 {row_index + 1}",
            "피해량": str(r.get("damage") or ""),
            "초당 피해량": str(r.get("dps") or ""),
            "피해량 지분": str(r.get("share") or ""),
            "백어택 적중률": str(r.get("back_hit") or "0"),
            "백어택 비중": str(r.get("back_share") or "0"),
            "헤드어택 적중률": str(r.get("head_hit") or "0"),
            "헤드어택 비중": str(r.get("head_share") or "0"),
            "치명타 적중률": str(r.get("crit_hit") or "0"),
            "치명타 비중": str(r.get("crit_share") or "0"),
            "사용 횟수": str(r.get("casts") or "0"),
            "쿨타임 비율": str(r.get("cooldown") or "0"),
            "_browser_ocr_v125": "1",
        }
        if row_index < len(meta_rows):
            item.update(meta_rows[row_index])
        out_rows.append(item)
    if not out_rows:
        return pd.DataFrame()
    return pd.DataFrame(out_rows)


def _apply_browser_ocr_kind_v125(kind: str, title: str, payload: dict[str, Any] | None, summary_image: Image.Image | None, attack_image: Image.Image | None, *, row_count: int, icon_threshold: float, name_threshold: float, aggregate_unmatched: bool) -> dict[str, Any]:
    payload = payload or {}
    report: dict[str, Any] = {"kind": kind, "title": title, "browser_payload": bool(payload)}
    key = f"{kind}_table"
    meta_key = f"{kind}_meta"
    t0 = time.perf_counter()
    if summary_image is not None and isinstance(payload.get("summary"), dict):
        meta = _meta_from_browser_summary_v125(payload.get("summary"))
        st.session_state[meta_key] = meta
        _sync_meta_to_manual_inputs(kind, meta)
        report["summary_browser_ms"] = payload.get("summary", {}).get("browser_ms")
    if attack_image is not None and isinstance(payload.get("attack"), dict):
        parsed = _browser_attack_to_df_v125(payload.get("attack"), attack_image, row_count)
        report["attack_browser_ms"] = payload.get("attack", {}).get("browser_ms")
        report["attack_rows_before_icon"] = 0 if parsed is None else len(parsed)
        if parsed is not None and not parsed.empty:
            try:
                candidates_full = get_ocr_skill_candidates_full()
                skill_icon_candidates = _skill_only_ocr_candidates(candidates_full)
            except Exception:
                skill_icon_candidates = []
            report["candidate_count"] = len(skill_icon_candidates)
            if skill_icon_candidates:
                t_icon = time.perf_counter()
                parsed = correct_battle_skill_names_with_icons(
                    parsed,
                    attack_image,
                    skill_icon_candidates,
                    threshold=float(name_threshold),
                    icon_threshold=float(icon_threshold),
                    drop_unmatched=False,
                )
                report["server_icon_match_ms"] = round((time.perf_counter() - t_icon) * 1000.0, 3)
                if aggregate_unmatched:
                    parsed = _aggregate_unmatched_rows_as_extra(parsed)
            parsed = sanitize_battle_table(parsed)
            parsed = _repair_battle_values_from_summary(parsed, st.session_state.get(meta_key, {}), elapsed_fallback=st.session_state.get(f"{kind}_elapsed_sec"))
            parsed = _drop_obvious_ocr_noise_rows(parsed)
            st.session_state[key] = prepare_battle_editor_df(sanitize_battle_table(parsed))
            _bump_table_version(kind)
            report["attack_rows_after"] = len(st.session_state[key])
    report["server_apply_ms"] = round((time.perf_counter() - t0) * 1000.0, 3)
    return report


def _render_client_ocr_component_v125(*, real_summary_up: Any, real_attack_up: Any, dummy_summary_up: Any, dummy_attack_up: Any, row_count: int) -> dict[str, Any] | None:
    if _CLIENT_OCR_COMPONENT_V125 is None:
        st.error("브라우저 OCR 컴포넌트를 불러오지 못했습니다.")
        return None
    return _CLIENT_OCR_COMPONENT_V125(
        real_summary_data_url=_upload_to_data_url_v125(real_summary_up),
        real_attack_data_url=_upload_to_data_url_v125(real_attack_up),
        dummy_summary_data_url=_upload_to_data_url_v125(dummy_summary_up),
        dummy_attack_data_url=_upload_to_data_url_v125(dummy_attack_up),
        row_count=int(row_count or 14),
        default=None,
        key="loa_client_ocr_component_v127",
    )


_ocr_tab_prev_v125_server = globals().get("ocr_tab")

def ocr_tab() -> None:  # type: ignore[override]
    st.header("전투분석기 입력")
    st.caption("종합정보에서 전투시간·총피해량·DPS를 자동으로 읽어옵니다.")
    _render_step_guide(2, "전투분석기 이미지 넣고 검수하기", [
        "OCR은 <b>방문자 CPU</b>에서 숫자 인식을 처리합니다.",
        "서버는 API 계산과 아이콘 보정만 담당합니다.",
        "서버 RapidOCR 폴백은 제거했습니다. 종합정보 숫자는 하단 원시 숫자 줄을 우선 읽습니다.",
    ])

    st.info("전투분석기 종합정보에서 전투시간·총피해량·DPS를 자동으로 읽어옵니다.")

    with st.container(border=True):
        st.markdown("#### 통합 이미지 업로드")
        c1, c2 = st.columns(2)
        with c1:
            st.markdown("**실전 전투분석기**")
            real_summary_up = st.file_uploader("실전 종합정보 이미지", type=["png", "jpg", "jpeg", "webp"], key="unified_real_summary_v126")
            real_attack_up = st.file_uploader("실전 공격정보 이미지", type=["png", "jpg", "jpeg", "webp"], key="unified_real_attack_v126")
        with c2:
            st.markdown("**허수/기준 전투분석기**")
            dummy_summary_up = st.file_uploader("허수/기준 종합정보 이미지", type=["png", "jpg", "jpeg", "webp"], key="unified_dummy_summary_v126")
            dummy_attack_up = st.file_uploader("허수/기준 공격정보 이미지", type=["png", "jpg", "jpeg", "webp"], key="unified_dummy_attack_v126")

        real_summary_img = _image_from_uploader_or_session_v115(real_summary_up, "real", "summary")
        real_attack_img = _image_from_uploader_or_session_v115(real_attack_up, "real", "attack")
        dummy_summary_img = _image_from_uploader_or_session_v115(dummy_summary_up, "dummy", "summary")
        dummy_attack_img = _image_from_uploader_or_session_v115(dummy_attack_up, "dummy", "attack")

        with st.expander("브라우저 OCR 설정", expanded=False):
            row_count = st.number_input("공격정보 인식 행 수", min_value=5, max_value=24, value=int(st.session_state.get("unified_row_count_v126", 14) or 14), step=1, key="unified_row_count_v126")
            icon_match_threshold = st.slider("아이콘 보정 신뢰도", min_value=45, max_value=95, value=int(st.session_state.get("unified_icon_threshold_v126", 74) or 74), step=1, key="unified_icon_threshold_v126")
            name_match_threshold = st.slider("스킬명 보정 민감도", min_value=40, max_value=90, value=int(st.session_state.get("unified_name_threshold_v126", 52) or 52), step=1, key="unified_name_threshold_v126")
            aggregate_unmatched = st.checkbox("낮은 신뢰도 행은 기타 추가딜로 합산", value=bool(st.session_state.get("unified_aggregate_unmatched_v126", False)), key="unified_aggregate_unmatched_v126")

        if not any([real_summary_up, real_attack_up, dummy_summary_up, dummy_attack_up]):
            st.info("화면 공유로 캡처하면 아래에서 바로 인식됩니다.")
        client_result = _render_client_ocr_component_v125(
            real_summary_up=real_summary_up,
            real_attack_up=real_attack_up,
            dummy_summary_up=dummy_summary_up,
            dummy_attack_up=dummy_attack_up,
            row_count=int(row_count),
        )
        if isinstance(client_result, dict):
            st.session_state["last_client_ocr_v126"] = client_result
            if client_result.get("error"):
                st.error(str(client_result.get("error")))
            elif client_result.get("done"):
                st.success(f"브라우저 OCR 완료 · {client_result.get('timings', {}).get('total_browser_ms', '-')} ms")
                if st.button("브라우저 OCR 결과를 표에 적용", type="primary", use_container_width=True, key="apply_client_ocr_v126"):
                    reports = []
                    if client_result.get("real"):
                        reports.append(_apply_browser_ocr_kind_v125(
                            "real", "실전 전투분석기", client_result.get("real"), real_summary_img, real_attack_img,
                            row_count=int(row_count), icon_threshold=float(icon_match_threshold)/100.0,
                            name_threshold=float(name_match_threshold)/100.0, aggregate_unmatched=bool(aggregate_unmatched),
                        ))
                    if client_result.get("dummy"):
                        reports.append(_apply_browser_ocr_kind_v125(
                            "dummy", "허수/기준 전투분석기", client_result.get("dummy"), dummy_summary_img, dummy_attack_img,
                            row_count=int(row_count), icon_threshold=float(icon_match_threshold)/100.0,
                            name_threshold=float(name_match_threshold)/100.0, aggregate_unmatched=bool(aggregate_unmatched),
                        ))
                    st.session_state["client_ocr_reports_v126"] = reports
                    st.success("브라우저 OCR 결과를 적용했습니다. 아래 표를 검수하세요.")
        reports = st.session_state.get("client_ocr_reports_v126") or []
        if reports:
            st.caption("브라우저 OCR 적용/서버 후처리 시간")
            show_cols = ["title", "summary_browser_ms", "attack_browser_ms", "server_icon_match_ms", "server_apply_ms", "attack_rows_after"]
            try:
                st.dataframe(pd.DataFrame(reports)[[c for c in show_cols if c in pd.DataFrame(reports).columns]], use_container_width=True, hide_index=True)
            except Exception:
                st.json(reports)

    real_tab, dummy_tab = st.tabs(["실전 검수", "허수/기준 검수"])
    with real_tab:
        _render_battle_review_only_v115("real", "실전 전투분석기")
    with dummy_tab:
        _render_battle_review_only_v115("dummy", "허수/기준 전투분석기")


_sidebar_controls_prev_v125 = globals().get("sidebar_controls")
def sidebar_controls() -> None:  # type: ignore[override]
    if callable(_sidebar_controls_prev_v125):
        _sidebar_controls_prev_v125()
    try:
        pass  # 버전 변경 안내 제거
        st.sidebar.caption("이미지 인식은 내 브라우저에서 바로 실행됩니다.")
    except Exception:
        pass



try:
    _make_real_integrated_debug_zip_prev_v125 = _make_real_integrated_debug_zip_v106
    def _make_real_integrated_debug_zip_v106(summary_image: Any, attack_image: Any) -> tuple[Path, dict[str, Any]]:  # type: ignore[override]
        zpath, context = _make_real_integrated_debug_zip_prev_v125(summary_image, attack_image)
        try:
            import zipfile as _zipfile_v125
            import json as _json_v125
            payload = st.session_state.get("last_client_ocr_v126") or {}
            reports = st.session_state.get("client_ocr_reports_v126") or []
            tmp = Path(str(zpath) + ".tmp")
            with _zipfile_v125.ZipFile(zpath, "r") as zin, _zipfile_v125.ZipFile(tmp, "w", compression=_zipfile_v125.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    zout.writestr(item, zin.read(item.filename))
                zout.writestr("context/client_ocr_v125.json", _json_v125.dumps({"payload": payload, "reports": reports}, ensure_ascii=False, indent=2, default=str))
            tmp.replace(zpath)
            context["client_ocr_v125"] = {"reports": reports, "has_payload": bool(payload)}
        except Exception as e:
            try:
                context.setdefault("errors", []).append({"stage": "v125_append", "error": repr(e)})
            except Exception:
                pass
        return zpath, context
except Exception:
    pass


# ==============================================================================
# v126: 서버 RapidOCR 제거 후 브라우저 OCR 결과 중심 통합 디버그
# - 기존 통합 디버그는 get_easyocr_reader()/RapidOCR을 호출할 수 있어 브라우저 전용 모드와 맞지 않습니다.
# - 여기서는 현재 session_state의 OCR 결과/표/메타/아이콘 보정 리포트만 저장합니다.
# ==============================================================================
def _make_real_integrated_debug_zip_v106(summary_image: Any, attack_image: Any) -> tuple[Path, dict[str, Any]]:  # type: ignore[override]
    import json as _json_v126
    import zipfile as _zipfile_v126
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    zpath = EXPORT_DIR / f"real_integrated_debug_browser_v127_{ts}.zip"
    context: dict[str, Any] = {
        "version": "v127_browser_summary_fix",
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "server_rapidocr_removed": True,
        "client_ocr_payload_present": bool(st.session_state.get("last_client_ocr_v126")),
        "client_ocr_reports_present": bool(st.session_state.get("client_ocr_reports_v126")),
        "errors": [],
    }
    def _write_df(zout, name: str, df: Any) -> None:
        try:
            if isinstance(df, pd.DataFrame):
                zout.writestr(name, df.to_csv(index=False).encode("utf-8-sig"))
        except Exception as e:
            context.setdefault("errors", []).append({"stage": f"write_{name}", "error": repr(e)})
    with _zipfile_v126.ZipFile(zpath, "w", compression=_zipfile_v126.ZIP_DEFLATED) as zout:
        zout.writestr("context/context.json", _json_v126.dumps(context, ensure_ascii=False, indent=2, default=str))
        zout.writestr("context/client_ocr_v127.json", _json_v126.dumps({
            "payload": st.session_state.get("last_client_ocr_v126") or {},
            "reports": st.session_state.get("client_ocr_reports_v126") or [],
            "real_meta": st.session_state.get("real_meta") or {},
            "dummy_meta": st.session_state.get("dummy_meta") or {},
        }, ensure_ascii=False, indent=2, default=str))
        _write_df(zout, "context/real_table.csv", st.session_state.get("real_table"))
        _write_df(zout, "context/dummy_table.csv", st.session_state.get("dummy_table"))
        try:
            if summary_image is not None:
                bio = io.BytesIO(); summary_image.save(bio, format="PNG"); zout.writestr("images/real_summary.png", bio.getvalue())
            if attack_image is not None:
                bio = io.BytesIO(); attack_image.save(bio, format="PNG"); zout.writestr("images/real_attack.png", bio.getvalue())
        except Exception as e:
            context.setdefault("errors", []).append({"stage": "write_images", "error": repr(e)})
    return zpath, context




# ==============================================================================
# v130: 브라우저 OCR 자동 적용 + 수정 기본 표 + 피해량 digit repair
# ==============================================================================
APP_CALC_VERSION = "v130_english_unit_templates"

try:
    import hashlib as _hashlib_v130
    import json as _json_v130
    import base64 as _base64_v130
    import streamlit.components.v1 as _components_v130
    _CLIENT_OCR_COMPONENT_V130 = _components_v130.declare_component(
        "loa_client_ocr_v148",
        path=str((Path(__file__).parent / "components" / "loa_client_ocr_v130").resolve()),
    )
except Exception:  # noqa: BLE001
    _hashlib_v130 = None  # type: ignore[assignment]
    _json_v130 = None  # type: ignore[assignment]
    _base64_v130 = None  # type: ignore[assignment]
    _CLIENT_OCR_COMPONENT_V130 = None  # type: ignore[assignment]


def _upload_to_data_url_v130(upload: Any) -> str:
    if upload is None or _base64_v130 is None:
        return ""
    try:
        data = upload.getvalue()
        name = str(getattr(upload, "name", "") or "").lower()
        mime = "image/png"
        if name.endswith(".jpg") or name.endswith(".jpeg"):
            mime = "image/jpeg"
        elif name.endswith(".webp"):
            mime = "image/webp"
        return f"data:{mime};base64," + _base64_v130.b64encode(data).decode("ascii")
    except Exception:
        return ""


class _CapturedImageUpload:
    """화면 공유 캡처 프레임을 업로드 객체처럼 쓰기 위한 래퍼.
    getvalue()/name만 제공하면 기존 업로드 파이프라인(데이터URL·그리드·시그니처)을 그대로 탑니다."""

    def __init__(self, png_bytes: bytes, name: str):
        self._data = png_bytes
        self.name = name

    def getvalue(self) -> bytes:
        return self._data


def _png_bytes_from_data_url_v137(data_url: str) -> bytes | None:
    """브라우저에서 온 'data:image/png;base64,...' 문자열을 PNG 바이트로 디코드."""
    if not data_url or _base64_v130 is None:
        return None
    try:
        s = str(data_url)
        if "," in s:
            s = s.split(",", 1)[1]
        return _base64_v130.b64decode(s)
    except Exception:
        return None


def _handle_screen_capture_result_v137(client_result: Any) -> bool:
    """컴포넌트가 '이미지 인식' 시 한 번에 보낸 캡처 배치를 세션에 저장합니다.

    v141: 캡처는 브라우저에 쌓아두고, '이미지 인식'을 누를 때 capture_batch로 한 번만
    보냅니다. 그래서 캡처마다 rerun하던 깜빡임/중복 저장이 사라집니다.
    저장 후 True를 반환하면 호출부가 한 번 rerun 하고, 그 렌더에서 자동 인식합니다.
    """
    if not isinstance(client_result, dict):
        return False
    batch = client_result.get("capture_batch")
    if not isinstance(batch, dict):
        return False
    stamp = str(client_result.get("ts") or "")
    # 같은 배치를 두 번 처리하지 않도록 스탬프로 중복 제거.
    if stamp and st.session_state.get("_last_capture_batch_ts") == stamp:
        return False
    stored = 0
    for kind in ("real_summary", "real_attack", "dummy_summary", "dummy_attack"):
        urls = batch.get(kind) or []
        if not isinstance(urls, list):
            continue
        imgs = []
        for u in urls:
            png = _png_bytes_from_data_url_v137(u)
            if png:
                imgs.append(png)
        if not imgs:
            continue
        if kind in ("real_attack", "dummy_attack"):
            key = f"cap_{kind}_list"
            lst = list(st.session_state.get(key) or [])
            for i, png in enumerate(imgs):
                lst.append(_CapturedImageUpload(png, f"capture_{kind}_{stamp}_{i}.jpg"))
            if len(lst) > 8:
                lst = lst[-8:]
            st.session_state[key] = lst
            stored += len(imgs)
        else:
            # 종합정보는 마지막 1장만 사용.
            st.session_state[f"cap_{kind}"] = _CapturedImageUpload(imgs[-1], f"capture_{kind}_{stamp}.jpg")
            stored += 1
    if not stored:
        return False
    st.session_state["_last_capture_batch_ts"] = stamp
    st.session_state["_auto_run_after_capture"] = True
    return True


def _handle_clear_capture_v144(client_result: Any) -> bool:
    """컴포넌트에서 특정 종류(실전/허수 · 공격/종합)의 캡처 취소 요청이 오면 해당 캡처만 지웁니다."""
    if not isinstance(client_result, dict):
        return False
    kind = str(client_result.get("clear_capture") or "").strip()
    if kind not in ("real_summary", "real_attack", "dummy_summary", "dummy_attack"):
        return False
    stamp = str(client_result.get("ts") or "")
    if stamp and st.session_state.get("_last_clear_capture_ts") == stamp:
        return False
    if kind in ("real_attack", "dummy_attack"):
        st.session_state.pop(f"cap_{kind}_list", None)
    else:
        st.session_state.pop(f"cap_{kind}", None)
    st.session_state["_last_clear_capture_ts"] = stamp
    return True


def _clear_captured_frames_v138() -> None:
    """캡처해 둔 화면 프레임을 세션에서 완전히 삭제합니다(이미지 인식 완료 후 호출)."""
    for _k in ["cap_real_attack_list", "cap_dummy_attack_list", "cap_real_summary", "cap_dummy_summary", "_last_capture_ts", "_last_capture_batch_ts", "_auto_run_after_capture"]:
        st.session_state.pop(_k, None)


def _input_sig_v130(*uploads: Any) -> str:
    try:
        h = _hashlib_v130.sha256() if _hashlib_v130 is not None else None
        if h is None:
            return ""
        for up in uploads:
            if up is None:
                h.update(b"<none>")
                continue
            h.update(str(getattr(up, "name", "") or "").encode("utf-8", "ignore"))
            data = up.getvalue()
            h.update(str(len(data)).encode("ascii"))
            h.update(data[:4096])
            h.update(data[-4096:])
        return h.hexdigest()[:24]
    except Exception:
        return ""


def _digits_only_v130(value: Any) -> str:
    return "".join(re.findall(r"\d", str(value or "")))


def _percent_from_digits_v130(value: Any) -> str:
    d = _digits_only_v130(value)
    if not d:
        return ""
    try:
        n = int(d)
    except Exception:
        return ""
    # 100.00 같은 값
    if d in {"10000", "100000"}:
        return "100.0"
    if len(d) <= 2:
        val = float(n)
    elif len(d) == 3:
        # 671 -> 6.71 또는 970 -> 9.70 처럼 보일 수 있으나, 전투분석기 비중은 대부분 소수 둘째 자리입니다.
        val = n / 10.0 if n > 1000 else n / 100.0
    else:
        val = n / 100.0
    # OCR이 77070처럼 앞/뒤 노이즈를 붙이면 0~100 범위로 접습니다.
    while val > 100.0 and len(d) > 2:
        d = d[:-1]
        try:
            val = int(d) / 100.0
        except Exception:
            break
    if val < 0:
        val = 0.0
    if val > 100.0:
        val = 100.0
    return str(round(val, 2))


def _casts_from_digits_v130(value: Any) -> str:
    d = _digits_only_v130(value)
    if not d:
        return ""
    try:
        n = int(d)
    except Exception:
        return ""
    # 사용 횟수는 보통 두 자리 이하. 긴 OCR 잡음은 앞쪽 두 자리만 사용합니다.
    if n > 999 and len(d) > 2:
        d = d[:2]
        try:
            n = int(d)
        except Exception:
            pass
    return str(n)


def _canonical_damage_unit_v130(value: Any) -> str:
    text = str(value or "").strip().lower()
    if text in {"man", "만"}:
        return "만"
    if text in {"jo", "조"}:
        return "조"
    if text in {"eok", "억"}:
        return "억"
    return "억"


def _unit_value_to_eok_v130(value: float, unit: str) -> float:
    if unit == "만":
        return value / 10000.0
    if unit == "조":
        return value * 10000.0
    return value


def _format_unit_value_v130(value: float, unit: str) -> str:
    return f"{float(value):,.2f}{unit}"


def _damage_candidate_ey_v130(digits: str, unit: str = "억") -> list[tuple[int, float, str]]:
    """Return candidates as (cut_count, eok_value_for_ordering, display_text).

    v130 uses uploaded unit templates. The filename-safe unit key is detected in the browser,
    then Python converts digits with the correct unit instead of assuming every cell is 억.
    """
    d = _digits_only_v130(digits)
    if not d:
        return []
    unit = _canonical_damage_unit_v130(unit)
    out: list[tuple[int, float, str]] = []
    # Tesseract가 단위/테두리를 숫자로 붙이면 뒤 1~2자리 제거 후보를 함께 둡니다.
    for cut in (0, 1, 2):
        dd = d[:-cut] if cut else d
        if not dd:
            continue
        try:
            val = int(dd) / 100.0
            out.append((cut, _unit_value_to_eok_v130(val, unit), _format_unit_value_v130(val, unit)))
        except Exception:
            pass
    uniq: list[tuple[int, float, str]] = []
    seen = set()
    for cut, val_ey, disp in out:
        key = (round(val_ey, 4), disp)
        if key not in seen:
            seen.add(key); uniq.append((cut, val_ey, disp))
    return uniq


def _repair_damage_digits_sequence_v130(rows: list[dict[str, Any]], total_damage: float | None = None) -> list[str]:
    total_ey = None
    try:
        if total_damage and float(total_damage) > 0:
            total_ey = float(total_damage) / 100_000_000.0
    except Exception:
        total_ey = None
    prev_ey = total_ey if total_ey else None
    repaired: list[str] = []
    for r in rows:
        raw_d = str(r.get("damage_digits") or r.get("damage") or "")
        unit = _canonical_damage_unit_v130(r.get("damage_unit_key") or r.get("damage_unit") or "억")
        cands = _damage_candidate_ey_v130(raw_d, unit)
        if not cands:
            repaired.append("")
            continue
        max_allowed = None
        if prev_ey and prev_ey > 0:
            # 전투분석기 공격정보는 피해량 내림차순입니다. 같은 행 주변 OCR 노이즈를 감안해 8% 여유를 둡니다.
            max_allowed = prev_ey * 1.08
        elif total_ey:
            max_allowed = total_ey * 1.02
        chosen = None
        if max_allowed:
            valid = [(cut, val_ey, disp) for cut, val_ey, disp in cands if val_ey <= max_allowed and val_ey > 0]
            if valid:
                valid.sort(key=lambda x: (x[0], -x[1]))
                chosen = valid[0]
        if chosen is None:
            cands.sort(key=lambda x: (x[0], -x[1]))
            chosen = cands[0]
        cut, ey, disp = chosen
        if ey > 0:
            prev_ey = ey
            repaired.append(disp)
        else:
            repaired.append("")
    return repaired


def _monotonic_unit_correct_v135(source_rows: list) -> dict[int, str]:
    """v135b: 공격정보 피해량은 내림차순이므로, 단위(억/만/조)가 잘못 잡혀 값이 '자릿수 단위로'
    어긋난 경우(예: 43.64억을 43.64만으로 1만배 작게 읽음)를 이전 행 기준으로 보정합니다.
    숫자 자체는 그대로 두고 단위만 고치며, OCR 값 노이즈(±수배)는 건드리지 않고
    이전 행 대비 1000배 이상 벗어날 때만 교정합니다."""
    import re as _re
    mults = {"조": 1e12, "억": 1e8, "만": 1e4}
    def parse(s):
        s = str(s or "").strip()
        m = _re.match(r"^(\d[\d,]*(?:\.\d+)?)", s)
        if not m:
            return None, "", s
        try:
            val = float(m.group(1).replace(",", ""))
        except Exception:
            return None, "", s
        um = _re.search(r"(조|억|만)", s)
        return val, (um.group(1) if um else ""), m.group(1)
    out: dict[int, str] = {}
    prev = None
    for idx, r in enumerate(source_rows):
        if not isinstance(r, dict):
            continue
        ri = int(r.get("row_index", idx) or idx)
        val, unit, numtxt = parse(r.get("damage"))
        if val is None or val <= 0:
            out[ri] = str(r.get("damage") or "")
            continue
        det = unit if unit in mults else "억"
        if prev is None:
            chosen = det
        else:
            lo, hi = prev / 1000.0, prev * 3.0
            if lo <= val * mults[det] <= hi:
                chosen = det
            else:
                inband = [u for u in mults if lo <= val * mults[u] <= hi]
                if inband:
                    le = [u for u in inband if val * mults[u] <= prev * 1.5]
                    chosen = max(le, key=lambda u: mults[u]) if le else min(inband, key=lambda u: abs(val * mults[u] - prev))
                else:
                    chosen = det
        prev = val * mults[chosen]
        out[ri] = f"{numtxt}{chosen}"
    return out


def _browser_attack_to_df_v130(attack_payload: dict[str, Any] | None, attack_image: Image.Image | None, row_count: int, total_damage: float | None = None, fill_missing: bool = True) -> pd.DataFrame:
    attack_payload = attack_payload or {}
    source_rows = attack_payload.get("rows") or []
    if not isinstance(source_rows, list) or not source_rows:
        return pd.DataFrame()
    meta_rows = _browser_attack_metadata_rows_v125(attack_image, row_count)
    corrected_dmg = _monotonic_unit_correct_v135(source_rows)  # v135b: 만/억 단위 단조성 보정
    out_rows: list[dict[str, Any]] = []
    for idx, r in enumerate(source_rows):
        if not isinstance(r, dict):
            continue
        row_index = int(r.get("row_index", idx) or idx)
        # v134 수정: 브라우저(v131+)가 이미 "2,083.44억"처럼 소수 2자리 + 단위까지 정확히 복구해
        # 보내므로, 그 값을 그대로 씁니다. (기존 v130은 damage_digits를 총피해량 기준으로 다시
        # 소수점 배치하다가, 총피해량 OCR이 틀리면 피해량이 통째로 10배 어긋났습니다.)
        # v135b: 단위가 자릿수 단위로 어긋난 행만 내림차순 규칙으로 보정한 값을 씁니다.
        dmg = corrected_dmg.get(row_index, str(r.get("damage") or ""))
        # v135c: 그 이미지 그리드가 '읽은' 컬럼만 값이 있고, 안 읽은 컬럼은 키가 없습니다.
        # fill_missing=False(멀티 병합 모드)면 안 읽은 컬럼은 None으로 둬서, 병합 시
        # 실제로 읽은 다른 이미지의 값을 채택합니다. True면 기존처럼 기본값으로 채웁니다.
        def _colp(base, dflt, kind="pct", na_on_blank=False):
            if (base in r) or ((base + "_digits") in r):
                # v135e: 방향 칸(백/헤드)이 원본 OCR상 '빈칸'(숫자 0개)이면 그 스킬은
                # 그 방향이 아예 없는 것(N/A)이므로 0이 아니라 None으로 둡니다.
                # 실제 "0.00%"(숫자 있음)는 그대로 0으로 들어옵니다.
                if na_on_blank:
                    _ro = str(r.get(base + "_raw_ocr") or "")
                    if not any(ch.isdigit() for ch in _ro):
                        return None
                raw = r.get(base + "_digits") or r.get(base)
                return _casts_from_digits_v130(raw) if kind == "casts" else _percent_from_digits_v130(raw)
            return dflt if fill_missing else None
        item: dict[str, Any] = {
            "이름": f"행 {row_index + 1}",
            "피해량": dmg,
            "초당 피해량": "",
            "피해량 지분": "",
            "백어택 적중률": _colp("back_hit", "0.0", na_on_blank=True),
            "백어택 비중": _colp("back_share", "", na_on_blank=True),
            "헤드어택 적중률": _colp("head_hit", "0.0", na_on_blank=True),
            "헤드어택 비중": _colp("head_share", "0.0", na_on_blank=True),
            "치명타 적중률": _colp("crit_hit", "0.0"),
            "치명타 비중": _colp("crit_share", ""),
            "사용 횟수": _colp("casts", "", kind="casts"),
            "쿨타임 비율": _colp("cooldown", ""),
            "_browser_ocr_v130": "1",
            "_damage_digits_raw": str(r.get("damage_digits") or r.get("damage") or ""),
            "_damage_unit_v130": str(r.get("damage_unit") or ""),
            "_damage_unit_score_v130": str(r.get("damage_unit_score") or ""),
        }
        if row_index < len(meta_rows):
            item.update(meta_rows[row_index])
        out_rows.append(item)
    return pd.DataFrame(out_rows) if out_rows else pd.DataFrame()


def _meta_from_browser_summary_v130(summary: dict[str, Any] | None) -> dict[str, Any]:
    meta = _meta_from_browser_summary_v125(summary)
    raw = meta.setdefault("raw", {})
    # v128/v130 digit payload를 우선 사용해서 쉼표가 빠져도 원시 숫자로 복구합니다.
    try:
        td_digits = _digits_only_v130((summary or {}).get("total_damage_raw_digits") or (summary or {}).get("total_damage_digits"))
        dps_digits = _digits_only_v130((summary or {}).get("dps_raw_digits") or (summary or {}).get("dps_digits"))
        # v134: 카드 아래 '큰 노란 억 숫자'는 자릿수 누락에 강한 보조값입니다.
        # 하단 원시 숫자(td_digits)가 노란값과 8% 이상 어긋나면, 원시 숫자가 자릿수를
        # 빠뜨렸을 가능성이 높으므로(예: 404.12억 vs 4,641.26억) 노란값을 채택합니다.
        yellow_total = None
        try:
            _yt = (summary or {}).get("total_damage_yellow_text")
            if _yt:
                yellow_total = parse_korean_number(str(_yt) + "억")
        except Exception:
            yellow_total = None
        if len(td_digits) >= 7:
            raw_total = float(int(td_digits))
            chosen = raw_total
            # v137: 노란 억 숫자는 소수점이 사라져(예: '4,641.26' → '4 641 26') 자릿수가
            # 1000배씩 틀어질 수 있습니다. 그래서 magnitude(비율)로 비교하면 안 되고,
            # '앞 유효숫자'만 비교합니다. 앞자리가 같으면 원시 흰 숫자(정밀값)를 신뢰하고,
            # 앞자리가 다를 때만(원시가 자릿수를 놓친 경우) 노란값을 보조로 씁니다.
            yd = _digits_only_v130((summary or {}).get("total_damage_yellow_digits") or (summary or {}).get("total_damage_yellow_text"))
            if yd and len(yd) >= 4:
                n = min(4, len(td_digits), len(yd))
                if td_digits[:n] != yd[:n] and yellow_total and yellow_total > 0:
                    # 앞자리 불일치 → 원시 숫자가 앞부분을 잘못 읽었을 가능성. 노란 유효숫자를
                    # 원시 숫자의 자릿수(magnitude)에 맞춰 재구성해 사용합니다(노란 magnitude는 불신).
                    _digits_target = len(td_digits)
                    _y = yd[:_digits_target].ljust(_digits_target, "0")
                    try:
                        chosen = float(int(_y))
                        raw["total_damage_source_v137"] = f"yellow-lead(raw {td_digits[:6]} vs yellow {yd[:6]})"
                    except Exception:
                        chosen = raw_total
            meta["total_damage"] = chosen
            meta["total_damage_text"] = format_korean_number(chosen)
            meta.setdefault("raw_display", {})["total_damage_text"] = meta["total_damage_text"]
            raw["total_damage_digits_v130"] = td_digits
        elif yellow_total and yellow_total > 0:
            meta["total_damage"] = yellow_total
            meta["total_damage_text"] = format_korean_number(yellow_total)
            meta.setdefault("raw_display", {})["total_damage_text"] = meta["total_damage_text"]
        if len(dps_digits) >= 5:
            meta["dps"] = float(int(dps_digits))
            meta["dps_text"] = format_korean_number(meta["dps"])
            meta.setdefault("raw_display", {})["dps_text"] = meta["dps_text"]
            raw["dps_digits_v130"] = dps_digits
    except Exception:
        pass
    return meta


def _select_primary_page_dfs_v136(dfs: list) -> list:
    """여러 공격정보 이미지 중 '스크롤 안 한 맨 위 페이지'의 이미지들만 고릅니다.

    사용자가 세로로 스크롤한 페이지들은 하위 행(낮은 딜 스킬)이 서로 어긋나 병합 때
    값이 꼬입니다(예: 기본 공격에 다른 행의 피해가 들어감). 그래서 기본 동작은
    '가로 열(같은 페이지의 다른 컬럼 화면)'만 합치고, 세로 스크롤 페이지는 버립니다.

    맨 위 페이지 = 첫 행 스킬이 '전체에서 피해량이 가장 큰 스킬'인 이미지들.
    (스크롤한 페이지는 첫 행이 그보다 낮은 스킬이라 자동으로 제외됩니다.)
    """
    valid = [d for d in dfs if isinstance(d, pd.DataFrame) and not d.empty]
    if len(valid) <= 1:
        return valid

    def first_row_name_dmg(d: pd.DataFrame) -> tuple[str, float]:
        try:
            for _, row in d.iterrows():
                dv = parse_korean_number(row.get("피해량")) or 0.0
                nm = str(row.get("이름") or "").strip()
                if dv > 0 or nm:
                    return nm, dv
        except Exception:
            pass
        return "", 0.0

    # 맨 위(스크롤 안 한) 페이지 = '첫 행'의 피해량이 가장 큰 이미지.
    # 주의: 전체 행 중 최대가 아니라 '첫 행'만 봐야 합니다. 하위 행이 OCR로 잘못
    # 부풀려져도(예: 파이널 249억→2,492억) 맨 위 스킬 판정이 흔들리지 않습니다.
    firsts = [first_row_name_dmg(d) for d in valid]
    top_idx = max(range(len(firsts)), key=lambda i: firsts[i][1])
    top_name, top_dmg = firsts[top_idx]
    if top_dmg <= 0:
        return valid

    keep = []
    for d, (nm, dv) in zip(valid, firsts):
        # 첫 행이 맨 위 스킬이면(이름 일치 또는 피해량이 top의 90% 이상) 같은 페이지로 간주.
        if (top_name and nm and nm == top_name) or (dv >= top_dmg * 0.9):
            keep.append(d)
    return keep or valid


def _merge_attack_dfs_by_skill_v135(dfs: list) -> pd.DataFrame:
    """v135c: 여러 공격정보 이미지(아이콘 매칭 완료)의 df를 스킬(이름) 기준으로 병합합니다.
    같은 스킬이 여러 장에 있으면 컬럼별로 '의미있는 값'을 채택(백 화면+헤드 화면 결합),
    다른 스킬이면 행을 추가(스크롤). 미매칭(행 N) 행은 합치지 않고 따로 둡니다. 피해량 내림차순 정렬."""
    import math as _math
    def is_real(v):
        if v is None:
            return False
        if isinstance(v, float) and _math.isnan(v):
            return False
        return str(v).strip() not in ("", "-", "nan", "None")
    valid = [d for d in dfs if isinstance(d, pd.DataFrame) and not d.empty]
    if not valid:
        return pd.DataFrame()
    all_cols: list[str] = []
    for d in valid:
        for c in d.columns:
            if c not in all_cols:
                all_cols.append(c)
    merged: dict[Any, dict[str, Any]] = {}
    order: list[Any] = []
    un = 0
    for d in valid:
        for _, row in d.iterrows():
            name = str(row.get("이름") or "").strip()
            if (not name) or name.startswith("행 ") or name in ("인식 안된 기타 추가딜", "기타 추가딜"):
                key = ("_un", un); un += 1
            else:
                key = name
            if key not in merged:
                merged[key] = {c: row.get(c) for c in all_cols}
                order.append(key)
            else:
                tgt = merged[key]
                for c in all_cols:
                    if not is_real(tgt.get(c)) and is_real(row.get(c)):
                        tgt[c] = row.get(c)
    out = pd.DataFrame([merged[k] for k in order], columns=all_cols)
    # v135e: 방향(백/헤드) 칸은 '빈칸=그 방향 없음(N/A)'을 유지해야 점수가 정상화됩니다.
    # 그래서 0.0으로 채우지 않고, 항상 존재하는 치명타 적중률만 0.0으로 보정합니다.
    for c in ["치명타 적중률"]:
        if c in out.columns:
            out[c] = out[c].apply(lambda v: v if is_real(v) else "0.0")
    try:
        out["_dmg_sort"] = out["피해량"].apply(lambda v: parse_korean_number(v) or 0.0)
        out = out.sort_values("_dmg_sort", ascending=False).drop(columns=["_dmg_sort"]).reset_index(drop=True)
    except Exception:
        pass
    return out


def _apply_browser_ocr_kind_v130(kind: str, title: str, payload: dict[str, Any] | None, summary_image: Image.Image | None, attack_images: Any, *, row_count: int, icon_threshold: float, name_threshold: float, aggregate_unmatched: bool) -> dict[str, Any]:
    payload = payload or {}
    report: dict[str, Any] = {"kind": kind, "title": title, "browser_payload": bool(payload), "v130": True}
    key = f"{kind}_table"
    meta_key = f"{kind}_meta"
    t0 = time.perf_counter()
    # 단일 이미지로 와도 리스트로 통일합니다.
    if attack_images is None:
        attack_images = []
    elif not isinstance(attack_images, (list, tuple)):
        attack_images = [attack_images]
    attack_images = [im for im in attack_images if im is not None]
    meta: dict[str, Any] = st.session_state.get(meta_key, {}) or {}
    if summary_image is not None and isinstance(payload.get("summary"), dict):
        meta = _meta_from_browser_summary_v130(payload.get("summary"))
        st.session_state[meta_key] = meta
        _sync_meta_to_manual_inputs(kind, meta)
        report["summary_browser_ms"] = payload.get("summary", {}).get("browser_ms")

    def _dbg_hit(df, tag):
        try:
            if isinstance(df, pd.DataFrame) and "백어택 적중률" in df.columns:
                report[f"dbg_백적중_{tag}"] = [str(x) for x in df["백어택 적중률"].head(3).tolist()]
                report[f"dbg_치적_{tag}"] = [str(x) for x in df.get("치명타 적중률", pd.Series(dtype=str)).head(3).tolist()]
            else:
                report[f"dbg_백적중_{tag}"] = "no_col"
        except Exception as _e:
            report[f"dbg_백적중_{tag}"] = f"err:{_e!r}"

    attack_payload = payload.get("attack")
    if attack_images and isinstance(attack_payload, dict):
        # 멀티 이미지: images 배열(없으면 단일 rows를 1장으로 취급)
        img_payloads = attack_payload.get("images")
        if not img_payloads:
            img_payloads = [{"rows": attack_payload.get("rows") or []}]
        multi = len(img_payloads) > 1
        report["attack_browser_ms"] = attack_payload.get("browser_ms")
        report["attack_image_count"] = len(img_payloads)
        try:
            skill_icon_candidates = _skill_only_ocr_candidates(get_ocr_skill_candidates_full())
        except Exception:
            skill_icon_candidates = []
        report["candidate_count"] = len(skill_icon_candidates)
        matched: list[pd.DataFrame] = []
        t_icon = time.perf_counter()
        for i, imgp in enumerate(img_payloads):
            img_i = attack_images[i] if i < len(attack_images) else attack_images[0]
            df_i = _browser_attack_to_df_v130({"rows": imgp.get("rows") or []}, img_i, row_count, total_damage=meta.get("total_damage"), fill_missing=(not multi))
            if df_i is None or df_i.empty:
                continue
            if skill_icon_candidates and img_i is not None:
                df_i = correct_battle_skill_names_with_icons(df_i, img_i, skill_icon_candidates, threshold=float(name_threshold), icon_threshold=float(icon_threshold), drop_unmatched=False)
            matched.append(df_i)
        report["server_icon_match_ms"] = round((time.perf_counter() - t_icon) * 1000.0, 3)
        report["attack_rows_before_icon"] = sum(len(d) for d in matched)
        if multi:
            # v136: 기본은 '가로 열만' 병합(맨 위 페이지의 컬럼 화면들만 합침).
            # 세로 스크롤 페이지를 합치려면 사이드바/OCR탭 옵션을 켜세요.
            merge_scroll = bool(st.session_state.get("attack_merge_scroll_pages", False))
            merge_input = matched if merge_scroll else _select_primary_page_dfs_v136(matched)
            report["attack_merge_mode"] = "scroll_union" if merge_scroll else "primary_page_only"
            report["attack_images_used"] = len(merge_input)
            parsed = _merge_attack_dfs_by_skill_v135(merge_input)
        else:
            parsed = matched[0] if matched else pd.DataFrame()
        _dbg_hit(parsed, "after_merge")
        if parsed is not None and not parsed.empty:
            if aggregate_unmatched:
                parsed = _aggregate_unmatched_rows_as_extra(parsed)
            parsed = sanitize_battle_table(parsed)
            parsed = _repair_battle_values_from_summary(parsed, st.session_state.get(meta_key, {}), elapsed_fallback=st.session_state.get(f"{kind}_elapsed_sec"))
            parsed = _drop_obvious_ocr_noise_rows(parsed)
            st.session_state[key] = prepare_battle_editor_df(sanitize_battle_table(parsed))
            _dbg_hit(st.session_state[key], "final_table")
            report["dbg_build_ver"] = "v135c_multi_merge"
            _bump_table_version(kind)
            report["attack_rows_after"] = len(st.session_state[key])
    report["server_apply_ms"] = round((time.perf_counter() - t0) * 1000.0, 3)
    return report


# v133: 공격정보 컬럼을 헤더 가이드 템플릿으로 자동 인식합니다.
# 사용자가 인게임에서 컬럼 순서를 바꾸지 않아도(순서 변경 불필요), 보이는 컬럼을 헤더 내용으로 찾아냅니다.
# 같은 위치에서 백/헤드 헤더가 둘 다 약하게 잡히면(…적중률 유사) 점수 높은 쪽만 남깁니다.
_ATTACK_COL_GUIDE_V133 = {
    "damage": "col_damage",
    "back_hit": "col_back_hit", "back_share": "col_back_share",
    "head_hit": "col_head_hit", "head_share": "col_head_share",
    "crit_hit": "col_crit_hit", "crit_share": "col_crit_share",
    "casts": "col_casts", "cooldown": "col_cooldown",
}
_ATTACK_COL_CACHE_V133: dict[str, Any] = {}


def _guide_dir_v133() -> Path:
    return Path(__file__).parent / "data" / "guide_templates"


def _load_guide_tpl_v133(name: str):
    try:
        import cv2 as _cv2
        import numpy as _np
        p = _guide_dir_v133() / f"{name}.png"
        if not p.exists():
            return None
        t = _cv2.imread(str(p), _cv2.IMREAD_UNCHANGED)
        if t is None:
            return None
        if t.ndim == 3 and t.shape[2] == 4:
            alpha = t[..., 3]
            bgr = t[..., :3].copy()
            bgr[alpha < 20] = 0
            return bgr
        return t[..., :3] if t.ndim == 3 else _cv2.cvtColor(t, _cv2.COLOR_GRAY2BGR)
    except Exception:
        return None


# v134: 기준(1920·타이틀 좌상단 858,78)에서의 공격정보 그리드 비율 상수.
_ATK_TITLE_REF_V134 = (858, 78)
_ATK_FIRST_C_V134 = 315   # 첫 행 중심 y(네이티브)
# v137: 행 간격 45는 게임 실제 간격보다 ~2% 커서 아래 행으로 갈수록 크롭이 밀려
# 마지막 행 숫자가 잘렸습니다(예: 13행 '27.21억' 상단 클리핑). 측정된 누적 밀림(약 16px/13행)
# 기준으로 44로 보정하고, 잔여 오차를 흡수하도록 셀 반높이를 20으로 살짝 키웠습니다.
_ATK_STEP_V134 = 44       # 행 간격(보정)
_ATK_HALF_H_V134 = 20     # 행 셀 반높이(드리프트 허용치 증가)
_ATK_NROWS_V134 = 14
_ATK_HDR_Y1_V134, _ATK_HDR_Y2_V134 = 245, 292  # 헤더 텍스트 띠(네이티브)
_ATK_COL_BASE_V134 = 0.76  # col_*.png 헤더 템플릿의 1920-네이티브 대비 스케일


def _title_template_path_v134() -> Path:
    return Path(__file__).parent / "data" / "title_template.png"


def _compute_attack_grid_v134(pil_image: "Image.Image | None") -> dict[str, Any] | None:
    """전투분석기 창을 타이틀 템플릿으로 '멀티스케일' 매칭해 창의 위치+스케일을 찾고,
    그 기준으로 행 y그리드와 (헤더 가이드 템플릿 매칭으로) 컬럼 x를 계산해 돌려줍니다.

    창은 해상도에 비례해 커지지 않고 UI 스케일에 따라 크기가 달라지므로(울트라와이드 등),
    스케일을 가정하지 않고 매칭으로 직접 찾습니다. 반환값은 모두 0~1 정규화 좌표입니다.
    """
    if pil_image is None:
        return None
    try:
        import cv2 as _cv2
        import numpy as _np
    except Exception:
        return None
    try:
        img = _cv2.cvtColor(_np.asarray(pil_image.convert("RGB")), _cv2.COLOR_RGB2BGR)
        H, W = img.shape[:2]
        # v134d 속도: 매칭은 다운스케일본에서 합니다. 결과 좌표는 0~1 정규화라 동일합니다.
        # (3440 같은 큰 이미지에서 매칭 시간을 ~4배 줄여 업로드시 화면 멈춤/어두워짐을 완화)
        if W > 1700:
            _ds = 1600.0 / W
            img = _cv2.resize(img, (int(W * _ds), int(H * _ds)), interpolation=_cv2.INTER_AREA)
            H, W = img.shape[:2]
        tp = _title_template_path_v134()
        if not tp.exists():
            return {"found": False, "reason": "no_title_template"}
        tpl = _cv2.imread(str(tp))
        if tpl is None:
            return {"found": False, "reason": "title_read_fail"}
        tpl = tpl[:, :, :3]
        th0, tw0 = tpl.shape[:2]
        # v134d 속도: 회색조로 매칭하면 컬러 대비 ~3배 빠릅니다(정확도 동일). 타이틀은 상단 30%만 봅니다.
        imgg = _cv2.cvtColor(img, _cv2.COLOR_BGR2GRAY)
        tplg = _cv2.cvtColor(tpl, _cv2.COLOR_BGR2GRAY)
        # 멀티스케일 타이틀 매칭. 화면 비율과 무관 — 창은 UI 스케일대로 균일 확대/축소되므로
        # '스케일 하나'만 찾으면 됩니다(극단 해상도/UI배율까지 넓게 탐색). coarse→fine 2단계.
        top = imgg[: int(H * 0.30)]

        def _title_scan(scales):
            b = (-1.0, 0, 0, 1.0)
            for s in scales:
                tw, th = int(tw0 * s), int(th0 * s)
                if th > top.shape[0] or tw > W or tw < 10:
                    continue
                res = _cv2.matchTemplate(top, _cv2.resize(tplg, (tw, th)), _cv2.TM_CCOEFF_NORMED)
                _, sc, _, loc = _cv2.minMaxLoc(res)
                if sc > b[0]:
                    b = (float(sc), int(loc[0]), int(loc[1]), float(s))
            return b
        coarse = _title_scan([round(float(x), 3) for x in _np.arange(0.45, 3.01, 0.12)])
        cs = coarse[3]
        fine = _title_scan([round(float(x), 3) for x in _np.arange(max(0.4, cs - 0.12), cs + 0.12, 0.03)])
        best = fine if fine[0] >= coarse[0] else coarse
        tscore, tx, ty, S = best
        if tscore < 0.5:
            return {"found": False, "title_score": round(tscore, 3), "reason": "title_low_score"}

        def my(ry: float) -> float:
            return ty + (ry - _ATK_TITLE_REF_V134[1]) * S

        # 헤더 띠(타이틀 기준) → 컬럼 헤더 매칭(스케일은 S*COL_BASE 근처)
        band_y1, band_y2 = int(my(_ATK_HDR_Y1_V134)), int(my(_ATK_HDR_Y2_V134))
        band_y1 = max(0, band_y1); band_y2 = min(H, band_y2)
        cand: list[dict[str, Any]] = []
        if band_y2 - band_y1 >= 6:
            band = imgg[band_y1:band_y2]
            base = S * _ATK_COL_BASE_V134
            for key, fn in _ATTACK_COL_GUIDE_V133.items():
                tt = _load_guide_tpl_v133(fn)
                if tt is None:
                    continue
                ttg = _cv2.cvtColor(tt, _cv2.COLOR_BGR2GRAY)
                h0, w0 = ttg.shape[:2]
                b2 = (-1.0, 0, 0)
                for cs in [round(float(x), 3) for x in _np.arange(base * 0.82, base * 1.20, 0.05)]:
                    tw, th = int(w0 * cs), int(h0 * cs)
                    if th > band.shape[0] or tw > W or tw < 8:
                        continue
                    res = _cv2.matchTemplate(band, _cv2.resize(ttg, (tw, th)), _cv2.TM_CCOEFF_NORMED)
                    _, s2, _, loc = _cv2.minMaxLoc(res)
                    if s2 > b2[0]:
                        b2 = (float(s2), loc[0] + tw // 2, tw)
                s2, xc, tw = b2
                if s2 >= 0.58:
                    cand.append({"key": key, "xc_px": xc, "halfw_px": tw / 2.0, "score": round(s2, 3)})
        # v143: 헤더는 '백어택 적중률/헤드어택 적중률'처럼 뒷부분('적중률/비중')이 같아
        # 서로 잘 헷갈립니다. 그래서 컬럼 종류(key)와 위치(x)를 '1:1'로 배정합니다.
        # 점수 높은 (key,위치)부터 확정하되, 이미 쓴 key나 이미 찬 위치는 건너뜁니다.
        # 이렇게 하면 head가 back 위치를 가로채거나(오라벨) 일부 컬럼이 통째로 빠지는(back_share/crit_hit
        # 미검출) 문제가 줄어듭니다.
        cand.sort(key=lambda c: -c["score"])
        claimed: list[tuple[float, float]] = []
        claimed_keys: set[str] = set()
        keep: list[dict[str, Any]] = []
        for c in cand:
            if c["key"] in claimed_keys:
                continue
            if any(abs(c["xc_px"] - x) < hw * 0.8 for x, hw in claimed):
                continue
            claimed.append((c["xc_px"], c["halfw_px"]))
            claimed_keys.add(c["key"])
            keep.append(c)
        keep.sort(key=lambda c: c["xc_px"])
        # v143 구조 보정: 방향/치명 컬럼은 항상 '적중률 → 비중' 순서로 '붙어서' 나옵니다.
        # 인접한 두 컬럼이 (적중률, 비중) 쌍인데 종류(백/헤드/치명)가 어긋나면, 더 확실한 쪽으로
        # 짝을 맞춥니다. 예: [head_hit, back_share] 처럼 짝이 깨지면 점수 높은 쪽 기준으로 통일.
        _pair = {"back_hit": "back_share", "head_hit": "head_share", "crit_hit": "crit_share"}
        _rpair = {v: k for k, v in _pair.items()}
        _type_of = {"back_hit": "back", "back_share": "back", "head_hit": "head", "head_share": "head", "crit_hit": "crit", "crit_share": "crit"}
        for i in range(len(keep) - 1):
            a, b = keep[i], keep[i + 1]
            ka, kb = a["key"], b["key"]
            # a=적중률, b=비중 이고 서로 붙어 있는데 종류가 다르면 보정
            if ka in _pair and kb in _rpair and _pair.get(ka) != kb:
                # 인접(가로로 바로 옆)일 때만
                if abs(b["xc_px"] - a["xc_px"]) < (a["halfw_px"] + b["halfw_px"]) * 1.6:
                    if a["score"] >= b["score"]:
                        b["key"] = _pair[ka]      # 적중률 종류에 맞춰 비중을 교체
                    else:
                        a["key"] = _rpair[kb]     # 비중 종류에 맞춰 적중률을 교체
        columns = [{
            "key": c["key"], "xc": round(c["xc_px"] / W, 4),
            "halfw": round(c["halfw_px"] / W, 4), "score": c["score"],
        } for c in keep]
        rows = {
            "first_yc": round(my(_ATK_FIRST_C_V134) / H, 4),
            "step": round(_ATK_STEP_V134 * S / H, 4),
            "half_h": round(_ATK_HALF_H_V134 * S / H, 4),
            "n": _ATK_NROWS_V134,
        }
        return {
            "found": True, "title_score": round(tscore, 3), "scale": round(S, 3),
            "origin": [tx, ty], "image_w": W, "image_h": H, "rows": rows, "columns": columns,
        }
    except Exception as e:
        return {"found": False, "reason": f"err:{e!r}"}


def _attack_columns_arg_v133(pil_image: "Image.Image | None", cache_key: str) -> str:
    """공격정보 그리드(멀티스케일 타이틀+헤더 매칭) 결과를 JSON 문자열로. 같은 이미지는 캐시."""
    import json as _json
    import hashlib as _hashlib
    if pil_image is None:
        return ""
    try:
        try:
            thumb = pil_image.convert("L").resize((32, 32))
            content_hash = _hashlib.md5(thumb.tobytes()).hexdigest()[:12]
        except Exception:
            content_hash = str(pil_image.size)
        sig = f"{cache_key}:{pil_image.size}:{content_hash}"
        cached = _ATTACK_COL_CACHE_V133.get(sig)
        if cached is None:
            cached = _compute_attack_grid_v134(pil_image) or {}
            _ATTACK_COL_CACHE_V133[sig] = cached
        return _json.dumps(cached, ensure_ascii=False)
    except Exception:
        return ""


def _compute_summary_anchor_v134(pil_image: "Image.Image | None") -> dict[str, Any] | None:
    """종합정보 카드도 같은 전투분석기 창이므로, 타이틀을 멀티스케일로 매칭해 창의
    위치+스케일만 돌려줍니다. 브라우저 parseSummary가 1920-기준 박스를 이 앵커로 변환합니다."""
    if pil_image is None:
        return None
    try:
        import cv2 as _cv2
        import numpy as _np
    except Exception:
        return None
    try:
        img = _cv2.cvtColor(_np.asarray(pil_image.convert("RGB")), _cv2.COLOR_RGB2BGR)
        H, W = img.shape[:2]
        orig_W = W
        # v134e: 매칭만 다운스케일본에서 빠르게 하고, origin/scale은 '원본 좌표'로 되돌립니다.
        # (브라우저 parseSummary는 원본 이미지에 origin/scale을 적용하므로, 여기서 원본 기준으로 환산해야 합니다.
        #  이전 v134d는 다운스케일 좌표를 그대로 줘서 박스가 어긋나 종합정보가 전부 빈값이었습니다.)
        ds_factor = 1.0
        if W > 1700:
            ds_factor = 1600.0 / W
            img = _cv2.resize(img, (int(W * ds_factor), int(H * ds_factor)), interpolation=_cv2.INTER_AREA)
            H, W = img.shape[:2]
        tp = _title_template_path_v134()
        if not tp.exists():
            return {"found": False}
        tpl = _cv2.imread(str(tp))
        if tpl is None:
            return {"found": False}
        tpl = tpl[:, :, :3]
        th0, tw0 = tpl.shape[:2]
        imgg = _cv2.cvtColor(img, _cv2.COLOR_BGR2GRAY)
        tplg = _cv2.cvtColor(tpl, _cv2.COLOR_BGR2GRAY)
        top = imgg[: int(H * 0.30)]

        def _scan(scales):
            b = (-1.0, 0, 0, 1.0)
            for s in scales:
                tw, th = int(tw0 * s), int(th0 * s)
                if th > top.shape[0] or tw > W or tw < 10:
                    continue
                res = _cv2.matchTemplate(top, _cv2.resize(tplg, (tw, th)), _cv2.TM_CCOEFF_NORMED)
                _, sc, _, loc = _cv2.minMaxLoc(res)
                if sc > b[0]:
                    b = (float(sc), int(loc[0]), int(loc[1]), float(s))
            return b
        coarse = _scan([round(float(x), 3) for x in _np.arange(0.45, 3.01, 0.12)])
        fine = _scan([round(float(x), 3) for x in _np.arange(max(0.4, coarse[3] - 0.12), coarse[3] + 0.12, 0.03)])
        tscore, tx, ty, S = fine if fine[0] >= coarse[0] else coarse
        if tscore < 0.5:
            return {"found": False, "title_score": round(tscore, 3)}
        # 다운스케일 좌표 → 원본 좌표 환산: origin/scale을 ds로 나눕니다.
        inv = 1.0 / ds_factor if ds_factor else 1.0
        return {"found": True, "title_score": round(tscore, 3), "scale": round(S * inv, 4),
                "origin": [int(round(tx * inv)), int(round(ty * inv))], "title_ref": [858, 78],
                "image_w": orig_W, "image_h": int(round(H * inv))}
    except Exception:
        return None


def _summary_anchor_arg_v134(pil_image: "Image.Image | None", cache_key: str) -> str:
    import json as _json
    import hashlib as _hashlib
    if pil_image is None:
        return ""
    try:
        try:
            thumb = pil_image.convert("L").resize((32, 32))
            content_hash = _hashlib.md5(thumb.tobytes()).hexdigest()[:12]
        except Exception:
            content_hash = str(pil_image.size)
        sig = f"sumanchor:{cache_key}:{pil_image.size}:{content_hash}"
        cached = _ATTACK_COL_CACHE_V133.get(sig)
        if cached is None:
            cached = _compute_summary_anchor_v134(pil_image) or {}
            _ATTACK_COL_CACHE_V133[sig] = cached
        return _json.dumps(cached, ensure_ascii=False)
    except Exception:
        return ""


def _render_client_ocr_component_v130(*, real_summary_up: Any, real_attack_ups: Any, dummy_summary_up: Any, dummy_attack_ups: Any, row_count: int, real_attack_imgs: Any = None, dummy_attack_imgs: Any = None, real_summary_img: "Image.Image | None" = None, dummy_summary_img: "Image.Image | None" = None) -> dict[str, Any] | None:
    if _CLIENT_OCR_COMPONENT_V130 is None:
        st.error("브라우저 OCR 컴포넌트를 불러오지 못했습니다.")
        return None
    import json as _json_c
    real_attack_ups = list(real_attack_ups or [])
    dummy_attack_ups = list(dummy_attack_ups or [])
    real_attack_imgs = list(real_attack_imgs or [])
    dummy_attack_imgs = list(dummy_attack_imgs or [])
    sig = _input_sig_v130(real_summary_up, *real_attack_ups, dummy_summary_up, *dummy_attack_ups)
    # v135c: 공격정보는 여러 장. 이미지별 그리드(멀티스케일 타이틀+헤더 매칭)를 배열로 넘깁니다.
    def _grids(imgs, tag):
        out = []
        for i, im in enumerate(imgs):
            try:
                out.append(_json_c.loads(_attack_columns_arg_v133(im, f"{tag}_{i}") or "{}"))
            except Exception:
                out.append({})
        return out
    real_grids = _grids(real_attack_imgs, "real_attack")
    dummy_grids = _grids(dummy_attack_imgs, "dummy_attack")
    real_urls = [_upload_to_data_url_v130(u) for u in real_attack_ups]
    dummy_urls = [_upload_to_data_url_v130(u) for u in dummy_attack_ups]
    real_summary_anchor = _summary_anchor_arg_v134(real_summary_img, "real_summary")
    dummy_summary_anchor = _summary_anchor_arg_v134(dummy_summary_img, "dummy_summary")
    # v144: 캡처 배치를 보낸 직후 렌더에서는 auto_run_after_capture=True를 넘겨,
    # 브라우저가 '이미지 인식'을 한 번 더 누르지 않아도 곧바로 OCR을 이어가게 합니다.
    _auto_run = bool(st.session_state.get("_auto_run_after_capture", False))
    _res = _CLIENT_OCR_COMPONENT_V130(
        real_summary_data_url=_upload_to_data_url_v130(real_summary_up),
        real_attack_data_urls=_json_c.dumps(real_urls),
        dummy_summary_data_url=_upload_to_data_url_v130(dummy_summary_up),
        dummy_attack_data_urls=_json_c.dumps(dummy_urls),
        row_count=int(row_count or 14),
        real_attack_grids=_json_c.dumps(real_grids, ensure_ascii=False),
        dummy_attack_grids=_json_c.dumps(dummy_grids, ensure_ascii=False),
        real_summary_anchor=real_summary_anchor,
        dummy_summary_anchor=dummy_summary_anchor,
        input_sig=sig,
        auto_run_after_capture=_auto_run,
        cap_counts=_json_c.dumps({
            "real_attack": len(st.session_state.get("cap_real_attack_list") or []),
            "dummy_attack": len(st.session_state.get("cap_dummy_attack_list") or []),
            "real_summary": 1 if st.session_state.get("cap_real_summary") else 0,
            "dummy_summary": 1 if st.session_state.get("cap_dummy_summary") else 0,
        }),
        default=None,
        key="loa_client_ocr_component_v148",
    )
    # 한 번 넘겨줬으면 플래그를 내려서, 이후 일반 렌더에서는 자동 실행이 재발동하지 않게 합니다.
    if _auto_run:
        st.session_state.pop("_auto_run_after_capture", None)
    return _res


def _auto_apply_client_result_v130(client_result: dict[str, Any], *, current_sig: str, real_summary_img: Image.Image | None, real_attack_imgs: Any, dummy_summary_img: Image.Image | None, dummy_attack_imgs: Any, row_count: int, icon_match_threshold: float, name_match_threshold: float, aggregate_unmatched: bool) -> list[dict[str, Any]] | None:
    if not isinstance(client_result, dict) or not client_result.get("done"):
        return None
    if str(client_result.get("input_sig") or "") != str(current_sig or ""):
        return None
    # v135c: 적용 식별자에 '서버 적용 코드 버전'을 포함해, app.py 로직을 고치면
    # (브라우저 OCR을 다시 안 돌려도) 다음 렌더에서 표가 새 로직으로 자동 재생성되게 합니다.
    _APPLY_CODE_VER = "v135e_direction_scoring"
    apply_id = f"{_APPLY_CODE_VER}|{client_result.get('started_at','')}|{client_result.get('input_sig','')}|{client_result.get('timings',{}).get('total_browser_ms','')}"
    if st.session_state.get("last_applied_client_ocr_v130") == apply_id:
        return None
    reports: list[dict[str, Any]] = []
    if client_result.get("real"):
        reports.append(_apply_browser_ocr_kind_v130(
            "real", "실전 전투분석기", client_result.get("real"), real_summary_img, real_attack_imgs,
            row_count=int(row_count), icon_threshold=float(icon_match_threshold)/100.0,
            name_threshold=float(name_match_threshold)/100.0, aggregate_unmatched=bool(aggregate_unmatched),
        ))
    if client_result.get("dummy"):
        reports.append(_apply_browser_ocr_kind_v130(
            "dummy", "허수/기준 전투분석기", client_result.get("dummy"), dummy_summary_img, dummy_attack_imgs,
            row_count=int(row_count), icon_threshold=float(icon_match_threshold)/100.0,
            name_threshold=float(name_match_threshold)/100.0, aggregate_unmatched=bool(aggregate_unmatched),
        ))
    # 화면 공유 캡처 프레임은 인식 완료 후 삭제되므로, 사이드바 디버그(통합/아이콘 매칭)에서
    # 다시 쓸 수 있도록 실전 공격/종합 이미지를 지속 키에 PNG 바이트로 저장해 둡니다.
    try:
        _ra_list = list(real_attack_imgs or [])
        if _ra_list and _ra_list[0] is not None:
            st.session_state["real_attack_image_bytes"] = _png_bytes(_ra_list[0])
        if real_summary_img is not None:
            st.session_state["real_summary_image_bytes"] = _png_bytes(real_summary_img)
    except Exception:
        pass
    # 아르카나 전용: 실전/허수 종합정보 이미지에서 '카드 사용 횟수'를 템플릿 매칭+OCR로 자동 감지.
    try:
        if _arcana_card_context():
            if real_summary_img is not None:
                _det = _detect_arcana_card_uses(real_summary_img)
                if _det and _det.get("value"):
                    st.session_state["arcana_card_uses_ocr"] = int(_det["value"])
                    st.session_state["arcana_card_uses_ocr_score"] = _det.get("score")
            if dummy_summary_img is not None:
                _detd = _detect_arcana_card_uses(dummy_summary_img)
                if _detd and _detd.get("value"):
                    st.session_state["arcana_card_uses_ocr_dummy"] = int(_detd["value"])
                    st.session_state["arcana_card_uses_ocr_dummy_score"] = _detd.get("score")
    except Exception:
        pass
    st.session_state["client_ocr_reports_v130"] = reports
    st.session_state["last_applied_client_ocr_v130"] = apply_id
    return reports


def _imgs_from_uploads_v135(ups: Any) -> list:
    """멀티 file_uploader 결과(리스트/단일)를 PIL 이미지 리스트로 변환합니다."""
    if ups is None:
        return []
    if not isinstance(ups, (list, tuple)):
        ups = [ups]
    out = []
    for u in ups:
        if u is None:
            continue
        try:
            out.append(Image.open(io.BytesIO(u.getvalue())).convert("RGB"))
        except Exception:
            pass
    return out


def ocr_tab() -> None:  # type: ignore[override]
    st.header("전투분석기 입력")
    st.caption("화면 공유로 전투분석기 종합정보·공격정보 화면을 캡처해 자동으로 인식합니다.")
    _render_step_guide(2, "전투분석기 화면 공유로 인식하기", [
        "아래에서 <b>화면 공유 시작</b>을 누르고, 전투분석기의 종합정보·공격정보 화면을 각각 캡처하세요.",
        "네 종류(허수 종합/공격, 실전 종합/공격)를 모두 캡처한 뒤 <b>이미지 인식</b>을 누르면 결과표에 자동 반영됩니다.",
        "잘못 캡처했으면 각 항목의 <b>취소</b> 버튼으로 지우고 다시 캡처할 수 있습니다.",
    ])

    st.info("화면 공유로 캡처하면 파일 업로드 없이 바로 인식됩니다. 인식이 끝나면 아래 검수표에 반영되며, 틀린 값은 표에서 직접 수정하세요.")

    with st.container(border=True):
        # 화면 공유 캡처 전용 입력. 파일 업로드는 사용하지 않습니다.
        real_attack_ups = []
        dummy_attack_ups = []
        real_summary_up = None
        dummy_summary_up = None
        # 화면 공유로 캡처한 프레임(세션 저장)을 입력으로 사용합니다.
        # 캡처는 서버 그리드+OCR 파이프라인을 그대로 타므로 정확도가 동일합니다.
        _cap_real_atk = list(st.session_state.get("cap_real_attack_list") or [])
        _cap_dummy_atk = list(st.session_state.get("cap_dummy_attack_list") or [])
        real_attack_ups = _cap_real_atk + real_attack_ups
        dummy_attack_ups = _cap_dummy_atk + dummy_attack_ups
        real_summary_up = real_summary_up or st.session_state.get("cap_real_summary")
        dummy_summary_up = dummy_summary_up or st.session_state.get("cap_dummy_summary")
        real_summary_img = _image_from_uploader_or_session_v115(real_summary_up, "real", "summary")
        dummy_summary_img = _image_from_uploader_or_session_v115(dummy_summary_up, "dummy", "summary")
        real_attack_imgs = _imgs_from_uploads_v135(real_attack_ups)
        dummy_attack_imgs = _imgs_from_uploads_v135(dummy_attack_ups)

        # v137: 이미지 인식 세부 설정은 화면에 노출하지 않고 기본값을 사용합니다.
        row_count = int(st.session_state.get("unified_row_count_v130", 14) or 14)
        icon_match_threshold = int(st.session_state.get("unified_icon_threshold_v130", 74) or 74)
        name_match_threshold = int(st.session_state.get("unified_name_threshold_v130", 52) or 52)
        aggregate_unmatched = bool(st.session_state.get("unified_aggregate_unmatched_v130", False))

        # v137: 화면 공유 캡처 현황 + 초기화
        _cap_count = len(_cap_real_atk) + len(_cap_dummy_atk) + (1 if st.session_state.get("cap_real_summary") else 0) + (1 if st.session_state.get("cap_dummy_summary") else 0)
        if _cap_count:
            cc1, cc2 = st.columns([3, 1])
            cc1.caption(f"🖥️ 화면 공유 캡처 {_cap_count}장 사용 중 · 실전 공격 {len(_cap_real_atk)} / 허수 공격 {len(_cap_dummy_atk)} / 종합 {(1 if st.session_state.get('cap_real_summary') else 0)+(1 if st.session_state.get('cap_dummy_summary') else 0)}")
            if cc2.button("캡처 초기화", key="clear_captures_v137"):
                for _k in ["cap_real_attack_list", "cap_dummy_attack_list", "cap_real_summary", "cap_dummy_summary", "_last_capture_ts", "_auto_run_after_capture"]:
                    st.session_state.pop(_k, None)
                st.rerun()

        if not any([real_summary_up, real_attack_ups, dummy_summary_up, dummy_attack_ups]):
            st.info("이미지를 올리거나, 아래 컴포넌트에서 '화면 공유'로 캡처해 인식할 수 있습니다.")
        current_sig = _input_sig_v130(real_summary_up, *real_attack_ups, dummy_summary_up, *dummy_attack_ups)
        client_result = _render_client_ocr_component_v130(
            real_summary_up=real_summary_up,
            real_attack_ups=real_attack_ups,
            dummy_summary_up=dummy_summary_up,
            dummy_attack_ups=dummy_attack_ups,
            row_count=int(row_count),
            real_attack_imgs=real_attack_imgs,
            dummy_attack_imgs=dummy_attack_imgs,
            real_summary_img=real_summary_img,
            dummy_summary_img=dummy_summary_img,
        )
        if isinstance(client_result, dict):
            st.session_state["last_client_ocr_v130"] = client_result
            # v137: 화면 공유 캡처 프레임이 오면 세션에 저장하고 rerun → 다음 렌더에서
            # 업로드와 동일하게 서버 그리드+OCR을 태웁니다.
            if _handle_screen_capture_result_v137(client_result):
                st.rerun()
            if _handle_clear_capture_v144(client_result):
                st.rerun()
            if client_result.get("error"):
                st.error(str(client_result.get("error")))
            elif client_result.get("done"):
                st.success(f"이미지 인식 완료 · {client_result.get('timings', {}).get('total_browser_ms', '-')} ms")
                reports_now = _auto_apply_client_result_v130(
                    client_result,
                    current_sig=current_sig,
                    real_summary_img=real_summary_img,
                    real_attack_imgs=real_attack_imgs,
                    dummy_summary_img=dummy_summary_img,
                    dummy_attack_imgs=dummy_attack_imgs,
                    row_count=int(row_count),
                    icon_match_threshold=float(icon_match_threshold),
                    name_match_threshold=float(name_match_threshold),
                    aggregate_unmatched=bool(aggregate_unmatched),
                )
                if reports_now is not None:
                    st.success("이미지 인식 결과를 표에 자동 적용했습니다. 아래 표에서 바로 검수/수정하세요.")
                    # v138: 인식 완료 후 캡처해 둔 프레임을 완전히 삭제(메모리 확보).
                    if any(st.session_state.get(k) for k in ["cap_real_attack_list", "cap_dummy_attack_list", "cap_real_summary", "cap_dummy_summary"]):
                        _clear_captured_frames_v138()
                        st.rerun()
        reports = st.session_state.get("client_ocr_reports_v130") or st.session_state.get("client_ocr_reports_v126") or []
        if reports:
            st.caption("이미지 인식 적용/후처리 시간")
            try:
                rdf = pd.DataFrame(reports)
                show_cols = ["title", "summary_browser_ms", "attack_browser_ms", "server_icon_match_ms", "server_apply_ms", "attack_rows_after"]
                st.dataframe(rdf[[c for c in show_cols if c in rdf.columns]], use_container_width=True, hide_index=True)
            except Exception:
                st.json(reports)

    real_tab, dummy_tab = st.tabs(["실전 검수", "허수/기준 검수"])
    with real_tab:
        _render_battle_review_only_v115("real", "실전 전투분석기")
    with dummy_tab:
        _render_battle_review_only_v115("dummy", "허수/기준 전투분석기")


def _render_arcana_card_review_v152(kind: str) -> None:
    """아르카나면 검수 탭(실전/허수)에 '카드 사용 횟수'(수정 가능)를 표시합니다.

    - 실전 탭 값(arcana_card_uses)이 확률 보정에 반영됩니다.
    - 허수 탭 값(arcana_card_uses_dummy)은 검수/기록용 참고 값입니다.
    - 실전 탭에는 각인(황제의 칙령/황후의 은총) 자동 감지가 틀릴 때를 대비한 수동 지정도 둡니다.
    """
    try:
        card_ctx = _arcana_card_context()
    except Exception:
        card_ctx = None
    if not card_ctx:
        return
    is_real = (kind == "real")
    uses_key = "arcana_card_uses" if is_real else "arcana_card_uses_dummy"
    ocr_key = "arcana_card_uses_ocr" if is_real else "arcana_card_uses_ocr_dummy"
    ocr_score_key = "arcana_card_uses_ocr_score" if is_real else "arcana_card_uses_ocr_dummy_score"
    _ocr = st.session_state.get(ocr_key)
    with st.container(border=True):
        eng = card_ctx.get("engraving") or "-"
        src = card_ctx.get("engraving_source") or ""
        knight_lv = card_ctx.get("knight_of_empress_level")
        _cards = list((card_ctx.get("cards") or {}).keys())
        st.markdown(f"**🃏 아르카나 카드 확률 보정 · {eng}**")
        _cap = f"각인 감지: {eng} ({src}) · 적용 카드: {', '.join(_cards) if _cards else '없음'}"
        if card_ctx.get("knight_of_empress_gated"):
            _cap += " · 황후의 기사=깨달음 아크패시브 미습득이라 제외"
        elif knight_lv:
            _cap += f" · 황후의 기사 노드 {int(knight_lv)}Lv"
        st.caption(_cap)
        # 최초 렌더 시 OCR 감지값으로 자동 채움(이후 사용자 수정 우선).
        if uses_key not in st.session_state:
            st.session_state[uses_key] = int(_ocr or 0)
        c1, c2 = st.columns(2)
        with c1:
            if _ocr:
                _sc = st.session_state.get(ocr_score_key)
                _sctxt = f" (매칭 {float(_sc):.0%})" if _sc is not None else ""
                st.caption(f"🔎 종합정보 이미지에서 감지: **{int(_ocr):,}회**{_sctxt}")
                if int(st.session_state.get(uses_key, 0) or 0) != int(_ocr):
                    if st.button(f"이미지 감지값 {int(_ocr):,}회 적용", key=f"apply_arcana_ocr_uses_{kind}", use_container_width=True):
                        st.session_state[uses_key] = int(_ocr)
                        st.rerun()
            st.number_input(
                "카드 사용 횟수",
                min_value=0,
                step=1,
                key=uses_key,
                help=(
                    "전투분석기 '종합 정보 → 추가 정보 → 카드 사용 횟수' 값입니다. "
                    "이미지 인식 시 자동으로 채워지고, 틀리면 직접 고치세요. "
                    + ("이 실전 값이 황제·황후의 기사 카드 확률 보정에 반영됩니다." if is_real else "허수/기준 검수용 참고 값입니다.")
                ),
            )
        if is_real:
            with c2:
                _opts = ["자동 감지", "황제의 칙령", "황후의 은총"]
                _cur = st.session_state.get("arcana_engraving_override", "") or "자동 감지"
                if _cur not in _opts:
                    _cur = "자동 감지"
                _sel = st.selectbox(
                    "각인 수동 지정",
                    _opts,
                    index=_opts.index(_cur),
                    key="arcana_engraving_override_sel",
                    help="각인 자동 감지가 틀리면 여기서 직접 지정하세요. '황제의 칙령'이면 황제 카드가 항상 포함됩니다.",
                )
                _newov = "" if _sel == "자동 감지" else _sel
                if _newov != (st.session_state.get("arcana_engraving_override", "") or ""):
                    st.session_state["arcana_engraving_override"] = _newov
                    st.rerun()


def _render_battle_review_only_v115(kind: str, title: str) -> None:  # type: ignore[override]
    """v130: 기본을 수정 가능 표로 둡니다. 내부 컬럼은 숨기고 분석용 컬럼만 표시합니다."""
    key = f"{kind}_table"
    meta_key = f"{kind}_meta"
    st.subheader(title)
    st.session_state.setdefault(key, empty_battle_table())
    show_summary_meta(st.session_state.get(meta_key, {}), title)
    _manual_meta_editor(kind, title)
    _render_arcana_card_review_v152(kind)

    editor_df_full = prepare_battle_editor_df(sanitize_battle_table(st.session_state.get(key, empty_battle_table())))
    display_cols = [c for c in _BATTLE_ANALYSIS_DISPLAY_COLS_V122 if c in editor_df_full.columns]
    if not display_cols:
        display_cols = [c for c in editor_df_full.columns if not str(c).startswith("_")]
    edit_base = editor_df_full[display_cols].copy() if display_cols else editor_df_full.copy()
    st.caption("표는 기본 수정 가능 상태입니다. OCR이 틀린 피해량/비중/횟수는 여기서 바로 고치면 결과 계산에 반영됩니다.")
    edited = render_compact_battle_editor(kind, edit_base)
    merged = editor_df_full.copy()
    for c in edited.columns:
        merged[c] = edited[c]
    st.session_state[key] = prepare_battle_editor_df(sanitize_battle_table(merged))

    if st.checkbox("계산용 숫자 변환 확인 열기", value=False, key=f"{kind}_norm_preview_v130"):
        try:
            norm_preview = normalize_battle_df(st.session_state[key]).copy()
            if not norm_preview.empty:
                show_cols = [c for c in ["name", "damage", "dps", "share_rate", "back_attack_share", "head_attack_share", "crit_share", "casts"] if c in norm_preview.columns]
                st.markdown(_v120_fast_html_table(norm_preview[show_cols], max_rows=30, max_cols=12), unsafe_allow_html=True)
            else:
                st.info("변환할 전투분석기 행이 없습니다.")
        except Exception as e:
            st.warning(f"숫자 변환 확인 중 오류: {e}")


_sidebar_controls_prev_v130 = globals().get("sidebar_controls")
def sidebar_controls() -> None:  # type: ignore[override]
    if callable(_sidebar_controls_prev_v130):
        _sidebar_controls_prev_v130()
    try:
        pass  # 버전 변경 안내 제거
        st.sidebar.caption("업로드한 억/만/조 템플릿을 영문 파일명으로 저장하고, 피해량 단위 매칭에 사용합니다.")
    except Exception:
        pass
    # 아르카나 카드 사용 횟수 입력은 ②전투분석기 입력 → 실전/허수 검수 탭으로 이동했습니다
    # (전투시간·총피해량·DPS 옆에서 각각 확인/수정). 사이드바 중복 입력은 제거합니다.


def main() -> None:  # type: ignore[override]
    init_state()
    sidebar_controls()
    inject_user_dashboard_css()
    _inject_ux_polish_css_v48()
    _inject_v120_static_css()
    _v121_inject_css()
    st.markdown(
        """
        <div class="loa-hero">
            <h1>⚔️ LOA 실전 효율 분석기</h1>
            <p>허수아비 기준 대비 실전 수행률을 치명타·확률 효과까지 보정해 분석합니다.</p>
        </div>
        """,
        unsafe_allow_html=True,
    )
    # 아이콘 매칭 디버그를 돌리면 다운로드 없이 여기서 바로 결과를 확인할 수 있습니다.
    # 기본적으로 숨김. 신규 캐릭터 작업 시 사이드바의 '🔧 디버그 도구 표시'를 켜면 다시 나타납니다.
    if st.session_state.get("show_debug_tools", False):
        _render_icon_match_inline_result_v150()
    page = st.radio(
        "화면",
        ["① 캐릭터 세팅", "② 전투분석기 입력", "③ 실력 분석 결과"],
        horizontal=True,
        label_visibility="collapsed",
        key="main_page_v153",
    )
    if page == "① 캐릭터 세팅":
        api_tab()
    elif page == "② 전투분석기 입력":
        ocr_tab()
    else:
        result_tab()


def _make_real_integrated_debug_zip_v106(summary_image: Any, attack_image: Any) -> tuple[Path, dict[str, Any]]:  # type: ignore[override]
    import json as _json_dbg_v130
    import zipfile as _zipfile_dbg_v130
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    zpath = EXPORT_DIR / f"real_integrated_debug_browser_v130_{ts}.zip"
    context: dict[str, Any] = {
        "version": "v130_english_unit_templates",
        "app_build": "v135c_multi_attack_merge",
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "server_rapidocr_removed": True,
        "unit_templates_english": True,
        "unit_template_files": ["attack_empty.png", "attack_unit_man.png", "attack_unit_eok.png", "attack_unit_jo.png", "summary_empty.png", "summary_unit_eok.png", "summary_unit_jo.png", "summary_percent.png"],
        "client_ocr_payload_present": bool(st.session_state.get("last_client_ocr_v130")),
        "client_ocr_reports_present": bool(st.session_state.get("client_ocr_reports_v130")),
        "errors": [],
    }
    def _write_df(zout, name: str, df: Any) -> None:
        try:
            if isinstance(df, pd.DataFrame):
                zout.writestr(name, df.to_csv(index=False).encode("utf-8-sig"))
        except Exception as e:
            context.setdefault("errors", []).append({"stage": f"write_{name}", "error": repr(e)})

    import base64 as _b64_dbg_v131
    import copy as _copy_dbg_v131

    def _decode_png_data_url(data_url: Any) -> bytes | None:
        try:
            s = str(data_url or "")
            if "," in s:
                s = s.split(",", 1)[1]
            if not s:
                return None
            return _b64_dbg_v131.b64decode(s)
        except Exception:
            return None

    def _write_browser_ocr_crops(zout, payload: dict[str, Any]) -> None:
        """v131/v135c: 브라우저 OCR 행별 crop(원본/이진화)+오버레이+진단표를 ocr_crops/로 풀어 씁니다.
        공격정보가 여러 장이면 img0/img1... 하위 폴더로 각각 씁니다."""
        def _write_one(base: str, dbg: dict, title: str = "공격정보") -> None:
            if not isinstance(dbg, dict) or not dbg:
                return
            rows = dbg.get("rows") or []
            ov = _decode_png_data_url(dbg.get("overlay"))
            if ov:
                zout.writestr(f"{base}/_row_overlay.png", ov)
            md = [
                f"# {title} OCR 진단 (v140)",
                "",
                f"- 이미지 크기: {dbg.get('image_size')}",
                f"- 그리드 기준: {dbg.get('grid_source')} / 컬럼: {dbg.get('col_source')}",
                f"- 서버 그리드/앵커: {dbg.get('server_grid')}",
                f"- 사용한 컬럼: {dbg.get('columns_used')}",
                "",
                "각 셀의 `_raw.png`(원본 crop)과 `_bin.png`(이진화)을 보고, 박스가 어긋났는지 / crop이 잘렸는지 확인하세요.",
                "",
                "| row | col | box(px) | raw_ocr | 결과 | 파일 |",
                "|---|---|---|---|---|---|",
            ]
            for r in rows:
                try:
                    i = int(r.get("idx"))
                except Exception:
                    i = r.get("idx")
                colname = str(r.get("col") or "damage")
                raw_png = _decode_png_data_url(r.get("crop_raw"))
                bin_png = _decode_png_data_url(r.get("crop_bin"))
                tag = f"{i:02d}" if isinstance(i, int) else str(i)
                fbase = f"row{tag}_{colname}"
                if raw_png:
                    zout.writestr(f"{base}/{fbase}_raw.png", raw_png)
                if bin_png:
                    zout.writestr(f"{base}/{fbase}_bin.png", bin_png)
                md.append(f"| {i} | {colname} | {r.get('box_px')} | `{r.get('raw_ocr') or ''}` | {r.get('reconstructed') or ''} | {fbase}_raw/bin.png |")
            zout.writestr(f"{base}/_diagnosis.md", ("\n".join(md)).encode("utf-8"))
        for kind in ("real", "dummy"):
            atk = ((payload.get(kind) or {}).get("attack") or {})
            images = atk.get("images")
            if images:
                for idx, imgp in enumerate(images):
                    _write_one(f"ocr_crops/{kind}_attack/img{idx}", (imgp or {}).get("debug") or {}, title="공격정보")
            else:
                _write_one(f"ocr_crops/{kind}_attack", atk.get("debug") or {}, title="공격정보")
            # v140: 종합정보(전투시간/총피해량/DPS) crop도 함께 씁니다.
            summ = ((payload.get(kind) or {}).get("summary") or {})
            _write_one(f"ocr_crops/{kind}_summary", summ.get("debug") or {}, title="종합정보")

    def _strip_crops_from_payload(payload: dict[str, Any]) -> dict[str, Any]:
        """client_ocr_v130.json이 base64 crop으로 비대해지지 않도록 crop 문자열만 제거한 사본을 만듭니다."""
        try:
            p = _copy_dbg_v131.deepcopy(payload)
        except Exception:
            return payload
        def _strip_dbg(dbg):
            if not isinstance(dbg, dict):
                return
            if dbg.get("overlay"):
                dbg["overlay"] = "<exported png>"
            for r in (dbg.get("rows") or []):
                if r.get("crop_raw"):
                    r["crop_raw"] = "<exported png>"
                if r.get("crop_bin"):
                    r["crop_bin"] = "<exported png>"
        for kind in ("real", "dummy"):
            atk = ((p.get(kind) or {}).get("attack") or {})
            if atk.get("images"):
                for imgp in atk.get("images"):
                    _strip_dbg((imgp or {}).get("debug") or {})
            else:
                _strip_dbg(atk.get("debug") or {})
            # v140: 종합정보 debug crop도 JSON에서 제거(용량 절약).
            _strip_dbg(((p.get(kind) or {}).get("summary") or {}).get("debug") or {})
        return p

    _browser_payload = st.session_state.get("last_client_ocr_v130") or {}
    with _zipfile_dbg_v130.ZipFile(zpath, "w", compression=_zipfile_dbg_v130.ZIP_DEFLATED) as zout:
        zout.writestr("context/context.json", _json_dbg_v130.dumps(context, ensure_ascii=False, indent=2, default=str))
        try:
            unit_manifest = (Path(__file__).parent / "data" / "unit_templates" / "manifest.json").read_text(encoding="utf-8")
            zout.writestr("context/unit_templates_manifest_v130.json", unit_manifest)
        except Exception as e:
            context.setdefault("errors", []).append({"stage": "write_unit_template_manifest", "error": repr(e)})
        try:
            _write_browser_ocr_crops(zout, _browser_payload)
        except Exception as e:
            context.setdefault("errors", []).append({"stage": "write_browser_ocr_crops", "error": repr(e)})
        zout.writestr("context/client_ocr_v130.json", _json_dbg_v130.dumps({
            "payload": _strip_crops_from_payload(_browser_payload),
            "reports": st.session_state.get("client_ocr_reports_v130") or [],
            "real_meta": st.session_state.get("real_meta") or {},
            "dummy_meta": st.session_state.get("dummy_meta") or {},
        }, ensure_ascii=False, indent=2, default=str))
        _write_df(zout, "context/real_table.csv", st.session_state.get("real_table"))
        _write_df(zout, "context/dummy_table.csv", st.session_state.get("dummy_table"))
        # v140: 종합정보 파싱 결과를 한눈에 검수할 수 있는 요약 CSV.
        try:
            _srows = []
            for _kind, _mk in (("real", "real_meta"), ("dummy", "dummy_meta")):
                _m = st.session_state.get(_mk) or {}
                _raw = _m.get("raw") or {}
                _srows.append({
                    "구분": _kind,
                    "전투시간(초)": _m.get("elapsed_seconds"),
                    "전투시간_raw": _raw.get("elapsed_text"),
                    "총피해량": _m.get("total_damage"),
                    "총피해량_표시": _m.get("total_damage_text"),
                    "총피해_흰raw": _raw.get("total_damage_text"),
                    "총피해_노랑raw": _raw.get("total_damage_yellow_text"),
                    "DPS": _m.get("dps"),
                    "DPS_표시": _m.get("dps_text"),
                    "DPS_흰raw": _raw.get("dps_text"),
                })
            _write_df(zout, "context/summary_values.csv", pd.DataFrame(_srows))
        except Exception as e:
            context.setdefault("errors", []).append({"stage": "write_summary_values", "error": repr(e)})
        try:
            if summary_image is not None:
                bio = io.BytesIO(); summary_image.save(bio, format="PNG"); zout.writestr("images/real_summary.png", bio.getvalue())
            if attack_image is not None:
                bio = io.BytesIO(); attack_image.save(bio, format="PNG"); zout.writestr("images/real_attack.png", bio.getvalue())
        except Exception as e:
            context.setdefault("errors", []).append({"stage": "write_images", "error": repr(e)})
    return zpath, context


def _public_mode_enabled_v153() -> bool:
    return str(_loa_early_os.environ.get("LOA_SHOW_INTERNAL_UI", "0")).strip().lower() not in {"1", "true", "yes", "on"}


def _lostark_api_token_v153() -> str:
    token = str(_loa_early_os.environ.get("LOSTARK_API_KEY") or _loa_early_os.environ.get("LOA_API_TOKEN") or "").strip()
    if token:
        return token
    try:
        token = str(st.secrets.get("lostark_api_key", "") or st.secrets.get("LOSTARK_API_KEY", "")).strip()
    except Exception:
        token = ""
    return token


_render_api_timing_caption_v117_private = globals().get("_render_api_timing_caption_v117")
def _render_api_timing_caption_v117(timing: dict[str, Any]) -> None:  # type: ignore[override]
    if _public_mode_enabled_v153():
        return
    if callable(_render_api_timing_caption_v117_private):
        _render_api_timing_caption_v117_private(timing)


_v121_inject_css_private_v153 = globals().get("_v121_inject_css")
def _v121_inject_css() -> None:  # type: ignore[override]
    if callable(_v121_inject_css_private_v153):
        _v121_inject_css_private_v153()
    st.markdown(
        """
        <style>
        div[role="radiogroup"] {
            width: min(100%, 980px);
            display: grid;
            grid-template-columns: repeat(3, minmax(230px, 1fr));
            gap: 18px;
            align-items: stretch;
            margin: 18px 0 28px;
        }
        div[role="radiogroup"] > label {
            min-width: 0;
            min-height: 64px;
            display: flex !important;
            align-items: center !important;
            justify-content: center !important;
            gap: 10px !important;
            padding: 0 22px !important;
            border: 1px solid rgba(120,150,210,.42);
            border-radius: 14px;
            background: linear-gradient(180deg, rgba(31,43,68,.92), rgba(18,25,42,.92));
            box-shadow: 0 8px 24px rgba(0,0,0,.18);
        }
        div[role="radiogroup"] > label:has(input:checked) {
            border-color: rgba(255,75,82,.86);
            background: linear-gradient(180deg, rgba(86,35,50,.98), rgba(33,43,74,.98));
        }
        div[role="radiogroup"] p {
            font-size: 17px !important;
            font-weight: 800 !important;
            line-height: 1.2 !important;
            white-space: nowrap !important;
            margin: 0 !important;
        }
        .loa-preset-card {
            border: 1px solid rgba(87,132,210,.48);
            border-radius: 12px;
            background: linear-gradient(135deg, rgba(18,45,78,.96), rgba(18,24,39,.98));
            padding: 20px 22px;
            margin: 8px 0 18px;
        }
        .loa-preset-card h3 {
            margin: 0 0 10px;
            font-size: 22px;
        }
        .loa-preset-card p, .loa-preset-card li {
            font-size: 16px;
            line-height: 1.65;
        }
        .loa-preset-card b {
            color: #ffd36e;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def sidebar_controls() -> None:  # type: ignore[override]
    if not _public_mode_enabled_v153():
        _sidebar_controls_prev_v130()
        return

    calc_config = load_yaml(str(CONFIG_DIR / "calculation_presets.yaml"))
    defaults = calc_config.get("defaults", {})
    st.session_state.setdefault("api_crit_basis", "백어택 기준(조건부 포함)")
    st.session_state.setdefault("target_crit_percent", float(defaults.get("target_crit_percent", 100.0)))
    st.session_state.setdefault("target_back_percent", float(defaults.get("target_back_attack_percent", 100.0)))
    st.session_state.setdefault("back_bonus_percent", float(defaults.get("back_attack_damage_bonus_percent", 5.0)))
    st.session_state["show_debug_tools"] = False

    with st.sidebar:
        st.markdown("### LOA 실전 분석")
        with st.form("api_search_form_public_v153", clear_on_submit=False):
            character_name = st.text_input("캐릭터명", value=st.session_state.get("character_name", ""), key="character_name")
            st.number_input(
                "목표 상승률(%)",
                value=float(st.session_state.get("target_gain_percent", defaults.get("target_gain_percent", 5.0))),
                min_value=0.0,
                max_value=100.0,
                step=0.5,
                key="target_gain_percent",
            )
            fetch = st.form_submit_button("검색", type="primary", use_container_width=True)

        if fetch:
            search_t0 = _time_v120_app.perf_counter()
            token = _lostark_api_token_v153()
            st.session_state["main_page_v153"] = "① 캐릭터 세팅"
            st.session_state["_search_click_t0_v120"] = search_t0
            if not character_name:
                st.warning("캐릭터명을 입력해주세요.")
            elif not token:
                st.error("API 키가 설정되지 않았습니다.")
            else:
                with st.spinner("캐릭터 세팅을 불러오는 중..."):
                    result, timing = _fetch_api_bundle_timed_v117(token, character_name)
                    t_assign = _time_v120_app.perf_counter()
                    st.session_state.api_bundle = result["api_bundle"]
                    st.session_state.api_summary = result["api_summary"]
                    timing["state_assign_ms"] = round((_time_v120_app.perf_counter() - t_assign) * 1000.0, 3)
                    timing["search_button_wall_before_render_ms_v121"] = round((_time_v120_app.perf_counter() - search_t0) * 1000.0, 3)
                    st.session_state.api_last_timing_v120 = timing
                    st.session_state.api_last_timing_v119 = timing
                    st.session_state.api_last_timing_v117 = timing
                    st.session_state.api_last_timing_v116 = timing
                    st.session_state.api_last_timing_v114 = timing

        st.session_state["show_debug_tools"] = False

        summary = st.session_state.get("api_summary") or {}
        profile = summary.get("profile_summary", {}) if isinstance(summary, dict) else {}
        if profile:
            st.divider()
            st.markdown(f"**{profile.get('캐릭터명') or st.session_state.get('character_name','')}**")
            st.caption(f"{profile.get('클래스명', profile.get('클래스', '-'))} · Lv.{profile.get('아이템레벨', '-')}")


_api_tab_private_v153 = globals().get("api_tab")
def api_tab() -> None:  # type: ignore[override]
    if not _public_mode_enabled_v153() and callable(_api_tab_private_v153):
        _api_tab_private_v153()
        return

    t0 = _time_v120_app.perf_counter()
    st.header("캐릭터 세팅 확인")
    st.caption("캐릭터 세팅과 계산에 사용된 값을 한눈에 확인할 수 있습니다.")
    summary = st.session_state.get("api_summary") or {}
    if not summary:
        st.info("왼쪽에서 캐릭터명을 넣고 검색하면 여기에 표시됩니다.")
        st.markdown('<div class="loa-v120-placeholder"></div><div class="loa-v120-placeholder"></div>', unsafe_allow_html=True)
        return
    if isinstance(summary, dict) and summary.get("estimator_error"):
        st.error(f"계산표 생성 오류: {summary.get('estimator_error')}")

    profile = summary.get("profile_summary", {}) if isinstance(summary, dict) else {}
    _v121_cards([
        ("캐릭터", profile.get("캐릭터명") or st.session_state.get("character_name", "-") or "-"),
        ("클래스", profile.get("클래스") or profile.get("클래스명") or "-"),
        ("직업", profile.get("직업") or "-"),
        ("아이템 레벨", profile.get("아이템레벨") or "-"),
    ])
    final_df = summary.get("arkgrid_final_skill_estimates")
    if not isinstance(final_df, pd.DataFrame) or final_df.empty:
        final_df = summary.get("skill_crit_estimates")
    if not isinstance(final_df, pd.DataFrame) or final_df.empty:
        _v121_render_table("최종 계산표", _v121_final_calc(summary), 1000, "예상 치명 확률은 백어택 기준/시너지/조건부를 포함한 최종 계산표 기준입니다.", full_height=True)
        _v121_render_table("전투/캐릭터 요약", _v121_combat_overview(summary), 80)
        _v121_render_table("전투 스킬 - 채용 스킬만", _v121_adopted_skills(summary), 60)
        timing = st.session_state.get("api_last_timing_v120") or {}
        if isinstance(timing, dict):
            timing["api_tab_render_ms_v153"] = round((_time_v120_app.perf_counter() - t0) * 1000.0, 3)
            timing["detail_calc_deferred_v155"] = True
            st.session_state.api_last_timing_v120 = timing
        return
    _v121_render_table(
        "최종 계산표",
        _v121_final_calc(summary),
        1000,
        "예상 치명 확률은 백어택 기준/시너지/조건부를 포함한 최종 계산표 기준입니다.",
        full_height=True,
    )
    _v121_render_table("전투/캐릭터 요약", _v121_combat_overview(summary), 80)
    _v121_render_table("전투 스킬 - 채용 스킬만", _v121_adopted_skills(summary), 60)

    timing = st.session_state.get("api_last_timing_v120") or {}
    if isinstance(timing, dict):
        timing["api_tab_render_ms_v153"] = round((_time_v120_app.perf_counter() - t0) * 1000.0, 3)
        st.session_state.api_last_timing_v120 = timing


_api_tab_restore_values_v156 = api_tab
def api_tab() -> None:  # type: ignore[override]
    if _public_mode_enabled_v153():
        ensure_api_summary_current()
    _api_tab_restore_values_v156()


_render_step_guide_private_v153 = globals().get("_render_step_guide")
def _render_step_guide(step: int, title: str, items: list[str]) -> None:  # type: ignore[override]
    if step == 2 and _public_mode_enabled_v153():
        st.markdown(
            """
            <div class="loa-preset-card">
              <h3>전투분석기 사전 세팅</h3>
              <p>전투분석기에서 <b>공격 정보</b> 탭을 열고 <b>순서 변경</b>을 눌러 아래 9개 항목이 상위 목록에 보이도록 설정해주세요.</p>
              <ul>
                <li>치명타 적중률, 치명타 비중</li>
                <li>백어택 적중률, 백어택 비중</li>
                <li>사용 횟수, 쿨타임 비율</li>
                <li>헤드어택 적중률, 헤드어택 비중</li>
                <li>피해량 지분</li>
              </ul>
              <p>순서는 바뀌어도 괜찮지만, 캡처 화면의 상위 항목 안에 필요한 값이 모두 보여야 인식이 안정적입니다.</p>
              <p>공격 정보가 한 페이지에 다 보이지 않으면 <b>2페이지까지 캡처</b>해주세요. 순서 변경한 리스트에 있는 스킬들이 실전/허수 공격 정보에 모두 등록되어야 합니다.</p>
            </div>
            """,
            unsafe_allow_html=True,
        )
        preset_path = Path(__file__).parent / "data" / "presetting.jpg"
        if preset_path.exists():
            with st.expander("사전 세팅 예시 이미지 보기", expanded=False):
                st.image(str(preset_path), use_container_width=True)
    if callable(_render_step_guide_private_v153):
        _render_step_guide_private_v153(step, title, items)


_result_tab_private_v153 = globals().get("result_tab")
class _HiddenJsonBlockV153:
    def __enter__(self):
        st.session_state["_hide_result_json_block_v153"] = True
        return self
    def __exit__(self, exc_type, exc, tb):
        st.session_state["_hide_result_json_block_v153"] = False
        return False


def result_tab() -> None:  # type: ignore[override]
    if not (_public_mode_enabled_v153() and callable(_result_tab_private_v153)):
        if callable(_result_tab_private_v153):
            _result_tab_private_v153()
        return
    orig_expander = st.expander
    orig_json = st.json
    orig_download = st.download_button

    def public_expander(label, *args, **kwargs):
        if "JSON" in str(label).upper():
            return _HiddenJsonBlockV153()
        return orig_expander(label, *args, **kwargs)

    def public_json(*args, **kwargs):
        if st.session_state.get("_hide_result_json_block_v153"):
            return None
        return orig_json(*args, **kwargs)

    def public_download_button(*args, **kwargs):
        if st.session_state.get("_hide_result_json_block_v153"):
            return False
        return orig_download(*args, **kwargs)

    ensure_api_summary_current()

    try:
        st.expander = public_expander  # type: ignore[assignment]
        st.json = public_json  # type: ignore[assignment]
        st.download_button = public_download_button  # type: ignore[assignment]
        _result_tab_private_v153()
    finally:
        st.expander = orig_expander  # type: ignore[assignment]
        st.json = orig_json  # type: ignore[assignment]
        st.download_button = orig_download  # type: ignore[assignment]
        st.session_state["_hide_result_json_block_v153"] = False


def _skill_row_by_keyword_v160(df: pd.DataFrame, keyword: str) -> dict[str, Any]:
    try:
        norm = normalize_battle_df(df)
    except Exception:
        norm = pd.DataFrame()
    if not isinstance(norm, pd.DataFrame) or norm.empty:
        return {}
    for _, row in norm.iterrows():
        if keyword in str(row.get("name") or ""):
            return row.to_dict()
    return {}


def _safe_float_v160(value: Any, default: float = 0.0) -> float:
    try:
        if value is None or pd.isna(value):
            return float(default)
    except Exception:
        pass
    try:
        return float(value)
    except Exception:
        return float(default)


_PRED231_BRUTAL_RESET_PROB_V161 = 0.75
_PRED231_HURRICANE_SELF_RESET_PROB_V161 = 0.125
_PRED231_STOMP_RESET_PROB_V161 = 0.125
_PRED231_START_BRUTAL_FIXED_COUNT_V161 = 1.0
_PRED231_STOMP_BASE_PERIOD_V161 = 5.0
_PRED231_BRUTAL_MOTION_TIME_V161 = 20.0 / 30.0
_PRED231_STOMP_MOTION_TIME_V161 = 14.0 / 30.0
_PRED231_FLAME_MOTION_TIME_V161 = 1.15


def _pred231_motion_defaults_v162() -> dict[str, float]:
    calc_config = load_yaml(str(CONFIG_DIR / "calculation_presets.yaml"))
    defaults = calc_config.get("defaults", {}) if isinstance(calc_config, dict) else {}
    return {
        "hurricane_motion_time": float(defaults.get("hurricane_motion_time", 1.3)),
        "hurricane_post_motion_cooldown": float(defaults.get("hurricane_post_motion_cooldown", 0.0)),
        "pred231_brutal_motion_time": float(defaults.get("pred231_brutal_motion_time", _PRED231_BRUTAL_MOTION_TIME_V161)),
        "pred231_stomp_motion_time": float(defaults.get("pred231_stomp_motion_time", _PRED231_STOMP_MOTION_TIME_V161)),
        "pred231_flame_motion_time": float(defaults.get("pred231_flame_motion_time", _PRED231_FLAME_MOTION_TIME_V161)),
        "pred231_stomp_base_period": float(defaults.get("pred231_stomp_base_period", _PRED231_STOMP_BASE_PERIOD_V161)),
    }


def _pred231_ensure_motion_defaults_v162() -> dict[str, float]:
    defaults = _pred231_motion_defaults_v162()
    for key, default in defaults.items():
        cur = _safe_float_v160(st.session_state.get(key), None)
        if cur is None or (key != "hurricane_post_motion_cooldown" and cur <= 0):
            st.session_state[key] = float(default)
    return defaults


def _pred231_contains_skill_v161(df: pd.DataFrame | None, keyword: str) -> bool:
    return bool(_skill_row_by_keyword_v160(df, keyword))


def _is_predator_231_context_v161() -> bool:
    real_table = st.session_state.get("real_table")
    dummy_table = st.session_state.get("dummy_table")
    has_231_skills = (
        (_pred231_contains_skill_v161(real_table, "허리케인") or _pred231_contains_skill_v161(dummy_table, "허리케인"))
        and (_pred231_contains_skill_v161(real_table, "브루탈") or _pred231_contains_skill_v161(dummy_table, "브루탈"))
    )
    context_parts: list[str] = []
    for key in ("api_summary_v120", "api_summary", "api_last_summary_v120", "character_summary", "selected_character"):
        try:
            context_parts.append(json.dumps(st.session_state.get(key), ensure_ascii=False, default=str))
        except Exception:
            context_parts.append(str(st.session_state.get(key) or ""))
    context = " ".join(context_parts)
    return has_231_skills and ("포식자" in context or "슬레이어" in context or _pred231_contains_skill_v161(real_table, "스톰프"))


def _pred231_row_stats_v161(df: pd.DataFrame | None, keywords: list[str]) -> dict[str, Any]:
    for keyword in keywords:
        row = _skill_row_by_keyword_v160(df, keyword)
        if row:
            damage = _safe_float_v160(row.get("damage"), 0.0)
            casts = _safe_float_v160(row.get("casts"), 0.0)
            return {
                "name": row.get("name") or keyword,
                "damage": damage,
                "casts": casts,
                "per_cast": damage / casts if casts > 0 else 0.0,
            }
    return {"name": keywords[0] if keywords else "-", "damage": 0.0, "casts": 0.0, "per_cast": 0.0}


def _pred231_table_stats_v161(df: pd.DataFrame | None, elapsed: float) -> dict[str, Any]:
    summary = summarize_battle(df if isinstance(df, pd.DataFrame) else pd.DataFrame(), elapsed)
    brutal = _pred231_row_stats_v161(df, ["브루탈"])
    hurricane = _pred231_row_stats_v161(df, ["허리케인"])
    flame = _pred231_row_stats_v161(df, ["플레임", "스파이럴", "파이널"])
    stomp = _pred231_row_stats_v161(df, ["스톰프"])
    total_damage = _safe_float_v160(summary.get("total_damage"), 0.0)
    tracked_damage = sum(_safe_float_v160(x.get("damage"), 0.0) for x in (brutal, hurricane, flame, stomp))
    other_damage = max(0.0, total_damage - tracked_damage)
    return {
        "elapsed": max(0.0, float(elapsed or 0.0)),
        "total_damage": total_damage,
        "other_damage": other_damage,
        "other_per_sec": other_damage / elapsed if elapsed > 0 else 0.0,
        "brutal": brutal,
        "hurricane": hurricane,
        "flame": flame,
        "stomp": stomp,
    }


def _pred231_display_scale_v161(*values: float) -> float:
    biggest = max([abs(float(v or 0.0)) for v in values] + [0.0])
    return 100_000_000.0 if biggest >= 1_000_000.0 else 1.0


def _pred231_fmt_num_v161(value: float, digits: int = 2) -> str:
    if value is None or not math.isfinite(float(value)):
        return "-"
    return f"{float(value):,.{digits}f}"


def _pred231_fmt_pct_v161(value: float | None) -> str:
    if value is None or not math.isfinite(float(value)):
        return "-"
    return f"{float(value) * 100.0:.2f}%"


def _pred231_binomial_distribution_v161(
    n_value: float,
    probability: float,
    elapsed: float,
    start_brutal_count: float,
    brutal_per_cast: float,
    hurricane_count: float,
    hurricane_per_cast: float,
    flame_count: float,
    flame_per_cast: float,
    stomp_count: float,
    stomp_per_cast: float,
    other_per_sec: float,
    display_scale: float,
    observed_dps: float | None = None,
) -> dict[str, Any]:
    n = max(0, int(round(float(n_value or 0.0))))
    p = min(0.999999, max(0.000001, float(probability or 0.0)))
    rows: list[dict[str, float]] = []
    cumulative = 0.0
    for k in range(n + 1):
        log_prob = (
            math.lgamma(n + 1)
            - math.lgamma(k + 1)
            - math.lgamma(n - k + 1)
            + k * math.log(p)
            + (n - k) * math.log(1.0 - p)
        )
        prob = math.exp(log_prob) if log_prob > -745 else 0.0
        cumulative += prob
        brutal_count = start_brutal_count + k
        total_damage = (
            brutal_count * brutal_per_cast
            + hurricane_count * hurricane_per_cast
            + flame_count * flame_per_cast
            + stomp_count * stomp_per_cast
            + other_per_sec * elapsed
        )
        dps = total_damage / elapsed / display_scale if elapsed > 0 else 0.0
        rows.append({"k": float(k), "prob": prob, "cum": min(1.0, cumulative), "dps": dps})
    mean_dps = sum(r["dps"] * r["prob"] for r in rows)
    nearest_prob = None
    if observed_dps is not None and rows:
        nearest = min(rows, key=lambda r: abs(r["dps"] - float(observed_dps)))
        nearest_prob = nearest["prob"]
    return {"rows": rows, "mean_dps": mean_dps, "nearest_prob": nearest_prob}


def _pred231_quantile_v161(dist: dict[str, Any], target: float) -> float:
    rows = dist.get("rows") or []
    if not rows:
        return 0.0
    selected = rows[0]
    for row in rows:
        if row.get("cum", 0.0) <= target:
            selected = row
        else:
            break
    return float(selected.get("dps") or 0.0)


def _pred231_chart_v161(title: str, dist: dict[str, Any], measured: float, mean_value: float, measured_label: str, mean_label: str) -> None:
    rows = [
        {"DPS": r["dps"], "발생확률": r["prob"] * 100.0, "구분": title}
        for r in (dist.get("rows") or [])
        if r.get("prob", 0.0) >= 0.0001
    ]
    if not rows:
        st.info("그래프를 만들 전투분석기 데이터가 부족합니다.")
        return
    if alt is None:
        chart_df = pd.DataFrame(rows).set_index("DPS")
        st.line_chart(chart_df[["발생확률"]])
        return
    points = pd.DataFrame([
        {"DPS": measured, "발생확률": max((dist.get("nearest_prob") or 0.0) * 100.0, 0.0), "구분": measured_label},
        {"DPS": mean_value, "발생확률": max(max(r["발생확률"] for r in rows), 0.0), "구분": mean_label},
    ])
    base = alt.Chart(pd.DataFrame(rows)).mark_line(size=3).encode(
        x=alt.X("DPS:Q", title="DPS(억/초)"),
        y=alt.Y("발생확률:Q", title="발생확률(%)"),
        color=alt.value("#4f7fd6"),
        tooltip=[alt.Tooltip("DPS:Q", format=".2f"), alt.Tooltip("발생확률:Q", format=".2f")],
    )
    point_layer = alt.Chart(points).mark_circle(size=90).encode(
        x="DPS:Q",
        y="발생확률:Q",
        color=alt.Color("구분:N", scale=alt.Scale(range=["#c00000", "#ed7d31"])),
        tooltip=[alt.Tooltip("구분:N"), alt.Tooltip("DPS:Q", format=".2f"), alt.Tooltip("발생확률:Q", format=".2f")],
    )
    st.altair_chart((base + point_layer).properties(title=title, height=320), use_container_width=True)


def _pred231_compute_v161() -> dict[str, Any]:
    real_meta = st.session_state.get("real_meta", {}) or {}
    dummy_meta = st.session_state.get("dummy_meta", {}) or {}
    real_elapsed = _safe_float_v160(real_meta.get("elapsed_seconds") or st.session_state.get("real_elapsed_sec"), 0.0)
    dummy_elapsed = _safe_float_v160(dummy_meta.get("elapsed_seconds") or st.session_state.get("dummy_elapsed_sec"), 0.0)
    real_stats = _pred231_table_stats_v161(st.session_state.get("real_table"), real_elapsed)
    dummy_stats = _pred231_table_stats_v161(st.session_state.get("dummy_table"), dummy_elapsed)
    display_scale = _pred231_display_scale_v161(real_stats["total_damage"], dummy_stats["total_damage"])

    _pred231_ensure_motion_defaults_v162()

    hurricane_motion_time = float(st.session_state.get("hurricane_motion_time") or 1.3)
    hurricane_post_motion_cooldown = float(st.session_state.get("hurricane_post_motion_cooldown") or 0.0)
    brutal_motion_time = float(st.session_state.get("pred231_brutal_motion_time") or _PRED231_BRUTAL_MOTION_TIME_V161)
    stomp_motion_time = float(st.session_state.get("pred231_stomp_motion_time") or _PRED231_STOMP_MOTION_TIME_V161)
    flame_motion_time = float(st.session_state.get("pred231_flame_motion_time") or _PRED231_FLAME_MOTION_TIME_V161)
    stomp_base_period = float(st.session_state.get("pred231_stomp_base_period") or _PRED231_STOMP_BASE_PERIOD_V161)
    hurricane_effective_time = hurricane_motion_time + hurricane_post_motion_cooldown
    self_reset = _PRED231_HURRICANE_SELF_RESET_PROB_V161
    brutal_prob = _PRED231_BRUTAL_RESET_PROB_V161
    stomp_prob = _PRED231_STOMP_RESET_PROB_V161
    chain_hurricane_expected = 1.0 / (1.0 - self_reset)
    chain_brutal_expected = brutal_prob / (1.0 - self_reset)
    chain_stomp_expected = stomp_prob / (1.0 - self_reset)
    chain_expected_time = (
        chain_hurricane_expected * hurricane_effective_time
        + chain_brutal_expected * brutal_motion_time
        + chain_stomp_expected * stomp_motion_time
    )
    base_stomp_count = math.floor(real_elapsed / stomp_base_period) if real_elapsed > 0 and stomp_base_period > 0 else 0.0
    flame_count = real_stats["flame"]["casts"]
    fixed_time = (
        _PRED231_START_BRUTAL_FIXED_COUNT_V161 * brutal_motion_time
        + flame_count * flame_motion_time
        + base_stomp_count * stomp_motion_time
    )
    chain_start_count = max(0.0, (real_elapsed - fixed_time) / chain_expected_time) if chain_expected_time > 0 else 0.0
    full_counts = {
        "brutal": _PRED231_START_BRUTAL_FIXED_COUNT_V161 + chain_start_count * chain_brutal_expected,
        "hurricane": chain_start_count * chain_hurricane_expected,
        "flame": flame_count,
        "stomp": base_stomp_count + chain_start_count * chain_stomp_expected,
        "other_seconds": real_elapsed,
    }
    real_avg_counts = {
        "brutal": _PRED231_START_BRUTAL_FIXED_COUNT_V161 + real_stats["hurricane"]["casts"] * brutal_prob,
        "hurricane": real_stats["hurricane"]["casts"],
        "flame": real_stats["flame"]["casts"],
        "stomp": real_stats["stomp"]["casts"],
        "other_seconds": real_elapsed,
    }

    def damage_from_counts(per_source: dict[str, Any], counts: dict[str, float]) -> float:
        return (
            counts["brutal"] * per_source["brutal"]["per_cast"]
            + counts["hurricane"] * per_source["hurricane"]["per_cast"]
            + counts["flame"] * per_source["flame"]["per_cast"]
            + counts["stomp"] * per_source["stomp"]["per_cast"]
            + counts["other_seconds"] * per_source["other_per_sec"]
        )

    full_dummy_damage = damage_from_counts(dummy_stats, full_counts)
    real_avg_dummy_damage = damage_from_counts(dummy_stats, real_avg_counts)
    real_original_mean_damage = damage_from_counts(real_stats, real_avg_counts)
    raid_measured_dps = real_stats["total_damage"] / real_elapsed / display_scale if real_elapsed > 0 else 0.0
    dummy_measured_dps = dummy_stats["total_damage"] / dummy_elapsed / display_scale if dummy_elapsed > 0 else 0.0

    dummy_dist = _pred231_binomial_distribution_v161(
        dummy_stats["hurricane"]["casts"], brutal_prob, dummy_elapsed,
        _PRED231_START_BRUTAL_FIXED_COUNT_V161,
        dummy_stats["brutal"]["per_cast"], dummy_stats["hurricane"]["casts"], dummy_stats["hurricane"]["per_cast"],
        dummy_stats["flame"]["casts"], dummy_stats["flame"]["per_cast"],
        dummy_stats["stomp"]["casts"], dummy_stats["stomp"]["per_cast"], dummy_stats["other_per_sec"],
        display_scale, dummy_measured_dps,
    )
    full_dist = _pred231_binomial_distribution_v161(
        full_counts["hurricane"], brutal_prob, real_elapsed,
        _PRED231_START_BRUTAL_FIXED_COUNT_V161,
        dummy_stats["brutal"]["per_cast"], full_counts["hurricane"], dummy_stats["hurricane"]["per_cast"],
        full_counts["flame"], dummy_stats["flame"]["per_cast"],
        full_counts["stomp"], dummy_stats["stomp"]["per_cast"], dummy_stats["other_per_sec"],
        display_scale, full_dummy_damage / real_elapsed / display_scale if real_elapsed > 0 else None,
    )
    real_dist = _pred231_binomial_distribution_v161(
        real_stats["hurricane"]["casts"], brutal_prob, real_elapsed,
        _PRED231_START_BRUTAL_FIXED_COUNT_V161,
        real_stats["brutal"]["per_cast"], real_stats["hurricane"]["casts"], real_stats["hurricane"]["per_cast"],
        real_stats["flame"]["casts"], real_stats["flame"]["per_cast"],
        real_stats["stomp"]["casts"], real_stats["stomp"]["per_cast"], real_stats["other_per_sec"],
        display_scale, raid_measured_dps,
    )
    return {
        "real_elapsed": real_elapsed,
        "dummy_elapsed": dummy_elapsed,
        "real_stats": real_stats,
        "dummy_stats": dummy_stats,
        "display_scale": display_scale,
        "hurricane_motion_time": hurricane_motion_time,
        "hurricane_post_motion_cooldown": hurricane_post_motion_cooldown,
        "hurricane_effective_time": hurricane_effective_time,
        "brutal_motion_time": brutal_motion_time,
        "stomp_motion_time": stomp_motion_time,
        "flame_motion_time": flame_motion_time,
        "stomp_base_period": stomp_base_period,
        "chain_hurricane_expected": chain_hurricane_expected,
        "chain_brutal_expected": chain_brutal_expected,
        "chain_stomp_expected": chain_stomp_expected,
        "chain_expected_time": chain_expected_time,
        "base_stomp_count": base_stomp_count,
        "flame_count": flame_count,
        "chain_start_count": chain_start_count,
        "full_counts": full_counts,
        "real_avg_counts": real_avg_counts,
        "full_dummy_damage": full_dummy_damage,
        "real_avg_dummy_damage": real_avg_dummy_damage,
        "real_original_mean_damage": real_original_mean_damage,
        "raid_measured_dps": raid_measured_dps,
        "dummy_measured_dps": dummy_measured_dps,
        "dummy_dist": dummy_dist,
        "full_dist": full_dist,
        "real_dist": real_dist,
    }


def _render_predator_231_result_v161() -> None:
    st.subheader("231 포식자 전용 결과")
    st.caption("브루탈 - 허리케인 - 플레임 블레이드 시작 사이클 기준으로, 엑셀의 이항분포 계산 구조를 앱 안에서 자동 계산합니다.")
    _pred231_ensure_motion_defaults_v162()
    c = st.columns([1, 1, 1])
    with c[0]:
        st.number_input(
            "허리케인 모션시간(초)",
            min_value=0.0,
            max_value=10.0,
            step=0.1,
            key="hurricane_motion_time",
            help="허리케인 소드가 실제로 시전되는 시간입니다.",
        )
    with c[1]:
        st.number_input(
            "허리케인 모션 후 남은 쿨타임(초)",
            min_value=0.0,
            max_value=30.0,
            step=0.1,
            key="hurricane_post_motion_cooldown",
            help="허리케인 모션이 끝난 뒤, 다음 허리케인을 다시 쓰기까지 기다려야 하는 시간입니다.",
        )
    data = _pred231_compute_v161()
    with c[2]:
        st.metric("허리케인 실제 1회 소요시간(초)", f"{data['hurricane_effective_time']:.2f}")
    st.markdown("**세부 모션 설정**")
    d1, d2, d3, d4 = st.columns(4)
    with d1:
        st.number_input(
            "브루탈 모션시간(초)",
            min_value=0.01,
            max_value=10.0,
            step=0.01,
            format="%.3f",
            key="pred231_brutal_motion_time",
            help="브루탈 임팩트가 실제로 시전되는 시간입니다. 기본값은 20/30초입니다.",
        )
    with d2:
        st.number_input(
            "스톰프 모션시간(초)",
            min_value=0.01,
            max_value=10.0,
            step=0.01,
            format="%.3f",
            key="pred231_stomp_motion_time",
            help="와일드 스톰프가 실제로 시전되는 시간입니다. 기본값은 14/30초입니다.",
        )
    with d3:
        st.number_input(
            "플레임 모션시간(초)",
            min_value=0.01,
            max_value=10.0,
            step=0.01,
            format="%.3f",
            key="pred231_flame_motion_time",
            help="플레임 블레이드가 실제로 시전되는 시간입니다.",
        )
    with d4:
        st.number_input(
            "스톰프 사용 주기(초)",
            min_value=0.1,
            max_value=60.0,
            step=0.1,
            format="%.1f",
            key="pred231_stomp_base_period",
            help="스톰프를 몇 초에 한 번씩 기본으로 쓰는지 입력합니다. 예: 5초마다 쓰면 5.0",
        )
    data = _pred231_compute_v161()
    st.info(
        "허리케인 모션 후 남은 쿨타임이 길수록 허수처럼 풀가동 가능한 허리케인 횟수가 줄어듭니다. "
        "그래서 같은 실전 시간이라도 허리케인 쿨타임이 긴 세팅은 이론상 풀가동 평균 DPS가 낮아질 수 있습니다."
    )
    st.caption(
        f"예: 허리케인 모션시간 {data['hurricane_motion_time']:.1f}초 + "
        f"모션 후 남은 쿨타임 {data['hurricane_post_motion_cooldown']:.1f}초 = "
        f"허리케인 실제 1회 소요시간 {data['hurricane_effective_time']:.1f}초"
    )
    if data["real_elapsed"] <= 0 or data["dummy_elapsed"] <= 0:
        st.warning("실전/허수아비 전투시간이 있어야 231 운용률과 분포를 계산할 수 있습니다.")
        return

    operation_rate = data["real_avg_dummy_damage"] / data["full_dummy_damage"] if data["full_dummy_damage"] > 0 else 0.0
    real_avg_dummy_dps = data["real_avg_dummy_damage"] / data["real_elapsed"] / data["display_scale"] if data["real_elapsed"] > 0 else 0.0
    full_dummy_dps = data["full_dummy_damage"] / data["real_elapsed"] / data["display_scale"] if data["real_elapsed"] > 0 else 0.0
    raid_mean_dps = data["real_dist"]["mean_dps"]
    measured_ratio = data["raid_measured_dps"] / raid_mean_dps if raid_mean_dps > 0 else 0.0
    measured_gap = data["raid_measured_dps"] - raid_mean_dps

    st.markdown("### 1) 핵심 결과")
    key_df = pd.DataFrame([
        {"항목": "허수처럼 풀가동 평균 피해량", "값": _pred231_fmt_num_v161(data["full_dummy_damage"] / data["display_scale"]), "단위": "억", "설명": "모션시간으로 산출한 허수 풀가동 기준", "비고": "100% 기준"},
        {"항목": "실전 평균 허수환산 피해량", "값": _pred231_fmt_num_v161(data["real_avg_dummy_damage"] / data["display_scale"]), "단위": "억", "설명": "허수 1회딜 × 실전 평균화 횟수", "비고": "운 평균화"},
        {"항목": "허수대비 실전 운용률", "값": _pred231_fmt_pct_v161(operation_rate), "단위": "%", "설명": "실전 평균 허수환산 피해량 ÷ 허수처럼 풀가동 평균 피해량", "비고": "핵심값"},
        {"항목": "허수처럼 풀가동 평균 DPS", "값": _pred231_fmt_num_v161(full_dummy_dps), "단위": "억/초", "설명": "허수처럼 풀가동 평균 피해량 ÷ 실전시간", "비고": ""},
        {"항목": "실전 원본 실측 DPS", "값": _pred231_fmt_num_v161(data["raid_measured_dps"]), "단위": "억/초", "설명": "실제 전분 총피해량 ÷ 실전시간", "비고": "그래프 점"},
        {"항목": "실전 원본 이항평균 DPS", "값": _pred231_fmt_num_v161(raid_mean_dps), "단위": "억/초", "설명": "실전 원본 1회딜 기준 평균 DPS", "비고": "그래프 평균"},
        {"항목": "실전 평균 허수환산 DPS", "값": _pred231_fmt_num_v161(real_avg_dummy_dps), "단위": "억/초", "설명": "실전 평균 허수환산 피해량 ÷ 실전시간", "비고": ""},
        {"항목": "실전 실측/평균 DPS 비율", "값": _pred231_fmt_pct_v161(measured_ratio), "단위": "%", "설명": "실측 DPS ÷ 실전 이항평균 DPS", "비고": "운 판정"},
        {"항목": "실전 실측 평균 DPS 대비", "값": _pred231_fmt_pct_v161(measured_ratio - 1.0), "단위": "%", "설명": "실측 DPS가 실전 이항평균 DPS보다 높은/낮은 정도", "비고": "운 판정"},
        {"항목": "실전 실측-평균 DPS 차이", "값": _pred231_fmt_num_v161(measured_gap), "단위": "억/초", "설명": "실측 DPS - 실전 이항평균 DPS", "비고": "운 판정"},
        {"항목": "실전 실측 DPS 발생확률", "값": _pred231_fmt_pct_v161(data["real_dist"].get("nearest_prob")), "단위": "%", "설명": "실측 DPS와 가장 가까운 이항분포 지점의 단일 발생확률", "비고": "발생확률"},
    ])
    st.dataframe(key_df, use_container_width=True, hide_index=True)

    left, right = st.columns(2)
    with left:
        _pred231_chart_v161(
            "실제 레이드 DPS 이항분포",
            data["real_dist"],
            data["raid_measured_dps"],
            raid_mean_dps,
            "실전 실측 DPS",
            "실전 이항평균 DPS",
        )
    with right:
        _pred231_chart_v161(
            "허수아비 원본 DPS 이항분포",
            data["dummy_dist"],
            data["dummy_measured_dps"],
            data["dummy_dist"]["mean_dps"],
            "허수 실측 DPS",
            "허수 이항평균 DPS",
        )

    st.markdown("### 2) 분포 구간 요약")
    quantile_rows = []
    for label, q in [("극저점 기준", 0.1), ("저점 기준", 0.2), ("평균 기준", 0.5), ("고점 기준", 0.7), ("극고점 기준", 0.9)]:
        quantile_rows.append({
            "구간": label,
            "허수아비 원본 DPS": _pred231_fmt_num_v161(_pred231_quantile_v161(data["dummy_dist"], q)),
            "허수처럼 풀가동 DPS": _pred231_fmt_num_v161(_pred231_quantile_v161(data["full_dist"], q)),
            "실전 원본 DPS": _pred231_fmt_num_v161(_pred231_quantile_v161(data["real_dist"], q)),
            "설명": "이항분포 누적확률 기준",
            "누적확률": _pred231_fmt_pct_v161(q),
        })
    st.dataframe(pd.DataFrame(quantile_rows), use_container_width=True, hide_index=True)

    st.markdown("### 3) 허수아비 이항분포 요약값")
    dummy_ratio = data["dummy_measured_dps"] / data["dummy_dist"]["mean_dps"] if data["dummy_dist"]["mean_dps"] > 0 else 0.0
    dummy_summary_df = pd.DataFrame([
        {"항목": "허수 실측 DPS", "값": _pred231_fmt_num_v161(data["dummy_measured_dps"]), "단위": "억/초", "설명": "허수 전분 원본 DPS"},
        {"항목": "허수 이항평균 DPS", "값": _pred231_fmt_num_v161(data["dummy_dist"]["mean_dps"]), "단위": "억/초", "설명": "허수 전분의 허리케인 횟수 기준 평균 쿨초 DPS"},
        {"항목": "허수 실측/평균 DPS 비율", "값": _pred231_fmt_pct_v161(dummy_ratio), "단위": "%", "설명": "허수 실측이 허수 평균보다 높은/낮은 정도"},
        {"항목": "허수 실측 DPS 발생확률", "값": _pred231_fmt_pct_v161(data["dummy_dist"].get("nearest_prob")), "단위": "%", "설명": "허수 실측 DPS와 가장 가까운 이항분포 지점의 단일 발생확률"},
    ])
    st.dataframe(dummy_summary_df, use_container_width=True, hide_index=True)

    with st.expander("자동 계산 상세", expanded=False):
        detail_df = pd.DataFrame([
            {"항목": "허리케인 모션시간", "값": _pred231_fmt_num_v161(data["hurricane_motion_time"], 3), "단위": "초", "설명": "허리케인 소드가 실제로 시전되는 시간", "비고": "사용자 설정"},
            {"항목": "허리케인 모션 후 남은 쿨타임", "값": _pred231_fmt_num_v161(data["hurricane_post_motion_cooldown"], 3), "단위": "초", "설명": "모션이 끝난 뒤 다음 허리케인을 쓰기까지 기다리는 시간", "비고": "사용자 설정"},
            {"항목": "브루탈 모션시간", "값": _pred231_fmt_num_v161(data["brutal_motion_time"], 3), "단위": "초", "설명": "브루탈 임팩트가 실제로 시전되는 시간", "비고": "사용자 설정"},
            {"항목": "스톰프 모션시간", "값": _pred231_fmt_num_v161(data["stomp_motion_time"], 3), "단위": "초", "설명": "와일드 스톰프가 실제로 시전되는 시간", "비고": "사용자 설정"},
            {"항목": "플레임 모션시간", "값": _pred231_fmt_num_v161(data["flame_motion_time"], 3), "단위": "초", "설명": "플레임 블레이드가 실제로 시전되는 시간", "비고": "사용자 설정"},
            {"항목": "스톰프 사용 주기", "값": _pred231_fmt_num_v161(data["stomp_base_period"], 1), "단위": "초", "설명": "스톰프를 몇 초에 한 번씩 기본으로 쓰는지", "비고": "사용자 설정"},
            {"항목": "체인당 허리케인 기대횟수", "값": _pred231_fmt_num_v161(data["chain_hurricane_expected"]), "단위": "회", "설명": "1/(1-허리케인 자기초기화 확률)", "비고": ""},
            {"항목": "체인당 브루탈 기대횟수", "값": _pred231_fmt_num_v161(data["chain_brutal_expected"]), "단위": "회", "설명": "브루탈 확률/(1-허리케인 자기초기화 확률)", "비고": ""},
            {"항목": "체인당 스톰프 추가 기대횟수", "값": _pred231_fmt_num_v161(data["chain_stomp_expected"]), "단위": "회", "설명": "스톰프 확률/(1-허리케인 자기초기화 확률)", "비고": ""},
            {"항목": "체인당 기대 소요시간", "값": _pred231_fmt_num_v161(data["chain_expected_time"]), "단위": "초", "설명": "허리케인 실제 1회 소요시간까지 포함", "비고": ""},
            {"항목": "실전 기준 기본 스톰프 횟수", "값": _pred231_fmt_num_v161(data["base_stomp_count"]), "단위": "회", "설명": "FLOOR(실전시간/스톰프 기본 주기)", "비고": "자동"},
            {"항목": "허수처럼 풀가동 플레임 횟수", "값": _pred231_fmt_num_v161(data["flame_count"]), "단위": "회", "설명": "실전 전투에서 사용된 플레임 횟수", "비고": "고정값"},
            {"항목": "허수처럼 풀가동 체인 시작횟수", "값": _pred231_fmt_num_v161(data["chain_start_count"]), "단위": "회", "설명": "남은 모션시간 ÷ 체인당 기대 소요시간", "비고": ""},
        ])
        st.dataframe(detail_df, use_container_width=True, hide_index=True)


_result_tab_prev_v160 = result_tab
def result_tab() -> None:  # type: ignore[override]
    if _is_predator_231_context_v161():
        _render_predator_231_result_v161()
        return
    _result_tab_prev_v160()


if __name__ == "__main__":
    main()
